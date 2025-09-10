# utils/helpers.py
"""
utils/helpers.py
Comprehensive helper functions for the DataViewer Application.
Contains utility functions from the existing utils.py and resource_utils.py.
Developed by Charlie Becquet.
"""

import numpy as np
import os
import re
import sys
import tempfile
import traceback
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label, Button, ttk
from pathlib import Path
from typing import Optional, Any, Union, Tuple, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

print("DEBUG: helpers.py - Starting import of utility functions")

# Global debug flag - change this to control ALL debug output across the app
DEBUG_ENABLED = True  # Set to True when debugging is needed

def debug_print(*args, **kwargs):
    """
    Global debug print function that can be imported by all modules.
    Only prints if DEBUG_ENABLED is True.
    
    Usage:
        from utils import debug_print
        debug_print("This is a debug message")
        debug_print(f"Variable value: {variable}")
    """
    if DEBUG_ENABLED:
        print(*args, **kwargs)

def set_debug_mode(enabled: bool):
    """
    Enable or disable debug mode globally.
    
    Args:
        enabled (bool): True to enable debug output, False to disable
    """
    global DEBUG_ENABLED
    DEBUG_ENABLED = enabled
    debug_print(f"Debug mode {'enabled' if enabled else 'disabled'}")

def is_debug_enabled() -> bool:
    """
    Check if debug mode is currently enabled.
    
    Returns:
        bool: True if debug is enabled, False otherwise
    """
    return DEBUG_ENABLED

print("DEBUG: helpers.py - Debug functions loaded")

# ============================================================================
# RESOURCE PATH FUNCTIONS (from resource_utils.py)
# ============================================================================

def get_resource_path(relative_path: str) -> str:
    """
    Get absolute path to resource, works for development and PyInstaller.
    Enhanced version combining functionality from utils.py and resource_utils.py.
    
    Args:
        relative_path (str): Relative path to the resource.
    
    Returns:
        str: Absolute path to the resource.
    """
    try:
        # Check if running as PyInstaller executable
        if hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
            debug_print(f"Using PyInstaller resource path: {base_path}")
        else:
            # Development mode - use current directory
            base_path = os.path.abspath(".")
            debug_print(f"Using development resource path: {base_path}")

        full_path = os.path.join(base_path, relative_path)
        # Normalize path separators for cross-platform compatibility
        full_path = os.path.normpath(full_path)
        
        debug_print(f"Resource path for {relative_path}: {full_path}")
        debug_print(f"Resource exists: {os.path.exists(full_path)}")
        
        return full_path
        
    except Exception as e:
        debug_print(f"ERROR: get_resource_path failed for {relative_path}: {e}")
        return relative_path

def get_resource_dir() -> str:
    """Get the base resource directory."""
    if hasattr(sys, '_MEIPASS'):
        return sys._MEIPASS
    else:
        return os.path.abspath(".")

def resource_exists(relative_path: str) -> bool:
    """Check if a resource exists."""
    return os.path.exists(get_resource_path(relative_path))

def resource_path(relative_path: str) -> str:
    """
    Legacy function name for compatibility.
    Get the absolute path to the resource, compatible with PyInstaller.
    """
    return get_resource_path(relative_path)

print("DEBUG: helpers.py - Resource path functions loaded")

# ============================================================================
# CORE UTILITY FUNCTIONS
# ============================================================================

def round_values(value, decimals=2):
    """Round the values to a specified number of decimal places."""
    try:
        return round(float(value), decimals)
    except (ValueError, TypeError):
        debug_print(f"Could not round value {value}, returning original")
        return value

def generate_temp_image(figure):
    """Generate a temporary image file from a matplotlib figure."""
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
        figure.savefig(tmpfile.name)
        debug_print(f"Generated temp image: {tmpfile.name}")
        return tmpfile.name

def get_save_path_dialog(default_extension: str = ".xlsx") -> Optional[str]:
    """
    Prompt the user to select a file save location and return the path.

    Args:
        default_extension (str): Default file extension for the save dialog.

    Returns:
        Optional[str]: The selected file path or None if canceled.
    """
    file_path = filedialog.asksaveasfilename(
        defaultextension=default_extension,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    debug_print(f"Save path selected: {file_path}")
    return file_path

print("DEBUG: helpers.py - Core utility functions loaded")

# ============================================================================
# DATAFRAME AND DATA PROCESSING FUNCTIONS
# ============================================================================

def clean_columns(data: pd.DataFrame) -> pd.DataFrame:
    """
    Clean column names in a DataFrame.
    Enhanced version that handles all edge cases.
    """
    if data.empty:
        debug_print("clean_columns - DataFrame is empty")
        return data
    
    original_columns = list(data.columns)
    debug_print(f"Original columns: {original_columns}")
    
    # Step 1: Convert all column names to strings and strip whitespace
    cleaned_data = data.copy()
    cleaned_data.columns = [str(col).strip().replace('\n', ' ').replace('\r', ' ') for col in cleaned_data.columns]
    
    # Step 2: Handle duplicate column names by appending numbers
    columns = list(cleaned_data.columns)
    seen = {}
    new_columns = []
    
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}.{seen[col]}")
        else:
            seen[col] = 0
            new_columns.append(col)
    
    cleaned_data.columns = new_columns
    
    # Step 3: Remove unnamed columns (optional - keep for compatibility)
    # cleaned_data = cleaned_data.loc[:, ~cleaned_data.columns.str.contains('^Unnamed')]
    
    # Step 4: Clean up generated column names that are just dots and numbers
    final_columns = []
    for col in cleaned_data.columns:
        cleaned_col = str(col)
        # If column name is just dots and numbers (like '.1', '.2'), convert to empty string
        if cleaned_col.startswith('.') and cleaned_col[1:].isdigit():
            cleaned_col = ''
            debug_print(f"Cleaned column name: '{col}' -> '{cleaned_col}'")
        final_columns.append(cleaned_col)
    
    cleaned_data.columns = final_columns
    
    # Step 5: Replace NaN values with empty strings
    cleaned_data = cleaned_data.fillna('')
    
    debug_print(f"clean_columns - {len(original_columns)} -> {len(cleaned_data.columns)} columns")
    debug_print(f"Final columns: {list(cleaned_data.columns)}")
    return cleaned_data

def remove_empty_columns(data: pd.DataFrame) -> pd.DataFrame:
    """
    Remove empty columns from a DataFrame. An empty column is defined as one where all values are NaN or 0.
    This process continues until a non-empty column is encountered.

    Args:
        data (pd.DataFrame): The input DataFrame.

    Returns:
        pd.DataFrame: The DataFrame with empty columns removed.
    """
    if data.empty:
        debug_print("remove_empty_columns - DataFrame is empty")
        return data
    
    original_shape = data.shape
    
    # Remove columns that are completely empty (all NaN or all zeros)
    mask = ~((data.isna() | (data == 0)).all())
    cleaned_data = data.loc[:, mask]
    
    debug_print(f"remove_empty_columns - {original_shape} -> {cleaned_data.shape}")
    return cleaned_data

print("DEBUG: helpers.py - DataFrame processing functions loaded")

# ============================================================================
# PATH AND FILE UTILITY FUNCTIONS
# ============================================================================

def get_save_path(file_path: str, suffix: str = "_processed") -> str:
    """Generate a save path with suffix."""
    path = Path(file_path)
    stem = path.stem
    extension = path.suffix
    parent = path.parent
    
    save_path = parent / f"{stem}{suffix}{extension}"
    
    debug_print(f"Generated save path: {file_path} -> {save_path}")
    return str(save_path)

def is_valid_excel_file(filename: str) -> bool:
    """
    Checks if the given filename is a valid Excel file that should be processed.
    Excludes temporary Excel files which often start with '~$'.
    
    Args:
        filename (str): The name of the file.
    
    Returns:
        bool: True if valid, False otherwise.
    """
    is_valid = filename.endswith('.xlsx') and not filename.startswith('~$')
    debug_print(f"is_valid_excel_file({filename}): {is_valid}")
    return is_valid

def ensure_directory_exists(directory_path: str) -> bool:
    """Ensure a directory exists, create if necessary."""
    try:
        Path(directory_path).mkdir(parents=True, exist_ok=True)
        debug_print(f"ensure_directory_exists - {directory_path} OK")
        return True
    except Exception as e:
        debug_print(f"ERROR: ensure_directory_exists failed for {directory_path}: {e}")
        return False

def get_file_extension(file_path: str) -> str:
    """Get file extension in lowercase."""
    ext = Path(file_path).suffix.lower()
    debug_print(f"get_file_extension({file_path}): {ext}")
    return ext

def is_file_accessible(file_path: str) -> bool:
    """Check if a file is accessible for reading."""
    try:
        path = Path(file_path)
        if not path.exists():
            debug_print(f"is_file_accessible({file_path}): File does not exist")
            return False
        
        if not path.is_file():
            debug_print(f"is_file_accessible({file_path}): Not a file")
            return False
        
        # Try to open file
        with open(file_path, 'rb') as f:
            f.read(1)  # Read just one byte
        
        debug_print(f"is_file_accessible({file_path}): ACCESSIBLE")
        return True
        
    except Exception as e:
        debug_print(f"is_file_accessible({file_path}): ERROR - {e}")
        return False

print("DEBUG: helpers.py - File utility functions loaded")

# ============================================================================
# EXCEL FILE PROCESSING FUNCTIONS
# ============================================================================

def load_excel_file(file_path: str) -> dict:
    """
    Load an Excel file and return its sheets.
    Enhanced with better error handling and debugging.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: Dictionary of sheet names and DataFrames.
    """
    try:
        debug_print(f"Loading Excel file: {file_path}")
        sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        debug_print(f"Successfully loaded {len(sheets)} sheets: {list(sheets.keys())}")
        return sheets
    except Exception as e:
        debug_print(f"ERROR: Failed to load Excel file {file_path}: {e}")
        raise ValueError(f"Error loading Excel file {file_path}: {e}")

def load_excel_file_with_formulas(file_path: str) -> dict:
    """
    Load Excel file and evaluate formulas to get calculated values.
    Uses openpyxl with data_only=True to evaluate formulas.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: Dictionary of sheet names and DataFrames.
    """
    try:
        debug_print(f"Loading Excel file with formula evaluation: {file_path}")
        
        # First, try to force Excel to recalculate by opening and saving the file
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # Force calculation of all formulas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # This is a formula - we need to force evaluation
                        debug_print(f"Found formula in {cell.coordinate}: {cell.value}")
        
        wb.close()
        
        # Now load with data_only=True to get calculated values
        wb_calc = openpyxl.load_workbook(file_path, data_only=True)
        sheets = {}
        
        for sheet_name in wb_calc.sheetnames:
            debug_print(f"Processing sheet: {sheet_name}")
            ws = wb_calc[sheet_name]
            
            # Convert worksheet to DataFrame row by row
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if data:
                # Create DataFrame with proper column handling
                df = pd.DataFrame(data)
                
                # Set the first row as headers if it contains strings
                if len(df) > 0:
                    # Use first row as headers
                    df.columns = [f"Unnamed: {i}" if pd.isna(col) else str(col) for i, col in enumerate(df.iloc[0])]
                    df = df[1:].reset_index(drop=True)
                
                sheets[sheet_name] = df
                debug_print(f"Sheet {sheet_name} loaded with shape: {df.shape}")
                
                # Debug: Check specific cells that should contain usage efficiency
                if len(df) > 1 and len(df.columns) > 8:
                    debug_print(f"Sample usage efficiency values in row 1:")
                    for col_idx in [8, 20, 32, 44, 56, 68]:  # Expected positions for 6 samples
                        if col_idx < len(df.columns):
                            val = df.iloc[1, col_idx] if len(df) > 1 else None
                            debug_print(f"Column {col_idx}: '{val}'")
            else:
                sheets[sheet_name] = pd.DataFrame()
                debug_print(f"Sheet {sheet_name} is empty")
        
        wb_calc.close()
        debug_print(f"Successfully loaded {len(sheets)} sheets with enhanced formula evaluation")
        return sheets
        
    except Exception as e:
        debug_print(f"ERROR: Failed to load Excel file with formula evaluation: {e}")
        # Fallback to original method
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            debug_print("Fallback to pandas read_excel successful")
            return sheets
        except Exception as fallback_error:
            raise ValueError(f"Error loading Excel file {file_path}: {fallback_error}")

def read_sheet_with_values_standards(file_path: str, sheet_name: Optional[str] = None):
    """Read Excel sheet with values exactly as they appear, evaluating formulas."""
    debug_print(f"Reading sheet with formula evaluation: {sheet_name}")
    
    # Use openpyxl with data_only=True to evaluate formulas
    wb = load_workbook(file_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Convert to DataFrame
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    if not data:
        return pd.DataFrame()
    
    # Create DataFrame and set first row as header
    df = pd.DataFrame(data)
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)
    
    wb.close()
    debug_print(f"Successfully read sheet {sheet_name} with formula evaluation")
    return df

def unmerge_all_cells(ws: Worksheet) -> None:
    """
    Unmerge all merged cells in the given worksheet.
    
    For each merged range, this function:
      - Retrieves the value from the top-left (master) cell.
      - Unmerges the range.
      - Sets every cell in the range to the retrieved value.
    
    Args:
        ws (Worksheet): The Openpyxl worksheet to process.
    """
    debug_print(f"Unmerging cells in worksheet: {ws.title}")
    
    # Create a list of merged ranges to avoid modifying the collection while iterating.
    merged_ranges = list(ws.merged_cells.ranges)
    
    debug_print(f"Found {len(merged_ranges)} merged ranges to process")
    
    for merged_range in merged_ranges:
        # Get the bounds of the merged range.
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        
        # Retrieve the value from the top-left (master) cell.
        master_value = ws.cell(row=min_row, column=min_col).value
        
        # Unmerge the cells.
        ws.unmerge_cells(str(merged_range))
        
        # Fill every cell in the previously merged range with the master value.
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=master_value)
        
        debug_print(f"Unmerged range {merged_range} with value: {master_value}")

print("DEBUG: helpers.py - Excel processing functions loaded")

# ============================================================================
# FILE FORMAT DETECTION FUNCTIONS
# ============================================================================

def is_standard_file(file_path: str) -> bool:
    """
    Check if a file follows the standard testing format.
    Enhanced logic to detect legacy vs standard files.
    """
    try:
        debug_print(f"Checking file format: {file_path}")
        
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            debug_print("Not an Excel file -> treating as standard")
            return True
        
        # Load workbook to check structure
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        num_sheets = len(sheet_names)
        
        debug_print(f"Found {num_sheets} sheets: {sheet_names}")
        
        # Legacy file criteria: single sheet named "Sheet1"
        if num_sheets == 1 and sheet_names[0].lower() == "sheet1":
            debug_print(f"Legacy sheet name: '{sheet_names[0]}'")
            wb.close()
            return False  # Return False for legacy files
        else:
            debug_print("File does not meet legacy criteria -> File is standard.")
            if num_sheets > 1:
                debug_print(f"Multiple sheets found ({num_sheets}), treating as standard")
            elif num_sheets == 1:
                debug_print(f"Single sheet found but name is '{sheet_names[0]}' (not 'Sheet1'), treating as standard")
            wb.close()
            return True   # Return True for standard files
            
    except Exception as e:
        debug_print(f"ERROR: Exception while checking file format: {e}")
        debug_print(f"Traceback: {traceback.format_exc()}")
        # On error, default to treating as standard file
        debug_print("Error occurred, defaulting to standard file format")
        return True

def plotting_sheet_test(sheet_name: str, data: pd.DataFrame = None) -> bool:
    """
    Determine if a sheet is a plotting sheet by searching the first few rows for 'puffs' and 'TPM'.
    Enhanced with both name-based and data-based detection.
    """
    try:
        if not sheet_name:
            return False
        
        # Check for legacy sheets first
        if "legacy" in sheet_name.lower():
            debug_print(f"plotting_sheet_test({sheet_name}): Legacy sheet detected")
            return True
        
        # Keywords that indicate a plotting sheet
        plotting_keywords = [
            'test', 'data', 'measurement', 'results', 'analysis',
            'tpm', 'resistance', 'pressure', 'efficiency'
        ]
        
        sheet_lower = sheet_name.lower()
        name_based_result = any(keyword in sheet_lower for keyword in plotting_keywords)
        
        # If we have data, also check the content
        if data is not None and not data.empty:
            if data.shape[0] < 3 or data.shape[1] < 2:
                debug_print(f"plotting_sheet_test({sheet_name}): Data too small for content analysis")
                return name_based_result

            # Dynamically search the first 5 rows for headers
            for i in range(min(5, len(data))):
                header_row = data.iloc[i].astype(str).str.lower()
                
                if header_row.str.contains("puffs").any() and header_row.str.contains("tpm").any():
                    debug_print(f"plotting_sheet_test({sheet_name}): Found plotting headers in row {i}")
                    return True

            debug_print(f"plotting_sheet_test({sheet_name}): No valid plotting headers found in data")
        
        debug_print(f"plotting_sheet_test({sheet_name}): {name_based_result} (name-based)")
        return name_based_result

    except Exception as e:
        debug_print(f"ERROR: plotting_sheet_test failed for {sheet_name}: {e}")
        return False

print("DEBUG: helpers.py - File format detection functions loaded")

# ============================================================================
# TEXT AND UI UTILITY FUNCTIONS
# ============================================================================

def wrap_text(text: str, width: int = 50) -> str:
    """
    Wrap text to specified width, preserving whole words.
    Enhanced version with better word wrapping logic.
    """
    if not text or width <= 0:
        return text
    
    words = text.split()
    lines = []
    current_line = []
    current_length = 0
    
    for word in words:
        # Check if adding this word would exceed the line length
        test_line_length = current_length + len(word) + len(current_line)  # +spaces
        
        if test_line_length <= width:
            # Word fits, add it to current line
            current_line.append(word)
            current_length += len(word)
        else:
            # Word doesn't fit
            if current_line:
                # Save current line and start new line with this word
                lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word)
            else:
                # Word itself is longer than max_width, we need to break it
                if len(word) > width:
                    # Break the long word at character boundaries
                    while len(word) > width:
                        lines.append(word[:width])
                        word = word[width:]
                    current_line = [word] if word else []
                    current_length = len(word) if word else 0
                else:
                    current_line = [word]
                    current_length = len(word)
    
    # Add the last line if it's not empty
    if current_line:
        lines.append(' '.join(current_line))
    
    wrapped = '\n'.join(lines)
    debug_print(f"wrap_text - {len(text)} chars -> {len(lines)} lines")
    return wrapped

def center_window(window: tk.Toplevel, width: int = None, height: int = None):
    """Center a window on screen."""
    try:
        window.update_idletasks()
        
        # Get window dimensions
        if width and height:
            w, h = width, height
        else:
            w = window.winfo_width()
            h = window.winfo_height()
        
        # Get screen dimensions
        screen_w = window.winfo_screenwidth()
        screen_h = window.winfo_screenheight()
        
        # Calculate center position
        x = (screen_w - w) // 2
        y = (screen_h - h) // 2
        
        # Set window geometry
        window.geometry(f"{w}x{h}+{x}+{y}")
        
        debug_print(f"center_window - {w}x{h} at ({x},{y}) on {screen_w}x{screen_h} screen")
        
    except Exception as e:
        debug_print(f"ERROR: center_window failed: {e}")

def show_success_message(title: str, message: str, parent: tk.Tk = None):
    """
    Show a success message dialog with enhanced styling.
    Enhanced version with custom dialog appearance.
    """
    try:
        # Simple version for now
        messagebox.showinfo(title, message, parent=parent)
        debug_print(f"show_success_message - {title}: {message}")
    except Exception as e:
        debug_print(f"ERROR: show_success_message failed: {e}")

print("DEBUG: helpers.py - UI utility functions loaded")

# ============================================================================
# DATA VALIDATION FUNCTIONS
# ============================================================================

def validate_dataframe(df: pd.DataFrame, min_rows: int = 1, min_cols: int = 1) -> tuple[bool, str]:
    """Validate a DataFrame meets minimum requirements."""
    if df is None:
        return False, "DataFrame is None"
    
    if df.empty:
        return False, "DataFrame is empty"
    
    if len(df) < min_rows:
        return False, f"DataFrame has {len(df)} rows, minimum {min_rows} required"
    
    if len(df.columns) < min_cols:
        return False, f"DataFrame has {len(df.columns)} columns, minimum {min_cols} required"
    
    debug_print(f"validate_dataframe - {len(df)} rows, {len(df.columns)} cols - VALID")
    return True, "DataFrame is valid"

def validate_sheet_data(data: pd.DataFrame, required_columns: list = None, required_rows: int = None) -> bool:
    """
    Validate the sheet data for common issues like empty DataFrame, missing columns, or missing rows.

    Args:
        data (pd.DataFrame): The DataFrame to validate.
        required_columns (list, optional): List of columns that must exist. Defaults to None.
        required_rows (int, optional): Minimum number of rows required. Defaults to None.

    Returns:
        bool: True if the sheet is valid, False otherwise.
    """
    if data.empty:
        debug_print("Sheet is empty.")
        return False

    if required_columns:
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            debug_print(f"Missing required columns: {missing_columns}")
            return False

    if required_rows and len(data) < required_rows:
        debug_print(f"Insufficient rows: {len(data)} < {required_rows}")
        return False

    debug_print("Sheet data validation passed")
    return True

def safe_numeric_conversion(value: Any, default: float = 0.0) -> float:
    """Safely convert a value to numeric, with fallback."""
    try:
        if pd.isna(value):
            return default
        
        # Try direct float conversion
        return float(value)
        
    except (ValueError, TypeError):
        try:
            # Try pandas numeric conversion
            result = pd.to_numeric(value, errors='coerce')
            return float(result) if not pd.isna(result) else default
        except:
            return default

print("DEBUG: helpers.py - Data validation functions loaded")

# ============================================================================
# HELPER FUNCTIONS FOR PROCESSING MODULE
# ============================================================================

def extract_meta_data(worksheet, patterns: dict) -> dict:
    """Extract meta_data from first 3 rows using regex patterns."""
    meta_data = {}
    debug_print(f"Extracting metadata using patterns: {list(patterns.keys())}")
    
    for row in range(1, 4):  # Rows 1-3 (1-indexed)
        for col in range(1, worksheet.max_column + 1):
            try:
                cell_value = str(worksheet.cell(row=row, column=col).value).lower()
                for key, pattern in patterns.items():
                    if re.search(pattern, cell_value, re.IGNORECASE):
                        meta_data[key] = worksheet.cell(row=row, column=col+1).value  # Get value from next cell
                        debug_print(f"Found metadata {key}: {meta_data[key]} at row {row}, col {col}")
                        break
            except Exception as e:
                debug_print(f"Error processing cell at row {row}, col {col}: {e}")
                continue
    
    debug_print(f"Extracted metadata: {meta_data}")
    return meta_data

def map_meta_data_to_template(template_ws, meta_data: dict):
    """Map extracted meta_data to template positions."""
    debug_print(f"Mapping metadata to template: {meta_data}")
    
    # This would contain the specific mapping logic for your template
    # Implementation depends on your template structure
    mapped_count = 0
    
    for key, value in meta_data.items():
        # Add your specific mapping logic here
        debug_print(f"Mapping {key}: {value}")
        mapped_count += 1
    
    debug_print(f"Mapped {mapped_count} metadata items to template")

def header_matches(header: str, target: str, threshold: float = 0.8) -> bool:
    """Check if header matches target with fuzzy matching."""
    if not header or not target:
        return False
    
    # Simple contains check for now
    similarity = target.lower() in header.lower()
    debug_print(f"header_matches('{header}', '{target}'): {similarity}")
    return similarity

def autofit_columns_in_excel(file_path: str):
    """
    Adjusts the column widths in the Excel file to fit the content automatically.
    """
    try:
        debug_print(f"Autofitting columns in: {file_path}")
        
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        wb.close()
        debug_print(f"Successfully autofitted columns in {file_path}")
        
    except Exception as e:
        debug_print(f"ERROR: Failed to autofit columns: {e}")

print("DEBUG: helpers.py - Processing helper functions loaded")

# ============================================================================
# FINAL INITIALIZATION
# ============================================================================

print("DEBUG: helpers.py - All utility functions loaded successfully")
print(f"DEBUG: helpers.py - Debug mode is {'ENABLED' if DEBUG_ENABLED else 'DISABLED'}")
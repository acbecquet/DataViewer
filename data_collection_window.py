"""
data_collection_window.py
Developed by Charlie Becquet.
Interface for rapid test data collection with enhanced saving and menu functionality.
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from tksheet import Sheet
import pandas as pd
import numpy as np
import os
import copy
import time
import openpyxl
import logging
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import statistics
from openpyxl.styles import PatternFill
from utils import FONT, debug_print, show_success_message
import threading
import subprocess

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("datacollection.log"),
        logging.StreamHandler()
    ]
)

# Standard Operating Procedures for each test
TEST_SOPS = {
    "User Test Simulation": """
SOP - User Test Simulation:

Day 1: Collect Initial Draw Resistance, TPM for 50 puffs, then use puff per day calculator to determine number of puffs. Make sure to enter initial oil mass.

Days 2-4: 10 puffs TPM after 

Day 5: Extended Test, every 20 puffs until device is empty

Day 6: Full Disassembly + Photos

Be sure to take detailed notes while conducting this test.

Key Points:
- Split testing: Phase 1 (0-50 puffs) and Phase 2 (extended puffs)
- No resistance measurements during extended testing
    """,
    
    "Intense Test": """
SOP - Intense Test:

Conduct intense testing at a draw pressure of 5kPa. Use either 200mL/3s/30s or 160mL/3s/30s regime.
Record TPM, draw pressure, and resistance at regular intervals.
Monitor for device overheating or failure.
Document any unusual observations in notes.

Key Points:
- 5kPa draw pressure target
- Standard 12-column data collection
- Record all measurements consistently
    """,

"User Simulation Test": """
SOP - User Test Simulation:

Day 1: Collect Initial Draw Resistance, TPM for 50 puffs, then use puff per day calculator to determine number of puffs. Make sure to enter initial oil mass.
Days 2-4: 10 puffs with TPM, then complete the daily number of puffs.
Day 5: Extended Test, every 10 puffs until device is empty
Day 6: Full Disassembly + Photos

Be sure to take detailed notes while conducting this test.

Key Points:
- Split testing: Phase 1 (0-50 puffs) and Phase 2 (extended puffs)
- No resistance measurements during extended testing
    """,
    
    "Big Headspace High Temperature": """
SOP - Big Headspace High Temperature Test:

Drain device to 30% remaining and then place in the oven at 40C.
After 1 hour, collect 10 puffs, tracking TPM and draw resistance for each puff. Repeat this 3 times, for 30 total puffs.
Be sure to take detailed notes on bubbling and any failure modes.

Key Points:
- 40C big headspace test
- Monitor for thermal effects
- Document temperature-related observations
    """,
    
    "Extended Test": """
SOP - Extended Test:

Long-duration testing to assess device lifetime and performance degradation.
Sessions of 10 puffs with a 60mL/3s/30s puffing regime. Rest 15 minutes between sessions.
Monitor for performance and consistency over time. Measure initial and final draw resistance, and measure TPM every 10 puffs.

Key Points:
- Extended duration testing
- Regular measurement intervals
- Performance degradation monitoring
- Comprehensive data collection
    """,
    
    "Quick Screening Test": """
SOP - Quick Screening Test:

Rapid assessment of device basic performance.
Focus on key performance indicators.
Streamlined testing for initial evaluation.

Key Points:
- Rapid testing protocol
- Key performance metrics only
- Initial device assessment
- Basic functionality verification
    """,
    
    "default": """
SOP - Standard Test Protocol:

Follow standard testing procedures for this test type.
Record all required measurements accurately.
Document any observations or anomalies.
Ensure proper test conditions throughout.

Key Points:
- Standard measurement protocol
- Accurate data recording
- Proper documentation
- Consistent test conditions
    """
}

class DataCollectionWindow:
    def __init__(self, parent, file_path, test_name, header_data, original_filename=None):
        """
        Initialize the data collection window.
    
        Args:
            parent (tk.Tk): The parent window.
            file_path (str): Path to the Excel file.
            test_name (str): Name of the test being conducted.
            header_data (dict): Dictionary containing header data.
            original_filename (str): original filename for .vap3 files
        """
        # Debug flag - set to False to disable all debug output
        self.DEBUG = False
        self.parent = parent
        self.root = parent.root
        self.file_path = file_path
        self.test_name = test_name
        self.header_data = header_data
        self.num_samples = header_data["num_samples"]
        self.result = None

        if hasattr(parent, 'root'):
            self.main_window_was_visible = parent.root.winfo_viewable()
            parent.root.withdraw()  # Hide main window
            debug_print("DEBUG: Main GUI window hidden")

        # Store the original filename for saving purposes
        self.original_filename = original_filename
        if self.original_filename:
            debug_print(f"DEBUG: DataCollectionWindow initialized with original filename: {self.original_filename}")
   
        # Validate num_samples
        if self.num_samples <= 0:
            self.num_samples = len(header_data.get('samples', []))
            if self.num_samples <= 0:
                self.num_samples = 1

        # Auto-save settings
        self.auto_save_interval = 5 * 60 * 1000  # 5 minutes in milliseconds
        self.auto_save_timer = None
        self.has_unsaved_changes = False
        self.last_save_time = None
    
        # Create the window
        self.window = tk.Toplevel(self.root)
        self.window.title(f"Data Collection - {test_name}")
        self.window.state('zoomed')
        self.window.minsize(1250, 625)
        
        self.window.lift()  # Bring to top of window stack
        self.window.focus_force()  # Force focus to this window
    
        # Make it stay on top temporarily
        self.window.attributes('-topmost', True)
        self.window.after(100, lambda: self.window.attributes('-topmost', False)) 

        # Default puff interval
        self.puff_interval = 10  # Default to 10

        # Set up keyboard shortcut flags
        self.hotkeys_enabled = True
        self.hotkey_bindings = {}
    
        # Create the style for ttk widgets
        self.style = ttk.Style()
        self.setup_styles()

        # Initialize data structures
        self.initialize_data()
    
        # Create the menu bar first
        self.create_menu_bar()
    
        # Create the UI
        self.create_widgets()

        self.load_existing_data_from_file()
    
        # Center the window
        self.center_window()
    
        # Set up event handlers
        self.setup_event_handlers()
    
        # Start auto-save timer
        self.start_auto_save_timer()

        self.ensure_initial_tpm_calculation()
    
        self.log(f"DataCollectionWindow initialized for {test_name} with {self.num_samples} samples", "debug")

    def create_menu_bar(self):
        """Create a comprehensive menu bar for the data collection window."""
        menubar = tk.Menu(self.window)
        self.window.config(menu=menubar)
    
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Save", accelerator="Ctrl+S", 
                            command=lambda: self.save_data(show_confirmation=False))
        file_menu.add_command(label="Save and Exit", 
                            command=lambda: self.save_data(exit_after=True))
        file_menu.add_separator()
        file_menu.add_command(label="Export CSV", command=self.export_csv)
        file_menu.add_separator()
        file_menu.add_command(label="Exit without Saving", command=self.exit_without_saving)
        menubar.add_cascade(label="File", menu=file_menu)
    
        # Edit menu  
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Edit Headers", command=self.edit_headers)
        edit_menu.add_command(label="Clear Current Sample", command=self.clear_current_sample)
        edit_menu.add_command(label="Clear All Data", command=self.clear_all_data)
        edit_menu.add_separator()
        edit_menu.add_command(label="Add Row", command=self.add_row)
        edit_menu.add_command(label="Add Rows", command=self.add_multiple_rows)
        edit_menu.add_command(label="Remove Last Row", command=self.remove_last_row)
        edit_menu.add_separator()
        edit_menu.add_command(label="Recalculate TPM", command=self.recalculate_all_tpm)
        menubar.add_cascade(label="Edit", menu=edit_menu)
    
        # Navigate menu
        navigate_menu = tk.Menu(menubar, tearoff=0)
        navigate_menu.add_command(label="Previous Sample", accelerator="Alt+Left", 
                                command=self.go_to_previous_sample)
        navigate_menu.add_command(label="Next Sample", accelerator="Alt+Right", 
                                command=self.go_to_next_sample)
        navigate_menu.add_separator()
        navigate_menu.add_command(label="Go to Sample...", command=self.go_to_sample_dialog)
        menubar.add_cascade(label="Navigate", menu=navigate_menu)
    
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="Change Puff Interval", command=self.change_puff_interval_dialog)
        tools_menu.add_command(label="Auto-Save Settings", command=self.auto_save_settings_dialog)
        tools_menu.add_separator()
        menubar.add_cascade(label="Tools", menu=tools_menu)
    
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Keyboard Shortcuts", command=self.show_keyboard_shortcuts)
        help_menu.add_command(label="About Data Collection", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)
    
        self.log("Menu bar created successfully", "debug")
    
    def setup_styles(self):
        """Set up styles for ttk widgets with enhanced visual separation."""
        # Configure ttk styles to use system defaults
        self.style.configure('TFrame', background='SystemButtonFace')
        self.style.configure('TLabel', background='SystemButtonFace')
        self.style.configure('TLabelframe', background='SystemButtonFace')
        self.style.configure('TLabelframe.Label', background='SystemButtonFace')
        self.style.configure('TNotebook', background='SystemButtonFace')
        self.style.configure('TNotebook.Tab', background='SystemButtonFace')

        # Create special styles for headers
        self.style.configure('Header.TLabel', font=("Arial", 14, "bold"), background='SystemButtonFace')
        self.style.configure('SubHeader.TLabel', font=("Arial", 12), background='SystemButtonFace')
        self.style.configure('Stats.TLabel', font=("Arial", 14, "bold"), background='SystemButtonFace')
        self.style.configure('SampleInfo.TLabel', font=("Arial", 11), background='SystemButtonFace')
                     
        # Enhanced Treeview styling with better visual separation
        self.style.configure('Treeview', 
                            background='white',
                            fieldbackground='white',
                            borderwidth=1,
                            rowheight=28,  # Taller rows for better separation
                            font=('Arial', 10))
    
        # Configure selection colors
        self.style.map('Treeview', 
                      background=[('selected', '#E8F4FD')],
                      foreground=[('selected', 'black')])

        # Style headers with enhanced borders
        self.style.configure('Treeview.Heading',
                            background='#D0D0D0',
                            foreground='black',
                            font=('Arial', 10, 'bold'),
                            borderwidth=2,
                            relief='raised')
    
        # Add hover effect
        self.style.map('Treeview.Heading',
                      background=[('active', '#C0C0C0')])

        debug_print("DEBUG: Enhanced styles configured with better visual separation")
    
    def on_window_resize_plot(self, event):
        """Handle window resize events with debouncing for plot updates."""
        debug_print(f"DEBUG: RESIZE EVENT DETECTED - Widget: {event.widget}, Window: {self.window}")
        debug_print(f"DEBUG: Event widget type: {type(event.widget)}")
        debug_print(f"DEBUG: Window type: {type(self.window)}")
        debug_print(f"DEBUG: Event widget == window? {event.widget == self.window}")
    
        # Only handle main window resize events, not child widgets
        if event.widget != self.window:
            debug_print(f"DEBUG: Ignoring resize event from child widget: {event.widget}")
            return
    
        debug_print("DEBUG: MAIN WINDOW RESIZE CONFIRMED - Processing...")
    
        # Get current window dimensions for verification
        current_width = self.window.winfo_width()
        current_height = self.window.winfo_height()
        debug_print(f"DEBUG: Current window dimensions: {current_width}x{current_height}")
    
        # Debounce rapid resize events
        if hasattr(self, '_resize_timer'):
            self.window.after_cancel(self._resize_timer)
            debug_print("DEBUG: Cancelled previous resize timer")
    
        # Schedule plot size update with a small delay to avoid excessive updates
        self._resize_timer = self.window.after(1000, self.update_plot_size_for_resize)
        debug_print("DEBUG: Scheduled plot resize update in 1000ms")

    def update_plot_size_for_resize(self):
        """Update plot size with artifact prevention and frame validation."""
        try:
            # Check if we have the necessary components
            if not hasattr(self, 'stats_canvas') or not self.stats_canvas.get_tk_widget().winfo_exists():
                debug_print("DEBUG: Stats canvas not available for resize")
                return
    
            if not hasattr(self, 'stats_fig') or not self.stats_fig:
                debug_print("DEBUG: Stats figure not available for resize")
                return
    
            # Wait for frame geometry to stabilize
            self.window.update_idletasks()
    
            # Use the actual stats container for sizing
            if hasattr(self, 'stats_frame_container') and self.stats_frame_container.winfo_exists():
                parent_for_sizing = self.stats_frame_container
            else:
                debug_print("DEBUG: Stats frame container not available, skipping resize")
                return
    
            parent_for_sizing.update_idletasks()
    
            # Validate that frames have reasonable dimensions before proceeding
            parent_width = parent_for_sizing.winfo_width()
            parent_height = parent_for_sizing.winfo_height()
    
            debug_print(f"DEBUG: Parent frame size for stats: {parent_width}x{parent_height}")
    
            if parent_width < 200 or parent_height < 200:
                debug_print("DEBUG: Parent frame size too small, skipping this resize update")
                return
    
            # Calculate new size based on validated frame dimensions
            new_width, new_height = self.calculate_dynamic_plot_size(parent_for_sizing)
    
            # Get current figure size for comparison
            current_width, current_height = self.stats_fig.get_size_inches()
    
            # Only update if change is significant
            width_diff = abs(new_width - current_width)
            height_diff = abs(new_height - current_height)
            threshold = 0.5  # Threshold to reduce excessive updates
    
            if width_diff > threshold or height_diff > threshold:
                debug_print(f"DEBUG: Significant size change detected - updating plot from {current_width:.2f}x{current_height:.2f} to {new_width:.2f}x{new_height:.2f}")
        
                # Apply the new size
                self.stats_fig.set_size_inches(new_width, new_height)
        
                # Redraw the canvas
                self.stats_canvas.draw_idle()
                debug_print("DEBUG: Plot resize completed")
            else:
                debug_print("DEBUG: Size change below threshold, skipping update to prevent artifacts")
        
        except Exception as e:
            debug_print(f"DEBUG: Error during plot resize: {str(e)}")
            import traceback
            debug_print(f"DEBUG: Full traceback: {traceback.format_exc()}")

    def calculate_dynamic_plot_size(self, parent_frame):
        """Calculate plot size that directly scales with window size."""
        debug_print("DEBUG: Starting dynamic plot size calculation")
    
        # Force geometry update to get current dimensions
        self.window.update_idletasks()
    
        if hasattr(self, 'stats_frame') and self.stats_frame.winfo_exists():
            parent_frame = self.stats_frame
            parent_frame.update_idletasks()
    
        # Get the actual dimensions
        available_width = parent_frame.winfo_width()
        available_height = parent_frame.winfo_height()
    
        debug_print(f"DEBUG: Parent frame: {parent_frame.__class__.__name__}")
        debug_print(f"DEBUG: Available dimensions - Width: {available_width}px, Height: {available_height}px")
    
        # Simple fallback for initial sizing or very small windows
        if available_width < 200 or available_height < 200:
            debug_print("DEBUG: Using fallback size for small window")
            return (6, 4)
    
        # Reserve space for stats text and controls
        text_space = 120  # Space for text statistics
        control_space = 50  # Space for labels and padding
    
        # Calculate available space for the plot
        plot_height_available = available_height - text_space - control_space
        plot_width_available = available_width - 40  # 20px margin on each side
    
        # Use most of the available space for the plot
        plot_width_pixels = max(plot_width_available, 200)
        plot_height_pixels = max(plot_height_available, 150)
    
        debug_print(f"DEBUG: Plot space in pixels - Width: {plot_width_pixels}px, Height: {plot_height_pixels}px")
    
        # Convert to inches for matplotlib (using standard 100 DPI)
        plot_width_inches = plot_width_pixels / 100.0
        plot_height_inches = plot_height_pixels / 100.0
    
        # Apply minimum and maximum sizes
        plot_width_inches = max(min(plot_width_inches, 12.0), 3.0)
        plot_height_inches = max(min(plot_height_inches, 8.0), 2.0)
    
        debug_print(f"DEBUG: FINAL plot size - Width: {plot_width_inches:.2f} inches, Height: {plot_height_inches:.2f} inches")
    
        return (plot_width_inches, plot_height_inches)

    def initialize_plot_size(self):
        """Initialize plot size after window is fully rendered."""
        debug_print("DEBUG: Initializing plot size after window render")
    
        # Give window time to fully render
        self.window.update_idletasks()
    
        # Update plot size if available
        if hasattr(self, 'update_plot_size_for_resize'):
            self.window.after(500, self.update_plot_size_for_resize)
    
        debug_print("DEBUG: Initial plot size set")

    def center_window(self):
        """Center the window on the screen."""
        # Force the window to update and calculate its actual size
        self.window.update()
    
        # Get the actual window dimensions
        width = self.window.winfo_width()
        height = self.window.winfo_height()
    
        # If width/height are still too small, use the requested size
        if width < 100 or height < 100:
            self.window.update_idletasks()
            width = self.window.winfo_reqwidth()
            height = self.window.winfo_reqheight()
    
        # Calculate center position
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
    
        # Ensure the window isn't positioned off-screen
        x = max(0, min(x, screen_width - width))
        y = max(0, min(y, screen_height - height))
    
        self.window.geometry(f"{width}x{height}+{x}+{y}")
    
        debug_print(f"DEBUG: Window centered at {x},{y} with size {width}x{height}")
    

    def create_widgets(self):
        """Create the data collection UI with a cleaner structure."""
        # Configure window
        self.window.configure(background='SystemButtonFace')

        # Create a main frame with padding
        main_frame = ttk.Frame(self.window, padding=10)
        main_frame.pack(fill="both", expand=True)

        # Create header section
        self.create_header_section(main_frame)

        # Create SOP section
        self.create_sop_section(main_frame)

        # Create main content area with horizontal split
        content_frame = ttk.PanedWindow(main_frame, orient="horizontal")
        content_frame.pack(fill="both", expand=True)

        # Create data entry area (left side)
        data_frame = ttk.Frame(content_frame)
        content_frame.add(data_frame, weight=2)  # 50% width

        # Create stats panel (right side)
        self.stats_frame = ttk.Frame(content_frame)
        content_frame.add(self.stats_frame, weight=2)  # 50% width

        # Create notebook with sample tabs
        self.notebook = ttk.Notebook(data_frame)
        self.notebook.pack(fill="both", expand=True, pady=(0, 10))  # Add bottom padding

        # Create sample tabs
        self.create_sample_tabs()

        # Create stats panel
        self.create_tpm_stats_panel()

        # Bind notebook tab change event
        self.notebook.bind("<<NotebookTabChanged>>", self.update_stats_panel)
   
    def log(self, message, level="info"):
        """Log a message with the specified level."""
        logger = logging.getLogger("DataCollectionWindow")
    
        if level.lower() == "debug":
            logger.debug(message)
        elif level.lower() == "info":
            logger.info(message)
        elif level.lower() == "warning":
            logger.warning(message)
        elif level.lower() == "error":
            logger.error(message)
        elif level.lower() == "critical":
            logger.critical(message)

    def create_sop_section(self, parent_frame):
        """Create the SOP (Standard Operating Procedure) section."""
        
        # Create a collapsible SOP frame
        sop_frame = ttk.LabelFrame(parent_frame, text="", style='TLabelframe')
        sop_frame.pack(fill="x", pady=(0, 10))
    
        # Create a header frame for button and title
        header_frame = ttk.Frame(sop_frame, style='TFrame')
        header_frame.pack(fill="x", padx=5, pady=5)
    
        # Add small toggle button on the left with minimal size
        self.sop_visible = False
        self.toggle_sop_button = ttk.Button(
            header_frame,
            text="Show",
            width=6,  # Even smaller width
            command=self.toggle_sop_visibility
        )
        self.toggle_sop_button.pack(side="left", padx=(0, 5))  # Small gap between button and title
    
        # Add title label after the button
        title_label = ttk.Label(header_frame, text="Standard Operating Procedure (SOP)", 
                               font=("Arial", 10, "bold"), style='TLabel')
        title_label.pack(side="left")
    
        # Get the SOP text for the current test
        sop_text = TEST_SOPS.get(self.test_name, TEST_SOPS["default"])
    
        # Create a text widget to display the SOP with centered text
        self.sop_text_widget = tk.Text(
            sop_frame, 
            height=6,  
            wrap=tk.WORD,
            font=("Arial", 9),
            bg="SystemButtonFace",
            fg="black",
            relief="flat",
            borderwidth=0,
            padx=10,
            pady=5
        )

        # Insert the SOP text and center it
        self.sop_text_widget.insert("1.0", sop_text.strip())
    
        # Center the text
        self.sop_text_widget.tag_configure("center", justify='center')
        self.sop_text_widget.tag_add("center", "1.0", "end")
    
        # Make it read-only
        self.sop_text_widget.config(state="disabled")
    
        debug_print("DEBUG: SOP section created successfully")

    def toggle_sop_visibility(self):
        """Toggle the visibility of the SOP text."""
        if self.sop_visible:
            # Hide the SOP
            self.sop_text_widget.pack_forget()
            self.toggle_sop_button.config(text="Show")
            self.sop_visible = False
            debug_print("DEBUG: SOP section hidden")
        else:
            # Show the SOP
            self.sop_text_widget.pack(fill="x", padx=10, pady=(0, 5))
            self.toggle_sop_button.config(text="Hide")
            self.sop_visible = True
            debug_print("DEBUG: SOP section shown")

    # Auto-save functionality
    def start_auto_save_timer(self):
        """Start the auto-save timer."""
        if self.auto_save_timer:
            self.window.after_cancel(self.auto_save_timer)
        
        self.auto_save_timer = self.window.after(self.auto_save_interval, self.auto_save)
        
    def auto_save(self):
        """Perform automatic save without user confirmation."""
        if self.has_unsaved_changes:            
            try:
                self.save_data(show_confirmation=False, auto_save=True)
                self.update_save_status(False)  # Mark as saved
                debug_print("DEBUG: Auto-save completed successfully")
            except Exception as e:
                debug_print(f"DEBUG: Auto-save failed: {e}")
                # Continue even if auto-save fails
        
        # Restart the timer
        self.start_auto_save_timer()
    
    def mark_unsaved_changes(self):
        """Mark that there are unsaved changes."""
        if not self.has_unsaved_changes:
            self.has_unsaved_changes = True
            self.update_save_status(True)
            
    def update_save_status(self, has_changes):
        """Update the save status indicator in the UI."""
        if has_changes:
            self.save_status_label.config(foreground="red", text="●")
            self.save_status_text.config(text="Unsaved changes")
        else:
            self.save_status_label.config(foreground="green", text="●")
            self.save_status_text.config(text="All changes saved")
            import datetime
            self.last_save_time = datetime.datetime.now()
    
    def exit_without_saving(self):
        """Exit without saving (with confirmation)."""
        if self.has_unsaved_changes:
            if not messagebox.askyesno("Confirm Exit", 
                                     "You have unsaved changes. Are you sure you want to exit without saving?"):
                return
        
        debug_print("DEBUG: Exiting without saving")
        self.result = "cancel"
        self.window.destroy()
    
    def export_csv(self):
        """Export current data to CSV files."""
        
        from tkinter import filedialog
        
        # Ask for directory to save CSV files
        directory = filedialog.askdirectory(title="Select directory to save CSV files")
        if not directory:
            return
        
        try:
            for i in range(self.num_samples):
                sample_id = f"Sample {i+1}"
                sample_name = self.header_data["samples"][i]["id"]
                
                # Create DataFrame for this sample
                df_data = {
                    "Puffs": self.data[sample_id]["puffs"],
                    "Before Weight (g)": self.data[sample_id]["before_weight"],
                    "After Weight (g)": self.data[sample_id]["after_weight"],
                    "Draw Pressure (kPa)": self.data[sample_id]["draw_pressure"],
                    "Smell": self.data[sample_id]["smell"],
                    "Notes": self.data[sample_id]["notes"],
                    "TPM (mg/puff)": self.data[sample_id]["tpm"]
                }
                
                df = pd.DataFrame(df_data)
                
                # Remove empty rows
                df = df.dropna(how='all', subset=["Before Weight (g)", "After Weight (g)"])
                
                csv_filename = f"{self.test_name}_{sample_name}_data.csv"
                csv_path = os.path.join(directory, csv_filename)
                df.to_csv(csv_path, index=False)
                
            show_success_message("Export Complete", f"Exported {self.num_samples} CSV files to {directory}", self.window)
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export CSV files: {e}")
    
    def edit_headers(self):
        """Open header editing dialog from within data collection."""
        debug_print("DEBUG: Edit Headers requested from data collection window")
        
        # Show the header data dialog in edit mode
        from header_data_dialog import HeaderDataDialog
        header_dialog = HeaderDataDialog(
            self.window, 
            self.file_path, 
            self.test_name, 
            edit_mode=True,
            current_data=self.header_data
        )
        result, new_header_data = header_dialog.show()
        
        if result:
            debug_print("DEBUG: Header data updated successfully")
            # Update our internal header data
            old_header_data = self.header_data.copy()
            self.header_data = new_header_data
            
            # Check if number of samples changed
            old_num_samples = old_header_data.get("num_samples", 0)
            new_num_samples = new_header_data.get("num_samples", 0)
            
            if old_num_samples != new_num_samples:
                debug_print(f"DEBUG: Number of samples changed from {old_num_samples} to {new_num_samples}")
                # Handle sample count change
                self.handle_sample_count_change(old_num_samples, new_num_samples)
            
            # Update the header display in all tabs
            self.update_header_display()
            
            # Apply changes to the Excel file
            self.apply_header_changes_to_file()
            
            # Mark as having unsaved changes
            self.mark_unsaved_changes()
            
            # Show success message
            self.show_temp_status_message("Headers updated successfully", 3000)
        else:
            debug_print("DEBUG: Header editing was cancelled")

    def handle_sample_count_change(self, old_count, new_count):
        """Handle changes in the number of samples."""
        debug_print(f"DEBUG: Handling sample count change: {old_count} -> {new_count}")

        if new_count > old_count:
            # Add new samples
            for i in range(old_count, new_count):
                sample_id = f"Sample {i+1}"
            
                # Initialize with the complete data structure (matching initialize_data)
                self.data[sample_id] = {
                    "current_row_index": 0, 
                    "avg_tpm": 0.0
                }

                # Add all standard columns (must match the structure in initialize_data)
                columns = ["puffs", "before_weight", "after_weight", "draw_pressure", "resistance", "smell", "clog", "notes", "tpm"]
                for column in columns:
                    self.data[sample_id][column] = []

                # Add special columns for User Test Simulation
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    self.data[sample_id]["chronography"] = []

                # Pre-initialize 50 rows for new sample
                for j in range(50):
                    puff = (j + 1) * self.puff_interval
                    self.data[sample_id]["puffs"].append(puff)
                
                    # Initialize other columns with empty values
                    for column in columns:
                        if column == "puffs":
                            continue  # Already initialized
                        elif column == "tpm":
                            self.data[sample_id][column].append(None)
                        else:
                            self.data[sample_id][column].append("")
                
                    # Add chronography initialization for User Test Simulation
                    if "chronography" in self.data[sample_id]:
                        self.data[sample_id]["chronography"].append("")
            
                debug_print(f"DEBUG: Added new sample {sample_id} with complete data structure")

            debug_print(f"DEBUG: Added {new_count - old_count} new samples")

        elif new_count < old_count:
            # Remove excess samples
            for i in range(new_count, old_count):
                sample_id = f"Sample {i+1}"
                if sample_id in self.data:
                    del self.data[sample_id]
                    debug_print(f"DEBUG: Removed sample {sample_id}")

            debug_print(f"DEBUG: Removed {old_count - new_count} samples")

        # Update the number of samples
        self.num_samples = new_count

        # Recreate the UI to reflect the new sample count
        self.recreate_sample_tabs()

    def recreate_sample_tabs(self):
        """Recreate all sample tabs when sample count changes."""
        debug_print("DEBUG: Recreating sample tabs")
    
        # Clear existing tabs
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
    
        # Just call the existing method
        self.create_sample_tabs()
    
        # Recreate the TPM stats panel
        self.create_tpm_stats_panel()
    
        debug_print(f"DEBUG: Recreated {self.num_samples} sample tabs")

    def update_header_display(self):
        """Update the header information displayed in the UI."""
        
        # Update the header labels (if they exist)
        try:
            # Update any header text that shows tester name, etc.
            for widget in self.window.winfo_children():
                if hasattr(widget, 'winfo_children'):
                    self.update_header_labels_recursive(widget)
            
            # Keep original tab labels unchanged
            for i in range(min(self.num_samples, len(self.sample_frames))):
                self.notebook.tab(i, text=f"Sample {i+1}")
                debug_print(f"DEBUG: Keeping tab {i} as 'Sample {i+1}'")
                    
        except Exception as e:
            debug_print(f"DEBUG: Error updating header display: {e}")

    def update_header_labels_recursive(self, widget):
        """Recursively update header labels in the widget tree."""
        try:
            if isinstance(widget, ttk.Label):
                text = widget.cget('text')
                if 'Tester:' in text:
                    new_text = f"Tester: {self.header_data['common']['tester']}"
                    widget.config(text=new_text)
            
            # Recurse through children
            for child in widget.winfo_children():
                self.update_header_labels_recursive(child)
                
        except Exception as e:
            debug_print(f"DEBUG: Error updating label: {e}")

    def apply_header_changes_to_file(self):
        """Apply header changes to the Excel file using the new sample-specific structure."""
        print("DEBUG: Applying header changes to Excel file")
    
        # For .vap3 files, we don't update the physical file directly
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            print("DEBUG: .vap3 file detected, updating header data in memory only")
            # The header data is already updated in self.header_data
            # Just mark that we have changes
            self.mark_unsaved_changes()
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path)
        
            if self.test_name not in wb.sheetnames:
                print(f"DEBUG: Sheet {self.test_name} not found")
                return
            
            ws = wb[self.test_name]
        
            # Get common data and samples data
            common_data = self.header_data.get('common', {})
            samples_data = self.header_data.get('samples', [])
        
            # Apply sample-specific data for each sample block
            num_samples = self.header_data.get('num_samples', 0)
            for i in range(num_samples):
                col_offset = i * 12
                sample_data = samples_data[i] if i < len(samples_data) else {}
            
                print(f"DEBUG: Applying header changes for sample {i+1} with offset {col_offset}")
            
                # Row 1, Column F (6) + offset: Sample ID  
                sample_id = sample_data.get('id', f'Sample {i+1}')
                ws.cell(row=1, column=6 + col_offset, value=sample_id)
            
                # Row 2, Column D (4) + offset: Resistance
                resistance = sample_data.get("resistance", "")
                if resistance:
                    try:
                        resistance_value = float(resistance)
                        ws.cell(row=2, column=4 + col_offset, value=resistance_value)
                    except ValueError:
                        ws.cell(row=2, column=4 + col_offset, value=resistance)
            
                # Row 3, Column D (4) + offset: Tester name
                tester_name = common_data.get("tester", "")
                if tester_name:
                    ws.cell(row=3, column=4 + col_offset, value=tester_name)
            
                # Row 2, Column B (2) + offset: Media
                media = sample_data.get("media", "")
                if media:
                    ws.cell(row=2, column=2 + col_offset, value=media)
            
                # Row 3, Column B (2) + offset: Viscosity
                viscosity = sample_data.get("viscosity", "")
                if viscosity:
                    try:
                        viscosity_value = float(viscosity)
                        ws.cell(row=3, column=2 + col_offset, value=viscosity_value)
                    except ValueError:
                        ws.cell(row=3, column=2 + col_offset, value=viscosity)
            
                # Row 3, Column F (6) + offset: Voltage
                voltage = sample_data.get("voltage", "")
                if voltage:
                    try:
                        voltage_value = float(voltage)
                        ws.cell(row=3, column=6 + col_offset, value=voltage_value)
                    except ValueError:
                        ws.cell(row=3, column=6 + col_offset, value=voltage)

                # Calculate and apply power to Row 2, Column F (6) + offset
                resistance = sample_data.get("resistance", "")
                device_type = self.header_data.get('common', {}).get('device_type', 'EVO')

                # Device type mapping for dR values
                device_dr_mapping = {
                    'T58G': 0.9,
                    'EVO': 0.15,
                    'EVOMAX': 0.15,
                    'T28': 0.1,
                    'T51': 0.8,
                    'other': 0.15,
                    None: 0.0
                }

                try:
                    if voltage and resistance:
                        voltage_val = float(voltage)
                        resistance_val = float(resistance)
                        dr_value = device_dr_mapping.get(device_type, 0.0)

                        print(f"DEBUG: Voltage: {voltage_val}V, Resistance: {resistance_val}Ω")
                        print(f"DEBUG: Device type lookup: '{device_type}' -> dR = {dr_value}")

        
                        # Explicit handling for None device type (backwards compatibility)
                        if device_type is None:
                            dr_value = 0.0
                            print(f"DEBUG: Using dR = 0 for backwards compatibility (device_type is None)")
        
                        calculated_power = (voltage_val ** 2) / (resistance_val + dr_value)
                        ws.cell(row=2, column=6 + col_offset, value=calculated_power)
                        print(f"DEBUG: Calculated and applied power {calculated_power:.3f}W for sample {i+1} (V={voltage_val}, R={resistance_val}, dR={dr_value}, device_type={device_type})")
                    else:
                        print(f"DEBUG: Cannot calculate power for sample {i+1} - missing voltage or resistance")
                except (ValueError, TypeError) as e:
                    print(f"DEBUG: Error calculating power for sample {i+1}: {e}")
            
                # Row 3, Column H (8) + offset: Oil Mass
                oil_mass = sample_data.get("oil_mass", "")
                if oil_mass:
                    try:
                        oil_mass_value = float(oil_mass)
                        ws.cell(row=3, column=8 + col_offset, value=oil_mass_value)
                    except ValueError:
                        ws.cell(row=3, column=8 + col_offset, value=oil_mass)
            
                # Row 2, Column H (8) + offset: Puffing Regime
                puffing_regime = sample_data.get("puffing_regime", "60mL/3s/30s")
                if puffing_regime:
                    ws.cell(row=2, column=8 + col_offset, value=puffing_regime)
        
            # Save the workbook
            wb.save(self.file_path)
            debug_print("DEBUG: Header changes applied successfully to Excel file")
        
        except Exception as e:
            debug_print(f"DEBUG: Error applying header changes to file: {e}")
            # For .vap3 files, this is expected - just update in memory
            if self.file_path.endswith('.vap3'):
                debug_print("DEBUG: .vap3 file format detected, header changes stored in memory only")
            else:
                # For actual Excel files, show the error
                from tkinter import messagebox
                messagebox.showerror("Error", f"Could not update header data in file: {e}")

    def show_temp_status_message(self, message, duration_ms):
        """Show a temporary status message."""
        # Create a temporary label
        temp_label = ttk.Label(
            self.window,
            text=message,
            font=("Arial", 10),
            foreground="green",
            background="SystemButtonFace"
        )
        temp_label.pack(side="bottom", fill="x")
        
        # Remove it after the specified duration
        self.window.after(duration_ms, lambda: temp_label.destroy() if temp_label.winfo_exists() else None)

    def clear_current_sample(self):
        """Clear data for the currently selected sample."""
        current_tab = self.notebook.index(self.notebook.select())
        sample_id = f"Sample {current_tab + 1}"

        if messagebox.askyesno("Confirm Clear", f"Are you sure you want to clear all data for {sample_id}?"):
            # Clear all data except puffs
            clear_keys = ["before_weight", "after_weight", "draw_pressure", "smell", "notes", "tpm"]
    
            # Add chronography for User Test Simulation
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                clear_keys.append("chronography")
            else:
                # Add resistance and clog for standard tests
                clear_keys.extend(["resistance", "clog"])
    
            for key in clear_keys:
                for i in range(len(self.data[sample_id][key])):
                    if key == "tpm":
                        self.data[sample_id][key][i] = None
                    else:
                        self.data[sample_id][key][i] = ""
    
            # Update the display
            if hasattr(self, 'sample_sheets') and current_tab < len(self.sample_sheets):
                sheet = self.sample_sheets[current_tab]
                self.update_tksheet(sheet, sample_id)
            self.update_stats_panel()
            self.mark_unsaved_changes()
    
            debug_print(f"DEBUG: Cleared data for {sample_id}")

    def clear_all_data(self):
        """Clear all data for all samples."""
        if messagebox.askyesno("Confirm Clear All", 
                             "Are you sure you want to clear ALL data for ALL samples?"):
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                clear_keys = ["before_weight", "after_weight", "draw_pressure", "smell", "notes", "tpm"]
        
                # Add chronography for User Test Simulation
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    clear_keys.append("chronography")
        
                for key in clear_keys:
                    for j in range(len(self.data[sample_id][key])):
                        if key == "tpm":
                            self.data[sample_id][key][j] = None
                        else:
                            self.data[sample_id][key][j] = ""
        
                # Update the display for this sample
                if hasattr(self, 'sample_sheets') and i < len(self.sample_sheets):
                    sheet = self.sample_sheets[i]
                    self.update_tksheet(sheet, sample_id)
    
            self.update_stats_panel()
            self.mark_unsaved_changes()
            debug_print("DEBUG: Cleared all data for all samples")
    
    def add_row(self):
        """Add a new row to all samples."""
        debug_print("DEBUG: Adding new row to all samples")
    
        # Get the current sample instead of checking all samples
        try:
            current_tab_index = self.notebook.index(self.notebook.select())
            current_sample_id = f"Sample {current_tab_index + 1}"
            debug_print(f"DEBUG: Current sample for add row: {current_sample_id}")
        except:
            current_sample_id = "Sample 1"
            current_tab_index = 0
            debug_print("DEBUG: Defaulting to Sample 1 for add row")
    
        # Find the last non-empty puff value in the current sample
        last_puff = 0
        for puff_val in reversed(self.data[current_sample_id]["puffs"]):
            if puff_val and str(puff_val).strip():
                try:
                    last_puff = int(float(puff_val))
                    debug_print(f"DEBUG: Found last puff in {current_sample_id}: {last_puff}")
                    break
                except (ValueError, TypeError):
                    continue
    
        # Calculate new puff value for current sample
        new_puff = last_puff + self.puff_interval
        debug_print(f"DEBUG: Calculated new_puff: {new_puff} (last puff: {last_puff}, interval: {self.puff_interval})")
    
        # Add row to all samples, but use appropriate puff values for each
        for i in range(self.num_samples):
            sample_id = f"Sample {i + 1}"
        
            # Ensure all required fields exist in the data structure (defensive check)
            required_fields = ["puffs", "before_weight", "after_weight", "draw_pressure", "resistance", "smell", "clog", "notes", "tpm"]
            for field in required_fields:
                if field not in self.data[sample_id]:
                    self.data[sample_id][field] = []
                    debug_print(f"DEBUG: Added missing field '{field}' to {sample_id}")
        
            # Add chronography for User Test Simulation if needed
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                if "chronography" not in self.data[sample_id]:
                    self.data[sample_id]["chronography"] = []
        
            if sample_id == current_sample_id:
                # Use calculated puff for current sample
                self.data[sample_id]["puffs"].append(new_puff)
            else:
                # For other samples, find their last puff and add interval
                last_sample_puff = 0
                for puff_val in reversed(self.data[sample_id]["puffs"]):
                    if puff_val and str(puff_val).strip():
                        try:
                            last_sample_puff = int(float(puff_val))
                            break
                        except (ValueError, TypeError):
                            continue
        
                other_sample_new_puff = last_sample_puff + self.puff_interval
                self.data[sample_id]["puffs"].append(other_sample_new_puff)
                debug_print(f"DEBUG: Added puff {other_sample_new_puff} to {sample_id}")
        
            # Add empty values for other columns
            self.data[sample_id]["before_weight"].append("")
            self.data[sample_id]["after_weight"].append("")
            self.data[sample_id]["draw_pressure"].append("")
            self.data[sample_id]["smell"].append("")
            self.data[sample_id]["notes"].append("")
            self.data[sample_id]["tpm"].append(None)
        
            # Add resistance and clog for standard tests
            if self.test_name not in ["User Test Simulation", "User Simulation Test"]:
                self.data[sample_id]["resistance"].append("")
                self.data[sample_id]["clog"].append("")
    
            # Add chronography for User Test Simulation
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                self.data[sample_id]["chronography"].append("")
    
        # Update all tksheets to show the new row
        for i in range(self.num_samples):
            sample_id = f"Sample {i + 1}"
            self.update_tksheet(sample_id)
    
        # Update TPM calculations and stats
        self.update_all_tpm_calculations()
    
        # Mark as having unsaved changes
        self.mark_unsaved_changes()
    
        debug_print(f"DEBUG: Added new row with puff count {new_puff} to {current_sample_id}")

    def add_multiple_rows(self):
        """Add multiple rows to all samples with user input."""
        debug_print("DEBUG: Adding multiple rows to all samples")
    
        # Create a dialog to get the number of rows
        dialog = tk.Toplevel(self.window)
        dialog.title("Add Multiple Rows")
        dialog.geometry("300x150")
        dialog.transient(self.window)
        dialog.grab_set()
        dialog.resizable(False, False)
    
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
    
        # Result variable
        result = {"rows": 0, "confirmed": False}
    
        # Create the dialog content
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill="both", expand=True)
    
        # Label
        ttk.Label(main_frame, text="Number of rows to add:", font=("Arial", 10)).pack(pady=(0, 10))
    
        # Entry with validation
        rows_var = tk.IntVar(value=1)
        entry = ttk.Spinbox(
            main_frame, 
            from_=1, 
            to=100, 
            textvariable=rows_var, 
            width=10,
            justify="center"
        )
        entry.pack(pady=(0, 20))
        entry.focus_set()
    
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")
    
        def on_ok():
            try:
                num_rows = rows_var.get()
                if num_rows < 1:
                    messagebox.showerror("Invalid Input", "Number of rows must be at least 1.")
                    return
                result["rows"] = num_rows
                result["confirmed"] = True
                dialog.destroy()
            except tk.TclError:
                messagebox.showerror("Invalid Input", "Please enter a valid number.")
    
        def on_cancel():
            result["confirmed"] = False
            dialog.destroy()
    
        # Buttons
        ttk.Button(button_frame, text="Cancel", command=on_cancel).pack(side="right")
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side="right", padx=(0, 5))
    
        # Bind Enter key to OK
        dialog.bind('<Return>', lambda e: on_ok())
        dialog.bind('<Escape>', lambda e: on_cancel())
    
        # Wait for dialog to close
        dialog.wait_window()
    
        # If user confirmed, add the rows
        if result["confirmed"] and result["rows"] > 0:
            num_rows_to_add = result["rows"]
            debug_print(f"DEBUG: User requested to add {num_rows_to_add} rows")
        
            # Get the current sample instead of checking all samples
            try:
                current_tab_index = self.notebook.index(self.notebook.select())
                current_sample_id = f"Sample {current_tab_index + 1}"
                debug_print(f"DEBUG: Current sample for add multiple rows: {current_sample_id}")
            except:
                current_sample_id = "Sample 1"
                current_tab_index = 0
                debug_print("DEBUG: Defaulting to Sample 1 for add multiple rows")
        
            # Find the last non-empty puff value in the current sample
            last_puff = 0
            for puff_val in reversed(self.data[current_sample_id]["puffs"]):
                if puff_val and str(puff_val).strip():
                    try:
                        last_puff = int(float(puff_val))
                        debug_print(f"DEBUG: Found last puff in {current_sample_id}: {last_puff}")
                        break
                    except (ValueError, TypeError):
                        continue
        
            # Add the specified number of rows
            for row_num in range(num_rows_to_add):
                # Calculate new puff value for current sample
                new_puff = last_puff + ((row_num + 1) * self.puff_interval)
                debug_print(f"DEBUG: Adding row {row_num + 1}/{num_rows_to_add} with puff {new_puff}")
            
                # Add row to all samples
                for i in range(self.num_samples):
                    sample_id = f"Sample {i + 1}"
                
                    # Ensure all required fields exist in the data structure (defensive check)
                    required_fields = ["puffs", "before_weight", "after_weight", "draw_pressure", "resistance", "smell", "clog", "notes", "tpm"]
                    for field in required_fields:
                        if field not in self.data[sample_id]:
                            self.data[sample_id][field] = []
                            debug_print(f"DEBUG: Added missing field '{field}' to {sample_id}")
                
                    # Add chronography for User Test Simulation if needed
                    if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                        if "chronography" not in self.data[sample_id]:
                            self.data[sample_id]["chronography"] = []
                
                    if sample_id == current_sample_id:
                        # Use calculated puff for current sample
                        self.data[sample_id]["puffs"].append(new_puff)
                    else:
                        # For other samples, find their last puff and add interval
                        last_sample_puff = 0
                        for puff_val in reversed(self.data[sample_id]["puffs"]):
                            if puff_val and str(puff_val).strip():
                                try:
                                    last_sample_puff = int(float(puff_val))
                                    break
                                except (ValueError, TypeError):
                                    continue
                
                        other_sample_new_puff = last_sample_puff + ((row_num + 1) * self.puff_interval)
                        self.data[sample_id]["puffs"].append(other_sample_new_puff)
                
                    # Add empty values for other columns
                    self.data[sample_id]["before_weight"].append("")
                    self.data[sample_id]["after_weight"].append("")
                    self.data[sample_id]["draw_pressure"].append("")
                    self.data[sample_id]["smell"].append("")
                    self.data[sample_id]["notes"].append("")
                    self.data[sample_id]["tpm"].append(None)
                
                    # Add resistance and clog for standard tests
                    if self.test_name not in ["User Test Simulation", "User Simulation Test"]:
                        self.data[sample_id]["resistance"].append("")
                        self.data[sample_id]["clog"].append("")
            
                    # Add chronography for User Test Simulation
                    if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                        self.data[sample_id]["chronography"].append("")
        
            # Update all tksheets to show the new rows
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                self.update_tksheet(sample_id)
        
            # Update TPM calculations and stats
            self.update_all_tpm_calculations()
        
            # Mark as having unsaved changes
            self.mark_unsaved_changes()
        
            debug_print(f"DEBUG: Successfully added {num_rows_to_add} rows to all samples")
        
            # Show success message
            self.show_temp_status_message(f"Added {num_rows_to_add} rows to all samples", 3000)
        else:
            debug_print("DEBUG: Add multiple rows cancelled by user")

    def remove_last_row(self):
        """Remove the last row from all samples."""
        if messagebox.askyesno("Confirm Remove", "Remove the last row from all samples?"):
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                for key in self.data[sample_id]:
                    if self.data[sample_id][key]:  # Only remove if list is not empty
                        self.data[sample_id][key].pop()
            
                if hasattr(self, 'sample_sheets') and i < len(self.sample_sheets):
                    sheet = self.sample_sheets[i]
                    self.update_tksheet(sheet, sample_id)
        
            self.update_stats_panel()
            self.mark_unsaved_changes()
            debug_print("DEBUG: Removed last row from all samples")
    
    def recalculate_all_tpm(self):
        """Recalculate TPM for all samples."""
        debug_print("DEBUG: Recalculating TPM for all samples")
        for i in range(self.num_samples):
            sample_id = f"Sample {i + 1}"
            self.calculate_tpm(sample_id)
        
        self.update_stats_panel()
        self.mark_unsaved_changes()
        show_success_message("Recalculation Complete", "TPM values have been recalculated for all samples.", self.window)
    
    def go_to_sample_dialog(self):
        """Show dialog to jump to a specific sample."""
        sample_names = [f"Sample {i+1} - {self.header_data['samples'][i]['id']}" for i in range(self.num_samples)]
        
        dialog = tk.Toplevel(self.window)
        dialog.title("Go to Sample")
        dialog.geometry("300x150")
        dialog.transient(self.window)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Select sample:").pack(pady=10)
        
        selected_sample = tk.StringVar()
        combo = ttk.Combobox(dialog, textvariable=selected_sample, values=sample_names, state="readonly")
        combo.pack(pady=5)
        combo.set(sample_names[0])
        
        def go_to_selected():
            index = sample_names.index(selected_sample.get())
            self.notebook.select(index)
            dialog.destroy()
        
        ttk.Button(dialog, text="Go", command=go_to_selected).pack(pady=10)
    
    def change_puff_interval_dialog(self):
        """Show dialog to change the puff interval."""
        new_interval = simpledialog.askinteger(
            "Change Puff Interval",
            f"Current puff interval: {self.puff_interval}\nEnter new interval:",
            initialvalue=self.puff_interval,
            minvalue=1,
            maxvalue=1000
        )
        
        if new_interval and new_interval != self.puff_interval:
            self.puff_interval = new_interval
            self.puff_interval_var.set(new_interval)
            
            # Update puff values for future rows
            debug_print(f"DEBUG: Changed puff interval to {new_interval}")
    
    def auto_save_settings_dialog(self):
        """Show dialog to configure auto-save settings."""
        current_minutes = self.auto_save_interval / 60 / 1000
        
        new_minutes = simpledialog.askfloat(
            "Auto-Save Settings",
            f"Current auto-save interval: {current_minutes} minutes\nEnter new interval (minutes):",
            initialvalue=current_minutes,
            minvalue=0.5,
            maxvalue=60
        )
        
        if new_minutes:
            self.auto_save_interval = int(new_minutes * 60 * 1000)
            self.start_auto_save_timer()  # Restart with new interval
            debug_print(f"DEBUG: Changed auto-save interval to {new_minutes} minutes")
    
    def show_keyboard_shortcuts(self):
        """Show keyboard shortcuts help."""
        shortcuts = """
Keyboard Shortcuts:

Navigation:
• Tab - Move to next cell
• Shift+Tab - Move to previous cell
• Arrow Keys - Navigate between cells
• Ctrl+Left/Right - Switch samples
• Enter - Move down one row

Editing:
• Double-click - Edit cell
• Type - Start editing selected cell
• Enter - Confirm edit and move down
• Escape - Cancel edit

File Operations:
• Ctrl+S - Quick save

General:
• F1 - Show this help
        """
        
        help_window = tk.Toplevel(self.window)
        help_window.title("Keyboard Shortcuts")
        help_window.geometry("400x500")
        help_window.transient(self.window)
        
        text_widget = tk.Text(help_window, wrap=tk.WORD, padx=10, pady=10)
        text_widget.pack(fill="both", expand=True)
        text_widget.insert("1.0", shortcuts)
        text_widget.config(state="disabled")
    
    def show_about(self):
        """Show about dialog."""
        about_text = f"""
Data Collection Window
Version 3.0

Test: {self.test_name}
Samples: {self.num_samples}
Auto-save: Every {self.auto_save_interval/60/1000:.1f} minutes

Features:
• Real-time TPM calculation
• Auto-save functionality
• Excel and CSV export
• Keyboard navigation
• Comprehensive data validation

Developed by Charlie Becquet
        """
        show_success_message("About Data Collection", about_text, self.window)
    
    def _save_to_excel(self):
        """Save data to the appropriate file format."""
        debug_print(f"DEBUG: _save_to_excel() starting - file: {self.file_path}")
    
        # Check if this is a .vap3 file or temporary file
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: Detected .vap3 file or non-existent file, saving to loaded sheets")
            self._save_to_loaded_sheets()
        else:
            debug_print("DEBUG: Detected Excel file, saving using openpyxl")
            self._save_to_excel_file()

        debug_print("DEBUG: Excel save completed, updating main GUI...")

        # For .vap3 files, the data is already updated in memory via _save_to_loaded_sheets
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: .vap3 file detected, skipping file-based update")
            # The _save_to_loaded_sheets method should have already updated the main GUI data
        else:
            debug_print("DEBUG: Excel file detected, updating from file")
            self._update_excel_data_in_main_gui()

    
    def _save_to_excel_file(self):
        """Save data to the Excel file."""
        debug_print(f"DEBUG: _save_to_excel() starting - file: {self.file_path}")

        # Load the workbook
        wb = openpyxl.load_workbook(self.file_path)
        debug_print(f"DEBUG: Loaded workbook, sheets: {wb.sheetnames}")

        # Get the sheet for this test
        if self.test_name not in wb.sheetnames:
            raise Exception(f"Sheet '{self.test_name}' not found in the file.")
    
        ws = wb[self.test_name]
        debug_print(f"DEBUG: Opened sheet '{self.test_name}'")

        # Define green fill for TPM cells
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Determine column layout based on test type
        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
            columns_per_sample = 8  # Including chronography
            debug_print(f"DEBUG: Using User Test Simulation format with 8 columns per sample")
        else:
            columns_per_sample = 12  # Standard format
            debug_print(f"DEBUG: Using standard format with 12 columns per sample")

        # Track how much data we're actually writing
        total_data_written = 0

        # For each sample, write the data
        for sample_idx in range(self.num_samples):
            sample_id = f"Sample {sample_idx+1}"
    
            # Calculate column offset based on test type
            col_offset = sample_idx * columns_per_sample
    
            debug_print(f"DEBUG: Writing data for {sample_id} at column offset {col_offset}")
    
            sample_data_written = 0
    
            # Write the data starting at row 5
            for i, puff in enumerate(self.data[sample_id]["puffs"]):
                row = i + 5  # Row 5 is the first data row
        
                # Only write if we have actual data (not just empty rows)
                has_data = (
                    self.data[sample_id]["before_weight"][i] or
                    self.data[sample_id]["after_weight"][i] or
                    self.data[sample_id]["draw_pressure"][i] or
                    self.data[sample_id]["smell"][i] or
                    self.data[sample_id]["notes"][i] or
                    self.data[sample_id]["tpm"][i] is not None
                )
            
                # For User Test Simulation, also check chronography
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    has_data = has_data or (i < len(self.data[sample_id]["chronography"]) and self.data[sample_id]["chronography"][i])
        
                if not has_data:
                    continue  # Skip empty rows
            
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    # User Test Simulation column layout (8 columns)
                    # Chronography column (A + offset)
                    if i < len(self.data[sample_id]["chronography"]) and self.data[sample_id]["chronography"][i]:
                        ws.cell(row=row, column=1 + col_offset, value=str(self.data[sample_id]["chronography"][i]))
                
                    # Puffs column (B + offset) 
                    ws.cell(row=row, column=2 + col_offset, value=puff)
                
                    # Before weight column (C + offset)
                    if self.data[sample_id]["before_weight"][i]:
                        try:
                            ws.cell(row=row, column=3 + col_offset, value=float(self.data[sample_id]["before_weight"][i]))
                        except:
                            ws.cell(row=row, column=3 + col_offset, value=self.data[sample_id]["before_weight"][i])
                
                    # After weight column (D + offset)
                    if self.data[sample_id]["after_weight"][i]:
                        try:
                            ws.cell(row=row, column=4 + col_offset, value=float(self.data[sample_id]["after_weight"][i]))
                        except:
                            ws.cell(row=row, column=4 + col_offset, value=self.data[sample_id]["after_weight"][i])
                
                    # Draw pressure column (E + offset)
                    if self.data[sample_id]["draw_pressure"][i]:
                        try:
                            ws.cell(row=row, column=5 + col_offset, value=float(self.data[sample_id]["draw_pressure"][i]))
                        except:
                            ws.cell(row=row, column=5 + col_offset, value=self.data[sample_id]["draw_pressure"][i])
                
                    # Skip resistance column (F + offset) - not used in User Test Simulation
                
                    # Failure column (G + offset)
                    if self.data[sample_id]["smell"][i]:
                        try:
                            ws.cell(row=row, column=6 + col_offset, value=float(self.data[sample_id]["smell"][i]))
                        except:
                            ws.cell(row=row, column=6 + col_offset, value=self.data[sample_id]["smell"][i])
                
                    # Notes column (H + offset)
                    if self.data[sample_id]["notes"][i]:
                        ws.cell(row=row, column=7 + col_offset, value=str(self.data[sample_id]["notes"][i]))
                    
                    debug_print(f"DEBUG: Saved User Test Simulation row {i} for {sample_id}")
                
                else:
                    # Standard column layout (12 columns)
                    # Puffs column (A + offset)
                    ws.cell(row=row, column=1 + col_offset, value=puff)
        
                    # Before weight column (B + offset)
                    if self.data[sample_id]["before_weight"][i]:
                        try:
                            ws.cell(row=row, column=2 + col_offset, value=float(self.data[sample_id]["before_weight"][i]))
                        except:
                            ws.cell(row=row, column=2 + col_offset, value=self.data[sample_id]["before_weight"][i])
        
                    # After weight column (C + offset)
                    if self.data[sample_id]["after_weight"][i]:
                        try:
                            ws.cell(row=row, column=3 + col_offset, value=float(self.data[sample_id]["after_weight"][i]))
                        except:
                            ws.cell(row=row, column=3 + col_offset, value=self.data[sample_id]["after_weight"][i])
        
                    # Draw pressure column (D + offset)
                    if self.data[sample_id]["draw_pressure"][i]:
                        try:
                            ws.cell(row=row, column=4 + col_offset, value=float(self.data[sample_id]["draw_pressure"][i]))
                        except:
                            ws.cell(row=row, column=4 + col_offset, value=self.data[sample_id]["draw_pressure"][i])
        
                    if self.data[sample_id]["resistance"][i]:
                        try:
                            ws.cell(row=row, column=5 + col_offset, value=float(self.data[sample_id]["resistance"][i]))
                        except:
                            ws.cell(row=row, column=5 + col_offset, value=self.data[sample_id]["resistance"][i])

                    # Smell column (F + offset)
                    if self.data[sample_id]["smell"][i]:
                        try:
                            ws.cell(row=row, column=6 + col_offset, value=float(self.data[sample_id]["smell"][i]))
                        except:
                            ws.cell(row=row, column=6 + col_offset, value=self.data[sample_id]["smell"][i])

                    if self.data[sample_id]["clog"][i]:
                        try:
                            ws.cell(row=row, column=7 + col_offset, value=float(self.data[sample_id]["clog"][i]))
                        except:
                            ws.cell(row=row, column=7 + col_offset, value=self.data[sample_id]["clog"][i])
        
                    # Notes column (H + offset)
                    if self.data[sample_id]["notes"][i]:
                        ws.cell(row=row, column=8 + col_offset, value=str(self.data[sample_id]["notes"][i]))
        
                    # TPM column (I + offset) - if calculated
                    if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None:
                        tpm_cell = ws.cell(row=row, column=9 + col_offset, value=float(self.data[sample_id]["tpm"][i]))
                        tpm_cell.fill = green_fill
        
                sample_data_written += 1
        
            total_data_written += sample_data_written
            debug_print(f"DEBUG: Wrote {sample_data_written} data rows for {sample_id}")

        # Save the workbook
        debug_print(f"DEBUG: Saving workbook with {total_data_written} total data rows written...")

        debug_print("DEBUG: About to save Excel file - verifying data to be saved:")
        for sample_idx in range(min(2, self.num_samples)):  # Check first 2 samples
            sample_id = f"Sample {sample_idx+1}"
            col_offset = sample_idx * columns_per_sample
        
            debug_print(f"DEBUG: Sample {sample_idx+1} data preview:")
            for i in range(min(3, len(self.data[sample_id]["puffs"]))):  # First 3 rows
                row = i + 5
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    chrono_val = ws.cell(row=row, column=1 + col_offset).value
                    puff_val = ws.cell(row=row, column=2 + col_offset).value
                    before_val = ws.cell(row=row, column=3 + col_offset).value
                    after_val = ws.cell(row=row, column=4 + col_offset).value
                    debug_print(f"DEBUG:   Row {i}: Chrono={chrono_val}, Puff={puff_val}, Before={before_val}, After={after_val}")
                else:
                    puff_val = ws.cell(row=row, column=1 + col_offset).value
                    before_val = ws.cell(row=row, column=2 + col_offset).value
                    after_val = ws.cell(row=row, column=3 + col_offset).value
                    tpm_val = ws.cell(row=row, column=9 + col_offset).value
                    debug_print(f"DEBUG:   Row {i}: Puff={puff_val}, Before={before_val}, After={after_val}, TPM={tpm_val}")

        wb.save(self.file_path)
        debug_print(f"DEBUG: Excel file saved successfully to {self.file_path}")
    
    def _save_to_loaded_sheets(self):
        """Save data to the loaded sheets in memory."""
        debug_print("DEBUG: _save_to_loaded_sheets() starting")
    
        # Use original filename if available, otherwise fall back to file_path
        display_filename = self.original_filename if self.original_filename else os.path.basename(self.file_path)
    
        debug_print(f"DEBUG: Saving to loaded sheets with display filename: {display_filename}")
    
        try:
            # Find the correct file data in all_filtered_sheets
            
            current_file_data = None
        
            # For .vap3 files, the matching logic needs to be more flexible
            if self.file_path.endswith('.vap3'):
                debug_print("DEBUG: .vap3 file detected, using flexible matching")
            
                # Try multiple matching strategies for .vap3 files
                for file_data in self.parent.all_filtered_sheets:
                    # First try original filename match
                    if (self.original_filename and 
                        (file_data.get("original_filename") == self.original_filename or
                         file_data.get("database_filename") == self.original_filename or
                         file_data.get("file_name") == self.original_filename)):
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching .vap3 file by original filename: {self.original_filename}")
                        break
                
                    # Then try display filename match
                    if (display_filename and 
                        (file_data.get("file_name") == display_filename or
                         file_data.get("display_filename") == display_filename)):
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching .vap3 file by display filename: {display_filename}")
                        break
            
                # If still no match, just use the first loaded file for .vap3 files
                if not current_file_data and self.parent.all_filtered_sheets:
                    current_file_data = self.parent.all_filtered_sheets[0]
                    debug_print("DEBUG: Using first loaded file as fallback for .vap3 file")
            else:
                # Regular Excel file matching
                for file_data in self.parent.all_filtered_sheets:
                    if file_data.get("file_path") == self.file_path:
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching Excel file: {self.file_path}")
                        break

            if not current_file_data:
                debug_print(f"ERROR: Could not find file data for {display_filename}")
                return

            debug_print(f"DEBUG: Saving data to loaded sheets for test: {self.test_name}")

            # Check if the sheet exists in loaded data
            if not hasattr(self.parent, 'filtered_sheets') or self.test_name not in self.parent.filtered_sheets:
                debug_print(f"ERROR: Sheet {self.test_name} not found in loaded data")
                raise Exception(f"Sheet '{self.test_name}' not found in loaded data")
        
            sheet_data = self.parent.filtered_sheets[self.test_name]['data'].copy()
            debug_print(f"DEBUG: Found loaded sheet data with shape: {sheet_data.shape}")
    
            # Determine format based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8  # [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                debug_print("DEBUG: Saving in User Test Simulation format")
            else:
                columns_per_sample = 12  # [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM, etc.]
                debug_print("DEBUG: Saving in standard format")
    
            total_data_written = 0
    
            # Save data for each sample
            for sample_idx in range(self.num_samples):
                sample_id = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample
        
                debug_print(f"DEBUG: Saving data for {sample_id} with column offset {col_offset}")
        
                sample_data_written = 0
                data_start_row = 4  # Data starts at row 5 (0-indexed row 4)
        
                # Clear existing data in this sample's area first
                for clear_row in range(data_start_row, len(sheet_data)):
                    for clear_col in range(col_offset, col_offset + columns_per_sample):
                        if clear_col < len(sheet_data.columns):
                            sheet_data.iloc[clear_row, clear_col] = ""
        
                # Write new data for this sample
                for i in range(len(self.data[sample_id]["puffs"])):
                    data_row_idx = data_start_row + i
                
                    # Ensure we don't exceed the DataFrame bounds
                    if data_row_idx >= len(sheet_data):
                        break
                
                    try:
                        # Prepare data values based on test type
                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User Test Simulation format: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                            values_to_write = [
                                self.data[sample_id]["chronography"][i] if i < len(self.data[sample_id]["chronography"]) else "",
                                self.data[sample_id]["puffs"][i] if i < len(self.data[sample_id]["puffs"]) else "",
                                self.data[sample_id]["before_weight"][i] if i < len(self.data[sample_id]["before_weight"]) else "",
                                self.data[sample_id]["after_weight"][i] if i < len(self.data[sample_id]["after_weight"]) else "",
                                self.data[sample_id]["draw_pressure"][i] if i < len(self.data[sample_id]["draw_pressure"]) else "",
                                self.data[sample_id]["smell"][i] if i < len(self.data[sample_id]["smell"]) else "",  # "Failure" stored in smell field
                                self.data[sample_id]["notes"][i] if i < len(self.data[sample_id]["notes"]) else "",
                                self.data[sample_id]["tpm"][i] if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None else ""
                            ]
                    
                        else:
                            # Standard format: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
                            values_to_write = [
                                self.data[sample_id]["puffs"][i] if i < len(self.data[sample_id]["puffs"]) else "",
                                self.data[sample_id]["before_weight"][i] if i < len(self.data[sample_id]["before_weight"]) else "",
                                self.data[sample_id]["after_weight"][i] if i < len(self.data[sample_id]["after_weight"]) else "",
                                self.data[sample_id]["draw_pressure"][i] if i < len(self.data[sample_id]["draw_pressure"]) else "",
                                self.data[sample_id]["resistance"][i] if i < len(self.data[sample_id]["resistance"]) else "",
                                self.data[sample_id]["smell"][i] if i < len(self.data[sample_id]["smell"]) else "",
                                self.data[sample_id]["clog"][i] if i < len(self.data[sample_id]["clog"]) else "",
                                self.data[sample_id]["notes"][i] if i < len(self.data[sample_id]["notes"]) else "",
                                self.data[sample_id]["tpm"][i] if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None else ""
                            ]
                
                        # Write values to the appropriate columns
                        for col_idx, value in enumerate(values_to_write):
                            target_col = col_offset + col_idx
                            if target_col < len(sheet_data.columns):
                                sheet_data.iloc[data_row_idx, target_col] = value
                
                        sample_data_written += 1
                
                    except Exception as e:
                        debug_print(f"DEBUG: Error writing row {i} for {sample_id}: {e}")
                        continue
        
                total_data_written += sample_data_written
                debug_print(f"DEBUG: Wrote {sample_data_written} data rows for {sample_id}")
    
            # Update the loaded sheet data in memory
            self.parent.filtered_sheets[self.test_name]['data'] = sheet_data
            debug_print(f"DEBUG: Updated loaded sheet data with {total_data_written} total data rows")
    
            # Also update the UI state in all_filtered_sheets
            if hasattr(self.parent, 'all_filtered_sheets'):
                for file_data in self.parent.all_filtered_sheets:
                    if self.test_name in file_data.get('filtered_sheets', {}):
                        file_data['filtered_sheets'][self.test_name]['data'] = sheet_data.copy()
                        debug_print("DEBUG: Updated all_filtered_sheets with new data")
                        break
    
            if self.test_name in self.parent.filtered_sheets:
                self.parent.filtered_sheets[self.test_name]['header_data'] = self.header_data
                debug_print("DEBUG: Stored header data in filtered_sheets for .vap3 persistence")
    
               
            debug_print("DEBUG: Successfully saved data to loaded sheets")
    
            debug_print("DEBUG: Loaded sheets save completed - main GUI data should now be current")
    
            # Ensure the main GUI's current filtered_sheets reflects the updated data
            if hasattr(self.parent, 'filtered_sheets') and self.test_name in self.parent.filtered_sheets:
                # The data should already be updated, but let's ensure it's current
                debug_print(f"DEBUG: Main GUI filtered_sheets for {self.test_name} is current")

        except Exception as e:
            debug_print(f"ERROR: Failed to save data to loaded sheets: {e}")
            import traceback
            traceback.print_exc()
            raise e

    def _update_application_state(self):
        """Update the main application's state if this is a VAP3 file."""
        # Check if the parent has methods to update state
        if hasattr(self.parent, 'filtered_sheets') and hasattr(self.parent, 'file_path'):
            if self.parent.file_path and self.parent.file_path.endswith('.vap3'):
                debug_print("DEBUG: Updating VAP3 file and application state")
                try:
                    # Import here to avoid circular imports
                    from vap_file_manager import VapFileManager
                    
                    vap_manager = VapFileManager()
                    
                    # Get current application state
                    image_crop_states = getattr(self.parent, 'image_crop_states', {})
                    plot_settings = {
                        'selected_plot_type': getattr(self.parent, 'selected_plot_type', tk.StringVar()).get()
                    }
                    
                    # Save to VAP3 file
                    success = vap_manager.save_to_vap3(
                        self.parent.file_path,
                        self.parent.filtered_sheets,
                        getattr(self.parent, 'sheet_images', {}),
                        getattr(self.parent, 'plot_options', []),
                        image_crop_states,
                        plot_settings
                    )
                    
                    if success:
                        debug_print("DEBUG: VAP3 file updated successfully")
                    else:
                        debug_print("DEBUG: Failed to update VAP3 file")
                        
                except Exception as e:
                    debug_print(f"DEBUG: Error updating VAP3 file: {e}")
    
    def create_sample_tabs(self):
        """Create tabs for all samples in the notebook."""
        # This method manages the overall process
        self.sample_frames = []
        self.sample_sheets = []
    
        # Calls create_sample_tab() for EACH sample
        for i in range(self.num_samples):
            sample_id = f"Sample {i+1}"
            sample_frame = ttk.Frame(self.notebook, padding=10, style='TFrame')
            sample_name = self.header_data['samples'][i].get('id', f"Sample {i+1}")
        
            self.notebook.add(sample_frame, text=f"Sample {i+1} - {sample_name}")
            self.sample_frames.append(sample_frame)
        
            # Create individual tab content
            sheet = self.create_sample_tab(sample_frame, sample_id, i)
            self.sample_sheets.append(sheet)

    def create_sample_tab(self, parent_frame, sample_id, sample_index):
        """Create a tab for a single sample with tksheet instead of treeview."""
        debug_print(f"DEBUG: Creating sample tab for {sample_id}")
        debug_print(f"DEBUG: Test name: '{self.test_name}'")

        # Create main container that will hold everything
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill="both", expand=True)

        # Configure main container to expand properly
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(0, weight=1)

        # Determine columns based on test type
        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
            debug_print(f"DEBUG: Setting up User Test Simulation columns with chronography")
            headers = ["Chronography", "Puffs", "Before Weight", "After Weight", "Draw Pressure", "Failure", "Notes", "TPM"]
            self.tree_columns = len(headers)

            # Ensure chronography data exists
            if "chronography" not in self.data[sample_id]:
                debug_print(f"DEBUG: Creating chronography data structure for {sample_id}")
                self.data[sample_id]["chronography"] = []
                existing_length = len(self.data[sample_id]["puffs"])
                for j in range(existing_length):
                    self.data[sample_id]["chronography"].append("")
    
            debug_print(f"DEBUG: User Test Simulation columns: {headers}")
        else:
            headers = ["Puffs", "Before Weight", "After Weight", "Draw Pressure", "Resistance", "Smell", "Clog", "Notes", "TPM"]
            self.tree_columns = len(headers)
            debug_print(f"DEBUG: Standard test columns: {headers}")

        # Create the tksheet widget
        sheet = Sheet(
            main_container,
            headers=headers,
            height=400,
            width=900,
            show_table=True,
            show_top_left=True,
            show_row_index=True,
            show_header=True,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            empty_horizontal=0,
            empty_vertical=50
        )

        # Configure sheet appearance
        sheet.set_options(
            table_bg="#FFFFFF",
            table_fg="#000000",
            header_bg="#D0D0D0", 
            header_fg="#000000",
            index_bg="#F0F0F0",
            index_fg="#000000",
            selected_cells_bg="#CCE5FF",
            selected_cells_fg="#000000",
            table_selected_cells_border_fg="#4472C4",
            table_selected_cells_bg="#CCE5FF",
            table_selected_rows_bg="#E8F4FD",
            table_selected_rows_fg="#000000"
        )

        sheet.enable_bindings([
            "single_select",
            "drag_select", 
            "column_select",
            "row_select",
            "arrowkeys",
            "tab",
            "delete",
            "backspace",
            "f2",
            "edit_cell",
            "begin_edit_cell",
            "ctrl_c",
            "ctrl_v",
            "ctrl_x",
            "column_width_resize",
            "row_height_resize",
            "right_click_popup_menu"
        ])

        # Grid the sheet to fill the container
        sheet.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)

        # Make TPM column read-only and pin it to the right
        tpm_col_idx = len(headers) - 1
        sheet.readonly_columns([tpm_col_idx])

        # CENTER ALIGN all columns
        for i in range(len(headers)):
            sheet.align_columns([i], align="center")

        # Set up enhanced event bindings for Excel-like behavior
        self.setup_enhanced_sheet_bindings(sheet, sample_id, sample_index)

        # Populate sheet with initial data  
        self.populate_tksheet_data(sheet, sample_id)

        self.old_sheet_data = sheet.get_sheet_data()

        debug_print(f"DEBUG: Sample tab created for {sample_id} with {len(headers)} columns using tksheet")
        return sheet

    def end_editing(self):
        """End any active editing in tksheets."""
        # For tksheet, we don't need to explicitly end editing like with treeview
        # This method is kept for compatibility with the save process
        debug_print("DEBUG: Ending any active editing (tksheet)")
        pass

    def on_edit_complete(self, sheet, sample_id, row_idx):
        """Handle actions when user completes editing a cell (Enter, Tab, etc.)"""
        try:
            debug_print(f"DEBUG: Edit completed for {sample_id} at row {row_idx}")
        
            # Check if current row has an after weight and next row needs before weight
            if row_idx < len(self.data[sample_id]["after_weight"]) - 1:  # Not the last row
                current_after_weight = self.data[sample_id]["after_weight"][row_idx]
                next_before_weight = self.data[sample_id]["before_weight"][row_idx + 1] if (row_idx + 1) < len(self.data[sample_id]["before_weight"]) else ""
            
                # If current row has after weight and next row's before weight is empty
                if current_after_weight and str(current_after_weight).strip() != "" and (not next_before_weight or str(next_before_weight).strip() == ""):
                    debug_print(f"DEBUG: Edit completion - auto-progressing weight from row {row_idx} ({current_after_weight}) to row {row_idx + 1}")
                
                    # Update the next row's before weight in data structure
                    self.data[sample_id]["before_weight"][row_idx + 1] = current_after_weight
                
                    # Update the sheet display
                    before_weight_col_idx = 2 if self.test_name in ["User Test Simulation", "User Simulation Test"] else 1
                    sheet.set_cell_data(row_idx + 1, before_weight_col_idx, str(current_after_weight))
                
                    debug_print(f"DEBUG: Edit completion - auto-set next row before_weight to {current_after_weight}")
                
                    # Mark as changed and recalculate TPM
                    self.mark_unsaved_changes()
                    self.calculate_tpm(sample_id)
                    self.update_tpm_in_sheet(sheet, sample_id)
                
        except Exception as e:
            debug_print(f"DEBUG: Error in on_edit_complete: {e}")

    def setup_enhanced_sheet_bindings(self, sheet, sample_id, sample_index):
        """Set up simplified event bindings for reliable cell editing."""
    
        # Simple event handler for when cells are modified
        def on_cell_modified(event):
            """Handle any cell modification."""
            try:
                debug_print(f"DEBUG: Cell modified in {sample_id}")
                # Small delay to ensure edit is complete
                self.window.after(50, lambda: self.process_sheet_changes(sheet, sample_id))
        
            except Exception as e:
                debug_print(f"DEBUG: Error in cell modified handler: {e}")

        # Handle selection changes and detect edit completion
        def on_selection_changed(event):
            try:
                selections = sheet.get_selected_cells()
                if selections:
                    new_row, col = selections[0]
                
                    # Check if we moved to a different row (edit completed on previous row)
                    if hasattr(self, '_current_edit_row') and self._current_edit_row is not None and self._current_edit_row != new_row:
                        debug_print(f"DEBUG: Edit completed, moved from row {self._current_edit_row} to row {new_row}")
                        # User moved to different row - edit on previous row is complete
                        self.on_edit_complete(sheet, sample_id, self._current_edit_row)
                
                    # Update current edit row
                    self._current_edit_row = new_row
                    debug_print(f"DEBUG: Selection changed to row {new_row}, col {col}")
                
            except Exception as e:
                debug_print(f"DEBUG: Error in selection handler: {e}")


        # Bind to the sheet's built-in events
        sheet.bind("<<SheetModified>>", on_cell_modified)
        sheet.bind("<<SheetSelectCell>>", on_selection_changed)
    
        debug_print(f"DEBUG: Simplified sheet bindings set up for {sample_id}")

    def process_sheet_changes(self, sheet, sample_id):
        """Process all changes to the sheet data."""
        debug_print(f"DEBUG: Processing sheet changes for {sample_id}")

        try:
            # Sync sheet data to internal structure
            self.sync_tksheet_to_data(sheet, sample_id)

            # Recalculate TPM
            self.calculate_tpm(sample_id)

            # Update TPM column in sheet
            self.update_tpm_in_sheet(sheet, sample_id)

            # Update stats panel
            self.window.after(100, self.update_stats_panel)

            # Mark as changed
            self.mark_unsaved_changes()

            # update old sheet data for next update
            self.old_sheet_data = sheet.get_sheet_data()

        except Exception as e:
            debug_print(f"DEBUG: Error processing sheet changes: {e}")


    def populate_tksheet_data(self, sheet, sample_id):
        """Populate tksheet with data from self.data"""
        debug_print(f"DEBUG: Populating tksheet for {sample_id}")
    
        data_rows = []
        data_length = len(self.data[sample_id]["puffs"])
    
        for i in range(data_length):
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                row = [
                    self.data[sample_id]["chronography"][i] if i < len(self.data[sample_id]["chronography"]) else "",
                    self.data[sample_id]["puffs"][i] if i < len(self.data[sample_id]["puffs"]) else "",
                    self.data[sample_id]["before_weight"][i] if i < len(self.data[sample_id]["before_weight"]) else "",
                    self.data[sample_id]["after_weight"][i] if i < len(self.data[sample_id]["after_weight"]) else "",
                    self.data[sample_id]["draw_pressure"][i] if i < len(self.data[sample_id]["draw_pressure"]) else "",
                    self.data[sample_id]["smell"][i] if i < len(self.data[sample_id]["smell"]) else "",  # Failure column
                    self.data[sample_id]["notes"][i] if i < len(self.data[sample_id]["notes"]) else "",
                    f"{self.data[sample_id]['tpm'][i]:.6f}" if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None else ""
                ]
            else:
                row = [
                    self.data[sample_id]["puffs"][i] if i < len(self.data[sample_id]["puffs"]) else "",
                    self.data[sample_id]["before_weight"][i] if i < len(self.data[sample_id]["before_weight"]) else "",
                    self.data[sample_id]["after_weight"][i] if i < len(self.data[sample_id]["after_weight"]) else "",
                    self.data[sample_id]["draw_pressure"][i] if i < len(self.data[sample_id]["draw_pressure"]) else "",
                    self.data[sample_id]["resistance"][i] if i < len(self.data[sample_id]["resistance"]) else "",
                    self.data[sample_id]["smell"][i] if i < len(self.data[sample_id]["smell"]) else "",
                    self.data[sample_id]["clog"][i] if i < len(self.data[sample_id]["clog"]) else "",
                    self.data[sample_id]["notes"][i] if i < len(self.data[sample_id]["notes"]) else "",
                    f"{self.data[sample_id]['tpm'][i]:.6f}" if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None else ""
                ]
            data_rows.append(row)
    
        # Add empty rows to reach the standard 50 rows
        while len(data_rows) < 50:
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                empty_row = ["", "", "", "", "", "", "", ""]
            else:
                empty_row = ["", "", "", "", "", "", "", "", ""]
            data_rows.append(empty_row)
    
        sheet.set_sheet_data(data_rows)
        sheet.set_all_cell_sizes_to_text(width=60)
        debug_print(f"DEBUG: Populated tksheet with {len(data_rows)} rows")

    def sync_tksheet_to_data(self, sheet, sample_id):
        """Sync tksheet data back to internal data structure"""
        debug_print(f"DEBUG: Syncing tksheet data to internal structure for {sample_id}")

        try:
            sheet_data = sheet.get_sheet_data()
            # Store old weight values to detect changes
            old_before_weights = self.data[sample_id]["before_weight"].copy()
            old_after_weights = self.data[sample_id]["after_weight"].copy()

            # Find last nonzero puff index AND value (for puff auto-population)
            last_nonzero_puff_index = None
            last_puff_value = 0
        
            for i, row in enumerate(sheet_data):
                try:
                    puff_val = float(row[0]) if row[0] not in (None, "") else 0
                    if puff_val != 0:
                        last_nonzero_puff_index = i
                        last_puff_value = int(puff_val)  # Store the actual value, not just index
                        debug_print(f"DEBUG: Found nonzero puff at row {i} with value {last_puff_value}")
                except (ValueError, IndexError):
                    continue

            # Find last nonzero AFTER WEIGHT index (for weight auto-progression) 
            last_nonzero_after_weight_index = None
        
            # Get correct column indices based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                before_weight_col = 2
                after_weight_col = 3
            else:
                before_weight_col = 1
                after_weight_col = 2
            
            for i, row in enumerate(sheet_data):
                try:
                    after_weight_val = str(row[after_weight_col]).strip() if len(row) > after_weight_col and row[after_weight_col] is not None else ""
                    if after_weight_val and after_weight_val != "":
                        last_nonzero_after_weight_index = i
                        debug_print(f"DEBUG: Found nonzero after weight at row {i} with value {after_weight_val}")
                except (ValueError, IndexError):
                    continue

            debug_print(f"DEBUG: Last puff index: {last_nonzero_puff_index}, Last after weight index: {last_nonzero_after_weight_index}")

            # Clear existing data arrays
            for key in self.data[sample_id]:
                if key not in ["current_row_index", "avg_tpm"]:
                    self.data[sample_id][key] = []

            # Track auto weight progression actions to execute later
            auto_weight_actions = []

            # Rebuild data from sheet
            for row_idx, row_data in enumerate(sheet_data):
                # Ensure row_data is a list and handle empty/None values
                if not isinstance(row_data, list):
                    row_data = list(row_data) if row_data is not None else []

                # Helper function to safely get cell value
                def safe_get_cell(data_row, col_idx):
                    if col_idx < len(data_row) and data_row[col_idx] is not None:
                        val = str(data_row[col_idx]).strip()
                        return val if val else ""
                    return ""

                # Detect before/after weight value change
                if self.old_sheet_data is not None:
                    old_row = self.old_sheet_data[row_idx] if row_idx < len(self.old_sheet_data) else []
                else:
                    old_row = sheet_data[row_idx] if row_idx < len(sheet_data) else []
            
                new_before = safe_get_cell(row_data, before_weight_col)
                new_after = safe_get_cell(row_data, after_weight_col)
                old_before = str(old_row[before_weight_col]).strip() if len(old_row) > before_weight_col and old_row[before_weight_col] is not None else ""
                old_after = str(old_row[after_weight_col]).strip() if len(old_row) > after_weight_col and old_row[after_weight_col] is not None else ""
                new_puffs = safe_get_cell(row_data, 0)

                # Check for auto weight progression condition - ONLY for after weight column changes
                should_trigger_auto_weight = False
                # CRITICAL: Only trigger if AFTER weight changed AND before weight did NOT change
                # This ensures we only trigger when user edited the after weight column, not before weight column
                after_weight_changed = (old_after == "" and new_after != "")
                before_weight_changed = (old_before == "" and new_before != "")
            
                if after_weight_changed and not before_weight_changed:  # Only after weight changed, not before weight
                    # Use AFTER WEIGHT index for weight progression logic
                    if last_nonzero_after_weight_index is None or row_idx >= last_nonzero_after_weight_index:
                        debug_print(f"DEBUG: AFTER weight column changed from empty to {new_after} at row {row_idx} - scheduling auto weight AND puff progression")
                        auto_weight_actions.append((row_idx, new_after))
                        should_trigger_auto_weight = True
                    else:
                        debug_print(f"DEBUG: After weight changed at row {row_idx} but this is before last after weight row {last_nonzero_after_weight_index} - no auto progression")
                elif before_weight_changed and not after_weight_changed:
                    debug_print(f"DEBUG: BEFORE weight column changed from empty to {new_before} at row {row_idx} - NO auto progression (only TPM update)")
                elif after_weight_changed and before_weight_changed:
                    debug_print(f"DEBUG: BOTH weight columns changed at row {row_idx} - this might be bulk data entry or auto-population, no auto progression")

                # Handle puff updates - USE PUFF INDEX for puff logic
                should_update_puffs = False
                debug_print(f"Should we update puffs? {old_row}, new before:{new_before}, new after:{new_after}")
            
                # Only update puffs if we're past the last known puffs row AND weights changed
                if last_nonzero_puff_index is not None and row_idx > last_nonzero_puff_index:
                    if (old_before == "" and new_before != "") or (old_after == "" and new_after != ""):
                        if new_puffs == "":
                            debug_print(f"Yes! New data detected at row {row_idx}, auto-filling puffs")
                            should_update_puffs = True
                        else:
                            debug_print(f"No - puffs already filled: {new_puffs}")
                    else:
                        debug_print(f"No - no weight changes detected")
                elif last_nonzero_puff_index is None:
                    # No existing puffs data, this is the first entry
                    if (old_before == "" and new_before != "") or (old_after == "" and new_after != ""):
                        if new_puffs == "":
                            debug_print(f"Yes! First data entry detected at row {row_idx}, auto-filling puffs")
                            should_update_puffs = True
                            # For first entry, set last_nonzero_puff_index to -1 so calculation works
                            last_nonzero_puff_index = -1
                            last_puff_value = 0
                else:
                    debug_print(f"No - editing existing puff data at row {row_idx} (last puffs row: {last_nonzero_puff_index})")
            
                if should_update_puffs:
                    # Calculate correct puff value: base + increments
                    puff_value = last_puff_value + (row_idx - last_nonzero_puff_index) * self.puff_interval
                    new_puffs = str(puff_value)
                    debug_print(f"DEBUG: Auto-filled puffs at row {row_idx} with value {new_puffs} (base: {last_puff_value}, distance: {row_idx - last_nonzero_puff_index}, interval: {self.puff_interval})")
                
                    # Update sheet visually
                    puff_col_idx = 1 if self.test_name in ["User Test Simulation", "User Simulation Test"] else 0
                    sheet.set_cell_data(row_idx, puff_col_idx, new_puffs)
                
                    # Populate all intermediate puff values from last_nonzero_puff_index + 1 to row_idx
                    for intermediate_row in range(last_nonzero_puff_index + 1, row_idx + 1):
                        if intermediate_row < len(sheet_data):
                            intermediate_puff_value = last_puff_value + (intermediate_row - last_nonzero_puff_index) * self.puff_interval
                            sheet.set_cell_data(intermediate_row, puff_col_idx, str(intermediate_puff_value))
                            debug_print(f"DEBUG: Auto-populated intermediate row {intermediate_row} with puffs value {intermediate_puff_value}")

                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    # Map columns to data structure for User Test Simulation
                    self.data[sample_id]["chronography"].append(safe_get_cell(row_data, 0))
                    self.data[sample_id]["puffs"].append(safe_get_cell(row_data, 1))
                    self.data[sample_id]["before_weight"].append(safe_get_cell(row_data, 2))
                    self.data[sample_id]["after_weight"].append(safe_get_cell(row_data, 3))
                    self.data[sample_id]["draw_pressure"].append(safe_get_cell(row_data, 4))
                    self.data[sample_id]["smell"].append(safe_get_cell(row_data, 5))  # Failure
                    self.data[sample_id]["notes"].append(safe_get_cell(row_data, 6))
                    self.data[sample_id]["tpm"].append(None)  # Will be recalculated
                else:
                    # Standard format mapping
                    self.data[sample_id]["puffs"].append(safe_get_cell(row_data, 0))
                    self.data[sample_id]["before_weight"].append(safe_get_cell(row_data, 1))
                    self.data[sample_id]["after_weight"].append(safe_get_cell(row_data, 2))
                    self.data[sample_id]["draw_pressure"].append(safe_get_cell(row_data, 3))
                    self.data[sample_id]["resistance"].append(safe_get_cell(row_data, 4))
                    self.data[sample_id]["smell"].append(safe_get_cell(row_data, 5))
                    self.data[sample_id]["clog"].append(safe_get_cell(row_data, 6))
                    self.data[sample_id]["notes"].append(safe_get_cell(row_data, 7))
                    self.data[sample_id]["tpm"].append(None)  # Will be recalculated

            # NOW execute auto weight progression actions after data arrays are fully rebuilt
            debug_print(f"DEBUG: Executing {len(auto_weight_actions)} auto weight progression actions")
            for action_row_idx, action_value in auto_weight_actions:
                debug_print(f"DEBUG: Executing auto weight progression for row {action_row_idx} with value {action_value}")
                # Update next row's before weight
                self.auto_progress_weight(sheet, sample_id, action_row_idx, action_value)
            
                # ALSO update next row's puffs if it doesn't have any
                next_row_idx = action_row_idx + 1
                if next_row_idx < len(self.data[sample_id]["puffs"]):
                    current_puffs = self.data[sample_id]["puffs"][next_row_idx] if next_row_idx < len(self.data[sample_id]["puffs"]) else ""
                    if not current_puffs or str(current_puffs).strip() == "":
                        # Calculate next puff value
                        current_puff_value = int(self.data[sample_id]["puffs"][action_row_idx]) if action_row_idx < len(self.data[sample_id]["puffs"]) and self.data[sample_id]["puffs"][action_row_idx] else 0
                        next_puff_value = current_puff_value + self.puff_interval
                    
                        # Update data structure
                        self.data[sample_id]["puffs"][next_row_idx] = next_puff_value
                    
                        # Update sheet display
                        puff_col_idx = 1 if self.test_name in ["User Test Simulation", "User Simulation Test"] else 0
                        sheet.set_cell_data(next_row_idx, puff_col_idx, str(next_puff_value))
                    
                        debug_print(f"DEBUG: Auto weight progression also updated next row puffs: row {next_row_idx} = {next_puff_value}")

            # Check if any weights changed
            weights_changed = (
                old_before_weights != self.data[sample_id]["before_weight"] or
                old_after_weights != self.data[sample_id]["after_weight"]
            )

            if weights_changed:
                debug_print(f"DEBUG: Weight values changed, triggering TPM recalculation")
                # Force immediate TPM calculation
                self.calculate_tpm(sample_id)
                # Force immediate plot update
                self.update_stats_panel()

            debug_print(f"DEBUG: Synced {len(sheet_data)} rows to internal data structure")

        except Exception as e:
            debug_print(f"DEBUG: Error syncing tksheet data: {e}")
            import traceback
            traceback.print_exc()

    def update_tpm_in_sheet(self, sheet, sample_id):
        """Update the TPM column in the sheet with calculated values"""
        # Get the number of headers to find TPM column index
        debug_print(f"DEBUG: attempting to calculate sheet header length")
        if hasattr(sheet, 'columns'):
            tpm_col_idx = len(sheet.columns) - 1  # TPM is always the last column
            debug_print(f"DEBUG: calculated sheet header length")
        else:
            # Fallback if headers attribute not available
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                tpm_col_idx = 7  # TPM column index for user simulation
            else:
                tpm_col_idx = 8  # TPM column index for standard tests
    
        for row_idx in range(len(self.data[sample_id]["tpm"])):
            tpm_value = self.data[sample_id]["tpm"][row_idx]
            if tpm_value is not None:
                sheet.set_cell_data(row_idx, tpm_col_idx, f"{tpm_value:.6f}")
            else:
                sheet.set_cell_data(row_idx, tpm_col_idx, "")

    def update_tksheet(self, sheet, sample_id):
        """Update the tksheet with current data (replaces update_treeview)"""
        debug_print(f"DEBUG: Updating tksheet for {sample_id}")
    
        try:
            # Populate the sheet with current data
            self.populate_tksheet_data(sheet, sample_id)
        
            # Highlight TPM cells that have calculated values
            #self.highlight_tpm_cells(sheet, sample_id)
        
            debug_print(f"DEBUG: tksheet updated for {sample_id}")
        except Exception as e:
            debug_print(f"DEBUG: Error updating tksheet: {e}")

    def highlight_tpm_cells(self, sheet, sample_id):
        """Highlight TPM cells that have calculated values (like the green highlighting in treeview)"""
        # Get TPM column index
        if hasattr(sheet, 'headers'):
            tpm_col_idx = len(sheet.headers) - 1
        else:
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                tpm_col_idx = 7
            else:
                tpm_col_idx = 8
    
        for row_idx in range(len(self.data[sample_id]["tpm"])):
            tpm_value = self.data[sample_id]["tpm"][row_idx]
            if tpm_value is not None:
                # Highlight the cell with green background (like your original TPM highlighting)
                sheet.highlight_cells(row=row_idx, column=tpm_col_idx, bg="#C6EFCE", fg="black")

    def refresh_main_gui_after_save(self):
        """Update the main GUI data structures after saving data."""
        try:
            debug_print("DEBUG: Refreshing main GUI after data collection save")
        
            # For .vap3 files or database files, the data is already updated in memory
            # We just need to ensure the main GUI reflects the current state
            if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
                self._update_vap3_data_in_main_gui()
            else:
                # For regular Excel files, reload from file
                self._update_excel_data_in_main_gui()
            
            # Mark this file as modified in the staging area
            self._mark_file_as_modified()
        
            debug_print("DEBUG: Main GUI refresh completed")
        
        except Exception as e:
            debug_print(f"ERROR: Failed to refresh main GUI: {e}")
            import traceback
            traceback.print_exc()

    def _update_excel_data_in_main_gui(self):
        """Update main GUI data for Excel files."""
        try:
            # Only do this for actual Excel files that exist
            if not os.path.exists(self.file_path) or not self.file_path.endswith(('.xlsx', '.xls')):
                debug_print("DEBUG: File doesn't exist or isn't Excel, skipping Excel update")
                return
            
            # Reload just this sheet from the Excel file
            import pandas as pd
            new_sheet_data = pd.read_excel(self.file_path, sheet_name=self.test_name)
        
            # Update the main GUI's filtered_sheets
            if hasattr(self.parent, 'filtered_sheets') and self.test_name in self.parent.filtered_sheets:
                self.parent.filtered_sheets[self.test_name]['data'] = new_sheet_data
                self.parent.filtered_sheets[self.test_name]['is_empty'] = new_sheet_data.empty
                debug_print(f"DEBUG: Updated sheet {self.test_name} in main GUI filtered_sheets")
        
            # Update all_filtered_sheets for the current file
            if hasattr(self.parent, 'all_filtered_sheets') and hasattr(self.parent, 'current_file'):
                for file_data in self.parent.all_filtered_sheets:
                    if file_data["file_name"] == self.parent.current_file:
                        if self.test_name in file_data["filtered_sheets"]:
                            file_data["filtered_sheets"][self.test_name]['data'] = new_sheet_data
                            file_data["filtered_sheets"][self.test_name]['is_empty'] = new_sheet_data.empty
                            debug_print(f"DEBUG: Updated sheet {self.test_name} in all_filtered_sheets for file {self.parent.current_file}")
                        break
        
        except Exception as e:
            debug_print(f"ERROR: Failed to update Excel data in main GUI: {e}")
            raise

    def _update_vap3_data_in_main_gui(self):
        """Update main GUI data for .vap3 files - data is already in memory."""
        try:
            debug_print("DEBUG: Updating .vap3 data in main GUI - data should already be current")
        
            # For .vap3 files, the data should have already been updated in the save process
            # The _save_to_loaded_sheets method should have updated the main GUI data structures
        
            # Just ensure the all_filtered_sheets stays synchronized
            if hasattr(self.parent, 'filtered_sheets') and hasattr(self.parent, 'all_filtered_sheets'):
                current_file = getattr(self.parent, 'current_file', None)
                if current_file:
                    for file_data in self.parent.all_filtered_sheets:
                        if file_data["file_name"] == current_file:
                            # Ensure the all_filtered_sheets data matches current filtered_sheets
                            file_data["filtered_sheets"] = copy.deepcopy(self.parent.filtered_sheets)
                            debug_print(f"DEBUG: Synchronized .vap3 data for file {current_file}")
                            break
        
            debug_print("DEBUG: .vap3 data update completed")
        
        except Exception as e:
            debug_print(f"ERROR: Failed to update .vap3 data in main GUI: {e}")
            raise

    def _mark_file_as_modified(self):
        """Mark the current file as modified in the staging area."""
        try:
            if hasattr(self.parent, 'all_filtered_sheets') and hasattr(self.parent, 'current_file'):
                for file_data in self.parent.all_filtered_sheets:
                    if file_data["file_name"] == self.parent.current_file:
                        file_data["is_modified"] = True
                        file_data["last_modified"] = time.time()
                        debug_print(f"DEBUG: Marked file {self.parent.current_file} as modified")
                        break
        
            # Update the main GUI window title to show modified status
            if hasattr(self.parent, 'root') and hasattr(self.parent, 'current_file'):
                current_title = self.parent.root.title()
                if not current_title.endswith(" *"):
                    self.parent.root.title(current_title + " *")
                    debug_print("DEBUG: Updated window title to show modified status")
                
        except Exception as e:
            debug_print(f"ERROR: Failed to mark file as modified: {e}")

    def calculate_tpm(self, sample_id):
        """Calculate TPM (Total Particulate Matter) for all rows with valid data."""
        valid_tpm_values = []
    
        for i in range(len(self.data[sample_id]["puffs"])):
            try:
                # Get weight values
                before_weight_str = self.data[sample_id]["before_weight"][i]
                after_weight_str = self.data[sample_id]["after_weight"][i]
            
                # Skip if either weight is missing
                if not before_weight_str or not after_weight_str:
                    continue
                
                # Convert to float
                before_weight = float(before_weight_str)
                after_weight = float(after_weight_str)
            
                # Validate weights
                if before_weight <= after_weight:
                    continue
                
                # Calculate puffs in this interval
                
                puff_interval = int(self.data[sample_id]["puffs"][i])
              
                puffs_in_interval = 10 # default 10
            
                if i > 0:
                    prev_puff = int(self.data[sample_id]["puffs"][i - 1])
                    puffs_in_interval = puff_interval - prev_puff
                
                # Skip if invalid puff interval
                if puffs_in_interval <= 0:
                    continue
                
                # Calculate TPM (mg/puff)
                weight_consumed = before_weight - after_weight  # in grams
                tpm = (weight_consumed * 1000) / puffs_in_interval  # Convert to mg per puff
            
                # Store result
                self.data[sample_id]["tpm"][i] = round(tpm, 3)
                valid_tpm_values.append(tpm)
                
            except Exception:
                # Ensure tpm list is long enough even for failed calculations
                while len(self.data[sample_id]["tpm"]) <= i:
                    self.data[sample_id]["tpm"].append(None)
    
        # Update average TPM
        self.data[sample_id]["avg_tpm"] = sum(valid_tpm_values) / len(valid_tpm_values) if valid_tpm_values else 0.0
    
        return len(valid_tpm_values) > 0

    def validate_weight_entry(self, sample_id, row_idx, column_name, value):
        """
        Validate weight entries to ensure data consistency.
        Returns True if valid, False otherwise.
        """
        try:
            if not value or not value.strip():
                return True  # Empty values are allowed
                
            weight_value = float(value.strip())
            
            # Check for reasonable weight values (between 0.001g and 100g)
            if weight_value < 0.001 or weight_value > 100:
                debug_print(f"WARNING: Weight value {weight_value}g seems unreasonable for row {row_idx}")
                return False
                
            # If this is an after_weight, check that it's less than before_weight
            if column_name == "after_weight":
                before_weight_str = self.data[sample_id]["before_weight"][row_idx]
                if before_weight_str and before_weight_str.strip():
                    try:
                        before_weight = float(before_weight_str.strip())
                        if weight_value >= before_weight:
                            debug_print(f"WARNING: After weight ({weight_value}g) should be less than before weight ({before_weight}g)")
                            return False
                    except ValueError:
                        pass
                        
            # If this is a before_weight, check that it's greater than after_weight
            elif column_name == "before_weight":
                after_weight_str = self.data[sample_id]["after_weight"][row_idx]
                if after_weight_str and after_weight_str.strip():
                    try:
                        after_weight = float(after_weight_str.strip())
                        if weight_value <= after_weight:
                            debug_print(f"WARNING: Before weight ({weight_value}g) should be greater than after weight ({after_weight}g)")
                            return False
                    except ValueError:
                        pass
                        
            return True
            
        except ValueError:
            debug_print(f"ERROR: Invalid weight value '{value}' - must be a number")
            return False
    
    def load_existing_data_from_loaded_sheets(self):
        """Load existing data from already-loaded sheet data (for .vap3 files)."""
        debug_print(f"DEBUG: Loading existing data from loaded sheets for test: {self.test_name}")
    
        try:
            # Get the loaded sheet data
            if not hasattr(self.parent, 'filtered_sheets') or self.test_name not in self.parent.filtered_sheets:
                debug_print(f"ERROR: Sheet {self.test_name} not found in loaded data")
                return
        
            sheet_data = self.parent.filtered_sheets[self.test_name]['data']
            debug_print(f"DEBUG: Found loaded sheet data with shape: {sheet_data.shape}")
    
            debug_print("DEBUG: First 10 rows and 16 columns of actual DataFrame data:")
            for i in range(min(10, len(sheet_data))):
                row_preview = []
                for j in range(min(16, len(sheet_data.columns))):
                    val = sheet_data.iloc[i, j]
                    if pd.isna(val):
                        row_preview.append("NaN")
                    else:
                        val_str = str(val)[:10]  # Truncate long values
                        row_preview.append(f"'{val_str}'")
                debug_print(f"  Row {i}: {row_preview}")

            # Determine the data format based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8
                debug_print("DEBUG: Loading User Test Simulation format")
            else:
                columns_per_sample = 12
                debug_print("DEBUG: Loading standard format")
    
            # Load data for each sample
            loaded_data_count = 0
        
            for sample_idx in range(self.num_samples):
                sample_id = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample
        
                debug_print(f"DEBUG: Loading data for {sample_id} with column offset {col_offset}")
            
                debug_print(f"DEBUG: Clearing existing template data for {sample_id}")
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    self.data[sample_id]["chronography"] = []
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["smell"] = []  
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []
                else:
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["resistance"] = []
                    self.data[sample_id]["smell"] = []
                    self.data[sample_id]["clog"] = []
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []
        
                # Load data starting from row 5 (DataFrame index 4)
                sample_had_data = False
                row_count = 0
        
                data_start_row = 4  # Data starts at row 5 (0-indexed row 4)
        
                for data_row_idx in range(data_start_row, min(len(sheet_data), 100)):  # Limit to reasonable number of rows
                    try:
                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User Test Simulation format: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                            chronography_col = 0 + col_offset
                            puffs_col = 1 + col_offset
                            before_weight_col = 2 + col_offset
                            after_weight_col = 3 + col_offset
                            draw_pressure_col = 4 + col_offset
                            failure_col = 5 + col_offset  # Stored in smell field
                            notes_col = 6 + col_offset
                            tpm_col = 7 + col_offset
                    
                            # Extract values if columns exist
                            values = {}
                    
                            # Helper function to clean values and strip quotes
                            def clean_value(raw_val):
                                if pd.isna(raw_val):
                                    return ""
                                val_str = str(raw_val).strip()
                                # Remove surrounding quotes if present
                                if val_str.startswith("'") and val_str.endswith("'"):
                                    val_str = val_str[1:-1]
                                elif val_str.startswith('"') and val_str.endswith('"'):
                                    val_str = val_str[1:-1]
                                return val_str
                        
                            debug_print(f"DEBUG: Processing row {data_row_idx} for {sample_id}")
                        
                            if chronography_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, chronography_col]
                                values['chronography'] = clean_value(raw_val)
                                debug_print(f"DEBUG: chronography raw: {raw_val}, cleaned: {values['chronography']}")
                            else:
                                values['chronography'] = ""
                        
                            if puffs_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, puffs_col]
                                values['puffs'] = clean_value(raw_val)
                                debug_print(f"DEBUG: puffs raw: {raw_val}, cleaned: {values['puffs']}")
                            else:
                                values['puffs'] = ""
                        
                            if before_weight_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, before_weight_col]
                                values['before_weight'] = clean_value(raw_val)
                                debug_print(f"DEBUG: before_weight raw: {raw_val}, cleaned: {values['before_weight']}")
                            else:
                                values['before_weight'] = ""
                        
                            if after_weight_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, after_weight_col]
                                values['after_weight'] = clean_value(raw_val)
                                debug_print(f"DEBUG: after_weight raw: {raw_val}, cleaned: {values['after_weight']}")
                            else:
                                values['after_weight'] = ""
                        
                            if draw_pressure_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, draw_pressure_col]
                                values['draw_pressure'] = clean_value(raw_val)
                            else:
                                values['draw_pressure'] = ""
                        
                            if failure_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, failure_col]
                                values['failure'] = clean_value(raw_val)
                            else:
                                values['failure'] = ""
                        
                            if notes_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, notes_col]
                                values['notes'] = clean_value(raw_val)
                            else:
                                values['notes'] = ""
                        
                            if tpm_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, tpm_col]
                                cleaned_val = clean_value(raw_val)
                                if cleaned_val and cleaned_val.strip():
                                    try:
                                        values['tpm'] = float(cleaned_val)
                                        debug_print(f"DEBUG: tpm raw: {raw_val}, cleaned: {cleaned_val}, float: {values['tpm']}")
                                    except (ValueError, TypeError):
                                        values['tpm'] = None
                                        debug_print(f"DEBUG: tpm conversion failed for: {cleaned_val}")
                                else:
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None
                    
                            # Check if this row has any meaningful data (not all empty)
                            if any([values['chronography'], values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Storing data for row {data_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                                # Append the data (building from scratch)
                                self.data[sample_id]["chronography"].append(values['chronography'])
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["smell"].append(values['failure']) 
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])
                        
                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Row {data_row_idx} has no meaningful data, skipping")
                        
                        else:
                            # Standard format: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
                            puffs_col = 0 + col_offset
                            before_weight_col = 1 + col_offset
                            after_weight_col = 2 + col_offset
                            draw_pressure_col = 3 + col_offset
                            resistance_col = 4 + col_offset
                            smell_col = 5 + col_offset
                            clog_col = 6 + col_offset
                            notes_col = 7 + col_offset
                            tpm_col = 8 + col_offset  # TPM is typically in column 9 for standard format
                    
                            # Extract values if columns exist
                            values = {}
                        
                            # Helper function to clean values and strip quotes
                            def clean_value(raw_val):
                                if pd.isna(raw_val):
                                    return ""
                                val_str = str(raw_val).strip()
                                # Remove surrounding quotes if present
                                if val_str.startswith("'") and val_str.endswith("'"):
                                    val_str = val_str[1:-1]
                                elif val_str.startswith('"') and val_str.endswith('"'):
                                    val_str = val_str[1:-1]
                                return val_str
                        
                            debug_print(f"DEBUG: Processing row {data_row_idx} for {sample_id}")
                    
                            for field, col_idx in [
                                ('puffs', puffs_col), ('before_weight', before_weight_col), 
                                ('after_weight', after_weight_col), ('draw_pressure', draw_pressure_col),
                                ('resistance', resistance_col), ('smell', smell_col), 
                                ('clog', clog_col), ('notes', notes_col)
                            ]:
                                if col_idx < len(sheet_data.columns):
                                    raw_val = sheet_data.iloc[data_row_idx, col_idx]
                                    values[field] = clean_value(raw_val)
                                    debug_print(f"DEBUG: {field} raw: {raw_val}, cleaned: {values[field]}")
                                else:
                                    values[field] = ""
                    
                            # Handle TPM separately
                            if tpm_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, tpm_col]
                                cleaned_val = clean_value(raw_val)
                                if cleaned_val and cleaned_val.strip():
                                    try:
                                        values['tpm'] = float(cleaned_val)
                                        debug_print(f"DEBUG: tpm raw: {raw_val}, cleaned: {cleaned_val}, float: {values['tpm']}")
                                    except (ValueError, TypeError):
                                        values['tpm'] = None
                                        debug_print(f"DEBUG: tpm conversion failed for: {cleaned_val}")
                                else:
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None
                    
                            # Check if this row has any meaningful data (not all empty)
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Storing data for row {data_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                                # Append the data (building from scratch)
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["resistance"].append(values['resistance'])
                                self.data[sample_id]["smell"].append(values['smell'])
                                self.data[sample_id]["clog"].append(values['clog'])
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])
                        
                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Row {data_row_idx} has no meaningful data, skipping")
                
                    except Exception as e:
                        debug_print(f"DEBUG: Error processing row {data_row_idx} for {sample_id}: {e}")
                        continue
        
                if sample_had_data:
                    loaded_data_count += 1
                    debug_print(f"DEBUG: {sample_id} - Loaded {row_count} rows of existing data (cleared template)")
                else:
                    debug_print(f"DEBUG: {sample_id} - No existing data found")
                    # If no data was found, ensure we still have empty arrays (not template data)
                    if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                        self.data[sample_id]["chronography"] = []
                        self.data[sample_id]["puffs"] = []
                        self.data[sample_id]["before_weight"] = []
                        self.data[sample_id]["after_weight"] = []
                        self.data[sample_id]["draw_pressure"] = []
                        self.data[sample_id]["smell"] = []
                        self.data[sample_id]["notes"] = []
                        self.data[sample_id]["tpm"] = []
                    else:
                        self.data[sample_id]["puffs"] = []
                        self.data[sample_id]["before_weight"] = []
                        self.data[sample_id]["after_weight"] = []
                        self.data[sample_id]["draw_pressure"] = []
                        self.data[sample_id]["resistance"] = []
                        self.data[sample_id]["smell"] = []
                        self.data[sample_id]["clog"] = []
                        self.data[sample_id]["notes"] = []
                        self.data[sample_id]["tpm"] = []
    
            debug_print(f"DEBUG: Successfully loaded existing data for {loaded_data_count} samples from loaded sheets")
    
            # Recalculate TPM values for all samples
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                self.calculate_tpm(sample_id)
                debug_print(f"DEBUG: Calculated TPM for {sample_id}")

            # update tksheets to show loaded data
            if hasattr(self, 'sample_sheets') and self.sample_sheets:
                for i, sample_sheet in enumerate(self.sample_sheets):
                    if i < self.num_samples:
                        sample_id = f"Sample {i + 1}"
                        self.update_tksheet(sample_sheet, sample_id)
                        debug_print(f"DEBUG: Updated tksheet for {sample_id}")
            else:
                debug_print("DEBUG: sample_sheets not available yet, data loading completed")

            # Update the stats panel
            self.update_stats_panel()

            debug_print("DEBUG: Existing data loading from loaded sheets completed successfully")
    
        except Exception as e:
            debug_print(f"ERROR: Failed to load existing data from loaded sheets: {e}")
            import traceback
            traceback.print_exc()

    def load_existing_data_from_file(self):
        """Load existing data from file or loaded sheets depending on file type."""
        debug_print(f"DEBUG: Loading existing data for file: {self.file_path}")
    
        # Check if this is a .vap3 file or if the file doesn't exist (temporary file)
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: Detected .vap3 file or non-existent file, loading from loaded sheets")
            self.load_existing_data_from_loaded_sheets()
        else:
            debug_print("DEBUG: Detected Excel file, loading from file using openpyxl")
            self.load_existing_data_from_excel_file()

    def load_existing_data_from_excel_file(self):
        """Load existing data from the Excel file into the data collection interface."""
        debug_print(f"DEBUG: Loading existing data from file: {self.file_path}")

        try:
            # Load the workbook and calculate formulas
            wb = openpyxl.load_workbook(self.file_path, data_only=True)  # data_only=True evaluates formulas

            # Check if the test sheet exists
            if self.test_name not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet '{self.test_name}' not found in file")
                return

            ws = wb[self.test_name]
            debug_print(f"DEBUG: Successfully opened sheet '{self.test_name}'")

            # Determine column layout based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8  # Including chronography
                debug_print(f"DEBUG: Loading User Test Simulation format with 8 columns per sample")
            else:
                columns_per_sample = 12  # Standard format
                debug_print(f"DEBUG: Loading standard format with 12 columns per sample")

            # Load data for each sample
            loaded_data_count = 0
            for sample_idx in range(self.num_samples):
                sample_id = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample

                debug_print(f"DEBUG: Loading data for {sample_id} with column offset {col_offset}")

                # CLEAR ALL EXISTING TEMPLATE DATA for this sample (like in load_existing_data_from_loaded_sheets)
                debug_print(f"DEBUG: Clearing existing template data for {sample_id}")
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    self.data[sample_id]["chronography"] = []
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["smell"] = []  # Used for failure in user simulation
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []
                else:
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["resistance"] = []
                    self.data[sample_id]["smell"] = []
                    self.data[sample_id]["clog"] = []
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []

                # Read data starting from row 5 (Excel row 5 = index 4)
                row_count = 0
                sample_had_data = False

                for excel_row_idx in range(5, min(ws.max_row + 1, 100)):  # Start from row 5, limit to 100 rows max
                    try:
                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User Test Simulation format: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                            chronography_col = 1 + col_offset  # Column A + offset
                            puffs_col = 2 + col_offset          # Column B + offset  
                            before_weight_col = 3 + col_offset  # Column C + offset
                            after_weight_col = 4 + col_offset   # Column D + offset
                            draw_pressure_col = 5 + col_offset  # Column E + offset
                            failure_col = 6 + col_offset        # Column F + offset (stored in smell field)
                            notes_col = 7 + col_offset          # Column G + offset
                            tpm_col = 8 + col_offset            # Column H + offset

                            # Extract values and clean them
                            def clean_excel_value(cell_value):
                                if cell_value is None:
                                    return ""
                                return str(cell_value).strip().strip('"').strip("'")

                            values = {}
                            values['chronography'] = clean_excel_value(ws.cell(row=excel_row_idx, column=chronography_col).value)
                            values['puffs'] = clean_excel_value(ws.cell(row=excel_row_idx, column=puffs_col).value)
                            values['before_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=before_weight_col).value)
                            values['after_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=after_weight_col).value)
                            values['draw_pressure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=draw_pressure_col).value)
                            values['failure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=failure_col).value)
                            values['notes'] = clean_excel_value(ws.cell(row=excel_row_idx, column=notes_col).value)
                        
                            # Handle TPM calculation
                            tpm_cell_value = ws.cell(row=excel_row_idx, column=tpm_col).value
                            if tpm_cell_value is not None:
                                try:
                                    values['tpm'] = float(tpm_cell_value)
                                except (ValueError, TypeError):
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Appending User Test Simulation data for Excel row {excel_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                            
                               
                                self.data[sample_id]["chronography"].append(values['chronography'])
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["smell"].append(values['failure'])  # failure stored in smell field
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Excel row {excel_row_idx} has no meaningful data, skipping")

                        else:
                            # Standard format: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
                            puffs_col = 1 + col_offset          # Column A + offset
                            before_weight_col = 2 + col_offset  # Column B + offset
                            after_weight_col = 3 + col_offset   # Column C + offset
                            draw_pressure_col = 4 + col_offset  # Column D + offset
                            resistance_col = 5 + col_offset     # Column E + offset
                            smell_col = 6 + col_offset          # Column F + offset
                            clog_col = 7 + col_offset           # Column G + offset
                            notes_col = 8 + col_offset          # Column H + offset
                            tpm_col = 9 + col_offset            # Column I + offset

                            # Extract values and clean them
                            def clean_excel_value(cell_value):
                                if cell_value is None:
                                    return ""
                                return str(cell_value).strip().strip('"').strip("'")

                            values = {}
                            values['puffs'] = clean_excel_value(ws.cell(row=excel_row_idx, column=puffs_col).value)
                            values['before_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=before_weight_col).value)
                            values['after_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=after_weight_col).value)
                            values['draw_pressure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=draw_pressure_col).value)
                            values['resistance'] = clean_excel_value(ws.cell(row=excel_row_idx, column=resistance_col).value)
                            values['smell'] = clean_excel_value(ws.cell(row=excel_row_idx, column=smell_col).value)
                            values['clog'] = clean_excel_value(ws.cell(row=excel_row_idx, column=clog_col).value)
                            values['notes'] = clean_excel_value(ws.cell(row=excel_row_idx, column=notes_col).value)

                            # Handle TPM calculation
                            tpm_cell_value = ws.cell(row=excel_row_idx, column=tpm_col).value
                            if tpm_cell_value is not None:
                                try:
                                    values['tpm'] = float(tpm_cell_value)
                                except (ValueError, TypeError):
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Appending standard data for Excel row {excel_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                            
                             
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["resistance"].append(values['resistance'])
                                self.data[sample_id]["smell"].append(values['smell'])
                                self.data[sample_id]["clog"].append(values['clog'])
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Excel row {excel_row_idx} has no meaningful data, skipping")

                    except Exception as e:
                        debug_print(f"DEBUG: Error processing Excel row {excel_row_idx} for {sample_id}: {e}")
                        continue

                if sample_had_data:
                    loaded_data_count += 1
                    debug_print(f"DEBUG: {sample_id} - Loaded {row_count} rows of existing data from Excel")
                else:
                    debug_print(f"DEBUG: {sample_id} - No existing data found in Excel")

            wb.close()
        
            if loaded_data_count > 0:
                debug_print(f"DEBUG: Successfully loaded existing data from Excel file for {loaded_data_count} samples")
            
                # Recalculate TPM values for all samples FIRST
                for i in range(self.num_samples):
                    sample_id = f"Sample {i + 1}"
                    self.calculate_tpm(sample_id)
                    debug_print(f"DEBUG: Calculated TPM for {sample_id}")

                # Update all treeviews to show the loaded data WITH calculated TPM
                if hasattr(self, 'sample_sheets') and self.sample_sheets:
                    for i, sample_sheet in enumerate(self.sample_sheets):
                        if i < self.num_samples:
                            sample_id = f"Sample {i + 1}"
                            self.update_tksheet(sample_sheet, sample_id)
                            debug_print(f"DEBUG: Updated tksheet for {sample_id} with TPM values")
                else:
                    debug_print("DEBUG: sample_sheets not available yet, TPM calculation completed")
                
                # Update the stats panel if available
                if hasattr(self, 'update_stats_panel'):
                    self.update_stats_panel()
            else:
                debug_print("DEBUG: No existing data found in Excel file")

        except Exception as e:
            debug_print(f"ERROR: Failed to load existing data from Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def ensure_initial_tpm_calculation(self):
        """Ensure TPM is calculated and displayed when window opens."""
        debug_print("DEBUG: Ensuring initial TPM calculation and display")
    
        for i in range(self.num_samples):
            sample_id = f"Sample {i + 1}"
        
            # Calculate TPM for this sample
            self.calculate_tpm(sample_id)
        
        if hasattr(self, 'sample_sheets') and i < len(self.sample_sheets):
            sheet = self.sample_sheets[i]
            self.update_tksheet(sheet, sample_id)
    
        # Update stats panel to show current TPM data
        self.update_stats_panel()
    
        debug_print("DEBUG: Initial TPM calculation and display completed")

    def setup_event_handlers(self):
        """Set up event handlers for the window."""
        # Window close protocol
        self.window.protocol("WM_DELETE_WINDOW", self.on_window_close)
    
        # Keyboard shortcuts
        self.setup_hotkeys()
    
        # Bind window resize events to update plot size
        self.window.bind('<Configure>', self.on_window_resize_plot, add=True)
        debug_print("DEBUG: Window resize event handler bound")
    
        # Initialize plot sizing after window is fully set up
        self.window.after(1000, self.initialize_plot_size)
    
        debug_print("DEBUG: Event handlers set up")
    
    def setup_hotkeys(self):
        """Set up keyboard shortcuts for navigation."""
        if not hasattr(self, 'hotkey_bindings'):
            self.hotkey_bindings = {}

        # Clear any existing bindings
        for key, binding_id in self.hotkey_bindings.items():
            try:
                self.window.unbind(key, binding_id)
            except:
                pass

        self.hotkey_bindings.clear()

        # Bind Ctrl+S for quick save
        self.window.bind("<Control-s>", lambda e: self.save_data(show_confirmation=False) if self.hotkeys_enabled else None)

        # Use Alt+Left/Right for navigation instead of Ctrl
        def handle_alt_nav(direction):
            if not self.hotkeys_enabled:
                return "break"
    
            if direction == "left":
                self.go_to_previous_sample()
            else:
                self.go_to_next_sample()
            return "break"

        # Bind Alt+Left/Right for sample navigation
        self.window.bind("<Alt-Left>", lambda e: handle_alt_nav("left"))
        self.window.bind("<Alt-Right>", lambda e: handle_alt_nav("right"))

        # Also bind with focus_set to ensure window has focus
        self.window.focus_set()

        debug_print("DEBUG: Hotkeys set up with Alt+Left/Right for sample navigation")
    
    def on_window_close(self):
        """Handle window close event with auto-save."""
        self.log("Window close event triggered", "debug")
 
        # Cancel auto-save timer
        if self.auto_save_timer:
            self.window.after_cancel(self.auto_save_timer)
    
        # Only handle unsaved changes if result is not already set
        if self.result is None:
            # Auto-save if there are unsaved changes
            if self.has_unsaved_changes:
                if messagebox.askyesno("Save Changes", 
                                     "You have unsaved changes. Save before closing?"):
                    try:
                        self.save_data(show_confirmation=False)
                        self.result = "load_file"
                    except Exception as e:
                        messagebox.showerror("Save Error", f"Failed to save: {e}")
                        return  # Don't close if save failed
                else:
                    self.result = "cancel"
            else:
                self.result = "load_file" if self.last_save_time else "cancel"
    
        # Show the main GUI window again before destroying
        if hasattr(self.parent, 'root') and self.main_window_was_visible:
            debug_print("DEBUG: Restoring main GUI window from data collection window")
            self.parent.root.deiconify()  # Show main window
            self.parent.root.state('zoomed')
            self.parent.root.lift()  # Bring to front
            self.parent.root.focus_set()  # Give focus to main window
        
            debug_print("DEBUG: Main GUI window restored")

        self.window.destroy()
    
    def save_data(self, exit_after=False, show_confirmation=True, auto_save=False):
        """Unified method for saving data."""
        # End any active editing
        self.end_editing()
    
        # For auto-save, we'll set a flag to avoid certain behaviors
        if auto_save:
            self._auto_save_in_progress = True
    
        # Confirm save if requested and not auto-saving
        if show_confirmation and not auto_save:
            if not messagebox.askyesno("Confirm Save", "Save the collected data to the file?"):
                return False
    
        try:
            # Calculate TPM values for all samples
            for i in range(self.num_samples):
                sample_id = f"Sample {i+1}"
                self.calculate_tpm(sample_id)
        
            # Save to Excel file
            self._save_to_excel()
        
            # Update application state if needed
            if hasattr(self.parent, 'filtered_sheets'):
                self._update_application_state()
        
            # Refresh the main GUI if not auto-saving
            if not auto_save:
                self.refresh_main_gui_after_save()
        
            # Mark as saved
            self.has_unsaved_changes = False
            self.update_save_status(False)
        
            # Show confirmation if requested (not for auto-save)
            if show_confirmation and not auto_save and not exit_after:
                show_success_message("Save Complete", "Data saved successfully.", self.window)
        
            # Clean up auto-save flag
            if auto_save:
                self._auto_save_in_progress = False
        
            # Exit if requested
            if exit_after:
                debug_print("DEBUG: Save and exit requested - calling on_window_close()")
                self.result = "load_file"
                # Call on_window_close() to properly restore main window before destroying
                self.on_window_close()
                return True

            return True
        
        except Exception as e:
            # Clean up auto-save flag
            if auto_save:
                self._auto_save_in_progress = False
            
            error_msg = f"Failed to save data: {e}"
            self.log(error_msg, "error")
        
            if not auto_save:  # Don't show errors for auto-save
                messagebox.showerror("Save Error", error_msg)
            
            return False
    
    def initialize_data(self):
        """Initialize data structures for all samples."""
        self.data = {}
        for i in range(self.num_samples):
            sample_id = f"Sample {i+1}"
    
            # Basic data structure
            self.data[sample_id] = {
                "current_row_index": 0, 
                "avg_tpm": 0.0
            }
    
            # Add standard columns (smell field will be used for both "smell" and "failure" depending on test type)
            columns = ["puffs", "before_weight", "after_weight", "draw_pressure", "resistance", "smell", "clog", "notes", "tpm"]
            for column in columns:
                self.data[sample_id][column] = []
    
            # Add special columns for User Test Simulation
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                self.data[sample_id]["chronography"] = []
    
            # Pre-initialize rows
            for j in range(50):
                puff = (j + 1) * self.puff_interval
                self.data[sample_id]["puffs"].append(puff)
        
                # Initialize other columns with empty values
                for column in columns:
                    if column == "puffs":
                        continue  # Already initialized
                    elif column == "tpm":
                        self.data[sample_id][column].append(None)
                    else:
                        self.data[sample_id][column].append("")
        
                # Initialize chronography if needed
                if "chronography" in self.data[sample_id]:
                    self.data[sample_id]["chronography"].append("")
            
            self.log(f"Initialized data for Sample {i+1}", "debug")

    def on_cancel(self):
        """Handle cancel button click or window close."""
        self.on_window_close()
    
    def show(self):
        """
        Show the window and wait for user input.
        
        Returns:
            str: "load_file" if data was saved and file should be loaded for viewing,
                 "cancel" if the user cancelled.
        """
        debug_print("DEBUG: Showing DataCollectionWindow")
        self.window.lift()
        self.window.focus_force()
        self.window.grab_set()  # Ensure it maintains focus
    
        self.window.wait_window()
        
        # Clean up auto-save timer
        if self.auto_save_timer:
            self.window.after_cancel(self.auto_save_timer)
        
        debug_print(f"DEBUG: DataCollectionWindow closed with result: {self.result}")
        return self.result  

    def create_tpm_stats_panel(self):
        """Create the enhanced TPM statistics panel with responsive resizing."""
        debug_print("DEBUG: Creating enhanced TPM statistics panel with resize support")

        # Clear existing widgets
        for widget in self.stats_frame.winfo_children():
            widget.destroy()

        # Create the main stats frame with proper expansion
        main_stats_frame = ttk.LabelFrame(self.stats_frame, text="TPM Statistics", padding=10)
        main_stats_frame.pack(fill='both', expand=True, padx=5, pady=5)
    
        # Store reference for resize handling
        self.plot_container = main_stats_frame
    
        # Sample selection for stats display (fixed height section)
        control_frame = ttk.Frame(main_stats_frame)
        control_frame.pack(side='top', fill='x', pady=(0, 5))
    
        # Current sample label
        self.current_sample_label = ttk.Label(control_frame, text="Sample 1: Unknown Sample", 
                                             font=('Arial', 12, 'bold'))
        self.current_sample_label.pack(anchor='w')
    
        # Stats text frame (fixed height section)  
        text_frame = ttk.Frame(main_stats_frame)
        text_frame.pack(side='top', fill='x', pady=(5, 10))
    
        # Stats display
        self.avg_tpm_label = ttk.Label(text_frame, text="Average TPM: 0.000", font=('Arial', 10))
        self.avg_tpm_label.pack(anchor='w')
    
        self.latest_tpm_label = ttk.Label(text_frame, text="Latest TPM: 0.000", font=('Arial', 10))
        self.latest_tpm_label.pack(anchor='w')
    
        self.std_dev_label = ttk.Label(text_frame, text="Std Dev (last 5 sessions): 0.000", font=('Arial', 10))
        self.std_dev_label.pack(anchor='w')
    
        self.current_puffs_label = ttk.Label(text_frame, text="Current Puffs: 0", font=('Arial', 10))
        self.current_puffs_label.pack(anchor='w')
    
        # Plot container frame (expandable section)
        plot_container = ttk.Frame(main_stats_frame)
        plot_container.pack(side='top', fill='both', expand=True)
    
        # Store reference to the container for proper sizing
        self.stats_frame_container = plot_container
    
        # Create the matplotlib plot with dynamic sizing
        self.setup_stats_plot_canvas(plot_container)
    
        # Initialize TPM labels dictionary if not exists
        if not hasattr(self, 'tpm_labels'):
            self.tpm_labels = {}
    
        # Update the statistics for the current sample
        self.update_stats_panel()
    
        debug_print("DEBUG: Enhanced TPM statistics panel created with resize support")

    def setup_stats_plot_canvas(self, parent):
        """Create the matplotlib canvas for the TPM plot with dynamic responsive sizing."""
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    
        # Calculate dynamic plot size based on available space
        dynamic_width, dynamic_height = self.calculate_dynamic_plot_size(parent)
    
        # Create figure with calculated responsive sizing
        self.stats_fig, self.stats_ax = plt.subplots(figsize=(dynamic_width, dynamic_height))
        self.stats_fig.patch.set_facecolor('white')
        self.stats_fig.subplots_adjust(left=0.15, right=0.95, top=0.90, bottom=0.15)
    
        # Create canvas with proper expansion configuration
        canvas_frame = ttk.Frame(parent)
        canvas_frame.pack(fill='both', expand=True)
    
        # Store reference to canvas_frame for resize handling
        self.canvas_frame = canvas_frame
    
        # Create canvas
        self.stats_canvas = FigureCanvasTkAgg(self.stats_fig, canvas_frame)
        canvas_widget = self.stats_canvas.get_tk_widget()
        canvas_widget.pack(fill='both', expand=True)
    
        # Initialize empty plot
        self.update_tpm_plot_for_current_sample()
    
        debug_print("DEBUG: Stats plot canvas created with responsive sizing")

    def update_tpm_plot_for_current_sample(self):
        """Update the TPM plot for the currently selected sample with smart y-axis bounds."""
        try:
            # Get currently selected sample
            current_tab_index = self.notebook.index(self.notebook.select())
            current_sample_id = f"Sample {current_tab_index + 1}"
        except:
            current_sample_id = "Sample 1"  
            current_tab_index = 0

        debug_print(f"DEBUG: Updating TPM plot for {current_sample_id}")
    
        # Clear the plot
        self.stats_ax.clear()
    
        # Get data for the sample
        tpm_values = [v for v in self.data[current_sample_id]["tpm"] if v is not None]
        puff_values = []
    
        # Get corresponding puff values for non-None TPM values
        for i, tpm in enumerate(self.data[current_sample_id]["tpm"]):
            if tpm is not None and i < len(self.data[current_sample_id]["puffs"]):
                puff_values.append(self.data[current_sample_id]["puffs"][i])
    
        if tpm_values and puff_values and len(tpm_values) == len(puff_values):
            # Plot TPM over puffs
            self.stats_ax.plot(puff_values, tpm_values, marker='o', linewidth=2, markersize=4, color='blue')
            self.stats_ax.set_xlabel('Puffs', fontsize=10)
            self.stats_ax.set_ylabel('TPM (mg/puff)', fontsize=10)
            self.stats_ax.set_title(f'TPM Over Time - {current_sample_id}', fontsize=11)
            self.stats_ax.grid(True, alpha=0.3)
        
            # Smart y-axis bounds logic
            if len(tpm_values) >= 2:
                sorted_tpm = sorted(tpm_values, reverse=True)  # Sort descending
                max_tpm = sorted_tpm[0]
                second_max_tpm = sorted_tpm[1]
            
                if max_tpm <= 20:
                    # Use max + 2 when max <= 20
                    y_max = max_tpm + 2
                    debug_print(f"DEBUG: Setting y-axis bounds 0 to {y_max} (max {max_tpm} + 2)")
                elif second_max_tpm <= 20:
                    # Use second largest value when max > 20 but second max <= 20
                    y_max = second_max_tpm + 2
                    debug_print(f"DEBUG: Setting y-axis bounds 0 to {y_max} (second max {second_max_tpm} + 2, ignoring outlier {max_tpm})")
                else:
                    # Both values > 20, default to 20
                    y_max = 20
                    debug_print(f"DEBUG: Setting y-axis bounds 0 to 20 (both max {max_tpm} and second max {second_max_tpm} > 20)")
            elif len(tpm_values) == 1:
                # Single value
                max_tpm = tpm_values[0]
                if max_tpm <= 20:
                    y_max = max_tpm + 2
                else:
                    y_max = 20  # Single outlier, cap at 20
                debug_print(f"DEBUG: Setting y-axis bounds 0 to {y_max} (single value {max_tpm})")
            else:
                # No values, use default
                y_max = 9
                debug_print(f"DEBUG: Setting default y-axis bounds 0 to 9")
        
            self.stats_ax.set_ylim(0, y_max)
        
            # Adjust tick label sizes for better fit
            self.stats_ax.tick_params(axis='both', which='major', labelsize=9)
        
        else:
            # Show empty plot with labels
            self.stats_ax.set_xlabel('Puffs', fontsize=10)
            self.stats_ax.set_ylabel('TPM (mg/puff)', fontsize=10)
            self.stats_ax.set_title(f'TPM Over Time - {current_sample_id}', fontsize=11)
            self.stats_ax.grid(True, alpha=0.3)
            self.stats_ax.text(0.5, 0.5, 'No TPM data available', 
                              transform=self.stats_ax.transAxes, ha='center', va='center', fontsize=10)
            self.stats_ax.set_ylim(0, 9)  # Default bounds for empty plot
    
        # Apply layout and draw
        self.stats_fig.tight_layout(pad=1.0)
        self.stats_canvas.draw()
    
        debug_print(f"DEBUG: TPM plot updated for {current_sample_id} with smart y-axis bounds")

    def update_stats_panel(self, event=None):
        """Update the TPM statistics panel to show only current sample with enhanced stats."""
        debug_print("DEBUG: Updating enhanced TPM statistics panel")

        # Get currently selected sample
        try:
            current_tab_index = self.notebook.index(self.notebook.select())
            current_sample_id = f"Sample {current_tab_index + 1}"
            debug_print(f"DEBUG: Updating stats for {current_sample_id}")
        except:
            current_sample_id = "Sample 1"  
            current_tab_index = 0

        # Get sample name
        sample_name = "Unknown Sample"
        if current_tab_index < len(self.header_data.get('samples', [])):
            sample_name = self.header_data['samples'][current_tab_index].get('id', f"Sample {current_tab_index + 1}")

        # Update sample label
        if hasattr(self, 'current_sample_label'):
            self.current_sample_label.config(text=f"{current_sample_id}: {sample_name}")

        # Calculate TPM values if needed
        self.calculate_tpm(current_sample_id)

        # Get TPM values and data for current sample (filtering out None values)
        tpm_values = [v for v in self.data[current_sample_id]["tpm"] if v is not None]

        if tpm_values:
            # Calculate statistics
            avg_tpm = sum(tpm_values) / len(tpm_values)
            latest_tpm = tpm_values[-1]

            # Calculate standard deviation of last 5 sessions (or all if < 5)
            recent_tpm_values = tpm_values[-5:] if len(tpm_values) >= 5 else tpm_values
            std_dev = statistics.stdev(recent_tpm_values) if len(recent_tpm_values) > 1 else 0.0

            # Find current puff count (furthest down row with after_weight data)
            current_puff_count = 0
            for i in range(len(self.data[current_sample_id]["after_weight"]) - 1, -1, -1):
                if (self.data[current_sample_id]["after_weight"][i] and 
                    str(self.data[current_sample_id]["after_weight"][i]).strip()):
                    current_puff_count = self.data[current_sample_id]["puffs"][i] if i < len(self.data[current_sample_id]["puffs"]) else 0
                    break

            self.data[current_sample_id]["avg_tpm"] = avg_tpm
        else:
            avg_tpm = 0.0
            latest_tpm = 0.0
            std_dev = 0.0
            current_puff_count = 0
            recent_tpm_values = []

        # Update labels
        if hasattr(self, 'avg_tpm_label'):
            self.avg_tpm_label.config(text=f"Average TPM: {avg_tpm:.6f}" if tpm_values else "Average TPM: N/A")
    
        if hasattr(self, 'latest_tpm_label'):
            self.latest_tpm_label.config(text=f"Latest TPM: {latest_tpm:.6f}" if tpm_values else "Latest TPM: N/A")
    
        if hasattr(self, 'std_dev_label'):
            sessions_text = f"(last {len(recent_tpm_values)} sessions)" if tpm_values else ""
            self.std_dev_label.config(text=f"Std Dev {sessions_text}: {std_dev:.6f}" if tpm_values else "Std Dev: N/A")
    
        if hasattr(self, 'current_puffs_label'):
            self.current_puffs_label.config(text=f"Current Puffs: {current_puff_count}")

        # Update TPM plot for current sample
        self.update_tpm_plot_for_current_sample()

        debug_print(f"DEBUG: Enhanced stats updated for {current_sample_id}")
    
    def go_to_previous_sample(self):
        """Navigate to the previous sample tab."""
        if not self.hotkeys_enabled:
            return
        current_tab = self.notebook.index(self.notebook.select())
        target_tab = (current_tab - 1) % len(self.sample_frames) 
        self.notebook.select(target_tab)

    def go_to_next_sample(self):
        """Navigate to the next sample tab."""
        if not self.hotkeys_enabled:
            return
        current_tab = self.notebook.index(self.notebook.select())
        target_tab = (current_tab + 1) % len(self.sample_frames)  
        self.notebook.select(target_tab)
    
    def get_column_name(self, col_idx):
        """Get the internal column name based on column index."""
        # Column mapping varies by test type
        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
            # User Test Simulation: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
            column_map = {
                0: "chronography",
                1: "puffs", 
                2: "before_weight", 
                3: "after_weight", 
                4: "draw_pressure", 
                5: "smell",  # "Failure" column maps to "smell" field in data structure
                6: "notes",
                7: "tpm"  # TPM column (read-only)
            }
        else:
            # Standard test: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
            column_map = {
                0: "puffs", 
                1: "before_weight", 
                2: "after_weight", 
                3: "draw_pressure", 
                4: "resistance",
                5: "smell", 
                6: "clog",
                7: "notes",
                8: "tpm"  # TPM column (read-only)
            }
    
        # Return appropriate column name or default to a safe value
        return column_map.get(col_idx, "notes")

    def create_header_section(self, parent_frame):
        """Create the header section with test information and save status."""
        header_frame = ttk.Frame(parent_frame, style='TFrame')
        header_frame.pack(fill="x", pady=(0, 10))
    
        # Left side of header
        header_left = ttk.Frame(header_frame, style='TFrame')
        header_left.pack(side="left", fill="x", expand=True)
    
        # Use styled labels
        ttk.Label(header_left, text=f"Test: {self.test_name}", style='Header.TLabel').pack(side="left")
        ttk.Label(header_left, text=f"Tester: {self.header_data['common']['tester']}", 
                 style='SubHeader.TLabel').pack(side="left", padx=(20, 0))
    
        # Right side of header - save status
        header_right = ttk.Frame(header_frame, style='TFrame')
        header_right.pack(side="right")
    
        self.save_status_label = ttk.Label(header_right, text="●", 
                                         style='SubHeader.TLabel', foreground="red")
        self.save_status_label.pack(side="right")
    
        self.save_status_text = ttk.Label(header_right, text="Unsaved changes", 
                                        style='SubHeader.TLabel')
        self.save_status_text.pack(side="right", padx=(0, 5))
    
        self.log("Header section created", "debug")

    def calculate_tpm_for_row(self, sample_id, changed_row_idx):
        """Calculate TPM for a specific row and related rows when data changes."""
        debug_print(f"DEBUG: Calculating TPM for {sample_id}, row {changed_row_idx}")
    
        # Calculate TPM for the changed row
        self.calculate_single_row_tpm(sample_id, changed_row_idx)
    
        # If puffs changed, recalculate all subsequent rows since intervals may have changed
        if changed_row_idx < len(self.data[sample_id]["puffs"]) - 1:
            for i in range(changed_row_idx + 1, len(self.data[sample_id]["puffs"])):
                self.calculate_single_row_tpm(sample_id, i)
    
        # Update average TPM
        valid_tpm_values = [v for v in self.data[sample_id]["tpm"] if v is not None]
        self.data[sample_id]["avg_tpm"] = sum(valid_tpm_values) / len(valid_tpm_values) if valid_tpm_values else 0.0

    def calculate_single_row_tpm(self, sample_id, row_idx):
        """Calculate TPM for a single row."""
        try:
            # Ensure TPM list is long enough
            while len(self.data[sample_id]["tpm"]) <= row_idx:
                self.data[sample_id]["tpm"].append(None)
        
            # Get weight values for this row
            before_weight_str = self.data[sample_id]["before_weight"][row_idx] if row_idx < len(self.data[sample_id]["before_weight"]) else ""
            after_weight_str = self.data[sample_id]["after_weight"][row_idx] if row_idx < len(self.data[sample_id]["after_weight"]) else ""
        
            # Skip if either weight is missing
            if not before_weight_str or not after_weight_str:
                self.data[sample_id]["tpm"][row_idx] = None
                return
            
            # Convert to float
            before_weight = float(before_weight_str)
            after_weight = float(after_weight_str)
        
            # Validate weights
            if before_weight <= after_weight:
                self.data[sample_id]["tpm"][row_idx] = None
                return
            
            # Calculate puffs in this interval
            current_puff = int(self.data[sample_id]["puffs"][row_idx]) if row_idx < len(self.data[sample_id]["puffs"]) else 0
            puffs_in_interval = current_puff
        
            if row_idx > 0:
                prev_puff = int(self.data[sample_id]["puffs"][row_idx - 1]) if (row_idx - 1) < len(self.data[sample_id]["puffs"]) else 0
                puffs_in_interval = current_puff - prev_puff
        
            # Skip if invalid puff interval
            if puffs_in_interval <= 0:
                self.data[sample_id]["tpm"][row_idx] = None
                return
            
            # Calculate TPM (mg/puff)
            weight_consumed = before_weight - after_weight  # in grams
            tpm = (weight_consumed * 1000) / puffs_in_interval  # Convert to mg per puff
        
            # Store result
            self.data[sample_id]["tpm"][row_idx] = round(tpm, 6)
        
            debug_print(f"DEBUG: Calculated TPM for {sample_id} row {row_idx}: {tpm:.6f} mg/puff")
        
        except (ValueError, TypeError, IndexError) as e:
            debug_print(f"DEBUG: Error calculating TPM for {sample_id} row {row_idx}: {e}")
            self.data[sample_id]["tpm"][row_idx] = None

    def convert_cell_value(self, value, column_name):
        """Convert user-entered value to appropriate type."""
        value = value.strip()
        if not value:
            return "" if column_name != "puffs" else 0

        try:
            if column_name == "puffs":
                return int(float(value))
            elif column_name in ["before_weight", "after_weight", "draw_pressure", "resistance", "smell", "clog"]:
                return float(value)
            elif column_name in ["chronography", "notes"]:
                return str(value)
        except ValueError:
            return value  # Keep as-is if conversion fails
        return value

    def auto_progress_weight(self, sheet, sample_id, row_idx, value):
        """Auto-fill the next row's before_weight with current after_weight."""
        if value in ["", 0, None]:
            return
        try:
            val = float(value)
            next_row = row_idx + 1
            if next_row < len(self.data[sample_id]["before_weight"]):
                self.data[sample_id]["before_weight"][next_row] = val
                # Update sheet display
                col_idx = 2 if self.test_name in ["User Test Simulation", "User Simulation Test"] else 1
                sheet.set_cell_data(next_row, col_idx, str(val))
                debug_print(f"DEBUG: Auto-set Sample {sample_id} row {next_row} before_weight to {val}")
        except (ValueError, TypeError):
            debug_print("DEBUG: Failed auto weight progression due to invalid value")

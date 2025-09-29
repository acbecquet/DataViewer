"""
Utility module for the DataViewer Application. Developed by Charlie Becquet.
Provides a Tkinter-based interface for interacting with Excel data, generating reports,
and plotting graphs.
"""
import numpy as np
import os
import re
import sys
import tempfile
import traceback
import pandas as pd
import openpyxl
from tkinter import filedialog, messagebox, Toplevel, Label, Button
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Global debug flag - change this to control ALL debug output across the app
DEBUG_ENABLED = True # Set to True when debugging is needed

def debug_print(*args, **kwargs):
    """
    Global debug print function that can be imported by all modules.
    Only prints if DEBUG_ENABLED is True.

    Usage:
        from utils import debug_print
        debug_print("This is a debug message")
        debug_print(f"Variable value: {variable}")
    """
    if DEBUG_ENABLED:
        print(*args, **kwargs)

def set_debug_mode(enabled):
    """
    Enable or disable debug mode globally.

    Args:
        enabled (bool): True to enable debug output, False to disable
    """
    global DEBUG_ENABLED
    DEBUG_ENABLED = enabled
    debug_print(f"Debug mode {'enabled' if enabled else 'disabled'}")

def is_debug_enabled():
    """
    Check if debug mode is currently enabled.

    Returns:
        bool: True if debug is enabled, False otherwise
    """
    return DEBUG_ENABLED


FONT = ('Arial', 10)
# Global constants
APP_BACKGROUND_COLOR = '#D3D3D3'
BUTTON_COLOR = '#4169E1'
PLOT_CHECKBOX_TITLE = "Click Checkbox to \nAdd/Remove Item \nFrom Plot"

def is_empty_sample(sample_data):
    """
    Check if a sample is empty based only on plotting data:
    - No TPM data
    - No Average TPM
    - No Draw Pressure
    - No Resistance

    Args:
        sample_data (dict or pd.Series): Sample data to check

    Returns:
        bool: True if sample is empty, False otherwise
    """
    try:
        # Convert to dict if it's a pandas Series
        if hasattr(sample_data, 'to_dict'):
            sample_dict = sample_data.to_dict()
        else:
            sample_dict = sample_data

        # Only check these plotting fields
        plotting_fields = ['Average TPM', 'Draw Pressure', 'Resistance']

        for field in plotting_fields:
            value = str(sample_dict.get(field, '')).strip()
            #debug_print(f"DEBUG: Checking field '{field}': '{value}'")

            # Skip completely empty values
            if not value or value in ['', 'nan', 'No data', 'None']:
                continue

            # Check if it's a meaningful numeric value
            try:
                numeric_val = float(value)
                # Remove pd.isna() call and use math.isnan() or simple check
                import math
                if numeric_val != 0 and not math.isnan(numeric_val):
                    #debug_print(f"DEBUG: Sample has plotting data in '{field}': {numeric_val}")
                    return False  # Has data, not empty
            except (ValueError, TypeError):
                # If it's not numeric but has meaningful content, it's not empty
                if len(value) > 0 and value not in ['nan', 'None', 'No data', '']:
                    #debug_print(f"DEBUG: Sample has non-numeric plotting data in '{field}': '{value}'")
                    return False


        debug_print("DEBUG: Sample has no plotting data - is empty")
        return True  # No plotting data found

    except Exception as e:
        debug_print(f"DEBUG: Error checking sample: {e}")
        return False  # If error, assume not empty

def filter_empty_samples_from_dataframe(df):
    """
    Filter out samples with no plotting data from processed DataFrame.
    """
    if df.empty:
        debug_print("DEBUG: DataFrame is already empty")
        return df

    try:
        debug_print(f"DEBUG: Checking {len(df)} samples for plotting data")

        # Create mask for samples with plotting data
        has_data_mask = []

        for index, row in df.iterrows():
            is_empty = is_empty_sample(row)
            has_data_mask.append(not is_empty)
            debug_print(f"DEBUG: Sample {index} ({'empty' if is_empty else 'has data'}): {row.get('Sample Name', 'Unknown')}")

        # Filter the dataframe
        filtered_df = df[has_data_mask].reset_index(drop=True)

        debug_print(f"DEBUG: Filtered from {len(df)} to {len(filtered_df)} samples with plotting data")
        return filtered_df

    except Exception as e:
        debug_print(f"DEBUG: Error filtering samples: {e}")
        import traceback
        traceback.print_exc()
        return df

def filter_empty_samples_from_full_data(full_sample_data, num_columns_per_sample=12):
    """
    Filter out samples with no TPM plotting data from full sample data.
    """
    if not pd or full_sample_data.empty:
        debug_print("DEBUG: No pandas or empty full_sample_data")
        return full_sample_data

    try:
        debug_print(f"DEBUG: Full data shape: {full_sample_data.shape}, columns per sample: {num_columns_per_sample}")

        # For User Test Simulation (8 columns), we need to match the processed data filtering
        if num_columns_per_sample == 8:
            debug_print("DEBUG: User Test Simulation detected - preserving all data since processed data was already filtered")
            # The processed data filtering already removed empty samples
            # So we keep the full data as-is for User Test Simulation
            return full_sample_data

        # For regular tests (12 columns), do the normal filtering
        num_samples = full_sample_data.shape[1] // num_columns_per_sample
        debug_print(f"DEBUG: Checking {num_samples} samples in full plotting data")

        if num_samples == 0:
            return full_sample_data

        # Find samples with actual TPM data
        samples_with_data = []

        for i in range(num_samples):
            start_col = i * num_columns_per_sample
            end_col = start_col + num_columns_per_sample
            sample_data = full_sample_data.iloc[:, start_col:end_col]

            debug_print(f"DEBUG: Checking sample {i+1} columns {start_col}-{end_col-1}")

            # Check TPM column (usually column 8 in 12-column format)
            has_tpm_data = False
            if num_columns_per_sample >= 9:
                tpm_col_idx = 8 if num_columns_per_sample == 12 else min(8, num_columns_per_sample - 1)
                if sample_data.shape[1] > tpm_col_idx and sample_data.shape[0] > 3:
                    # Get TPM data from row 3 onwards
                    tpm_data = sample_data.iloc[3:, tpm_col_idx]

                    # Convert to numeric and check for real values
                    numeric_tpm = pd.to_numeric(tpm_data, errors='coerce')
                    valid_tpm = numeric_tpm.dropna()

                    if len(valid_tpm) > 0 and (valid_tpm > 0).any():
                        has_tpm_data = True
                        debug_print(f"DEBUG: Sample {i+1} has TPM data: {valid_tpm.head().tolist()}")

            if has_tpm_data:
                samples_with_data.append(i)

        debug_print(f"DEBUG: Found {len(samples_with_data)} samples with TPM data out of {num_samples}")

        # If no samples have data, return empty DataFrame
        if not samples_with_data:
            debug_print("DEBUG: No samples with plotting data, returning empty DataFrame")
            return pd.DataFrame()

        # Reconstruct with only samples that have data
        filtered_columns = []
        for sample_idx in samples_with_data:
            start_col = sample_idx * num_columns_per_sample
            end_col = start_col + num_columns_per_sample
            sample_cols = list(range(start_col, min(end_col, full_sample_data.shape[1])))
            filtered_columns.extend(sample_cols)
            debug_print(f"DEBUG: Including sample {sample_idx+1} columns {start_col}-{end_col-1}")

        filtered_data = full_sample_data.iloc[:, filtered_columns]
        debug_print(f"DEBUG: Filtered plotting data from {full_sample_data.shape[1]} to {filtered_data.shape[1]} columns")
        return filtered_data

    except Exception as e:
        debug_print(f"DEBUG: Error filtering plotting data: {e}")
        import traceback
        traceback.print_exc()
        return full_sample_data

def round_values(value, decimals=2):
    """Round the values to a specified number of decimal places."""
    try:
        return round(float(value), decimals)
    except (ValueError, TypeError):
        return value

def get_save_path(self, default_extension: str = ".xlsx") -> Optional[str]:
    """
    Prompt the user to select a file save location and return the path.

    Args:
        default_extension (str): Default file extension for the save dialog.

    Returns:
        Optional[str]: The selected file path or None if canceled.
    """
    return filedialog.asksaveasfilename(
        defaultextension=default_extension,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

def generate_temp_image(self, figure):
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
        figure.savefig(tmpfile.name)
        return tmpfile.name

def remove_empty_columns(data: pd.DataFrame) -> pd.DataFrame:
    """
    Remove empty columns from a DataFrame. An empty column is defined as one where all values are NaN or 0.
    This process continues until a non-empty column is encountered.

    Args:
        data (pd.DataFrame): The input DataFrame.

    Returns:
        pd.DataFrame: The DataFrame with empty columns removed.
    """
    # Iterate over the columns and drop any empty columns (filled with NaN or 0)
    while True:
        # Check if all values in the first column are NaN or 0
        if data.iloc[:, 0].replace(0, pd.NA).isna().all():
            # If the column is empty, drop it
            data = data.iloc[:, 1:]  # Drop the first column
        else:
            # If we find a non-empty column, break the loop
            break
        # Stop if we have dropped all columns (to avoid errors with empty DataFrames)
        if data.shape[1] == 0:
            break
    return data

def clean_columns(data):
    """
    Clean column names and remove only truly empty columns.
    This function:
    - Retains columns with valid headers, even if they are empty.
    - Renames columns with NaN headers to their column index (1-based).
    - Ensures all column names are unique by appending a counter to duplicates.
    """
    print(f"DEBUG: clean_columns starting with {len(data.columns)} columns")
    print(f"DEBUG: Original columns: {list(data.columns)}")

    # Step 1: Replace NaN headers with column index (1-based)
    new_column_names = []
    for i, col in enumerate(data.columns):
        if pd.isna(col):
            new_column_names.append(str(i + 1))
        else:
            new_column_names.append(str(col))

    data.columns = new_column_names
    data.columns = data.columns.astype(str)
    print(f"DEBUG: After renaming NaN headers: {list(data.columns)}")

    # Step 2: Only remove columns that are both unnamed AND completely empty
    columns_to_keep = []
    for i, col in enumerate(data.columns):
        is_unnamed = col.isdigit() or col.startswith('Unnamed')

        # Fix: Ensure each condition returns a single boolean value
        try:
            column_data = data.iloc[:, i]
            is_all_na = bool(column_data.isna().all())
            is_all_empty_string = bool((column_data == '').all())
            is_all_zero = bool((column_data == 0).all())

            is_completely_empty = is_all_na or is_all_empty_string or is_all_zero
        except Exception as e:
            print(f"DEBUG: Error checking column {i} ({col}), keeping it: {e}")
            is_completely_empty = False

        # Keep column if it's named OR if it has any data
        if not (is_unnamed and is_completely_empty):
            columns_to_keep.append(i)
            #print(f"DEBUG: Keeping column {i} ({col}): unnamed={is_unnamed}, empty={is_completely_empty}")
        else:
            print(f"DEBUG: Removing column {i} ({col}): unnamed={is_unnamed}, empty={is_completely_empty}")

    data = data.iloc[:, columns_to_keep]
    print(f"DEBUG: After removing empty unnamed columns: {list(data.columns)}")

    # Step 3: Deduplicate column names
    new_columns = pd.Series(data.columns)
    for dup in new_columns[new_columns.duplicated()].unique():
        dups = new_columns[new_columns == dup].index
        new_columns[dups] = [f"{dup}.{i}" if i != 0 else dup for i in range(len(dups))]
    data.columns = new_columns

    # Step 4: Clean up generated column names that are just dots and numbers
    final_columns = []
    for col in data.columns:
        cleaned_col = str(col)
        # If column name is just dots and numbers (like '.1', '.2'), convert to empty string
        if cleaned_col.startswith('.') and cleaned_col[1:].isdigit():
            cleaned_col = ''
            print(f"DEBUG: Cleaned column name: '{col}' -> '{cleaned_col}'")
        final_columns.append(cleaned_col)

    data.columns = final_columns

    # Step 5: Replace NaN values with empty strings
    data = data.fillna('')

    print(f"DEBUG: Final columns: {list(data.columns)}")
    return data

def wrap_text(text, max_width=None):
    """
    Wrap text for labels to fit within a given width, preserving whole words.
    Args:
        text (str): The text to wrap.
        max_width (int): The maximum number of characters per line. If None, it will calculate dynamically.
    Returns:
        str: The wrapped text with line breaks.
    """
    if max_width is None:
        # Dynamically calculate max_width based on widget size
        axes_width = self.figure.get_size_inches()[0] * self.figure.dpi  # Width in pixels
        button_width = axes_width * 0.12  # Fraction of the figure allocated to checkboxes (from `add_axes`)
        char_width = 8  # Approximate width of a character in pixels (adjust as needed)
        max_width = int(button_width / char_width)

    if len(text) <= max_width:
        return text

    # Word-preserving wrapping
    words = text.split()
    lines = []
    current_line = ""

    for word in words:
        # Check if adding this word would exceed the line length
        test_line = current_line + " " + word if current_line else word

        if len(test_line) <= max_width:
            # Word fits, add it to current line
            current_line = test_line
        else:
            # Word doesn't fit
            if current_line:
                # Save current line and start new line with this word
                lines.append(current_line)
                current_line = word
            else:
                # Word itself is longer than max_width, we need to break it
                if len(word) > max_width:
                    # Break the long word at character boundaries
                    while len(word) > max_width:
                        lines.append(word[:max_width])
                        word = word[max_width:]
                    current_line = word if word else ""
                else:
                    current_line = word

    # Add the last line if it's not empty
    if current_line:
        lines.append(current_line)

    return '\n'.join(lines)

def autofit_columns_in_excel(file_path):
    """
    Adjusts the column widths in the Excel file to fit the content automatically.
    """
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter
                for cell in column:
                    try:  # Ensure cells with no value don't cause issues
                        max_length = max(max_length, len(str(cell.value) if cell.value else ""))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add padding for readability
                sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(file_path)
    except Exception as e:
        print(f"Error while adjusting column widths: {e}")

def is_standard_file(file_path: str) -> bool:
    """
    Determine if the file is standard by checking if it meets legacy file criteria.

    Legacy files are ONLY detected if:
    1. The file contains exactly 1 sheet, AND
    2. That sheet has the default name 'Sheet1'

    All other files are considered standard files.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        bool: True if the file is standard (should use standard processing),
              False if it is legacy (meets the specific legacy criteria).
    """
    try:
        print(f"DEBUG: Checking file format for: {file_path}")

        # Load all sheet names to check structure
        sheets_dict = pd.read_excel(file_path, sheet_name=None, header=None, nrows=1)
        sheet_names = list(sheets_dict.keys())
        num_sheets = len(sheet_names)

        debug_print(f"DEBUG: File contains {num_sheets} sheet(s): {sheet_names}")

        # Check for legacy file criteria
        # Legacy file MUST have exactly 1 sheet AND that sheet must be named 'Sheet1'
        if num_sheets == 1 and sheet_names[0] == 'Sheet1':
            debug_print("DEBUG: File meets legacy criteria: 1 sheet named 'Sheet1' -> File is legacy.")
            debug_print(f"DEBUG: Legacy sheet name: '{sheet_names[0]}'")
            return False  # Return False for legacy files
        else:
            debug_print("DEBUG: File does not meet legacy criteria -> File is standard.")
            if num_sheets > 1:
                debug_print(f"DEBUG: Multiple sheets found ({num_sheets}), treating as standard")
            elif num_sheets == 1:
                debug_print(f"DEBUG: Single sheet found but name is '{sheet_names[0]}' (not 'Sheet1'), treating as standard")
            return True   # Return True for standard files

    except Exception as e:
        print(f"ERROR: Exception while checking file format: {e}")
        debug_print(f"DEBUG: Traceback: {traceback.format_exc()}")
        # On error, default to treating as standard file
        debug_print("DEBUG: Error occurred, defaulting to standard file format")
        return True

def plotting_sheet_test(sheet_name, data):
    """
    Determine if a sheet is a plotting sheet by searching the first few rows for 'puffs' and 'tpm'.
    """
    try:
        if "legacy" in sheet_name.lower():
            return True

        if data.shape[0] < 3 or data.shape[1] < 2:
            return False

        # Dynamically search the first 5 rows for headers
        for i in range(min(5, len(data))):
            header_row = data.iloc[i].astype(str).str.lower()
            # print(f"Checking row {i} for headers: {header_row.values}")

            if header_row.str.contains("puffs").any() and header_row.str.contains("tpm").any():
                debug_print(f"Found plotting headers in row {i}")
                return True

        debug_print("No valid plotting headers found")
        return False

    except Exception as e:
        print(f"Error in plotting_sheet_test: {e}")
        return False

# Helper functions
def extract_meta_data(worksheet, patterns):
    """Extract meta_data from first 3 rows using regex patterns."""
    meta_data = {}
    for row in range(1, 3):  # Rows 1-3 (1-indexed)
        for col in range(1, worksheet.max_column + 1):
            cell_value = str(worksheet.cell(row=row, column=col).value).lower()
            for key, pattern in patterns.items():
                if re.search(pattern, cell_value, re.IGNORECASE):
                    meta_data[key] = worksheet.cell(
                        row=row,
                        column=col+1
                    ).value  # Get value from next cell
                    break
    return meta_data

def map_meta_data_to_template(template_ws, meta_data):
    """Map extracted meta_data to template positions."""
    mapping = {
        "sample_name": (1, 6),   # Row 1, Column F
        "voltage": (4, 6),       # Row 4, Column F
        "viscosity": (4, 2),     # Row 4, Column B
        "resistance": (3, 4),    # Row 3, Column D
        "puff_regime": (3, 8),   # Row 3, Column H
        "initial_oil": (3, 8),   # Row 3, Column H
        "date": (1, 4),         # Row 1, Column D
        "media": (2, 2)         # Row 2, Column B
    }
    for key, (row, col) in mapping.items():
        if key in meta_data:
            template_ws.cell(row=row, column=col, value=meta_data[key])

def copy_data_rows(src_ws, dest_ws):
    """Copy data rows (4+) exactly, maintaining structure."""
    for row in src_ws.iter_rows(min_row=4):
        for cell in row:
            dest_ws.cell(
                row=cell.row,
                column=cell.column,
                value=cell.value
            )

def get_plot_sheet_names():
    """
    Get the names of plot sheets.

    Returns:
        list: List of plot sheet names.
    """
    return [
        "Quick Screening Test", "Lifetime Test", "Device Life Test", "Horizontal Puffing Test", "Extended Test", "Long Puff Test",
        "Rapid Puff Test", "Intense Test", "Big Headspace Low T Test", "Big Headspace High T Test", "Big Headspace Serial Test",
        "Viscosity Compatibility", "Upside Down Test", "Big Headspace Pocket Test",
        "Low Temperature Stability","Vacuum Test", "Negative Pressure Test", "Viscosity Compatibility", "User Test Simulation", "User Simulation Test","Various Oil Compatibility", "Sheet1"
    ]

def read_sheet_with_values(file_path: str, sheet_name: Optional[str] = None):
    # Load workbook with computed values
    wb = load_workbook(file_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    data = ws.values  # This is a generator over rows
    # Convert generator to DataFrame
    df = pd.DataFrame(data)
    return df

def load_excel_file_with_formulas(file_path):
    """
    Load an Excel file with formula evaluation and return its sheets.
    Uses openpyxl with data_only=True to evaluate formulas.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: Dictionary of sheet names and DataFrames.
    """
    try:
        debug_print(f"DEBUG: Loading Excel file with formula evaluation: {file_path}")

        # First, try to force Excel to recalculate by opening and saving the file
        wb = openpyxl.load_workbook(file_path, data_only=False)

        # Force calculation of all formulas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # This is a formula - we need to force evaluation
                        debug_print(f"DEBUG: Found formula in {cell.coordinate}: {cell.value}")

        wb.close()

        # Now load with data_only=True to get calculated values
        wb_calc = openpyxl.load_workbook(file_path, data_only=True)
        sheets = {}

        for sheet_name in wb_calc.sheetnames:
            debug_print(f"DEBUG: Processing sheet: {sheet_name}")
            ws = wb_calc[sheet_name]

            # Convert worksheet to DataFrame row by row
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            if data:
                # Create DataFrame with proper column handling
                df = pd.DataFrame(data)

                # Set the first row as headers if it contains strings
                if len(df) > 0:
                    # Use first row as headers
                    df.columns = [f"Unnamed: {i}" if pd.isna(col) else str(col) for i, col in enumerate(df.iloc[0])]
                    df = df[1:].reset_index(drop=True)

                sheets[sheet_name] = df
                debug_print(f"DEBUG: Sheet {sheet_name} loaded with shape: {df.shape}")

                # Debug: Check specific cells that should contain usage efficiency
                if len(df) > 1 and len(df.columns) > 8:
                    debug_print(f"DEBUG: Sample usage efficiency values in row 1:")
                    for col_idx in [8, 20, 32, 44, 56, 68]:  # Expected positions for 6 samples
                        if col_idx < len(df.columns):
                            val = df.iloc[1, col_idx] if len(df) > 1 else None
                            debug_print(f"DEBUG: Column {col_idx}: '{val}'")
            else:
                sheets[sheet_name] = pd.DataFrame()
                debug_print(f"DEBUG: Sheet {sheet_name} is empty")

        wb_calc.close()
        debug_print(f"DEBUG: Successfully loaded {len(sheets)} sheets with enhanced formula evaluation")
        return sheets

    except Exception as e:
        print(f"ERROR: Failed to load Excel file with formula evaluation: {e}")
        # Fallback to original method
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            debug_print("DEBUG: Fallback to pandas read_excel successful")
            return sheets
        except Exception as fallback_error:
            raise ValueError(f"Error loading Excel file {file_path}: {fallback_error}")

def read_sheet_with_values_standards(file_path: str, sheet_name: Optional[str] = None):
    """Read Excel sheet with values exactly as they appear, evaluating formulas."""
    debug_print(f"DEBUG: Reading sheet with formula evaluation: {sheet_name}")

    # Use openpyxl with data_only=True to evaluate formulas
    wb = load_workbook(file_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Convert to DataFrame
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    if not data:
        return pd.DataFrame()

    # Create DataFrame and set first row as header
    df = pd.DataFrame(data)
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    wb.close()
    debug_print(f"DEBUG: Successfully read sheet {sheet_name} with formula evaluation")
    return df

def unmerge_all_cells(ws: Worksheet) -> None:
    """
    Unmerge all merged cells in the given worksheet.

    For each merged range, this function:
      - Retrieves the value from the top-left (master) cell.
      - Unmerges the range.
      - Sets every cell in the range to the retrieved value.

    Args:
        ws (Worksheet): The Openpyxl worksheet to process.
    """
    # Create a list of merged ranges to avoid modifying the collection while iterating.
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        # Get the bounds of the merged range.
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

        # Retrieve the value from the top-left (master) cell.
        master_value = ws.cell(row=min_row, column=min_col).value

        # Unmerge the cells.
        ws.unmerge_cells(str(merged_range))

        # Fill every cell in the previously merged range with the master value.
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=master_value)

def resource_path(relative_path):
    """
    Get the absolute path to the resource, compatible with PyInstaller.

    Args:
        relative_path (str): Relative path to the resource.

    Returns:
        str: Absolute path to the resource.
    """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def is_valid_excel_file(filename: str) -> bool:
    """
    Checks if the given filename is a valid Excel file that should be processed.
    Excludes temporary Excel files which often start with '~$'.

    Args:
        filename (str): The name of the file.

    Returns:
        bool: True if valid, False otherwise.
    """
    return filename.endswith('.xlsx') and not filename.startswith('~$')

def load_excel_file(file_path):
    """
    Load an Excel file and return its sheets.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: Dictionary of sheet names and DataFrames.
    """
    try:
        sheets = pd.read_excel(file_path, sheet_name = None, engine='openpyxl')
        return sheets
    except Exception as e:
        raise ValueError(f"Error loading Excel file {file_path}: {e}")

def validate_sheet_data(data, required_columns=None, required_rows=None):
    """
    Validate the sheet data for common issues like empty DataFrame, missing columns, or missing rows.

    Args:
        data (pd.DataFrame): The DataFrame to validate.
        required_columns (list, optional): List of columns that must exist. Defaults to None.
        required_rows (int, optional): Minimum number of rows required. Defaults to None.

    Returns:
        bool: True if the sheet is valid, False otherwise.
    """
    if data.empty:
        print("Sheet is empty.")
        return False

    # Check required columns
    if required_columns:
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            print(f"Missing required columns: {missing_columns}")
            return False

    # Check required rows
    if required_rows is not None and data.shape[0] < required_rows:
        print(f"Not enough rows. Expected at least {required_rows}, found {data.shape[0]}.")
        return False

    return True

def is_sheet_empty_or_zero(data: pd.DataFrame) -> bool:
    """
    Check if the specified sheet has any data in columns C through H (columns 2 to 7)
    in rows 5 through 20 (indices 4 to 19 in Python).

    Args:
        data (pd.DataFrame): The sheet data to check.

    Returns:
        bool: True if the sheet is empty or all zero, False otherwise.
    """
    # Define the range of rows and columns to check (rows 5 to 20, columns C to H)
    rows_to_check = data.iloc[4:20, 2:8]  # Rows 5-20, Columns C-H (Python uses 0-based index)

    # Check if all values in the selected range are NaN or 0
    if rows_to_check.isna().all().all() or (rows_to_check == 0).all().all():
        return True  # The sheet is considered empty or all zero
    else:
        return False  # There is some valid data in the specified range

def header_matches(cell_value, pattern: str) -> bool:
    """
    Check if the given cell value matches the regex pattern.

    Args:
        cell_value: The value from the cell.
        pattern (str): The regex pattern to search for.

    Returns:
        bool: True if the pattern is found, False otherwise.
    """
    if pd.isna(cell_value):
        return False
    return re.search(pattern, str(cell_value), re.IGNORECASE) is not None

def clean_display_suffixes(data):
    """
    Remove pandas-generated suffixes (.1, .2, .3, etc.) from data for display purposes,
    but preserve actual decimal numbers.
    """

    def is_pandas_suffix(value_str):
        """
        Check if a string ends with a pandas-generated suffix.
        Pandas suffixes are single digits (.1, .2, .3) on non-numeric text.
        """
        # Check if it ends with a period followed by a single digit
        suffix_match = re.search(r'\.(\d+)$', value_str)
        if not suffix_match:
            return False

        # Get the part before the suffix
        base_part = value_str[:suffix_match.start()]
        suffix_part = suffix_match.group(1)

        # If the suffix is a single digit and the base part is not purely numeric,
        # it's likely a pandas suffix
        if len(suffix_part) == 1:  # Single digit suffix
            try:
                # If the base part is a number, this is probably a decimal, not a pandas suffix
                float(base_part)
                return False  # It's a decimal number, don't remove
            except ValueError:
                # Base part is not a number, so .1, .2, etc. is likely a pandas suffix
                return True

        return False

    def clean_cell_value(value):
        if pd.isna(value) or value == "":
            return value

        value_str = str(value).strip()

        # Only remove if it looks like a pandas suffix
        if is_pandas_suffix(value_str):
            # Remove the suffix
            cleaned_value = re.sub(r'\.\d+$', '', value_str)
            print(f"DEBUG: Cleaned pandas suffix: '{value_str}' -> '{cleaned_value}'")
            return cleaned_value

        return value

    # Clean column names
    if hasattr(data, 'columns'):
        cleaned_columns = []
        for col in data.columns:
            col_str = str(col)
            if is_pandas_suffix(col_str):
                cleaned_col = re.sub(r'\.\d+$', '', col_str)
                print(f"DEBUG: Cleaned column name: '{col_str}' -> '{cleaned_col}'")
                cleaned_columns.append(cleaned_col)
            else:
                cleaned_columns.append(col_str)
        data.columns = cleaned_columns

    # Clean cell values
    for col in data.columns:
        data[col] = data[col].apply(clean_cell_value)

    return data


def show_success_message(title, message, parent=None):
    """Show a success message without system sound."""
    import tkinter as tk
    from tkinter import ttk

    # Create a custom dialog
    success_dialog = tk.Toplevel(parent) if parent else tk.Toplevel()
    success_dialog.title(title)
    success_dialog.resizable(False, False)
    success_dialog.transient(parent) if parent else None
    success_dialog.grab_set()

    # Hide initially to prevent stutter
    success_dialog.withdraw()

    # Create main frame
    main_frame = ttk.Frame(success_dialog, padding="20")
    main_frame.pack(fill="both", expand=True)

    # Success icon (using Unicode checkmark)
    icon_label = ttk.Label(main_frame, text="✓", font=("Arial", 24), foreground="green")
    icon_label.pack(pady=(0, 10))

    # Title
    title_label = ttk.Label(main_frame, text=title, font=("Arial", 12, "bold"))
    title_label.pack(pady=(0, 5))

    # Message
    message_label = ttk.Label(main_frame, text=message, font=("Arial", 10),
                             wraplength=400, justify="center")
    message_label.pack(pady=(0, 15))

    # OK button
    def on_ok():
        success_dialog.destroy()

    ok_button = ttk.Button(main_frame, text="OK", command=on_ok)
    ok_button.pack()

    # Bind Enter key to OK
    success_dialog.bind('<Return>', lambda e: on_ok())
    success_dialog.bind('<Escape>', lambda e: on_ok())

    # Center and show the dialog
    success_dialog.update_idletasks()
    width = success_dialog.winfo_reqwidth()
    height = success_dialog.winfo_reqheight()

    if parent:
        # Center relative to parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
    else:
        # Center on screen
        screen_width = success_dialog.winfo_screenwidth()
        screen_height = success_dialog.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

    success_dialog.geometry(f"{width}x{height}+{x}+{y}")
    success_dialog.deiconify()

    # Focus the OK button
    ok_button.focus_set()

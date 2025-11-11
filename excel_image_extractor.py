"""
Excel Image Extractor Module for DataViewer Application

This module handles extraction of embedded images from Excel files and integrates
them into the loaded images display.
"""

import os
import io
import tempfile
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from utils import debug_print


class ExcelImageExtractor:
    """Extracts embedded images from Excel files and integrates them into the GUI."""
    
    def __init__(self, gui):
        """Initialize the Excel image extractor.
        
        Args:
            gui: Reference to the main DataViewer GUI instance
        """
        self.gui = gui
        self.temp_dir = tempfile.mkdtemp(prefix="excel_images_")
        self.image_sample_mapping = {}  # Maps image paths to sample numbers
        debug_print(f"DEBUG: Created temp directory for extracted images: {self.temp_dir}")
    
    def extract_images_from_excel(self, excel_path, target_sheet_name=None):
        """Extract all embedded images from an Excel file with sample detection.
        
        Args:
            excel_path: Path to the Excel file
            target_sheet_name: Optional specific sheet to extract from
            
        Returns:
            Dictionary mapping sheet names to lists of extracted image paths
        """
        debug_print(f"DEBUG: Extracting images from Excel file: {excel_path}")
        
        extracted_images = {}
        
        try:
            workbook = load_workbook(excel_path, data_only=True)
            
            debug_print(f"DEBUG: Loaded workbook with {len(workbook.sheetnames)} sheets")
            
            for sheet_name in workbook.sheetnames:
                if target_sheet_name and sheet_name != target_sheet_name:
                    continue
                
                sheet = workbook[sheet_name]
                
                if not hasattr(sheet, '_images') or not sheet._images:
                    debug_print(f"DEBUG: No images found in sheet: {sheet_name}")
                    continue
                
                debug_print(f"DEBUG: Found {len(sheet._images)} images in sheet: {sheet_name}")
                
                sheet_images = []
                
                for idx, image in enumerate(sheet._images):
                    try:
                        # Get column position from image anchor
                        column_index = 0
                        if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                            column_index = image.anchor._from.col
                        
                        # Calculate sample number (each sample is 12 columns)
                        sample_number = (column_index // 12) + 1
                        
                        extracted_path = self._extract_single_image(image, sheet_name, idx)
                        if extracted_path:
                            sheet_images.append(extracted_path)
                            # Store sample mapping
                            self.image_sample_mapping[extracted_path] = sample_number
                            debug_print(f"DEBUG: Extracted image {idx+1} from {sheet_name} - Column {column_index} (Sample {sample_number})")
                    
                    except Exception as e:
                        debug_print(f"DEBUG: Failed to extract image {idx+1} from {sheet_name}: {e}")
                        continue
                
                if sheet_images:
                    extracted_images[sheet_name] = sheet_images
                    debug_print(f"DEBUG: Extracted {len(sheet_images)} images from {sheet_name}")
            
            workbook.close()
            
            debug_print(f"DEBUG: Total extraction complete: {sum(len(imgs) for imgs in extracted_images.values())} images from {len(extracted_images)} sheets")
            
            return extracted_images
        
        except Exception as e:
            debug_print(f"ERROR: Failed to extract images from Excel: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def _extract_single_image(self, image_obj, sheet_name, index):
        """Extract a single image from Excel to a temporary file.
        
        Args:
            image_obj: Openpyxl image object
            sheet_name: Name of the sheet containing the image
            index: Index of the image in the sheet
            
        Returns:
            Path to the extracted image file, or None if extraction failed
        """
        try:
            image_data = image_obj._data()
            pil_image = Image.open(io.BytesIO(image_data))
            
            image_format = pil_image.format or 'PNG'
            extension = image_format.lower()
            
            safe_sheet_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
            filename = f"{safe_sheet_name}_image_{index+1}.{extension}"
            filepath = os.path.join(self.temp_dir, filename)
            
            pil_image.save(filepath)
            
            debug_print(f"DEBUG: Saved extracted image to: {filepath}")
            debug_print(f"DEBUG: Image dimensions: {pil_image.size}, format: {image_format}")
            
            return filepath
        
        except Exception as e:
            debug_print(f"ERROR: Failed to extract single image: {e}")
            return None
    
    def integrate_extracted_images_to_gui(self, extracted_images, current_sheet=None, file_name=None):
        """Integrate extracted images into the GUI's image loader with sample labels.

        Args:
            extracted_images: Dictionary mapping sheet names to image paths
            current_sheet: Optional current sheet name to integrate images for
            file_name: Optional explicit file name to use instead of gui.current_file
        """
        debug_print("DEBUG: Integrating extracted images into GUI with sample labels")

        try:
            if not hasattr(self.gui, 'sheet_images'):
                self.gui.sheet_images = {}
    
            current_file = file_name if file_name else self.gui.current_file
            if not current_file:
                debug_print("DEBUG: No current file set and no file_name provided")
                return
    
            debug_print(f"DEBUG: Using file name: {current_file}")
    
            if current_file not in self.gui.sheet_images:
                self.gui.sheet_images[current_file] = {}
    
            if not hasattr(self.gui, 'image_crop_states'):
                self.gui.image_crop_states = {}
        
            # Store sample mapping at GUI level, not ImageLoader level
            if not hasattr(self.gui, 'image_sample_mapping'):
                self.gui.image_sample_mapping = {}
            self.gui.image_sample_mapping.update(self.image_sample_mapping)
            debug_print(f"DEBUG: Stored {len(self.image_sample_mapping)} image sample mappings at GUI level")
    
            # Add images for each sheet
            for sheet_name, image_paths in extracted_images.items():
                if sheet_name not in self.gui.sheet_images[current_file]:
                    self.gui.sheet_images[current_file][sheet_name] = []
        
                for img_path in image_paths:
                    if img_path not in self.gui.sheet_images[current_file][sheet_name]:
                        self.gui.sheet_images[current_file][sheet_name].append(img_path)
                
                        if img_path not in self.gui.image_crop_states:
                            self.gui.image_crop_states[img_path] = False
        
                debug_print(f"DEBUG: Integrated {len(image_paths)} images for sheet: {sheet_name}")
    
            # If current sheet has extracted images, update the display
            if current_sheet and current_sheet in extracted_images:
                if hasattr(self.gui, 'image_loader') and self.gui.image_loader:
                    debug_print(f"DEBUG: Updating image display for current sheet: {current_sheet}")
            
                    # Pass sample mapping to image loader from GUI
                    self.gui.image_loader.image_sample_mapping = self.gui.image_sample_mapping
            
                    self.gui.image_loader.load_images_from_list(
                        self.gui.sheet_images[current_file][current_sheet]
                    )
            
                    for img_path in self.gui.sheet_images[current_file][current_sheet]:
                        if img_path in self.gui.image_crop_states:
                            self.gui.image_loader.image_crop_states[img_path] = self.gui.image_crop_states[img_path]
            
                    self.gui.image_loader.display_images()
                    self.gui.image_frame.update_idletasks()
    
            debug_print("DEBUG: Image integration complete with sample labels")

        except Exception as e:
            debug_print(f"ERROR: Failed to integrate extracted images: {e}")
            import traceback
            traceback.print_exc()
    
    def cleanup_temp_directory(self):
        """Clean up temporary directory containing extracted images."""
        try:
            if os.path.exists(self.temp_dir):
                import shutil
                shutil.rmtree(self.temp_dir)
                debug_print(f"DEBUG: Cleaned up temp directory: {self.temp_dir}")
        except Exception as e:
            debug_print(f"DEBUG: Error cleaning up temp directory: {e}")


def extract_and_load_excel_images(gui, excel_path, current_sheet=None):
    """Convenience function to extract and load images from Excel file.
    
    Args:
        gui: Main DataViewer GUI instance
        excel_path: Path to Excel file
        current_sheet: Optional current sheet name
        
    Returns:
        Number of images extracted
    """
    debug_print(f"DEBUG: Starting Excel image extraction for: {excel_path}")
    
    try:
        if not hasattr(gui, 'excel_image_extractor') or gui.excel_image_extractor is None:
            gui.excel_image_extractor = ExcelImageExtractor(gui)
        
        extractor = gui.excel_image_extractor
        
        extracted_images = extractor.extract_images_from_excel(excel_path, current_sheet)
        
        if not extracted_images:
            debug_print("DEBUG: No images found in Excel file")
            return 0
        
        file_name = os.path.basename(excel_path)
        debug_print(f"DEBUG: Using file name for integration: {file_name}")
        
        extractor.integrate_extracted_images_to_gui(extracted_images, current_sheet, file_name=file_name)
        
        total_images = sum(len(imgs) for imgs in extracted_images.values())
        debug_print(f"DEBUG: Successfully extracted and loaded {total_images} images")
        
        return total_images
    
    except Exception as e:
        debug_print(f"ERROR: Failed to extract and load Excel images: {e}")
        import traceback
        traceback.print_exc()
        return 0
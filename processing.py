"""
Processing module for the DataViewer Application. Developed by Charlie Becquet.
"""
import re
import math
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import os
from typing import Optional, List, Tuple, Any
from openpyxl import load_workbook
from pptx.util import Inches
from utils import (
    round_values,
    remove_empty_columns,
    wrap_text,
    plotting_sheet_test,
    extract_meta_data,
    map_meta_data_to_template,
    read_sheet_with_values,
    unmerge_all_cells,
    validate_sheet_data,
    header_matches,
    load_excel_file,
    debug_print,
    load_excel_file_with_formulas
)

def get_valid_plot_options(plot_options: List[str], full_sample_data: pd.DataFrame) -> List[str]:
    """
    Check which plot options have valid, non-empty data and return the valid options.

    Args:
        plot_options (list): List of plot types to check.
        full_sample_data (pd.DataFrame): DataFrame containing the sample data.

    Returns:
        list: Valid plot options.
    """
    valid_options = []
    for plot_type in plot_options:
        y_data = get_y_data_for_plot_type(full_sample_data, plot_type)
        if y_data.dropna().astype(bool).any():  # Ensure there are non-NaN, non-zero values
            valid_options.append(plot_type)
    return valid_options

def get_y_data_for_plot_type(sample_data, plot_type):
    """
    Extract y-data for the specified plot type.
    Always calculates TPM and Power Efficiency on the fly for consistency.

    Args:
        sample_data (pd.DataFrame): DataFrame containing the sample data.
        plot_type (str): The type of plot to generate.

    Returns:
        pd.Series: y-data for plotting.
    """
    if plot_type == "TPM":
        #debug_print("DEBUG: Calculating TPM from weight differences with puffing intervals")
        
        # Get puffs, before weights, and after weights
        puffs = pd.to_numeric(sample_data.iloc[3:, 0], errors='coerce')  # Column 0 for Extended Test
        before_weights = pd.to_numeric(sample_data.iloc[3:, 1], errors='coerce')
        after_weights = pd.to_numeric(sample_data.iloc[3:, 2], errors='coerce')
        
        # Calculate puffing intervals
        puffing_intervals = pd.Series(index=puffs.index, dtype=float)
        for i, idx in enumerate(puffs.index):
            if i == 0:  # First row: current_puffs - 0
                current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else 10  # Default to 10 if NaN
                puffing_intervals.loc[idx] = current_puffs
            else:  # Subsequent rows: current_puffs - previous_puffs
                prev_idx = puffs.index[i-1]
                current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else 0
                prev_puffs = puffs.loc[prev_idx] if pd.notna(puffs.loc[prev_idx]) else 0
                puff_interval = current_puffs - prev_puffs
                
                # If interval is negative, zero, or current puffs is NaN, use current puffs or default
                if puff_interval <= 0 or pd.isna(current_puffs):
                    if pd.isna(current_puffs) or current_puffs == 0:
                        # Both puffs and interval are invalid, default to 10
                        fallback_puffs = 10
                        #debug_print(f"DEBUG: Row {i} - Both puffs ({current_puffs}) and interval ({puff_interval}) invalid, using default {fallback_puffs}")
                        puffing_intervals.loc[idx] = fallback_puffs
                    else:
                        # Use current puffs value
                        #debug_print(f"DEBUG: Row {i} - Interval {puff_interval} <= 0, using current puffs {current_puffs}")
                        puffing_intervals.loc[idx] = current_puffs
                else:
                    puffing_intervals.loc[idx] = puff_interval
        
        # Calculate TPM: (before - after) / puffing_interval * 1000 (mg/puff)
        weight_diff = (before_weights - after_weights) * 1000  # Convert g to mg
        calculated_tpm = weight_diff / puffing_intervals
        # Only calculate TPM where puffing_intervals > 0
        calculated_tpm = pd.Series(index=weight_diff.index, dtype=float)
        for idx in weight_diff.index:
            if pd.notna(puffing_intervals.loc[idx]) and puffing_intervals.loc[idx] > 0:
                calculated_tpm.loc[idx] = weight_diff.loc[idx] / puffing_intervals.loc[idx]
            else:
                calculated_tpm.loc[idx] = np.nan
                #debug_print(f"DEBUG: Skipping TPM calculation for row with invalid interval: {puffing_intervals.loc[idx]}")
        return calculated_tpm

    elif plot_type == "Normalized TPM":
        debug_print("DEBUG: Calculating Normalized TPM by dividing TPM by puff time")
        
        # Get regular TPM data first
        tpm_data = get_y_data_for_plot_type(sample_data, "TPM")
        tpm_numeric = pd.to_numeric(tpm_data, errors='coerce')
        debug_print(f"DEBUG: Got {len(tpm_numeric.dropna())} TPM values for normalization")
        
        # Extract puffing regime from this sample's position: row 1 (index 0), column 8 (index 7)
        puff_time = None
        puffing_regime = None
        
        try:
            puffing_regime_cell = sample_data.iloc[0, 7]  # Row 1, Column 8 (H) for this sample
            if pd.notna(puffing_regime_cell):
                puffing_regime = str(puffing_regime_cell).strip()
                debug_print(f"DEBUG: Found puffing regime at [0,7]: '{puffing_regime}'")
                
                # Extract puff time using regex pattern - FIXED for case sensitivity
                import re
                pattern = r'mL/(\d+(?:\.\d+)?)s/'
                match = re.search(pattern, puffing_regime, re.IGNORECASE)  # ADDED re.IGNORECASE
                if match:
                    puff_time = float(match.group(1))
                    debug_print(f"DEBUG: Extracted puff time: {puff_time}s from '{puffing_regime}'")
                else:
                    debug_print(f"DEBUG: Could not extract puff time from pattern: '{puffing_regime}'")
            else:
                debug_print("DEBUG: No puffing regime found at expected position [0,7]")
                
        except (ValueError, IndexError, TypeError, AttributeError) as e:
            debug_print(f"DEBUG: Error extracting puffing regime from [0,7]: {e}")
        
        # Apply normalization if puff time was found, otherwise use default
        if puff_time is not None and puff_time > 0:
            normalized_tpm = tpm_numeric / puff_time
            debug_print(f"DEBUG: Successfully normalized TPM by puff time {puff_time}s")
            debug_print(f"DEBUG: Normalized TPM values: {normalized_tpm.dropna().tolist()}")
            debug_print(f"DEBUG: Original TPM range: {tpm_numeric.min():.3f} - {tpm_numeric.max():.3f} mg/puff")
            debug_print(f"DEBUG: Normalized TPM range: {normalized_tpm.min():.3f} - {normalized_tpm.max():.3f} mg/s")
            return normalized_tpm
        else:
            debug_print("DEBUG: Cannot calculate normalized TPM - puff time not found or invalid")
            debug_print("DEBUG: Using default puff time of 3.0s for normalization")
            default_puff_time = 3.0
            normalized_tpm = tpm_numeric / default_puff_time
            debug_print(f"DEBUG: Normalized TPM with default {default_puff_time}s: {normalized_tpm.dropna().tolist()}")
            return normalized_tpm       

    elif plot_type == "Power Efficiency":
        debug_print("DEBUG: Calculating Power Efficiency from TPM/Power")
        
        # Get TPM first
        tpm_data = get_y_data_for_plot_type(sample_data, "TPM")
        tpm_numeric = pd.to_numeric(tpm_data, errors='coerce')
        
        # Extract voltage and resistance from metadata
        voltage = None
        resistance = None
        
        try:
            voltage_val = sample_data.iloc[1, 5]  # Adjust column index as needed
            if pd.notna(voltage_val):
                voltage = float(voltage_val)
                debug_print(f"DEBUG: Extracted voltage: {voltage}V")
        except (ValueError, IndexError, TypeError):
            debug_print("DEBUG: Could not extract voltage")
            
        try:
            resistance_val = sample_data.iloc[0, 3]  # Adjust column index as needed
            if pd.notna(resistance_val):
                resistance = float(resistance_val)
                debug_print(f"DEBUG: Extracted resistance: {resistance}Ω")
        except (ValueError, IndexError, TypeError):
            debug_print("DEBUG: Could not extract resistance")
        
        # Calculate power and power efficiency
        if voltage and resistance and voltage > 0 and resistance > 0:
            power = (voltage ** 2) / resistance
            debug_print(f"DEBUG: Calculated power: {power:.3f}W")
            calculated_power_eff = tpm_numeric / power
            debug_print(f"DEBUG: Calculated Power Efficiency values: {calculated_power_eff.dropna().tolist()}")
            return calculated_power_eff
        else:
            debug_print("DEBUG: Cannot calculate power efficiency - missing or invalid voltage/resistance")
            return pd.Series(dtype=float)
            
    elif plot_type == "Draw Pressure":
        return pd.to_numeric(sample_data.iloc[3:, 3], errors='coerce')
        
    elif plot_type == "Resistance":
        return pd.to_numeric(sample_data.iloc[3:, 4], errors='coerce')
        
    else:
        # Default to TPM
        return get_y_data_for_plot_type(sample_data, "TPM")

def get_y_label_for_plot_type(plot_type):
    """
    Get the y-axis label for the specified plot type.

    Args:
        plot_type (str): The type of plot.

    Returns:
        str: y-axis label.
    """
    y_label_mapping = {
        "TPM": 'TPM (mg/puff)',
        "Normalized TPM": 'Normalized TPM (mg/s)',
        "Draw Pressure": 'Draw Pressure (kPa)',
        "Resistance": 'Resistance (Ohms)',
        "Power Efficiency": 'Power Efficiency (mg/W)',
    }
    return y_label_mapping.get(plot_type, 'TPM (mg/puff)')  # Default to TPM

def plot_user_test_simulation_samples(full_sample_data: pd.DataFrame, num_columns_per_sample: int, plot_type: str, sample_names: List[str] = None) -> Tuple[plt.Figure, List[str]]:
    """
    Generate split plots for User Test Simulation.
    Creates two separate plots: one for puffs 0-50 (first 5 rows) and one for remaining puffs (9th row onwards).
    
    Args:
        full_sample_data (pd.DataFrame): DataFrame containing sample data.
        num_columns_per_sample (int): Number of columns per sample.
        plot_type (str): Type of plot to generate.
        sample_names (List[str], optional): List of sample names to use in legend.
    """
    debug_print(f"DEBUG: plot_user_test_simulation_samples called with data shape: {full_sample_data.shape}")
    debug_print(f"DEBUG: Provided sample_names: {sample_names}")
    debug_print(f"DEBUG: Full sample data first few rows:")
    debug_print(full_sample_data.iloc[:5, :15].to_string())
    
    num_samples = full_sample_data.shape[1] // num_columns_per_sample
    debug_print(f"DEBUG: User Test Simulation - Number of samples: {num_samples}")

    # Check if this should be a bar chart
    if plot_type == "TPM (Bar)":
        debug_print("DEBUG: Creating User Test Simulation bar chart")
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        extracted_sample_names = plot_user_test_simulation_bar_chart(ax1, ax2, full_sample_data, num_samples, num_columns_per_sample, sample_names)
        
        # Mark this as a split plot with bar chart data
        fig.is_split_plot = True
        fig.is_bar_chart = True
        
        # Store bar references for checkbox functionality
        fig.phase1_bars = ax1.patches
        fig.phase2_bars = ax2.patches
        
        debug_print(f"DEBUG: Successfully created User Test Simulation bar chart with {len(extracted_sample_names)} samples")
        return fig, extracted_sample_names
    
    # Original line plot logic for other plot types
    debug_print(f"DEBUG: User Test Simulation - Number of samples: {num_samples}")
    
    # Replace 0 with NaN for cleaner plotting
    full_sample_data = full_sample_data.replace(0, np.nan)
    
    # Create subplots for split plotting
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    extracted_sample_names = []
    phase1_lines = []
    phase2_lines = []
    y_max = 0
    
    for i in range(num_samples):
        start_col = i * num_columns_per_sample
        sample_data = full_sample_data.iloc[:, start_col:start_col + num_columns_per_sample]
        
        debug_print(f"DEBUG: Processing sample {i+1} columns {start_col} to {start_col + num_columns_per_sample - 1}")
        
        # Extract puffs data (column index 1 in User Test Simulation)
        x_data = pd.to_numeric(sample_data.iloc[3:, 1], errors='coerce').dropna()
        debug_print(f"DEBUG: Sample {i+1} puffs data length: {len(x_data)}, values: {x_data.head().tolist()}")

        # Extract y-data based on plot type
        y_data = get_y_data_for_user_test_simulation_plot_type(sample_data, plot_type)
        y_data = pd.to_numeric(y_data, errors='coerce').dropna()
        debug_print(f"DEBUG: Sample {i+1} y_data length: {len(y_data)}, values: {y_data.head().tolist()}")
        
        # Ensure x and y data have common indices
        common_index = x_data.index.intersection(y_data.index)
        if common_index.empty:
            debug_print(f"DEBUG: Sample {i+1} SKIPPED - no common data points")
            continue
            
        x_data = x_data.loc[common_index]
        y_data = y_data.loc[common_index]
        
        x_data = fix_x_axis_sequence(x_data)
        debug_print(f"DEBUG: Sample {i+1} fixed puffs data length: {len(x_data)}, values: {x_data.head().tolist()}")
        

        # Use provided sample names if available, otherwise use default
        if sample_names and i < len(sample_names):
            sample_name = sample_names[i]
            debug_print(f"DEBUG: Using provided sample name: '{sample_name}'")
        else:
            sample_name = f"Sample {i+1}"
            debug_print(f"DEBUG: Using default sample name: '{sample_name}'")
            
        extracted_sample_names.append(sample_name)
        
        # Split data into two phases
        data_start_index = 3  # Data starts at row 3
        
        # Phase 1: First 5 data rows (puffs 0-50)
        phase1_end_index = data_start_index + 5  # First 5 data rows
        phase1_mask = x_data.index < phase1_end_index
        phase1_x = x_data[phase1_mask]
        phase1_y = y_data[phase1_mask]
        
        # Phase 2: From 9th data row onwards (extended puffs)
        phase2_start_index = data_start_index + 8  # 9th data row (0-indexed: 8th)
        phase2_mask = x_data.index >= phase2_start_index
        phase2_x = x_data[phase2_mask]
        phase2_y = y_data[phase2_mask]
        
        debug_print(f"DEBUG: Sample {i+1} Phase 1 data points: {len(phase1_x)}, Phase 2 data points: {len(phase2_x)}")
        
        # Plot Phase 1 (0-50 puffs) and store line reference
        if not phase1_x.empty and not phase1_y.empty:
            line1 = ax1.plot(phase1_x, phase1_y, marker='o', label=sample_name)[0]
            phase1_lines.append(line1)
            y_max = max(y_max, phase1_y.max())
            debug_print(f"DEBUG: Plotted Phase 1 for {sample_name} with {len(phase1_x)} points")
        else:
            # Add placeholder line for consistency
            line1 = ax1.plot([], [], marker='o', label=sample_name)[0]
            phase1_lines.append(line1)
            debug_print(f"DEBUG: Added placeholder Phase 1 line for {sample_name}")
        
        # Plot Phase 2 (remaining puffs) and store line reference
        if not phase2_x.empty and not phase2_y.empty:
            line2 = ax2.plot(phase2_x, phase2_y, marker='o', label=sample_name)[0]
            phase2_lines.append(line2)
            y_max = max(y_max, phase2_y.max())
            debug_print(f"DEBUG: Plotted Phase 2 for {sample_name} with {len(phase2_x)} points")
        else:
            # Add placeholder line for consistency
            line2 = ax2.plot([], [], marker='o', label=sample_name)[0]
            phase2_lines.append(line2)
            debug_print(f"DEBUG: Added placeholder Phase 2 line for {sample_name}")

    debug_print(f"DEBUG: Final sample names for legend: {extracted_sample_names}")

    # Configure Phase 1 plot
    ax1.set_xlabel('Puffs')
    ax1.set_ylabel(get_y_label_for_plot_type(plot_type))
    ax1.set_title(f'{plot_type} - Phase 1 (Puffs 0-50)')
    ax1.legend(loc='upper right')
    ax1.set_xlim(0, 60)  # Slightly wider than 50 for better visualization
    prevent_x_label_overlap(ax1)

    # Configure Phase 2 plot  
    ax2.set_xlabel('Puffs')
    ax2.set_ylabel(get_y_label_for_plot_type(plot_type))
    ax2.set_title(f'{plot_type} - Phase 2 (Extended Puffs)')
    ax2.legend(loc='upper right')
    prevent_x_label_overlap(ax2)

    # Set consistent y-axis limits
    if plot_type == "Normalized TPM":
        ax1.set_ylim(-0.2, 4)
        ax2.set_ylim(-0.2, 4)
        debug_print("DEBUG: Set User Test Simulation Normalized TPM y-limits to -0.2 to 4")
    elif y_max > 9 and y_max <= 50:
        ax1.set_ylim(0, y_max)
        ax2.set_ylim(0, y_max)
    else:
        ax1.set_ylim(0, 9)
        ax2.set_ylim(0, 9)

    plt.tight_layout()
    
    # Store both sets of lines for checkbox functionality
    fig.phase1_lines = phase1_lines
    fig.phase2_lines = phase2_lines
    fig.is_split_plot = True
    
    debug_print(f"DEBUG: Successfully created User Test Simulation split plot with {len(extracted_sample_names)} samples")
    return fig, extracted_sample_names

def get_y_data_for_user_test_simulation_plot_type(sample_data, plot_type):
    """
    Extract y-data for the specified plot type for User Test Simulation.
    Always calculates TPM and Power Efficiency on the fly for consistency.
    Adjusted for 8-column layout instead of 12-column.
    """
    if plot_type == "TPM":
        debug_print("DEBUG: User Test Simulation - Calculating TPM from weight differences with puffing intervals")
        
        puffs = pd.to_numeric(sample_data.iloc[3:, 1], errors='coerce')  # Column 1 for User Test Simulation
        before_weights = pd.to_numeric(sample_data.iloc[3:, 2], errors='coerce') 
        after_weights = pd.to_numeric(sample_data.iloc[3:, 3], errors='coerce')  
        
        puffing_intervals = pd.Series(index=puffs.index, dtype=float)
        for i, idx in enumerate(puffs.index):
            if i == 0:  # First row: current_puffs - 0
                current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else 10  # Default to 10 if NaN
                puffing_intervals.loc[idx] = current_puffs
            else:  # Subsequent rows: current_puffs - previous_puffs
                prev_idx = puffs.index[i-1]
                current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else 0
                prev_puffs = puffs.loc[prev_idx] if pd.notna(puffs.loc[prev_idx]) else 0
                puff_interval = current_puffs - prev_puffs
                
                # If interval is negative, zero, or current puffs is NaN, use current puffs or default
                if puff_interval <= 0 or pd.isna(current_puffs):
                    if pd.isna(current_puffs) or current_puffs == 0:
                        # Both puffs and interval are invalid, default to 10
                        fallback_puffs = 10
                        debug_print(f"DEBUG: Row {i} - Both puffs ({current_puffs}) and interval ({puff_interval}) invalid, using default {fallback_puffs}")
                        puffing_intervals.loc[idx] = fallback_puffs
                    else:
                        # Use current puffs value
                        debug_print(f"DEBUG: Row {i} - Interval {puff_interval} <= 0, using current puffs {current_puffs}")
                        puffing_intervals.loc[idx] = current_puffs
                else:
                    puffing_intervals.loc[idx] = puff_interval
        
        # Calculate TPM: (before - after) / puffing_interval * 1000 (mg/puff)
        weight_diff = (before_weights - after_weights) * 1000  # Convert g to mg
        calculated_tpm = weight_diff / puffing_intervals
        
        for idx in weight_diff.index:
            if pd.notna(puffing_intervals.loc[idx]) and puffing_intervals.loc[idx] > 0:
                calculated_tpm.loc[idx] = weight_diff.loc[idx] / puffing_intervals.loc[idx]
            else:
                calculated_tpm.loc[idx] = np.nan
                debug_print(f"DEBUG: User Test Simulation - Skipping TPM calculation for row with invalid interval: {puffing_intervals.loc[idx]}")

        return calculated_tpm
        
    elif plot_type == "Power Efficiency":
        debug_print("DEBUG: User Test Simulation - Calculating Power Efficiency from TPM/Power")
        
        # Get TPM first
        tpm_data = get_y_data_for_user_test_simulation_plot_type(sample_data, "TPM")
        tpm_numeric = pd.to_numeric(tpm_data, errors='coerce')
        
        # Extract voltage and resistance from metadata (adjust for User Test Simulation layout)
        voltage = None
        resistance = None
        
        try:
            voltage_val = sample_data.iloc[0, 5]  # Adjust as needed
            if pd.notna(voltage_val):
                voltage = float(voltage_val)
                debug_print(f"DEBUG: User Test Simulation - Extracted voltage: {voltage}V")
        except (ValueError, IndexError, TypeError):
            debug_print("DEBUG: User Test Simulation - Could not extract voltage")
            
        try:
            resistance_val = sample_data.columns[3]  # Adjust as needed
            if pd.notna(resistance_val):
                resistance = float(resistance_val)
                debug_print(f"DEBUG: User Test Simulation - Extracted resistance: {resistance}Ω")
        except (ValueError, IndexError, TypeError):
            debug_print("DEBUG: User Test Simulation - Could not extract resistance")
        
        # Calculate power and power efficiency
        if voltage and resistance and voltage > 0 and resistance > 0:
            power = (voltage ** 2) / resistance
            debug_print(f"DEBUG: User Test Simulation - Calculated power: {power:.3f}W")
            calculated_power_eff = tpm_numeric / power
            debug_print(f"DEBUG: User Test Simulation - Calculated Power Efficiency values: {calculated_power_eff.dropna().tolist()}")
            return calculated_power_eff
        else:
            debug_print("DEBUG: User Test Simulation - Cannot calculate power efficiency - missing or invalid voltage/resistance")
            return pd.Series(dtype=float)
            
    elif plot_type == "Draw Pressure":
        # For User Test Simulation, Draw Pressure is in column 4 (adjust as needed)
        return pd.to_numeric(sample_data.iloc[3:, 4], errors='coerce')
        
    else:
        # Default to TPM for any other plot type
        return get_y_data_for_user_test_simulation_plot_type(sample_data, "TPM")

def plot_user_test_simulation_bar_chart(ax1, ax2, full_sample_data, num_samples, num_columns_per_sample, sample_names=None):
    """
    Generate split bar charts for User Test Simulation average TPM with error bars.
    Creates two bar charts: Phase 1 (0-50 puffs) and Phase 2 (extended puffs).
    Each bar shows average TPM with standard deviation as error bars.
    """
    debug_print(f"DEBUG: plot_user_test_simulation_bar_chart called with {num_samples} samples")
    
    phase1_averages = []
    phase1_std_devs = []
    phase2_averages = []
    phase2_std_devs = []
    labels = []
    extracted_sample_names = []

    for i in range(num_samples):
        start_col = i * num_columns_per_sample
        end_col = start_col + num_columns_per_sample
        sample_data = full_sample_data.iloc[:, start_col:end_col]

        debug_print(f"DEBUG: Processing sample {i+1} for bar chart, columns {start_col} to {end_col-1}")

        # Check if sample has valid data (check puffs column - column 1 for User Test Simulation)
        if sample_data.shape[0] <= 3 or pd.isna(sample_data.iloc[3, 1]):
            debug_print(f"DEBUG: Sample {i+1} has no valid data, skipping")
            continue

        # Calculate TPM values using the same method as the line plots
        tpm_data = get_y_data_for_user_test_simulation_plot_type(sample_data, "TPM")
        tpm_numeric = pd.to_numeric(tpm_data, errors='coerce').dropna()
        
        debug_print(f"DEBUG: Sample {i+1} TPM data length: {len(tpm_numeric)}, values: {tpm_numeric.head().tolist()}")
        
        # Extract puffs data to understand the sequence
        puffs_data = pd.to_numeric(sample_data.iloc[3:, 1], errors='coerce').dropna()
        if not puffs_data.empty:
            # Fix the puff sequence
            fixed_puffs = fix_x_axis_sequence(puffs_data)
            debug_print(f"DEBUG: Sample {i+1} fixed puffs sequence for phase splitting")
        
        if tpm_numeric.empty:
            debug_print(f"DEBUG: Sample {i+1} has no valid TPM data, skipping")
            continue
        
        # Split TPM data into phases based on the same logic as line plots
        # Phase 1: First 5 data points (corresponding to 0-50 puffs)
        phase1_tpm = tpm_numeric.iloc[:5] if len(tpm_numeric) >= 5 else tpm_numeric
        # Phase 2: From 9th data point onwards (index 8+, corresponding to extended puffs)
        phase2_tpm = tpm_numeric.iloc[8:] if len(tpm_numeric) > 8 else pd.Series(dtype=float)
        
        # Calculate averages and standard deviations
        phase1_avg = phase1_tpm.mean() if not phase1_tpm.empty else 0
        phase1_std = phase1_tpm.std() if len(phase1_tpm) > 1 else 0
        phase2_avg = phase2_tpm.mean() if not phase2_tpm.empty else 0
        phase2_std = phase2_tpm.std() if len(phase2_tpm) > 1 else 0
        
        phase1_averages.append(phase1_avg)
        phase1_std_devs.append(phase1_std)
        phase2_averages.append(phase2_avg)
        phase2_std_devs.append(phase2_std)


        if sample_names and i < len(sample_names):
            sample_name = sample_names[i]
            debug_print(f"DEBUG: Using provided sample name: '{sample_name}'")
        else:
            sample_name = f"Sample {i+1}"
            debug_print(f"DEBUG: Using default sample name: '{sample_name}'")
        
        extracted_sample_names.append(sample_name)
        wrapped_name = wrap_text(text=sample_name, max_width=10)
        labels.append(wrapped_name)
        
        debug_print(f"DEBUG: Sample {i+1} - Phase 1: avg={phase1_avg:.3f}, std={phase1_std:.3f}, Phase 2: avg={phase2_avg:.3f}, std={phase2_std:.3f}")

    if not phase1_averages:
        debug_print("DEBUG: No valid samples found for bar chart")
        return []

    # Create colormaps for unique colors
    num_bars = len(phase1_averages)
    colors = plt.cm.get_cmap('tab10', num_bars)(np.linspace(0, 1, num_bars))

    # Create numeric positions for bars to prevent overlapping
    x_positions = np.arange(len(phase1_averages))

    # Plot Phase 1 bar chart with error bars
    bars1 = ax1.bar(x_positions, phase1_averages, yerr=phase1_std_devs, 
                     color=colors, capsize=5, error_kw={'elinewidth': 2, 'capthick': 2}, width=0.6)
    ax1.set_xlabel('Samples')
    ax1.set_ylabel('Average TPM (mg/puff)')
    ax1.set_title('Average TPM - Phase 1 (Puffs 0-50)')

    # Set x-axis ticks and labels properly
    ax1.set_xticks(x_positions)
    ax1.set_xticklabels(labels, rotation=45, ha='right')
    ax1.tick_params(axis='x', labelsize=8)

    # Set y-axis to start from 0 and add some padding
    ax1.set_ylim(0, max(phase1_averages) * 1.2 if phase1_averages else 1)

    # Plot Phase 2 bar chart with error bars
    bars2 = ax2.bar(x_positions, phase2_averages, yerr=phase2_std_devs, 
                     color=colors, capsize=5, error_kw={'elinewidth': 2, 'capthick': 2}, width=0.6)
    ax2.set_xlabel('Samples')
    ax2.set_ylabel('Average TPM (mg/puff)')
    ax2.set_title('Average TPM - Phase 2 (Extended Puffs)')

    # Set x-axis ticks and labels properly
    ax2.set_xticks(x_positions)
    ax2.set_xticklabels(labels, rotation=45, ha='right')
    ax2.tick_params(axis='x', labelsize=8)

    # Set y-axis to start from 0 and add some padding
    ax2.set_ylim(0, max(phase2_averages) * 1.2 if phase2_averages else 1)

    debug_print(f"DEBUG: Created User Test Simulation bar charts with {len(extracted_sample_names)} samples")
    return extracted_sample_names

def fix_x_axis_sequence(x_data):
    """
    Fix x-axis data to ensure it's always increasing.
    If x2-x1 < 0, set x2 = x1+10 to avoid plotting issues.
    This is critical for User Test Simulation where puff counts can reset.
    
    Args:
        x_data (pd.Series): The x-axis data (usually puffs)
        
    Returns:
        pd.Series: Fixed x-axis data with monotonic increasing values
    """
    try:
        if len(x_data) <= 1:
            return x_data
            
        # Convert to numeric first if it's not already
        if x_data.dtype == 'object':  # String data
            x_numeric = pd.to_numeric(x_data, errors='coerce').dropna()
            debug_print(f"DEBUG: Converted string x_data to numeric: {len(x_numeric)} values")
        else:
            x_numeric = x_data.copy()
            
        if len(x_numeric) <= 1:
            return x_numeric
            
        # Convert to list for easier manipulation
        x_values = x_numeric.values.copy()
        
        # Fix any decreasing sequences
        fixes_applied = 0
        for i in range(1, len(x_values)):
            if x_values[i] - x_values[i-1] < 0:
                # If current value is less than previous, set it to previous + 10
                old_value = x_values[i]
                x_values[i] = x_values[i-1] + 10
                fixes_applied += 1
                #debug_print(f"DEBUG: Fixed x-axis sequence at index {i}: {old_value} -> {x_values[i]} (was decreasing)")
            elif x_values[i] == x_values[i-1]:
                # If values are equal, add small increment to maintain increasing sequence
                old_value = x_values[i]
                x_values[i] = x_values[i-1] + 5
                fixes_applied += 1
                #debug_print(f"DEBUG: Fixed duplicate x-axis value at index {i}: {old_value} -> {x_values[i]}")
        
        if fixes_applied > 0:
            debug_print(f"DEBUG: Applied {fixes_applied} fixes to x-axis sequence")
        
        # Return as pandas Series with original index
        return pd.Series(x_values, index=x_numeric.index)
        
    except Exception as e:
        debug_print(f"DEBUG: Error fixing x-axis sequence: {e}")
        return x_data  # Return original data if fixing fails

def plot_all_samples(full_sample_data: pd.DataFrame, plot_type: str, num_columns_per_sample: int = 12, sample_names: List[str] = None) -> Tuple[plt.Figure, List[str]]:
    """
    Generate plots for all samples in the provided data.

    Args:
        full_sample_data (pd.DataFrame): DataFrame containing sample data.
        plot_type (str): Type of plot to generate.
        num_columns_per_sample (int): Number of columns per sample.
        sample_names (List[str], optional): List of sample names to use in legend.

    Returns:
        tuple: (matplotlib.figure.Figure, list of sample names)
    """
    # ADD THESE DEBUG LINES:
    debug_print(f"DEBUG: plot_all_samples - full_sample_data shape: {full_sample_data.shape}")
    debug_print(f"DEBUG: plot_all_samples - provided sample_names: {sample_names}")
    debug_print(f"DEBUG: plot_all_samples - first 5 rows, first 15 columns:")
    debug_print(full_sample_data.iloc[:5, :15].to_string())
    debug_print("=" * 80)

    # ENSURE num_columns_per_sample is an integer
    try:
        num_columns_per_sample = int(num_columns_per_sample)
        debug_print(f"DEBUG:   num_columns_per_sample after int conversion: {num_columns_per_sample}")
    except (ValueError, TypeError) as e:
        print(f"ERROR: Could not convert num_columns_per_sample to int: {e}")
        raise ValueError(f"num_columns_per_sample must be convertible to int, got: {num_columns_per_sample} (type: {type(num_columns_per_sample)})")
    
    
    # Check if this is User Test Simulation (8 columns per sample)
    if num_columns_per_sample == 8:
        debug_print("DEBUG: Detected User Test Simulation - using split plotting")
        return plot_user_test_simulation_samples(full_sample_data, num_columns_per_sample, plot_type, sample_names)

    # Original logic for standard tests (12 columns per sample)
    num_samples = full_sample_data.shape[1] // num_columns_per_sample
    full_sample_data = full_sample_data.replace(0, np.nan)
    fig, ax = plt.subplots(figsize=(8, 6))
    extracted_sample_names = []

    if plot_type == "TPM (Bar)":
        extracted_sample_names = plot_tpm_bar_chart(ax, full_sample_data, num_samples, num_columns_per_sample, sample_names)
    else:
        y_max = 0
        for i in range(num_samples):
            start_col = i * num_columns_per_sample
            sample_data = full_sample_data.iloc[:, start_col:start_col + num_columns_per_sample]

            #debug_print(f"DEBUG: Sample {i+1} data shape: {sample_data.shape}")
            #debug_print(f"DEBUG: Sample {i+1} columns: {sample_data.columns.tolist()}")
    
            x_data = sample_data.iloc[3:, 0].dropna()
            #debug_print(f"DEBUG: Sample {i+1} x_data (puffs) length: {len(x_data)}, first few values: {x_data.head().tolist()}")

            y_data = get_y_data_for_plot_type(sample_data, plot_type)
            #debug_print(f"DEBUG: Sample {i+1} raw y_data length: {len(y_data)}, first few values: {y_data.head().tolist()}")
    
            y_data = pd.to_numeric(y_data, errors='coerce').dropna()
            #debug_print(f"DEBUG: Sample {i+1} numeric y_data length: {len(y_data)}, first few values: {y_data.head().tolist()}")

            common_index = x_data.index.intersection(y_data.index)
            #debug_print(f"DEBUG: Sample {i+1} common_index length: {len(common_index)}")
    
            x_data = x_data.loc[common_index]
            y_data = y_data.loc[common_index]
    
            x_data = fix_x_axis_sequence(x_data)
            #debug_print(f"DEBUG: Sample {i+1} fixed x_data length: {len(x_data)}, first few values: {x_data.head().tolist()}")

            #debug_print(f"DEBUG: Sample {i+1} final x_data length: {len(x_data)}, y_data length: {len(y_data)}")

            if not x_data.empty and not y_data.empty:
                # Use provided sample name if available, otherwise extract from data
                if sample_names and i < len(sample_names):
                    sample_name = sample_names[i]
                    debug_print(f"DEBUG: Using provided sample name: '{sample_name}'")
                else:
                    sample_name = sample_data.columns[5]
                    debug_print(f"DEBUG: Using extracted sample name: '{sample_name}'")
                
                ax.plot(x_data, y_data, marker='o', label=sample_name)
                extracted_sample_names.append(sample_name)
                y_max = max(y_max, y_data.max())
            else:
                debug_print(f"DEBUG: Sample {i+1} SKIPPED - x_data empty: {x_data.empty}, y_data empty: {y_data.empty}")

        ax.set_xlabel('Puffs')
        ax.set_ylabel(get_y_label_for_plot_type(plot_type))
        ax.set_title(plot_type)
        ax.legend(loc='upper right')

        prevent_x_label_overlap(ax)

        # Set y-axis limits based on plot type
        if plot_type == "Normalized TPM":
            ax.set_ylim(-0.2, 4)
            debug_print("DEBUG: Set Normalized TPM y-limits to -0.2 to 4")
        elif y_max > 9 and y_max <= 50:
            ax.set_ylim(0, y_max)
        else:
            ax.set_ylim(0, 9)

    return fig, extracted_sample_names

def plot_tpm_bar_chart(ax, full_sample_data, num_samples, num_columns_per_sample, sample_names=None):
    """
    Generate a bar chart of the average TPM for each sample with wrapped sample names.

    Args:
        ax (matplotlib.axes.Axes): Matplotlib axis to draw on.
        full_sample_data (pd.DataFrame): DataFrame of sample data.
        num_samples (int): Number of samples.
        num_columns_per_sample (int): Columns per sample.
        sample_names (list): Optional list of sample names.

    Returns:
        list: Sample names.
    """
    averages = []
    labels = []
    extracted_sample_names = []

    for i in range(num_samples):
        start_col = i * num_columns_per_sample
        end_col = start_col + num_columns_per_sample
        sample_data = full_sample_data.iloc[:, start_col:end_col]

        if pd.isna(sample_data.iloc[3, 1]):
            continue

        tpm_data = sample_data.iloc[3:, 8].dropna()  # TPM column (adjust if needed)
        average_tpm = tpm_data.mean()
        averages.append(average_tpm)

        # Use provided sample name if available, otherwise extract from data
        if sample_names and i < len(sample_names):
            sample_name = sample_names[i]
            debug_print(f"DEBUG: Using provided sample name: '{sample_name}'")
        else:
            sample_name = sample_data.columns[5] if len(sample_data.columns) > 5 else f"Sample {i+1}"
            debug_print(f"DEBUG: Using extracted sample name: '{sample_name}'")
        
        extracted_sample_names.append(sample_name)
        wrapped_name = wrap_text(text=sample_name, max_width=10)  # Use `wrap_text` to dynamically wrap names
        labels.append(wrapped_name)

    if not averages:
        debug_print("DEBUG: No valid samples found for bar chart")
        return []

    # Create numeric positions for bars
    x_positions = np.arange(len(averages))
    
    # Create a colormap with unique colors
    num_bars = len(averages)
    colors = plt.cm.get_cmap('tab10', num_bars)(np.linspace(0, 1, num_bars))

    # Plot the bar chart with proper positioning
    bars = ax.bar(x_positions, averages, color=colors, width=0.6)
    
    # Set the x-axis labels and ticks
    ax.set_xticks(x_positions)
    ax.set_xticklabels(labels, rotation=45, ha='right')
    
    ax.set_xlabel('Samples')
    ax.set_ylabel('Average TPM (mg/puff)')
    ax.set_title('Average TPM per Sample')
    ax.tick_params(axis='x', labelsize=8)
    
    # Set y-axis to start from 0 and add some padding
    ax.set_ylim(0, max(averages) * 1.2 if averages else 1)

    return extracted_sample_names

def prevent_x_label_overlap(ax):
    """
    Hide overlapping x-axis labels to improve readability.
    Simple solution that doesn't modify tick positions, just visibility.
    """
    try:
        # Get current labels and their positions
        labels = ax.get_xticklabels()
        if len(labels) <= 1:
            return
        
        # Calculate approximate label widths
        fig = ax.get_figure()
        if fig is None:
            return
            
        # Get the axis width in points
        bbox = ax.get_window_extent()
        axis_width_points = bbox.width
        
        # Estimate average label width (rough approximation)
        avg_label_length = sum(len(label.get_text()) for label in labels) / len(labels)
        estimated_label_width = avg_label_length * 12  # roughly 12 points per character
        
        # Calculate how many labels can fit without overlap
        if estimated_label_width > 0:
            max_labels = max(1, int(axis_width_points / (estimated_label_width * 1.3)))  # 1.2 for spacing
            
            if len(labels) > max_labels:
                # Calculate step to show evenly distributed labels
                step = max(1, len(labels) // max_labels)
                
                # Hide labels that would cause overlap
                for i, label in enumerate(labels):
                    if i % step != 0:
                        label.set_visible(False)
                
                debug_print(f"DEBUG: Prevented x-label overlap - showing every {step} labels ({max_labels} total)")
            else:
                debug_print(f"DEBUG: No x-label overlap detected - showing all {len(labels)} labels")
    except Exception as e:
        debug_print(f"DEBUG: Error in prevent_x_label_overlap: {e}")
        # If anything goes wrong, just leave labels as they are
        pass

def clean_presentation_tables(presentation):
    """
    Clean all tables in the given PowerPoint presentation by:
    - Removing rows where all cells are empty.
    - Removing columns where all cells, including the header, are empty.

    Args:
        presentation (pptx.Presentation): The PowerPoint presentation to process.
    """
    for slide in presentation.slides:
        for shape in slide.shapes:
            if not shape.has_table:
                continue  # Skip if the shape is not a table
            
            table = shape.table

            # Get the number of rows and columns
            num_rows = len(table.rows)
            num_cols = len(table.columns)

            # Step 1: Remove empty rows
            rows_to_keep = []
            for row_idx in range(num_rows):
                is_empty_row = all(
                    not cell.text.strip() for cell in table.rows[row_idx].cells
                )
                if not is_empty_row:
                    rows_to_keep.append(row_idx)

            # Keep only the non-empty rows
            for row_idx in reversed(range(num_rows)):
                if row_idx not in rows_to_keep:
                    table._tbl.remove(table.rows[row_idx]._tr)  # Remove row from XML

            # Step 2: Remove empty columns
            num_rows = len(table.rows)  # Updated number of rows after cleaning
            cols_to_keep = []
            for col_idx in range(num_cols):
                # Check if all cells in the column (including the header) are empty or 'nan'
                is_empty_col = all(
                    not table.cell(row_idx, col_idx).text.strip() or table.cell(row_idx, col_idx).text.strip() == "nan"
                    for row_idx in range(num_rows)
                )
                if not is_empty_col:
                    cols_to_keep.append(col_idx)

            # Keep only the non-empty columns
            for col_idx in reversed(range(num_cols)):
                if col_idx not in cols_to_keep:
                    for row in table.rows:
                        row._tr.remove(row.cells[col_idx]._tc)  # Remove column cell from XML

# ==================== DATA PROCESSING FUNCTIONS ====================

def process_sheet(data, headers_row=3, data_start_row=4):
    """
    Process a sheet by extracting headers and data.
    
    Args:
        data (pd.DataFrame): Input data from the sheet.
        headers_row (int): Row index for headers.
        data_start_row (int): Row index where data starts.
        
    Returns:
        tuple: (processed_data, table_data)
    """
    # data = remove_empty_columns(data) - for now, let's not remove empty columns for non-plotting sheets.
    headers = data.iloc[headers_row, :].tolist()
    table_data = data.iloc[data_start_row:, :]
    table_data = table_data.replace(0, np.nan)
    processed_data = pd.DataFrame(table_data.values, columns=headers)
    return processed_data, table_data


def process_plot_sheet(data, headers_row=3, data_start_row=4, num_columns_per_sample=12, custom_extracted_data_fn=None):
    """
    Process plotting sheets with fixed burn/clog/leak extraction from raw data.
    """
    debug_print(f"DEBUG: process_plot_sheet_fixed called with data shape: {data.shape}")
    
    # Store reference to raw data before cleaning
    raw_data = data.copy()
    
    # For data collection, allow minimal data (less strict validation)
    min_required_rows = max(headers_row + 1, 3)
    if data.shape[0] < min_required_rows:
        debug_print(f"DEBUG: Data has {data.shape[0]} rows, minimum required is {min_required_rows}")
        return create_empty_plot_structure(data, headers_row, num_columns_per_sample)
    
    try:
        # Clean up the data for processing but keep raw_data unchanged
        data = remove_empty_columns(data).replace(0, np.nan)
        debug_print(f"DEBUG: Data after cleaning: {data.shape}")

        samples = []
        full_sample_data = []
        sample_arrays = {}

        # Calculate the number of samples
        num_samples = data.shape[1] // num_columns_per_sample
        debug_print(f"DEBUG: Calculated {num_samples} samples")

        if num_samples == 0:
            debug_print("DEBUG: No samples detected, creating empty structure")
            return create_empty_plot_structure(data, headers_row, num_columns_per_sample)

        for i in range(num_samples):
            start_col = i * num_columns_per_sample
            end_col = start_col + num_columns_per_sample
            sample_data = data.iloc[:, start_col:end_col]

            if sample_data.empty:
                debug_print(f"DEBUG: Sample {i+1} is empty. Skipping.")
                continue

            # Extract plotting data with error handling
            try:
                sample_arrays[f"Sample_{i+1}_Puffs"] = sample_data.iloc[3:, 0].to_numpy()
                sample_arrays[f"Sample_{i+1}_TPM"] = sample_data.iloc[3:, 8].to_numpy()

                # Use custom function if provided, otherwise use our fixed function
                if custom_extracted_data_fn:
                    # For compatibility, try to pass raw data if the function supports it
                    try:
                        extracted_data = custom_extracted_data_fn(sample_data, raw_data, i)
                    except TypeError:
                        # Fallback to old signature
                        extracted_data = custom_extracted_data_fn(sample_data)
                else:
                    # Check if sample has sufficient data before processing
                    if sample_data.shape[0] < 4 or sample_data.shape[1] < 3:
                        debug_print(f"DEBUG: Sample {i+1} has insufficient data shape {sample_data.shape}, creating placeholder")
                        placeholder_data = {
                            "Sample Name": f"Sample {i+1}",
                            "Media": "",
                            "Viscosity": "",
                            "Voltage, Resistance, Power": "",
                            "Average TPM": "No data",
                            "Standard Deviation": "No data",
                            "Initial Oil Mass": "",
                            "Usage Efficiency": "",
                            "Burn?": "",
                            "Clog?": "",
                            "Leak?": ""
                        }
                        samples.append(placeholder_data)
                        full_sample_data.append(sample_data)
                        continue
                    
                    # If we get here, sample has sufficient data
                    extracted_data = updated_extracted_data_function_with_raw_data(sample_data, raw_data, i)

                samples.append(extracted_data)
                full_sample_data.append(sample_data)
                debug_print(f"DEBUG: Successfully processed sample {i+1}")
                
            except IndexError as e:
                debug_print(f"DEBUG: Index error for sample {i+1}: {e}. Creating placeholder.")
                # Create placeholder data for data collection
                placeholder_data = {
                    "Sample Name": f"Sample {i+1}",
                    "Media": "",
                    "Viscosity": "",
                    "Voltage, Resistance, Power": "",
                    "Average TPM": "No data",
                    "Standard Deviation": "No data",
                    "Initial Oil Mass": "",
                    "Usage Efficiency": "",
                    "Burn?": "",
                    "Clog?": "",
                    "Leak?": ""
                }
                samples.append(placeholder_data)
                full_sample_data.append(sample_data)
                continue

        # Create processed data and full sample data
        if samples:
            processed_data = pd.DataFrame(samples)
            full_sample_data_df = pd.concat(full_sample_data, axis=1) if full_sample_data else data
        else:
            debug_print("DEBUG: No valid samples found, creating minimal structure")
            processed_data = pd.DataFrame([{
                "Sample Name": "Sample 1",
                "Media": "",
                "Viscosity": "",
                "Voltage, Resistance, Power": "",
                "Average TPM": "No data",
                "Standard Deviation": "No data",
                "Initial Oil Mass": "",
                "Usage Efficiency": "",
                "Burn?": "",
                "Clog?": "",
                "Leak?": ""
            }])
            full_sample_data_df = data

        debug_print(f"DEBUG: Final processed_data shape: {processed_data.shape}")
        debug_print(f"DEBUG: Final full_sample_data shape: {full_sample_data_df.shape}")
        return processed_data, sample_arrays, full_sample_data_df
        
    except Exception as e:
        debug_print(f"DEBUG: Error processing plot sheet: {e}")
        import traceback
        traceback.print_exc()
        return create_empty_plot_structure(data, headers_row, num_columns_per_sample)

def no_efficiency_extracted_data(sample_data, raw_data, sample_index):
    """
    Custom function to extract data without efficiency metrics, using enhanced sample name extraction with suffix support.
    """
    print(f"DEBUG: Processing sample {sample_index + 1} (no efficiency) with enhanced name extraction")
    
    # Extract sample name with old format support (same logic as above)
    sample_name = f"Sample {sample_index + 1}"  # Default fallback
    
    if sample_data.shape[1] > 8:  # Ensure we have enough columns
        headers = sample_data.columns.astype(str)
        
        # Look for old format patterns with suffix support
        project_value = None
        sample_value = None
        
        for i, header in enumerate(headers):
            header_lower = str(header).lower().strip()
            
            # Enhanced pattern matching to handle suffixed headers
            if header_lower.startswith("project:") and i + 1 < len(headers):
                project_value = str(headers[i + 1]).strip()
                # Remove the suffix from the project value if it exists
                if project_value.endswith(f'.{sample_index}'):
                    project_value = project_value[:-len(f'.{sample_index}')]
            elif header_lower.startswith("sample:") and i + 1 < len(headers):
                sample_value = str(headers[i + 1]).strip()
        
        # Combine project and sample values if found (old format)
        if project_value and sample_value:
            if project_value.lower() not in ['nan', 'none', ''] and sample_value.lower() not in ['nan', 'none', '']:
                sample_name = f"{project_value} {sample_value}".strip()
        elif project_value and project_value.lower() not in ['nan', 'none', '']:
            sample_name = project_value.strip()
        elif sample_value and sample_value.lower() not in ['nan', 'none', '']:
            sample_name = sample_value.strip()
        else:
            # Try new format - Sample ID at column 5
            if len(headers) > 5:
                sample_id_value = str(headers[5]).strip()
                if sample_id_value and sample_id_value.lower() not in ['nan', 'none', '', 'unnamed: 5']:
                    sample_name = sample_id_value
    
    tpm_data = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce').dropna()
    avg_tpm = round_values(tpm_data.mean()) if not tpm_data.empty else None
    std_tpm = round_values(tpm_data.std()) if not tpm_data.empty else None
    
    # Extract burn/clog/leak using the new method
    burn, clog, leak = extract_burn_clog_leak_from_raw_data(raw_data, sample_index)
    
    return {
        "Sample Name": sample_name,  # Now uses enhanced extraction
        "Media": sample_data.iloc[0, 1] if sample_data.shape[0] > 0 else "",
        "Viscosity": sample_data.iloc[1, 1] if sample_data.shape[0] > 1 else "",
        "Voltage, Resistance, Power": f"{sample_data.iloc[1, 5]} V, "
                                       f"{round_values(sample_data.iloc[0, 3])} ohm, "
                                       f"{round_values(sample_data.iloc[0, 5])} W"
                                       if sample_data.shape[0] > 1 and sample_data.shape[1] > 5 else "",
        "Average TPM": avg_tpm,
        "Standard Deviation": std_tpm,
        "Burn?": burn,
        "Clog?": clog,
        "Leak?": leak
    }

def process_generic_sheet(data, headers_row=3, data_start_row=4):
    """
    Process data for any sheet that doesn't match predefined sheet names.
    
    Args:
        data (pd.DataFrame): Input data from the Excel sheet.
    
    Returns:
        tuple: (display_df, additional_output, full_sample_data)
            display_df (pd.DataFrame): Processed table data for display.
            additional_output (dict): Any additional information that might be needed.
            full_sample_data (pd.DataFrame): Full sample data concatenated for all samples.
    """
    # Check if data is empty or invalid
    if data.empty or data.isna().all().all():
        return pd.DataFrame(), {}, pd.DataFrame()  # Return empty data structures

    try:
        # Check for sufficient rows
        if data.shape[0] <= headers_row or data.shape[0] <= data_start_row:
            #debug_print(f"Insufficient rows for processing (headers_row={headers_row}, data_start_row={data_start_row})")
            return pd.DataFrame(), {}, pd.DataFrame()

        # Extract headers and validate them
        column_names = data.iloc[headers_row].fillna("").tolist()
        
        # Slice the data starting from the data_start_row and convert to strings
        table_data = data.iloc[data_start_row:].copy()
        table_data.columns = column_names
        table_data = table_data.astype(str)  # Convert all data to strings

        # Create a DataFrame with the extracted headers and data
        display_df = table_data.copy()

        # Collect all data into a single DataFrame for further use
        full_sample_data_df = pd.concat([table_data], axis=1)
        
        return display_df, {}, full_sample_data_df
    except Exception as e:
        #print(f"Error processing generic sheet: {e}")
        return pd.DataFrame(), {}, pd.DataFrame()


def create_empty_plot_structure(data, headers_row=3, num_columns_per_sample=12):
    """
    Create an empty structure for data collection when sheet has minimal data.
    
    Args:
        data (pd.DataFrame): Original data
        headers_row (int): Row index for headers
        num_columns_per_sample (int): Number of columns per sample
        
    Returns:
        tuple: (processed_data, sample_arrays, full_sample_data)
    """
    debug_print("DEBUG: Creating empty plot structure for data collection")
    
    # Create minimal processed data structure
    processed_data = pd.DataFrame([{
        "Sample Name": "Sample 1",
        "Media": "",
        "Viscosity": "",
        "Voltage, Resistance, Power": "",
        "Average TPM": "No data",
        "Standard Deviation": "No data",
        "Initial Oil Mass": "",
        "Usage Efficiency": "",
        "Burn?": "",
        "Clog?": "",
        "Leak?": ""
    }])
    
    # Empty sample arrays
    sample_arrays = {}
    
    # Use original data or create minimal structure
    if data.empty:
        # Create a minimal data structure
        full_sample_data = pd.DataFrame(
            index=range(10),  # 10 rows for basic structure
            columns=range(num_columns_per_sample)  # Standard column count
        )
    else:
        full_sample_data = data
    
    debug_print(f"DEBUG: Created empty structure - processed: {processed_data.shape}, full: {full_sample_data.shape}")
    return processed_data, sample_arrays, full_sample_data

# ==================== SHEET PROCESSING DISPATCHER ====================

def get_processing_function(sheet_name):
    """
    Get the appropriate processing function for a sheet based on its name.
    
    Args:
        sheet_name (str): Name of the sheet.
        
    Returns:
        callable: Function to process the sheet.
    """
    functions = get_processing_functions()
    if "legacy" in sheet_name.lower():
        return process_legacy_test
    return functions.get(sheet_name, functions["default"])

def get_processing_functions():
    """
    Get a dictionary mapping sheet names to their processing functions.
    
    Returns:
        dict: Map of sheet names to processing functions.
    """
    return {
        "Test Plan": process_test_plan,
        "Initial State Inspection": process_initial_state_inspection,
        "Quick Screening Test": process_quick_screening_test,
        "Lifetime Test": process_device_life_test,
        "Device Life Test": process_device_life_test,
        "Aerosol Temperature": process_aerosol_temp_test,
        "User Test - Full Cycle": process_user_test,
        "User Test Simulation": process_user_test_simulation,
        "User Simulation Test": process_user_test_simulation,
        "Horizontal Puffing Test": process_horizontal_test,
        "Extended Test": process_extended_test,
        "Long Puff Test": process_long_puff_test,
        "Rapid Puff Test": process_rapid_puff_test,
        "Intense Test": process_intense_test,
        "Big Headspace Low T Test": process_big_head_low_t_test,
        "Big Headspace Serial Test": process_big_head_serial_test,
        "Anti-Burn Protection Test": process_burn_protection_test,
        "Big Headspace High T Test": process_big_head_high_t_test,
        "Upside Down Test": process_upside_down_test,
        "Big Headspace Pocket Test": process_pocket_test,
        "Temperature Cycling Test": process_temperature_cycling_test,
        "High T High Humidity Test": process_high_t_high_humidity_test,
        "Low Temperature Stability": process_cold_storage_test,
        "Vacuum Test": process_vacuum_test,
        "Negative Pressure Test": process_vacuum_test,
        "Viscosity Compatibility": process_viscosity_compatibility_test,
        "Various Oil Compatibility": process_various_oil_test,
        "Quick Sensory Test": process_quick_sensory_test,
        "Off-odor Score": process_off_odor_score,
        "Sensory Consistency": process_sensory_consistency,
        "Heavy Metal Leaching Test": process_leaching_test,
        "Sheet1": process_sheet1,
        "default": process_generic_sheet
    }

# ==================== SHEET-SPECIFIC PROCESSING FUNCTIONS ====================
# These functions maintain compatibility with the rest of the codebase
# Many are now simplified wrappers around the more general processing functions

def process_test_plan(data):
    """Process data for the Test Plan sheet."""
    return process_generic_sheet(data, headers_row=5, data_start_row=6)

def process_initial_state_inspection(data):
    """Process data for the Initial State Inspection sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_quick_screening_test(data):
    """Process data for the Quick Screening Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_device_life_test(data):
    """Process data for the Device Life Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12
    )

def process_aerosol_temp_test(data):
    """Process data for the Aerosol Temperature sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_off_odor_score(data):
    """Process data for the Off-odor Score sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_sensory_consistency(data):
    """Process data for the Sensory Consistency sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_user_test(data):
    """Process data for the User Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_horizontal_test(data):
    """Process data for the Horizontal Puffing Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12
    )

def process_extended_test(data):
    """Process data for the Extended Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12
    )

def process_long_puff_test(data):
    """Process data for the Long Puff Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_rapid_puff_test(data):
    """Process data for the Rapid Puff Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_intense_test(data):
    """Process data for the Intense Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12
    )

def process_legacy_test(data, headers_row=3, data_start_row=4, num_columns_per_sample=12):
    """
    Process legacy test data (already converted to standard format) similarly to other plotting sheets.
    Legacy files (after conversion) contain only the key columns (e.g. puffs and TPM (mg/puff))
    and minimal meta_data. This function uses the same logic as process_plot_sheet to generate the
    processed data, sample arrays, and concatenated full_sample_data.
    
    Args:
        data (pd.DataFrame): The legacy file's data in standard format.
        headers_row (int): The row containing headers.
        data_start_row (int): The row at which the data starts.
        num_columns_per_sample (int): Number of columns per sample.
        
    Returns:
        tuple: (processed_data, sample_arrays, full_sample_data)
            processed_data (pd.DataFrame): Processed data for display.
            sample_arrays (dict): Sample arrays for plotting.
            full_sample_data (pd.DataFrame): Concatenated data for all samples.
    """
    processed_data, sample_arrays, full_sample_data = process_plot_sheet(data, headers_row, data_start_row, num_columns_per_sample)
    #print("Legacy data debug:")
    #print("Puffs array:", sample_arrays.get("Sample_1_Puffs"))
    #print("TPM array:", sample_arrays.get("Sample_1_TPM"))
    return processed_data, sample_arrays, full_sample_data

def process_big_head_low_t_test(data):
    """Process data for the Big Headspace Low T Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_burn_protection_test(data):
    """Process data for the Anti-Burn Protection Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_big_head_high_t_test(data):
    """Process data for the Big Headspace High T Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_big_head_serial_test(data):
    """Process data for the Big Headspace Serial Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_upside_down_test(data):
    """Process data for the Upside Down Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_pocket_test(data):
    """Process data for the Big Headspace Pocket Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_temperature_cycling_test(data):
    """Process data for the Temperature Cycling Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_high_t_high_humidity_test(data):
    """Process data for the High T High Humidity Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_cold_storage_test(data):
    """Process data for the Low Temperature Stability sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_vacuum_test(data):
    """Process data for the Vacuum Test sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_viscosity_compatibility_test(data):
    """Process data for the Viscosity Compatibility sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_various_oil_test(data):
    """Process data for the Various Oil Compatibility sheet."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12,
        custom_extracted_data_fn=no_efficiency_extracted_data
    )

def process_quick_sensory_test(data):
    """Process data for the Quick Sensory Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_leaching_test(data):
    """Process data for the Heavy Metal Leaching Test sheet."""
    return process_generic_sheet(data, headers_row=3, data_start_row=4)

def process_sheet1(data):
    """Process data for 'Sheet1' similarly to 'process_extended_test'."""
    return process_plot_sheet(
        data,
        headers_row=3,
        data_start_row=4,
        num_columns_per_sample=12
    )

def process_user_test_simulation(data):
    """
    Process data for the User Test Simulation sheet.
    This test has unique characteristics:
    - 8 columns per sample instead of 12
    - Puffs column is in position 1 (second column) instead of 0
    - Data needs to be split for plotting (0-50 puffs and remaining puffs)
    - Sample ID taken from column header 5 (index 4)
    - Extracts metadata from specific header locations
    """
    debug_print(f"DEBUG: process_user_test_simulation called with data shape: {data.shape}")
    
    # For data collection, allow minimal data (less strict validation)
    min_required_rows = max(3 + 1, 3)  # At least header row + 1
    if data.shape[0] < min_required_rows:
        debug_print(f"DEBUG: Data has {data.shape[0]} rows, minimum required is {min_required_rows}")
        return create_empty_user_test_simulation_structure(data)
    
    if not validate_sheet_data(data, required_rows=min_required_rows):
        debug_print("DEBUG: Sheet validation failed, creating empty structure")
        return create_empty_user_test_simulation_structure(data)

    try:
        # Clean up the data
        data = remove_empty_columns(data).replace(0, np.nan)
        debug_print(f"DEBUG: Data after cleaning: {data.shape}")

        samples = []
        full_sample_data = []
        sample_arrays = {}

        # Calculate the number of potential samples (8 columns per sample)
        num_columns_per_sample = 8
        potential_samples = data.shape[1] // num_columns_per_sample
        debug_print(f"DEBUG: Potential samples based on columns: {potential_samples}")

        # Process each sample block
        for i in range(potential_samples):
            start_col = i * num_columns_per_sample
            end_col = start_col + num_columns_per_sample
            sample_data = data.iloc[:, start_col:end_col]
            
            debug_print(f"DEBUG: Processing sample {i+1} in columns {start_col} to {end_col-1}")
            
            # Check if this sample block has real measurement data
            # Look for numeric data in puffs column (index 1) starting from row 3
            has_real_data = False
            if sample_data.shape[0] > 3 and sample_data.shape[1] > 7:
                puffs_data = pd.to_numeric(sample_data.iloc[3:, 1], errors='coerce').dropna()
                
                if not puffs_data.empty and len(puffs_data) > 0:
                    has_real_data = True
                    debug_print(f"DEBUG: Sample {i+1} has real measurement data - {len(puffs_data)} puff values")
                else:
                    debug_print(f"DEBUG: Sample {i+1} has no real measurement data")
            
            if not has_real_data:
                debug_print(f"DEBUG: Skipping sample {i+1} - no measurement data found")
                continue
                
            # Extract sample name from row 0, column 5 (Sample ID location)
            sample_name = "Sample " + str(len(samples) + 1)  # Default name
            if sample_data.shape[0] > 0 and sample_data.shape[1] > 5:
                sample_id_value = sample_data.columns[5]  # Row 0, Column 5
                if sample_id_value and str(sample_id_value).strip() and str(sample_id_value).strip().lower() != 'nan':
                    sample_name = str(sample_id_value).strip()
                    debug_print(f"DEBUG: Extracted sample name '{sample_name}' from Sample ID location (row 0, col 5)")
                else:
                    debug_print(f"DEBUG: Sample ID location empty or invalid, using default name '{sample_name}'")
            
            # Extract metadata from specific header locations
            media = ""
            voltage = ""
            resistance = ""
            power = ""
            initial_oil_mass = ""
            
            try:
                # Media: Row 1, Column 1
                if sample_data.shape[0] > 1 and sample_data.shape[1] > 1:
                    media_val = sample_data.iloc[0, 1]
                    if media_val and str(media_val).strip().lower() != 'nan':
                        media = str(media_val).strip()
                        debug_print(f"DEBUG: Extracted media: '{media}'")
                
                # Voltage: Row 1, Column 5 
                if sample_data.shape[0] > 1 and sample_data.shape[1] > 5:
                    voltage_val = sample_data.iloc[0, 5]
                    if voltage_val and str(voltage_val).strip().lower() != 'nan':
                        voltage = str(voltage_val).strip()
                        debug_print(f"DEBUG: Extracted voltage: '{voltage}'")
                
                # Initial Oil Mass: Row 0, Column 7
                if sample_data.shape[0] > 0 and sample_data.shape[1] > 7:
                    oil_mass_val = sample_data.columns[7]
                    if oil_mass_val and str(oil_mass_val).strip().lower() != 'nan':
                        initial_oil_mass = str(oil_mass_val).strip()
                        debug_print(f"DEBUG: Extracted initial oil mass: '{initial_oil_mass}'")
                
                # Power: Row 1, Column 7
                if sample_data.shape[0] > 1 and sample_data.shape[1] > 7:
                    power_val = sample_data.iloc[0, 7]
                    if power_val and str(power_val).strip().lower() != 'nan' and str(power_val).strip() != '#DIV/0!':
                        power = str(power_val).strip()
                        debug_print(f"DEBUG: Extracted power: '{power}'")
                
                
                # Resistance: Row 0, Column 3
                if sample_data.shape[0] > 1 and sample_data.shape[1] > 3:
                    resistance_val = sample_data.columns[3]
                    if resistance_val and str(resistance_val).strip().lower() != 'nan' and str(resistance_val).strip() != '#DIV/0!':
                        resistance = str(resistance_val).strip()
                        debug_print(f"DEBUG: Extracted resistance: '{resistance}'")
                
                
            except Exception as e:
                debug_print(f"DEBUG: Error extracting metadata for sample {i+1}: {e}")
            
            try:
                # Extract plotting data for User Test Simulation
                sample_arrays[f"Sample_{len(samples)+1}_Puffs"] = sample_data.iloc[3:, 1].to_numpy()  # Column 1 for puffs
                sample_arrays[f"Sample_{len(samples)+1}_TPM"] = sample_data.iloc[3:, 7].to_numpy()   # Column 7 for TPM
                
                # Calculate TPM statistics from all available data
                tpm_data = pd.to_numeric(sample_data.iloc[3:, 7], errors='coerce')
                valid_tpm = tpm_data.dropna()
                
                if len(valid_tpm) > 0:
                    avg_tpm = round_values(valid_tpm.mean())
                    std_tpm = round_values(valid_tpm.std()) if len(valid_tpm) > 1 else 0
                else:
                    avg_tpm = "No data"
                    std_tpm = "No data"
                
                debug_print(f"DEBUG: Sample '{sample_name}' TPM stats - Avg: {avg_tpm}, Std: {std_tpm}")
                
                usage_efficiency = ""
                try:
                    if sample_data.shape[0] > 1 and sample_data.shape[1] > 8:
                        efficiency_val = sample_data.iloc[1, 8]  # I3 = row 1, column 8 with -1 row indexing
                        if not pd.isna(efficiency_val) and str(efficiency_val).strip().lower() != 'nan':
                            # Handle both percentage and decimal formats
                            efficiency_str = str(efficiency_val).strip()
                            if '%' not in efficiency_str:
                                # If it's a decimal, convert to percentage
                                try:
                                    efficiency_num = float(efficiency_str)
                                    usage_efficiency = f"{round_values(efficiency_num * 100):.1f}%"
                                except:
                                    usage_efficiency = efficiency_str
                            else:
                                usage_efficiency = efficiency_str
                            debug_print(f"DEBUG: User Test Simulation - Extracted usage efficiency from I3: {usage_efficiency}")
                except Exception as e:
                    debug_print(f"DEBUG: User Test Simulation - Error extracting usage efficiency from I3: {e}")
                
                # Create voltage, resistance, power combined string
                voltage_resistance_power = ""
                components = []
                if voltage:
                    components.append(f"V: {voltage}")
                if resistance:
                    components.append(f"R: {resistance}")
                if power:
                    components.append(f"P: {power}")
                voltage_resistance_power = ", ".join(components)
                
                # Create extracted data for this sample
                extracted_data = {
                    "Sample Name": sample_name,
                    "Media": media,
                    "Viscosity": "",  # Not visible in the header image
                    "Voltage, Resistance, Power": voltage_resistance_power,
                    "Average TPM": avg_tpm,
                    "Standard Deviation": std_tpm,
                    "Initial Oil Mass": initial_oil_mass,
                    "Usage Efficiency": usage_efficiency,
                    "Test Type": "User Test Simulation"
                }

                samples.append(extracted_data)
                full_sample_data.append(sample_data)
                debug_print(f"DEBUG: Successfully processed User Test Simulation sample {len(samples)}: '{sample_name}'")
                debug_print(f"DEBUG: Sample data - Media: '{media}', Voltage: '{voltage}', Power: '{power}', Oil Mass: '{initial_oil_mass}'")
                
            except Exception as e:
                debug_print(f"DEBUG: Error processing sample {i+1}: {e}")
                continue

        # Create processed data and full sample data  
        if samples:
            processed_data = pd.DataFrame(samples)
            full_sample_data_df = pd.concat(full_sample_data, axis=1) if full_sample_data else pd.DataFrame()
        else:
            debug_print("DEBUG: No valid samples processed, creating minimal structure")
            processed_data = pd.DataFrame([{
                "Sample Name": "Sample 1",
                "Media": "",
                "Viscosity": "",
                "Voltage, Resistance, Power": "",
                "Average TPM": "No data",
                "Standard Deviation": "No data",
                "Initial Oil Mass": "",
                "Usage Efficiency": "",
                "Test Type": "User Test Simulation"
            }])
            full_sample_data_df = data

        debug_print(f"DEBUG: Final User Test Simulation processed_data shape: {processed_data.shape}")
        debug_print(f"DEBUG: Final User Test Simulation full_sample_data shape: {full_sample_data_df.shape}")
        debug_print(f"DEBUG: process_plot_sheet - using concatenated data: {bool(samples)}")
        debug_print(f"DEBUG: process_plot_sheet - samples count: {len(samples) if samples else 0}")
        return processed_data, sample_arrays, full_sample_data_df
        
    except Exception as e:
        debug_print(f"DEBUG: Error processing User Test Simulation sheet: {e}")
        debug_print(f"DEBUG: Error traceback: {traceback.format_exc()}")
        debug_print(f"DEBUG: process_plot_sheet - using concatenated data: {bool(samples)}")
        debug_print(f"DEBUG: process_plot_sheet - samples count: {len(samples) if samples else 0}")
        return create_empty_user_test_simulation_structure(data)

def create_empty_user_test_simulation_structure(data):
    """
    Create an empty structure for User Test Simulation data collection.
    """
    debug_print("DEBUG: Creating empty User Test Simulation structure for data collection")
    
    # Create minimal processed data structure
    processed_data = pd.DataFrame([{
        "Sample Name": "Sample 1",
        "Media": "",
        "Viscosity": "",
        "Voltage, Resistance, Power": "",
        "Average TPM": "No data",
        "Standard Deviation": "No data",
        "Initial Oil Mass": "",
        "Usage Efficiency": "",
        "Test Type": "User Test Simulation"
    }])
    
    # Empty sample arrays
    sample_arrays = {}
    
    # Use original data or create minimal structure
    num_columns_per_sample = 8  # User Test Simulation uses 8 columns
    if data.empty:
        # Create a minimal data structure
        full_sample_data = pd.DataFrame(
            index=range(10),  # 10 rows for basic structure
            columns=range(num_columns_per_sample)  # 8 columns for User Test Simulation
        )
    else:
        full_sample_data = data
    
    debug_print(f"DEBUG: Created empty User Test Simulation structure - processed: {processed_data.shape}, full: {full_sample_data.shape}")
    return processed_data, sample_arrays, full_sample_data

def extract_burn_clog_leak_from_raw_data(data, sample_index, num_columns_per_sample=12):
    """
    Simple, direct extraction of burn/clog/leak data.
    - Burn: ALWAYS from column header at K (column 10 + offset)
    - Clog: ALWAYS from row 0, column K (column 10 + offset) 
    - Leak: ALWAYS from row 1, column K (column 10 + offset)
    """
    col_offset = sample_index * num_columns_per_sample
    target_col = 10 + col_offset  # K column + offset
    
    print(f"DEBUG: Simple extraction for sample {sample_index + 1}, target column: {target_col}")
    
    def safe_get_cell(row, col, default=""):
        try:
            if row < len(data) and col < len(data.columns):
                val = data.iloc[row, col]
                if pd.isna(val):
                    return ""
                return str(val).strip()
            return default
        except:
            return default
    
    def safe_get_header(col, default=""):
        try:
            if col < len(data.columns):
                header = data.columns[col]
                if pd.isna(header):
                    return ""
                header_str = str(header).strip()
                # If it's "Unnamed", return empty - otherwise return the actual value
                if header_str.startswith('Unnamed'):
                    return ""
                return header_str
            return default
        except:
            return default
    
    # Direct extraction - no searching, no guessing
    burn = safe_get_header(target_col)     # Column header
    clog = safe_get_cell(0, target_col)    # Row 0, same column  
    leak = safe_get_cell(1, target_col)    # Row 1, same column
    
    print(f"  - Column header: '{data.columns[target_col] if target_col < len(data.columns) else 'N/A'}'")
    print(f"  - Row 0 value: '{safe_get_cell(0, target_col)}'")
    print(f"  - Row 1 value: '{safe_get_cell(1, target_col)}'")
    print(f"  - Extracted burn (header): '{burn}'")
    print(f"  - Extracted clog (row 0): '{clog}'")
    print(f"  - Extracted leak (row 1): '{leak}'")
    
    return burn, clog, leak

def updated_extracted_data_function_with_raw_data(sample_data, raw_data, sample_index):
    """
    Updated extraction function with new header structure.
    Enhanced to handle both old and new template formats for sample names with suffix support.
    Now includes Usage Efficiency, Normalized TPM, and Initial Oil Mass.
    """
    print(f"DEBUG: Processing sample {sample_index + 1} with enhanced name extraction and new fields")
    print(f"DEBUG: Sample {sample_index + 1} data shape: {sample_data.shape}")
    
    # Check if sample has sufficient data before processing
    if sample_data.shape[0] < 4 or sample_data.shape[1] < 3:
        print(f"DEBUG: Sample {sample_index + 1} has insufficient data shape {sample_data.shape}, skipping")
        return {
            "Sample Name": f"Sample {sample_index + 1}",
            "Media": "",
            "Viscosity": "",
            "Puffing Regime": "",
            "Voltage, Resistance, Power": "",
            "Average TPM": "No data",
            "Standard Deviation": "No data",
            "Normalized TPM": "",
            "Draw Pressure": "",
            "Usage Efficiency": "",
            "Initial Oil Mass": "",
            "Burn": "",
            "Clog": "",
            "Notes": ""
        }
    
    # Extract sample name - use the same logic that works in plotting functions
    sample_name = f"Sample {sample_index + 1}"  # Default fallback
    try:
        if sample_data.shape[1] > 5:
            sample_name_candidate = sample_data.columns[5]
            if pd.notna(sample_name_candidate):
                sample_name_str = str(sample_name_candidate).strip()
                if sample_name_str and sample_name_str.lower() not in ['nan', 'none', '', 'unnamed: 5']:
                    sample_name = sample_name_str
                    print(f"DEBUG: Extracted sample name from columns[5]: '{sample_name}'")
                else:
                    print(f"DEBUG: Sample name at columns[5] was invalid: '{sample_name_str}', using default")
            else:
                print(f"DEBUG: Sample name at columns[5] was NaN, using default")
        else:
            print(f"DEBUG: Not enough columns for sample name extraction, using default")
    except Exception as e:
        print(f"DEBUG: Error extracting sample name for sample {sample_index + 1}: {e}")
        sample_name = f"Sample {sample_index + 1}"

    print(f"DEBUG: Final sample name for sample {sample_index + 1}: '{sample_name}'")
    
    
    # Extract TPM data for calculations
    tpm_data = pd.Series(dtype=float)
    if sample_data.shape[0] > 3 and sample_data.shape[1] > 8:
        tpm_data = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce').dropna()
    avg_tpm = round_values(tpm_data.mean()) if not tpm_data.empty else None
    std_tpm = round_values(tpm_data.std()) if not tpm_data.empty else None
    
    # Extract draw pressure (average from column 3, similar to TPM)
    draw_pressure_data = pd.Series(dtype=float)
    if sample_data.shape[0] > 3 and sample_data.shape[1] > 3:
        draw_pressure_data = pd.to_numeric(sample_data.iloc[3:, 3], errors='coerce').dropna()
    avg_draw_pressure = round_values(draw_pressure_data.mean()) if not draw_pressure_data.empty else ""
    
    # Extract burn/clog/leak using existing method
    burn, clog, leak = extract_burn_clog_leak_from_raw_data(raw_data, sample_index)
    
    # Extract puffing regime from header data (row 1, column 7)
    puffing_regime = ""
    if sample_data.shape[0] > 1 and sample_data.shape[1] > 7:
        puffing_regime = str(sample_data.iloc[0, 7]) if pd.notna(sample_data.iloc[0, 7]) else ""
    
    # Extract notes from header data (check multiple locations)
    notes = ""
    if sample_data.shape[0] > 2 and sample_data.shape[1] > 7:
        # Check row 2, column 7 for notes
        notes_candidate = str(sample_data.iloc[2, 7]) if pd.notna(sample_data.iloc[2, 7]) else ""
        if notes_candidate and not notes_candidate.lower().startswith('unnamed'):
            notes = notes_candidate
    
    # Keep the existing combined format for voltage, resistance, power
    voltage_resistance_power = ""
    if sample_data.shape[0] > 1 and sample_data.shape[1] > 5:
        voltage = sample_data.iloc[1, 5]
        resistance = round_values(sample_data.iloc[0, 3]) if sample_data.shape[1] > 3 else ""
        power = round_values(sample_data.iloc[0, 5])
        voltage_resistance_power = f"{voltage} V, {resistance} ohm, {power} W"
    
    # NEW: Calculate the three missing fields
    initial_oil_mass = extract_initial_oil_mass(sample_data)
    usage_efficiency = calculate_usage_efficiency_for_sample(sample_data)
    normalized_tpm = calculate_normalized_tpm_for_sample(sample_data, tpm_data)
    
    print(f"DEBUG: Calculated new fields for sample {sample_index + 1}:")
    print(f"  - Initial Oil Mass: '{initial_oil_mass}'")
    print(f"  - Usage Efficiency: '{usage_efficiency}'")
    print(f"  - Normalized TPM: '{normalized_tpm}'")
    
    return {
        "Sample Name": sample_name,
        "Media": sample_data.iloc[0, 1] if sample_data.shape[0] > 0 else "",
        "Viscosity": sample_data.iloc[1, 1] if sample_data.shape[0] > 1 else "",
        "Puffing Regime": puffing_regime,
        "Voltage, Resistance, Power": voltage_resistance_power,  # Combined as requested
        "Average TPM": avg_tpm,
        "Standard Deviation": std_tpm,
        "Normalized TPM": normalized_tpm,  # NEW FIELD
        "Draw Pressure": avg_draw_pressure,
        "Usage Efficiency": usage_efficiency,  # NEW FIELD
        "Initial Oil Mass": initial_oil_mass,  # NEW FIELD
        "Burn": burn,
        "Clog": clog,
        "Notes": notes
    }

def updated_extracted_data_function_with_raw_data_old(sample_data, raw_data, sample_index):
    """
    Updated extraction function that gets burn/clog/leak from raw data.
    Enhanced to handle both old and new template formats for sample names with suffix support.
    """
    print(f"DEBUG: Processing sample {sample_index + 1} with enhanced name extraction")
    print(f"DEBUG: Sample {sample_index + 1} data shape: {sample_data.shape}")
    
    # Check if sample has sufficient data before processing
    if sample_data.shape[0] < 4 or sample_data.shape[1] < 3:
        print(f"DEBUG: Sample {sample_index + 1} has insufficient data shape {sample_data.shape}, skipping")
        return {
            "Sample Name": f"Sample {sample_index + 1}",
            "Media": "",
            "Viscosity": "",
            "Voltage, Resistance, Power": "",
            "Average TPM": "No data",
            "Standard Deviation": "No data", 
            "Initial Oil Mass": "",
            "Usage Efficiency": "",
            "Burn?": "",
            "Clog?": "",
            "Leak?": ""
        }
    
    # Extract sample name with enhanced suffix handling
    sample_name = f"Sample {sample_index + 1}"  # Default fallback
    
    if sample_data.shape[1] > 4:  # Ensure we have enough columns
        headers = sample_data.columns.astype(str)
        
        # Look for sample ID patterns with suffix support
        project_value = None
        sample_value = None
        
        for i, header in enumerate(headers):
            header_lower = str(header).lower().strip()
            
            # Enhanced pattern matching to handle suffixed headers
            # Check for "sample id:" patterns (handles "sample id:", "sample id:.1", etc.)
            if (header_lower == "sample id:" or 
                (header_lower.startswith("sample id:.") and header_lower[11:].isdigit())):
                if i + 1 < len(headers):
                    sample_value = str(headers[i + 1]).strip()
                    break
            
            # Check for old format "project:" patterns
            elif (header_lower == "project:" or 
                  (header_lower.startswith("project:.") and header_lower[9:].isdigit())):
                if i + 1 < len(headers):
                    project_value = str(headers[i + 1]).strip()
                    # Remove pandas suffix from the project value if it exists
                    if '.' in project_value and project_value.split('.')[-1].isdigit():
                        project_value = '.'.join(project_value.split('.')[:-1])
            
            # Check for old format "sample:" patterns  
            elif (header_lower == "sample:" or
                  (header_lower.startswith("sample:.") and header_lower[8:].isdigit())):
                if i + 1 < len(headers):
                    temp_sample_value = str(headers[i + 1]).strip()
                    # Don't overwrite if we already found one from Sample ID
                    if not sample_value:
                        sample_value = temp_sample_value
        
        # Determine final sample name based on what we found
        if sample_value and sample_value.lower() not in ['nan', 'none', '', f'unnamed: {5}']:
            if project_value and project_value.lower() not in ['nan', 'none', '']:
                # Old format: combine project + sample
                sample_name = f"{project_value} {sample_value}".strip()
            else:
                # New format or standalone sample value
                sample_name = sample_value.strip()
        elif project_value and project_value.lower() not in ['nan', 'none', '']:
            sample_name = project_value.strip()
        else:
            # Fallback: try direct column 5 access (for new format)
            if len(headers) > 5:
                fallback_value = str(headers[5]).strip()
                if fallback_value and fallback_value.lower() not in ['nan', 'none', '', 'unnamed: 5']:
                    sample_name = fallback_value
    
    print(f"DEBUG: Final sample name for sample {sample_index + 1}: '{sample_name}'")
    
    # Extract burn/clog/leak from raw data with proper offsets
    burn, clog, leak = extract_burn_clog_leak_from_raw_data(raw_data, sample_index)
    
    # Calculate TPM statistics from sample_data with bounds checking
    tpm_data = pd.Series(dtype=float)
    if sample_data.shape[0] > 3 and sample_data.shape[1] > 8:
        tpm_data = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce').dropna()
    avg_tpm = round_values(tpm_data.mean()) if not tpm_data.empty else None
    std_tpm = round_values(tpm_data.std()) if not tpm_data.empty else None
    
    # Extract media and viscosity with bounds checking
    media = ""
    viscosity = ""
    if sample_data.shape[0] > 1 and sample_data.shape[1] > 1:
        media = str(sample_data.iloc[0, 1]) if not pd.isna(sample_data.iloc[0, 1]) else ""
        viscosity = str(sample_data.iloc[1, 1]) if not pd.isna(sample_data.iloc[1, 1]) else ""
    
    # Extract voltage, resistance, power with bounds checking
    voltage = ""
    resistance = ""
    power = ""
    if sample_data.shape[0] > 1 and sample_data.shape[1] > 5:
        voltage = str(sample_data.iloc[1, 5]) if not pd.isna(sample_data.iloc[1, 5]) else ""
        if sample_data.shape[1] > 3:
            resistance = round_values(sample_data.iloc[0, 3]) if not pd.isna(sample_data.iloc[0, 3]) else ""
        power = round_values(sample_data.iloc[0, 5]) if not pd.isna(sample_data.iloc[0, 5]) else ""
    
    # Calculate usage efficiency using the actual Excel formula logic
    usage_efficiency = ""
    calculated_efficiency = None

    try:
        if sample_data.shape[0] > 3 and sample_data.shape[1] > 8:
            # Get initial oil mass from H3 (column 7, row 1 with -1 indexing)
            initial_oil_mass_val = None
            if sample_data.shape[1] > 7:
                initial_oil_mass_val = sample_data.iloc[1, 7]
        
            # Get puffs values from column A (column 0) starting from row 4 (row 3 with -1 indexing)
            puffs_values = pd.to_numeric(sample_data.iloc[3:, 0], errors='coerce').dropna()
        
            # Get TPM values from column I (column 8) starting from row 4 (row 3 with -1 indexing)
            tpm_values = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce').dropna()
        
            if (initial_oil_mass_val is not None and not pd.isna(initial_oil_mass_val) and 
                len(puffs_values) > 0 and len(tpm_values) > 0):
            
                try:
                    initial_oil_mass_mg = float(initial_oil_mass_val) * 1000  # Convert g to mg
                
                    # Calculate total aerosol mass following Excel formula logic
                    total_aerosol_mass_mg = 0
                
                    # Align the arrays to same length
                    min_length = min(len(puffs_values), len(tpm_values))
                    puffs_aligned = puffs_values.iloc[:min_length]
                    tpm_aligned = tpm_values.iloc[:min_length]
                
                    for i in range(min_length):
                        tpm_val = tpm_aligned.iloc[i]
                        puffs_val = puffs_aligned.iloc[i]
                    
                        if not pd.isna(tpm_val) and not pd.isna(puffs_val):
                            if i == 0:
                                # First measurement: TPM * total puffs
                                aerosol_mass = tpm_val * puffs_val
                            else:
                                # Subsequent measurements: TPM * (current_puffs - previous_puffs)
                                previous_puffs = puffs_aligned.iloc[i-1]
                                if not pd.isna(previous_puffs):
                                    incremental_puffs = puffs_val - previous_puffs
                                    aerosol_mass = tpm_val * incremental_puffs
                                else:
                                    aerosol_mass = tpm_val * puffs_val
                        
                            total_aerosol_mass_mg += aerosol_mass
                
                    # Calculate usage efficiency: (total aerosol mass / initial oil mass) * 100
                    if initial_oil_mass_mg > 0:
                        calculated_efficiency = (total_aerosol_mass_mg / initial_oil_mass_mg) * 100
                        usage_efficiency = f"{round_values(calculated_efficiency):.1f}%"
                    
                        print(f"DEBUG: Usage efficiency calculation for sample {sample_index + 1}:")
                        print(f"  - Initial oil mass: {initial_oil_mass_val}g ({initial_oil_mass_mg}mg)")
                        print(f"  - Data points: {min_length}")
                        print(f"  - Total aerosol mass: {round_values(total_aerosol_mass_mg):.2f}mg")
                        print(f"  - Calculated usage efficiency: {usage_efficiency}")
                    else:
                        print(f"DEBUG: Invalid initial oil mass for sample {sample_index + 1}: {initial_oil_mass_val}")
                    
                except (ValueError, TypeError) as e:
                    print(f"DEBUG: Error converting values for sample {sample_index + 1}: {e}")
            else:
                print(f"DEBUG: Missing data for sample {sample_index + 1}: oil_mass={initial_oil_mass_val}, puffs_count={len(puffs_values)}, tpm_count={len(tpm_values)}")
                
    except Exception as e:
        print(f"DEBUG: Error calculating usage efficiency for sample {sample_index + 1}: {e}")

    
    return {
        "Sample Name": sample_name,
        "Media": media,
        "Viscosity": viscosity,
        "Voltage, Resistance, Power": f"{voltage} V, {resistance} ohm, {power} W" if voltage or resistance or power else "",
        "Average TPM": avg_tpm,
        "Standard Deviation": std_tpm,
        "Initial Oil Mass": initial_oil_mass_val,
        "Usage Efficiency": usage_efficiency,
        "Burn?": burn,
        "Clog?": clog,
        "Leak?": leak
    }

def round_values(value, decimals=2):
    """Helper function to safely round numeric values."""
    try:
        if pd.isna(value):
            return ""
        return round(float(value), decimals)
    except (ValueError, TypeError):
        return value

# ==================== AGGREGATION FUNCTIONS ====================

def aggregate_sheet_metrics(full_sample_data: pd.DataFrame, num_columns_per_sample: int = 12) -> pd.DataFrame:
    """
    Given full_sample_data from a single sheet (which contains multiple samples laid out in groups of 12 columns),
    this function computes, for each sample, several metrics:
      - Average TPM (from column index 8, rows 3 onward)
      - TPM Standard Deviation (from column index 8, rows 3 onward)
      - Total Puffs: determined by locating the last non-empty row in column index 2 (weight) and then taking
        the value from the puffs column (index 0) at that row.
      - Draw Pressure: from the same last valid row in column index 3.
      - Smell: from the same last valid row in column index 5.
      - Date: from meta_data at row index 0, column index 3.
      - Puff Regime: from meta_data at row index 1, column index 7.
      - Initial Oil Mass: from meta_data at row index 2, column index 7.
      - Burn: from meta_data at row index 0, column index 9.
      - Clog: from meta_data at row index 1, column index 9.
      - Leak: from meta_data at row index 2, column index 9.
      - Sample Name: from column index 5 (assumed to contain the sample name).
    
    Returns a DataFrame with one row per sample.
    """
    num_samples = full_sample_data.shape[1] // num_columns_per_sample
    aggregates = []
    for i in range(num_samples):
        start_col = i * num_columns_per_sample
        end_col = start_col + num_columns_per_sample
        sample_data = full_sample_data.iloc[:, start_col:end_col]
        
        # Convert values to numeric; rows before index 3 are assumed header/info rows.
        try:
            tpm_series = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce')
            avg_tpm = tpm_series.mean()
            std_tpm = tpm_series.std()

        except Exception as e:
            avg_tpm = None
            std_tpm = None

        # For the total number of puffs, instead of summing, we locate the last non-empty row in the third column (index 2)
        # and then take the value in the puffs column (index 0) at that row.
        # for draw pressure, take the value corresponding to the last non-empty row in after weight as well. If empty set value to nan
        # for smell, do the same. If empty just set the value to nan
        try:
            # Select rows starting at index 3 for the third column
            weight_series = sample_data.iloc[3:, 2]
            # Find the last valid (non-NaN and non-empty) index
            last_valid_index = weight_series.last_valid_index()
            if last_valid_index is not None:
                total_puffs = pd.to_numeric(
                    sample_data.loc[last_valid_index, sample_data.columns[0]],
                    errors='coerce'
                )
                draw_pressure = pd.to_numeric(
                    sample_data.loc[last_valid_index, sample_data.columns[3]], 
                    errors='coerce'
                )
                smell = pd.to_numeric(
                    sample_data.loc[last_valid_index, sample_data.columns[5]], 
                    errors='coerce'
                )
            else:
                total_puffs = np.nan
                draw_pressure = np.nan
                smell = np.nan

        except Exception as e:
            total_puffs = None
            draw_pressure = None
            smell = None

        # Extract meta_data from the first three rows (indices 0, 1, and 2)
        try:
            date_val = sample_data.iloc[0, 3]  # Date from column 4 (index 3) at row 1 (index 0)
        except Exception:
            date_val = None

        try:
            puff_regime = sample_data.iloc[1, 7]  # Puff regime from column 8 (index 7) at row 2 (index 1)
        except Exception:
            puff_regime = None

        try:
            initial_oil_mass = sample_data.iloc[2, 7]  # Initial oil mass from column 8 (index 7) at row 3 (index 2)
        except Exception:
            initial_oil_mass = None

        try:
            burn = sample_data.iloc[0, 9]  # Burn from column 10 (index 9) at row 1 (index 0)
        except Exception:
            burn = None

        try:
            clog = sample_data.iloc[1, 9]  # Clog from column 10 (index 9) at row 2 (index 1)
        except Exception:
            clog = None

        try:
            leak = sample_data.iloc[2, 9]  # Leak from column 10 (index 9) at row 3 (index 2)
        except Exception:
            leak = None

        # Retrieve the sample name from column index 5
        sample_name = sample_data.columns[5] if len(sample_data.columns) > 5 else f"Sample {i+1}"
        
        aggregates.append({
            "Sample Name": sample_name,
            "Average TPM": avg_tpm,
            "TPM Std Dev": std_tpm,
            "Total Puffs": total_puffs,
            "Draw Pressure": draw_pressure,
            "Smell": smell,
            "Date": date_val,
            "Puff Regime": puff_regime,
            "Initial Oil Mass": initial_oil_mass,
            "Burn": burn,
            "Clog": clog,
            "Leak": leak
        })
        
    return pd.DataFrame(aggregates)

def plot_aggregate_trends(aggregate_df: pd.DataFrame) -> plt.Figure:
    """
    Given a DataFrame of aggregate metrics (with columns "Sample Name", "Average TPM" and "Total Puffs"),
    this function generates a plot that shows both the average TPM (as a bar chart) and the total number of puffs 
    (as a line plot on a secondary axis) across samples.
    """
    fig, ax1 = plt.subplots(figsize=(8, 6))
    samples = aggregate_df["Sample Name"]
    tpm = aggregate_df["Average TPM"]
    puffs = aggregate_df["Total Puffs"]

    # Plot average TPM as bars
    color_tpm = 'tab:blue'
    ax1.set_xlabel("Samples")
    ax1.set_ylabel("Average TPM (mg/puff)", color=color_tpm)
    bars = ax1.bar(samples, tpm, color=color_tpm, alpha=0.6, label="Average TPM")
    ax1.tick_params(axis='y', labelcolor=color_tpm)
    plt.xticks(rotation=45, ha='right')

    # Create a second y-axis for total puffs
    ax2 = ax1.twinx()
    color_puffs = 'tab:red'
    ax2.set_ylabel("Total Puffs", color=color_puffs)
    ax2.plot(samples, puffs, color=color_puffs, marker='o', label="Total Puffs")
    ax2.tick_params(axis='y', labelcolor=color_puffs)

    fig.tight_layout()
    plt.title("Aggregate Metrics per Sample")
    plt.legend(loc='upper left')
    return fig

# ==================== LEGACY FILE PROCESSING FUNCTIONS ====================



def extract_samples_from_old_file(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Extract samples from an old Excel file by robustly locating the headers "puffs" and "TPM (mg/puff)" anywhere in the sheet.
    For each sample header found, a new set of meta_data (from up to 3 rows above) and column data is extracted.
    
    Returns a list of dictionaries, one per sample, each containing:
        - 'sample_name': The value found near the puffs header.
        - 'puffs': A pandas Series of the puffs column.
        - 'tpm': A pandas Series of the TPM column.
        - 'header_row': The index of the header row.
        - Additional meta_data if detected.
    """
    # Read the sheet without assuming a header row.
    df = read_sheet_with_values(file_path, sheet_name)
    samples = []
    nrows, ncols = df.shape

    # Regex patterns for meta_data and data headers.
    meta_data_patterns = {
        "sample_name": [
            r"(cart(ridge)?\s*#|sample\s*(name|id))",  # e.g. "Cart #", "Sample ID"
            r"puffing\s*data\s*for\s*:?\s*" # e.g puffing data for:
        ],
        "resistance": [
            r"\bri\s*\(?\s*ohms?\s*\)?\s*:?\s*",  # e.g. "Ri (Ohms)"
            r"resistance\s*\(?ohms?\)?\s*:?\s*"
        ],
        "voltage": [r"voltage\s*:?\s*"],
        "viscosity": [r"viscosity\b\s*:?\s*"],
        "puffing_regime": [r"\b(puff(ing)?\s*regime|puff\s*settings?)\s*:?\s*"],
        "initial_oil_mass": [r"initial\s*oil\s*mass\b\s*:?\s*"],
        "date": [r"date\s*:?\s*"],
        "media": [r"media\s*:?\s*"]
    }

    data_header_patterns = {
        "puffs": r"puffs",
        "tpm": r"tpm\s*\(mg\s*/\s*puff\)",
        "before_weight": r"before\s*weight/g",
        "after_weight": r"after\s*weight/g",
        "draw_pressure": r"pv1|draw\s*pressure\s*\(kpa\)",
        "smell": r"smell",
        "notes": r"notes"
    }

    numeric_columns = {"puffs", "tpm", "before_weight", "after_weight", "draw_pressure"}

    # Track processed columns for sample data and meta_data separately.
    processed_cols = {row: [] for row in range(nrows)}
    processed_meta_data = {row: [] for row in range(nrows)}
    # Set a threshold: cells within this number of columns are considered too close.
    proximity_threshold = 1

    # Loop over every cell in the sheet.
    for row in range(nrows):
        for col in range(ncols):
            # Skip this cell if it's near a previously processed sample header in the same row.
            if any(abs(col - proc_col) < proximity_threshold for proc_col in processed_cols[row]):
                continue

            cell_val = df.iat[row, col]
            if header_matches(cell_val, data_header_patterns["puffs"]):
                # New sample header found.
                sample = {"sample_name": str(cell_val).strip(), "header_row": row}

                # Mark this column as processed for sample data.
                processed_cols[row].append(col)

                # -----------------------------
                # Extract meta_data (up to 3 rows above)
                # -----------------------------
                start_search_row = max(0, row - 3)
                meta_data_found = {}
                for r in range(row - 1, start_search_row - 1, -1):
                    for c in range(ncols):
                        # Skip if this cell in meta_data row was already used.
                        if any(abs(c - pm) < proximity_threshold for pm in processed_meta_data[r]):
                            continue
                        cell_val_above = df.iat[r, c]
                        for key, patterns in meta_data_patterns.items():
                            if key in meta_data_found:
                                continue  # Already found this key for the current sample.
                            for pattern in patterns:
                                if header_matches(cell_val_above, pattern):
                                    value = df.iat[r, c + 1] if (c + 1 < ncols) else None
                                    meta_data_found[key] = value
                                    # Mark this column as used for meta_data in row r.
                                    processed_meta_data[r].append(c)
                                    break
                            if key in meta_data_found:
                                break
                sample.update(meta_data_found)

                # -----------------------------
                # Extract column data for this sample.
                # -----------------------------
                data_cols = {}
                for key, pattern in data_header_patterns.items():
                    # Look to the right from the current header cell.
                    for j in range(col, ncols):
                        if header_matches(df.iat[row, j], pattern):
                            raw_data = df.iloc[row + 1:, j]
                            if key in numeric_columns:
                                data_cols[key] = pd.to_numeric(raw_data, errors='coerce').dropna()
                            else:
                                data_cols[key] = raw_data.astype(str).replace("nan", "")
                            break
                # Only add the sample if both "puffs" and "tpm" data were found.
                if "puffs" in data_cols and "tpm" in data_cols:
                    sample.update(data_cols)
                    samples.append(sample)
    return samples

def convert_legacy_file_using_template(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Converts a legacy Excel file to the standardized template format.
    - Uses 12-column blocks per sample.
    - Maps meta_data into exact template positions.
    - For each sample block, only rows up to the first empty 'After weight/g'
      are written and all cells below that row (within that block) are cleared.
    """
    # Determine the template path.
    if template_path is None:
        template_path = os.path.join(os.path.abspath("."), "resources", 
                                     "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Load legacy samples using our extraction function.
    legacy_samples = extract_samples_from_old_file(legacy_file_path)
    if not legacy_samples:
        raise ValueError("No valid legacy sample data found.")

    # meta_data mapping: each key maps to a list of regex patterns and the target (row, col offset).
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"], 
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"], 
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping relative to the sample block start.
    # Note: "after_weight" is at column offset 2.
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True),
        "after_weight": (2, True),
        "draw_pressure": (3, True),
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each legacy sample.
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        #debug_print(f"\nProcessing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # --- 1. meta_data Handling ---
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # --- 2. Data Column Handling with Row Slicing ---
        # Determine the cutoff row for this sample block based on "after_weight".
        raw_after_weight = sample.get("after_weight", pd.Series(dtype=object))
        raw_after_weight = raw_after_weight.reset_index(drop=True)
        cutoff = len(raw_after_weight)
        for i, val in enumerate(raw_after_weight):
            if pd.isna(val) or str(val).strip() == "":
                cutoff = i
                break
        #debug_print(f"Sample {sample_idx + 1}: Writing {cutoff} data rows based on 'After weight/g' column.")

        # Now loop over each key in DATA_COL_MAPPING and write only rows up to the cutoff.
        for key, (rel_col, is_numeric) in DATA_COL_MAPPING.items():
            raw_data = sample.get(key, pd.Series(dtype=object))
            raw_data = raw_data.reset_index(drop=True)
            sliced_data = raw_data.iloc[:cutoff]
            if is_numeric:
                data = pd.to_numeric(sliced_data, errors="coerce")
            else:
                data = sliced_data.astype(str).replace("nan", "").replace("None", "")
            for row_offset, value in enumerate(data):
                target_row = 5 + row_offset  # Data starts at row 5.
                target_col = col_offset + rel_col
                try:
                    ws.cell(row=target_row, column=target_col, 
                            value=float(value) if is_numeric else str(value))
                except Exception as e:
                    #debug_print(f"Error writing {key} at ({target_row},{target_col}): {e}")
                    ws.cell(row=target_row, column=target_col, value="ERROR")

        # --- 3. Clearing Extra Rows in the Block ---
        # For rows below the cutoff (in the sample block), clear all cells in the block's columns.
        # We assume the block occupies 12 columns starting at col_offset.
        start_clear_row = 5 + cutoff  # first row to clear in the block.
        # Use the maximum row in the worksheet as the ending point.
        for row_clear in range(start_clear_row, ws.max_row + 1):
            for col_clear in range(col_offset, col_offset + 12):
                ws.cell(row=row_clear, column=col_clear).value = None

    # --- 4. Final Cleanup ---
    # Delete any extra columns beyond the last sample block.
    last_sample_col = len(legacy_samples) * 12
    if ws.max_column > last_sample_col:
        ws.delete_cols(last_sample_col + 1, ws.max_column - last_sample_col)

    # Delete all sheets except the template sheet.
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters).
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"{base_name} Data"[:31]
    ws.title = new_sheet_name
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    
    new_file_name = f"{base_name} Legacy.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)
    #debug_print(f"\nSaved processed file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]

def convert_legacy_standards_using_template(legacy_file_path: str, template_path: str = None) -> dict:
    """
    Converts a legacy standards Excel file to the standardized template format.
    """
    from openpyxl.cell.cell import MergedCell
    
    if template_path is None:
        template_path = os.path.join(os.path.abspath("."), "resources", 
                                    "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
    wb_template = load_workbook(template_path, read_only=False)
    legacy_wb = load_workbook(legacy_file_path, read_only=False)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    new_file_name = f"{base_name} Legacy Standards.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    
    SIMPLE_meta_data_MAPPING = {
        "sample_name": r"(cart(ridge)?\s*#|sample\s*(name|id))",
        "voltage": r"voltage",
        "viscosity": r"viscosity",
        "resistance": r"(ri\s*\(?\s*ohms?\s*\)?|resistance)",
        "puff_regime": r"puff(ing)?\s*regime",
        "initial_oil": r"initial\s*oil\s*mass",
        "date": r"date",
        "media": r"media"
    }
    
    for sheet_name in wb_template.sheetnames:
        # Create a new sheet to avoid merged cell issues
        new_sheet_name = f"{sheet_name}_new"
        new_ws = wb_template.create_sheet(title=new_sheet_name)
        
        if sheet_name in legacy_wb.sheetnames:
            legacy_ws = legacy_wb[sheet_name]
            legacy_df = read_sheet_with_values(legacy_file_path, sheet_name)
            
            # Check if this is a plotting sheet
            is_plotting = plotting_sheet_test(sheet_name, legacy_df)
            
            # Process based on sheet type
            if is_plotting:
                # For plotting sheets: copy the entire legacy sheet
                processed_df = legacy_df.copy()
                meta_data = extract_meta_data(legacy_ws, SIMPLE_meta_data_MAPPING)
                
                # Write to the new sheet
                for i, row in processed_df.iterrows():
                    for j, value in enumerate(row):
                        new_ws.cell(row=i + 1, column=j + 1, value=value)
            else:
                # For non-plotting sheets: copy the legacy sheet
                processed_df = legacy_df.copy()
                for i, row in processed_df.iterrows():
                    for j, value in enumerate(row):
                        new_ws.cell(row=i + 1, column=j + 1, value=value)
        else:
            # If sheet not in legacy file, copy from template sheet
            ws = wb_template[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                for col_idx, value in enumerate(row, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove original sheets and rename new ones
    for sheet_name in list(wb_template.sheetnames):
        if not sheet_name.endswith('_new'):
            del wb_template[sheet_name]
    
    for sheet_name in list(wb_template.sheetnames):
        if sheet_name.endswith('_new'):
            original_name = sheet_name[:-4]  # Remove _new suffix
            wb_template[sheet_name].title = original_name
    
    wb_template.save(new_file_path)
    return load_excel_file(new_file_path)

def extract_samples_from_old_file_v2(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Enhanced version of extract_samples_from_old_file that handles both old and new template formats.
    Old format differences:
    - Has separate "Project:" and "Sample:" fields instead of "Sample ID:"
    - Has "Ri (Ohms)" and "Rf (Ohms)" instead of just "Resistance (Ohms)"
    - Test type is in sheet name, not top-left corner
    
    Returns a list of dictionaries, one per sample, each containing:
        - 'sample_name': Combined from Project + Sample for old format
        - 'puffs': A pandas Series of the puffs column
        - 'tpm': A pandas Series of the TPM column
        - 'header_row': The index of the header row
        - Additional metadata if detected
    """
    print(f"DEBUG: Starting enhanced extraction from file: {file_path}")
    print(f"DEBUG: Target sheet: {sheet_name}")
    
    # Read the sheet without assuming a header row
    df = read_sheet_with_values(file_path, sheet_name)
    samples = []
    nrows, ncols = df.shape
    
    print(f"DEBUG: Sheet dimensions: {nrows} rows x {ncols} columns")

    # Enhanced regex patterns for metadata - now includes both old and new format patterns
    meta_data_patterns = {
        "sample_name": [
            r"(cart(ridge)?\s*#|sample\s*(name|id))",  # New format: "Cart #", "Sample ID"
            r"puffing\s*data\s*for\s*:?\s*",  # New format: "puffing data for:"
            r"^sample\s*:?\s*(?:\.\d+)?$"  # Old format: "Sample:", "Sample:.1", "Sample:.2", etc.
        ],
        "project": [
            r"^project\s*:?\s*(?:\.\d+)?$"  # Old format: "Project:", "Project:.1", "Project:.2", etc.
        ],
        "resistance": [
            r"\bri\s*\(?\s*ohms?\s*\)?\s*:?\s*(?:\.\d+)?",  # Old format: "Ri (Ohms)", "Ri (Ohms).1", etc.
            r"resistance\s*\(?ohms?\)?\s*:?\s*(?:\.\d+)?"   # New format: "Resistance (Ohms)", "Resistance (Ohms).1", etc.
        ],
        "voltage": [r"voltage\s*:?\s*(?:\.\d+)?"],  # "voltage:", "voltage:.1", etc.
        "viscosity": [r"viscosity\b\s*:?\s*(?:\.\d+)?"],  # "viscosity:", "viscosity:.1", etc.
        "puffing_regime": [r"\b(puff(ing)?\s*regime|puff\s*settings?)\s*:?\s*(?:\.\d+)?"]  # With suffixes
    }

    # FIXED: More specific data header patterns to avoid false matches
    data_header_patterns = {
        "puffs": r"^puffs?$",  # EXACT match for "puff" or "puffs" - not part of another phrase
        "tpm": r"\btpm\b",
        "before_weight": r"before.{0,10}weight",
        "after_weight": r"after.{0,10}weight",
        "draw_pressure": r"draw.{0,10}pressure",
        "smell": r"\bsmell\b",
        "notes": r"\bnotes?\b"
    }

    numeric_columns = {"puffs", "tpm", "before_weight", "after_weight", "draw_pressure"}

    # Track which rows and columns have been used for metadata to avoid duplication
    processed_meta_data = {r: [] for r in range(nrows)}
    processed_cols = {r: [] for r in range(nrows)}
    proximity_threshold = 8  # Increased to 8 to account for 12-column sample blocks

    print(f"DEBUG: Starting cell-by-cell scanning...")

    # Cell-by-cell scanning for "puffs" headers (indicates sample start)
    for row in range(nrows):
        for col in range(ncols):
            # Skip this cell if it's near a previously processed sample header in the same row
            if any(abs(col - proc_col) < proximity_threshold for proc_col in processed_cols[row]):
                continue

            cell_val = df.iat[row, col]
            # More strict matching - check if it's exactly "puffs" and not part of another phrase
            if header_matches(cell_val, data_header_patterns["puffs"]):
                print(f"DEBUG: Found puffs header at row {row}, col {col}")
                
                # Additional validation: ensure this is actually the start of a sample block
                # Check if this column is at expected sample positions (0, 12, 24, 36, ...)
                if col % 12 != 0:
                    print(f"DEBUG: Skipping col {col} - not at expected sample start position (should be multiple of 12)")
                    continue
                
                # New sample header found
                sample = {"sample_name": str(cell_val).strip(), "header_row": row}

                # Mark this column as processed for sample data
                processed_cols[row].append(col)

                # Calculate the expected column range for this sample (12 columns per sample)
                sample_start_col = col
                sample_end_col = col + 12
                
                print(f"DEBUG: Searching for metadata from row {max(0, row - 3)} to {row - 1} within columns {sample_start_col} to {sample_end_col}")

                # Extract metadata (up to 3 rows above) within this sample's column range
                start_search_row = max(0, row - 3)
                meta_data_found = {}
                project_value = None
                sample_value = None
                
                # FIXED: More thorough metadata search with better debugging
                for r in range(row - 1, start_search_row - 1, -1):
                    for c in range(sample_start_col, min(sample_end_col, ncols)):
                        # Skip if this cell in metadata row was already used
                        if any(abs(c - pm) < 2 for pm in processed_meta_data[r]):  # Reduced threshold for metadata
                            continue
                        
                        cell_val_above = df.iat[r, c]
                        cell_str = str(cell_val_above).strip().lower()
                        
                        # ENHANCED: More explicit project pattern matching
                        if header_matches(cell_val_above, meta_data_patterns["project"][0]):
                            # Look for the value in the next cell (to the right)
                            if c + 1 < ncols:
                                value = df.iat[r, c + 1]
                                if value and str(value).strip().lower() not in ['nan', 'none', '']:
                                    project_value = str(value).strip()
                                    print(f"DEBUG: Found project value: {project_value}")
                                    processed_meta_data[r].append(c)
                        
                        # ENHANCED: More explicit sample pattern matching
                        elif header_matches(cell_val_above, meta_data_patterns["sample_name"][2]):  # Use the old format pattern
                            # Look for the value in the next cell (to the right)
                            if c + 1 < ncols:
                                value = df.iat[r, c + 1]
                                if value and str(value).strip().lower() not in ['nan', 'none', '']:
                                    sample_value = str(value).strip()
                                    print(f"DEBUG: Found sample value: {sample_value}")
                                    processed_meta_data[r].append(c)
                        
                        # Check for other metadata patterns
                        else:
                            for key, patterns in meta_data_patterns.items():
                                if key in meta_data_found or key in ["project", "sample_name"]:
                                    continue  # Already found this key or handled above
                                for pattern in patterns:
                                    if header_matches(cell_val_above, pattern):
                                        value = df.iat[r, c + 1] if (c + 1 < ncols) else None
                                        meta_data_found[key] = value
                                        processed_meta_data[r].append(c)
                                        break
                                if key in meta_data_found:
                                    break

                # Combine project and sample for old format sample names
                if project_value and sample_value:
                    combined_sample_name = f"{project_value} {sample_value}"
                    print(f"DEBUG: Combined old format sample name: {combined_sample_name}")
                    meta_data_found["sample_name"] = combined_sample_name
                elif project_value:
                    meta_data_found["sample_name"] = project_value
                    print(f"DEBUG: Using project as sample name: {project_value}")
                elif sample_value:
                    meta_data_found["sample_name"] = sample_value
                    print(f"DEBUG: Using sample as sample name: {sample_value}")
                else:
                    # Use fallback sample name based on position
                    fallback_name = f"Sample {len(samples) + 1}"
                    print(f"DEBUG: Using fallback sample name: {fallback_name}")
                    meta_data_found["sample_name"] = fallback_name

                print(f"DEBUG: Final metadata found for sample: {meta_data_found}")
                sample.update(meta_data_found)

                # Extract column data for this sample within the sample's column range
                data_cols = {}
                print(f"DEBUG: Extracting column data starting from col {col}")
                
                for key, pattern in data_header_patterns.items():
                    # Look to the right from the current header cell within the sample range
                    for j in range(col, min(sample_end_col, ncols)):
                        if header_matches(df.iat[row, j], pattern):
                            print(f"DEBUG: Found {key} column at position {j}")
                            raw_data = df.iloc[row + 1:, j]
                            if key in numeric_columns:
                                data_cols[key] = pd.to_numeric(raw_data, errors='coerce').dropna()
                            else:
                                data_cols[key] = raw_data.astype(str).replace("nan", "")
                            break

                # Only add the sample if both "puffs" and "tpm" data were found
                if "puffs" in data_cols and "tpm" in data_cols:
                    sample.update(data_cols)
                    samples.append(sample)
                    print(f"DEBUG: Successfully added sample {len(samples)}: {meta_data_found.get('sample_name', 'Unknown')}")
                else:
                    print(f"DEBUG: Skipping sample - missing required data. Found: {list(data_cols.keys())}")

    print(f"DEBUG: Extraction complete. Found {len(samples)} valid samples")
    for i, sample in enumerate(samples):
        print(f"DEBUG: Sample {i+1}: {sample.get('sample_name', 'Unknown')} - {len(sample.get('puffs', []))} puffs")
    
    return samples

def is_legacy_sample_empty(sample):
    """
    Check if a legacy sample has meaningful data.
    Enhanced to be more graceful with different data formats.
    
    Args:
        sample (dict): Legacy sample data with fields like 'puffs', 'tpm', etc.
        
    Returns:
        bool: True if sample is empty/invalid, False if it has meaningful data
    """
    try:
        # Check for TPM data (primary indicator of valid sample)
        tpm_data = sample.get('tpm', [])
        if hasattr(tpm_data, '__len__') and len(tpm_data) > 0:
            # Convert to list if it's a pandas Series
            if hasattr(tpm_data, 'tolist'):
                tpm_values = tpm_data.tolist()
            else:
                tpm_values = list(tpm_data)
            
            # Check if there are any non-zero, non-NaN TPM values
            valid_tpm_count = 0
            for val in tpm_values:
                try:
                    numeric_val = float(val)
                    if numeric_val > 0 and not math.isnan(numeric_val):
                        valid_tpm_count += 1
                except (ValueError, TypeError):
                    continue
            
            if valid_tpm_count > 0:
                print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has {valid_tpm_count} valid TPM values")
                return False  # Has meaningful data
        
        # Check puffs data as secondary indicator
        puffs_data = sample.get('puffs', [])
        if hasattr(puffs_data, '__len__') and len(puffs_data) > 2:  # Need at least 3 data points
            # If we have substantial puffs data, consider it valid even without TPM
            print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has {len(puffs_data)} puffs data points")
            return False
        
        # Check if this is just metadata without data
        metadata_fields = ['sample_name', 'voltage', 'viscosity', 'resistance', 'media', 'date', 'puffing_regime', 'initial_oil_mass']
        has_metadata = any(sample.get(field, '') for field in metadata_fields)
        
        if has_metadata and not any(sample.get(field, []) for field in ['puffs', 'tpm', 'before_weight', 'after_weight']):
            print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has metadata but no data arrays - treating as empty")
            return True
        
        print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' appears empty - no meaningful data")
        return True  # No meaningful data found
        
    except Exception as e:
        print(f"DEBUG: Error checking legacy sample: {e}")
        return False  # If error, assume not empty to be safe


def filter_legacy_samples(legacy_samples):
    """
    Filter legacy samples to only include those with meaningful data.
    
    Args:
        legacy_samples (list): List of legacy sample dictionaries
        
    Returns:
        list: Filtered list containing only samples with meaningful data
    """
    import math
    
    filtered_samples = []
    
    for i, sample in enumerate(legacy_samples):
        sample_name = sample.get('sample_name', f'Sample {i+1}')
        
        if not is_legacy_sample_empty(sample):
            filtered_samples.append(sample)
            print(f"DEBUG: Keeping sample {len(filtered_samples)}: {sample_name}")
        else:
            print(f"DEBUG: Filtering out empty sample: {sample_name}")
    
    print(f"DEBUG: Filtered from {len(legacy_samples)} to {len(filtered_samples)} meaningful samples")
    return filtered_samples


def convert_legacy_file_using_template_v2(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Enhanced version that handles both old and new template formats.
    Uses the enhanced extraction function and properly handles:
    - Combined Project + Sample fields from old format
    - Ri vs Rf resistance selection
    - Test type detection from sheet name for old format
    - Filters out empty samples before writing to template
    """
    import math  # Add this import for the filtering functions
    print(f"DEBUG: Starting enhanced template conversion for: {legacy_file_path}")
    
    # Determine the template path
    if template_path is None:
        template_path = os.path.join(os.path.abspath("."), "resources", 
                                     "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Load legacy samples using our enhanced extraction function
    all_legacy_samples = extract_samples_from_old_file_v2(legacy_file_path)
    if not all_legacy_samples:
        raise ValueError("No valid legacy sample data found.")

    print(f"DEBUG: Successfully extracted {len(all_legacy_samples)} samples")
    
    # Filter samples to only include those with meaningful data
    legacy_samples = filter_legacy_samples(all_legacy_samples)
    
    if not legacy_samples:
        raise ValueError("No samples with meaningful data found after filtering.")
    
    print(f"DEBUG: Processing {len(legacy_samples)} filtered samples for template conversion")

    # Enhanced metadata mapping with old format support
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"], 
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"], 
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping relative to the sample block start
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True),
        "after_weight": (2, True),
        "draw_pressure": (3, True),
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each FILTERED legacy sample
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        print(f"DEBUG: Processing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # Metadata Handling
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        # Write metadata to template
        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            if value:
                print(f"DEBUG: Set {template_key} to '{value}' at row {tpl_row}, col {col_offset + tpl_col_offset}")
                ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # Data Handling
        for data_key, (data_col_offset, is_numeric) in DATA_COL_MAPPING.items():
            if data_key in sample:
                data_series = sample[data_key]
                target_col = col_offset + data_col_offset
                
                # Convert pandas Series to list for consistent handling
                if hasattr(data_series, 'tolist'):
                    data_values = data_series.tolist()
                elif hasattr(data_series, '__iter__'):
                    data_values = list(data_series)
                else:
                    data_values = [data_series] if data_series else []
                
                # Remove NaN and empty values
                clean_values = []
                for val in data_values:
                    if pd.notna(val) and str(val).strip() != '':
                        clean_values.append(val)
                
                print(f"DEBUG: Writing {data_key} data to column {target_col} ({len(clean_values)} values)")
                
                # Write the data starting from row 4 (data_start_row)
                for row_idx, value in enumerate(clean_values, start=4):
                    if is_numeric:
                        try:
                            numeric_value = float(value)
                            ws.cell(row=row_idx, column=target_col, value=numeric_value)
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=target_col, value=value)
                    else:
                        ws.cell(row=row_idx, column=target_col, value=str(value))

    # Keep only the Intense Test sheet
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"Legacy_{base_name}"[:31]
    ws.title = new_sheet_name
    
    # Ensure legacy data directory exists
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    new_file_name = f"{base_name} Legacy.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)
    
    print(f"DEBUG: Saved processed file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]


def detect_template_format(file_path: str, sheet_name: Optional[str] = None) -> str:
    """
    Detect whether this is an old, new, or cart template format.
    
    Returns:
        "cart_format": Cart format with Cart # in A2, Media in A3, Ri in C2
        "old": Old format with Project/Sample fields and Ri/Rf
        "new": New format with Sample ID and single Resistance
        "unknown": Could not determine format
    """
    try:
        print(f"DEBUG: Detecting template format for {file_path}")
        df = read_sheet_with_values(file_path, sheet_name)
        nrows, ncols = df.shape
        
        # First check for cart format (most specific)
        if nrows > 3 and ncols > 8:
            # Check specific positions for cart format indicators
            a2_val = str(df.iloc[1, 0]).lower() if pd.notna(df.iloc[1, 0]) else ""
            a3_val = str(df.iloc[2, 0]).lower() if pd.notna(df.iloc[2, 0]) else ""
            c2_val = str(df.iloc[1, 2]).lower() if pd.notna(df.iloc[1, 2]) else ""
            
            print(f"DEBUG: Checking cart format - A2: '{a2_val}', A3: '{a3_val}', C2: '{c2_val}'")
            
            # Cart format has "Cart #" in A2, "Media" in A3, and "Ri" in C2
            if ("cart" in a2_val and "#" in a2_val and 
                "media" in a3_val and 
                "ri" in c2_val):
                print("DEBUG: Detected cart format")
                return "cart_format"
        
        # Look for old vs new format indicators in first few rows
        old_format_indicators = 0
        new_format_indicators = 0
        
        for row in range(min(5, nrows)):
            for col in range(min(10, ncols)):
                cell_val = str(df.iat[row, col]).lower().strip()
                
                # Old format indicators
                if re.search(r"project\s*:", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Project:' at row {row}, col {col}")
                if re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Ri (Ohms)' at row {row}, col {col}")
                if re.search(r"rf\s*\(\s*ohms?\s*\)", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Rf (Ohms)' at row {row}, col {col}")
                
                # New format indicators
                if re.search(r"sample\s*(id|name)\s*:", cell_val):
                    new_format_indicators += 1
                    print(f"DEBUG: Found 'Sample ID/Name:' at row {row}, col {col}")
                if re.search(r"resistance\s*\(\s*ohms?\s*\)\s*:", cell_val) and "ri" not in cell_val and "rf" not in cell_val:
                    new_format_indicators += 1
                    print(f"DEBUG: Found 'Resistance (Ohms):' at row {row}, col {col}")
        
        print(f"DEBUG: Format detection - Old indicators: {old_format_indicators}, New indicators: {new_format_indicators}")
        
        if old_format_indicators > new_format_indicators:
            return "old"
        elif new_format_indicators > old_format_indicators:
            return "new"
        else:
            return "unknown"
            
    except Exception as e:
        print(f"DEBUG: Error detecting template format: {e}")
        return "unknown"

def process_legacy_file_auto_detect(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Automatically detect template format and use appropriate processing function.
    Updated to handle cart format.
    """
    print(f"DEBUG: Auto-detecting format for {legacy_file_path}")
    
    format_type = detect_template_format(legacy_file_path)
    print(f"DEBUG: Detected format: {format_type}")
    
    if format_type == "cart_format":
        print("DEBUG: Using cart format processing")
        return convert_cart_format_to_template(legacy_file_path, template_path)
    elif format_type == "old":
        print("DEBUG: Using enhanced processing for old format")
        return convert_legacy_file_using_template_v2(legacy_file_path, template_path)
    else:
        print("DEBUG: Using standard processing for new/unknown format")
        return convert_legacy_file_using_template(legacy_file_path, template_path)

def extract_samples_from_cart_format(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Extract samples from the cart format legacy files.
    
    This format has:
    - Cart # in A2, value in B2
    - Media in A3, value in B3  
    - Ri in C2, value in D2
    - Viscosity in G2, value in H2
    - Voltage in G3, value in H3
    - PV1-PV5 instead of draw pressure
    - TPM in last column (no avg/std dev after)
    """
    print(f"DEBUG: Extracting samples from cart format: {file_path}")
    
    try:
        if sheet_name is None:
            wb = load_workbook(file_path)
            sheet_name = wb.sheetnames[0]
            wb.close()
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        print(f"DEBUG: Loaded sheet with shape: {df.shape}")
        
        samples = []
        nrows, ncols = df.shape
        
        # Find the data start row (look for "Puffs" or numbers in first column)
        data_start_row = None
        header_row = None
        for row in range(min(15, nrows)):
            cell_val = str(df.iloc[row, 0]).lower() if pd.notna(df.iloc[row, 0]) else ""
            if "puff" in cell_val:
                data_start_row = row + 1  # Data starts one row after headers
                header_row = row  # Headers are at this row
                print(f"DEBUG: Found headers at row {row}, data starts at row {row + 1}")
                break
        
        if data_start_row is None or header_row is None:
            print("ERROR: Could not find header/data start row")
            return []
        
        # Print all headers for debugging
        print("DEBUG: Headers found:")
        for col in range(min(ncols, 20)):  # Check first 20 columns
            header_val = str(df.iloc[header_row, col]) if pd.notna(df.iloc[header_row, col]) else ""
            if header_val.strip():
                print(f"  Col {col}: '{header_val}'")
        
        # Find TPM columns by scanning the header row
        tpm_columns = []
        for col in range(ncols):
            header_val = str(df.iloc[header_row, col]).lower() if pd.notna(df.iloc[header_row, col]) else ""
            if "tpm" in header_val:
                tpm_columns.append(col)
                print(f"DEBUG: Found TPM at column {col}: '{df.iloc[header_row, col]}'")
        
        if not tpm_columns:
            print("ERROR: No TPM columns found")
            return []
        
        # For cart format, we expect one main sample. TPM should be the last data column
        # Let's find the sample boundaries more systematically
        
        # Find sample blocks by looking for the standard pattern: Puffs, Before weight, After weight, etc.
        sample_blocks = []
        
        # Look for "Puffs" columns which indicate sample starts
        puffs_columns = []
        for col in range(ncols):
            header_val = str(df.iloc[header_row, col]).lower() if pd.notna(df.iloc[header_row, col]) else ""
            if header_val.strip() == "puffs" or header_val.strip() == "puff":
                puffs_columns.append(col)
                print(f"DEBUG: Found Puffs column at {col}")
        
        if not puffs_columns:
            # If no explicit "Puffs" column, assume first column is puffs for cart format
            puffs_columns = [0]
            print("DEBUG: No 'Puffs' header found, assuming column 0 is puffs")
        
        # For each puffs column, find the corresponding TPM column
        for puffs_col in puffs_columns:
            # Find the nearest TPM column to the right
            nearest_tpm = None
            for tpm_col in tpm_columns:
                if tpm_col > puffs_col:
                    if nearest_tpm is None or tpm_col < nearest_tpm:
                        nearest_tpm = tpm_col
            
            if nearest_tpm is not None:
                sample_blocks.append((puffs_col, nearest_tpm))
                print(f"DEBUG: Sample block from column {puffs_col} to {nearest_tpm}")
        
        # If no sample blocks found, create one assuming standard cart format layout
        if not sample_blocks and tpm_columns:
            sample_blocks = [(0, tpm_columns[0])]
            print(f"DEBUG: Created default sample block from 0 to {tpm_columns[0]}")
        
        print(f"DEBUG: Found {len(sample_blocks)} sample blocks: {sample_blocks}")
        
        # Extract shared metadata first
        shared_metadata = {}
        
        # Cart # / Sample ID from B2 (row 1, col 1)
        sample_id = str(df.iloc[1, 1]) if pd.notna(df.iloc[1, 1]) else ""
        if sample_id and sample_id != 'nan':
            shared_metadata['sample_name'] = sample_id
        
        # Media from B3 (row 2, col 1)
        media = str(df.iloc[2, 1]) if pd.notna(df.iloc[2, 1]) else ""
        shared_metadata['media'] = media
        
        # Ri (resistance) from D2 (row 1, col 3)
        resistance = str(df.iloc[1, 3]) if pd.notna(df.iloc[1, 3]) else ""
        shared_metadata['resistance'] = resistance
        
        # Viscosity from H2 (row 1, col 7)
        viscosity = str(df.iloc[1, 7]) if (ncols > 7 and pd.notna(df.iloc[1, 7])) else ""
        shared_metadata['viscosity'] = viscosity
        
        # Voltage from H3 (row 2, col 7)
        voltage = str(df.iloc[2, 7]) if (ncols > 7 and pd.notna(df.iloc[2, 7])) else ""
        shared_metadata['voltage'] = voltage
        
        print(f"DEBUG: Extracted shared metadata - Sample: '{shared_metadata.get('sample_name', '')}', Media: '{shared_metadata.get('media', '')}', Resistance: '{shared_metadata.get('resistance', '')}', Viscosity: '{shared_metadata.get('viscosity', '')}', Voltage: '{shared_metadata.get('voltage', '')}'")
        
        # Extract each sample
        for sample_idx, (start_col, tpm_col) in enumerate(sample_blocks):
            print(f"DEBUG: Processing sample {sample_idx + 1} from col {start_col} to {tpm_col}")
            
            sample = shared_metadata.copy()
            if sample_idx > 0:  # For multiple samples, add index to name
                base_name = shared_metadata.get('sample_name', 'Sample')
                sample['sample_name'] = f"{base_name}_{sample_idx + 1}"
            
            # Set default values for missing metadata
            sample.setdefault('date', '')
            sample.setdefault('puffing_regime', '')
            sample.setdefault('initial_oil_mass', '')
            
            # Define expected column positions based on standard format
            puffs_col = start_col
            before_weight_col = start_col + 1
            after_weight_col = start_col + 2
            
            # Find PV1 column (look for pressure column between after_weight and TPM)
            pv1_col = None
            for check_col in range(start_col + 3, tpm_col):
                header_val = str(df.iloc[header_row, check_col]).lower() if pd.notna(df.iloc[header_row, check_col]) else ""
                if "pv1" in header_val or "pressure" in header_val:
                    pv1_col = check_col
                    print(f"DEBUG: Found pressure column at {check_col}: '{df.iloc[header_row, check_col]}'")
                    break
            
            # If no specific PV1 found, use the first column after after_weight
            if pv1_col is None and start_col + 3 < tpm_col:
                pv1_col = start_col + 3
                print(f"DEBUG: Using default pressure column at {pv1_col}")
            
            # Extract data arrays
            data_found = False
            try:
                # Puffs data
                puffs_data = df.iloc[data_start_row:, puffs_col]
                puffs_clean = pd.to_numeric(puffs_data, errors='coerce').dropna()
                
                # Before weight data  
                before_weight_data = df.iloc[data_start_row:, before_weight_col] if before_weight_col < ncols else pd.Series(dtype=float)
                before_weight_clean = pd.to_numeric(before_weight_data, errors='coerce').dropna()
                
                # After weight data
                after_weight_data = df.iloc[data_start_row:, after_weight_col] if after_weight_col < ncols else pd.Series(dtype=float)
                after_weight_clean = pd.to_numeric(after_weight_data, errors='coerce').dropna()
                
                # PV1 data (for draw pressure)
                if pv1_col is not None and pv1_col < ncols:
                    pv1_data = df.iloc[data_start_row:, pv1_col]
                    pv1_clean = pd.to_numeric(pv1_data, errors='coerce').dropna()
                else:
                    pv1_clean = pd.Series(dtype=float)
                
                # TPM data
                tpm_data = df.iloc[data_start_row:, tpm_col]
                tpm_clean = pd.to_numeric(tpm_data, errors='coerce').dropna()
                
                print(f"DEBUG: Data extraction results:")
                print(f"  Puffs: {len(puffs_clean)} values")
                print(f"  Before weight: {len(before_weight_clean)} values") 
                print(f"  After weight: {len(after_weight_clean)} values")
                print(f"  PV1/Pressure: {len(pv1_clean)} values")
                print(f"  TPM: {len(tpm_clean)} values")
                
                # Only add sample if we have meaningful data
                if len(puffs_clean) > 0 and len(tpm_clean) > 0:
                    sample['puffs'] = puffs_clean
                    sample['before_weight'] = before_weight_clean
                    sample['after_weight'] = after_weight_clean
                    sample['draw_pressure'] = pv1_clean  # Use PV1 for draw pressure
                    sample['tpm'] = tpm_clean
                    
                    # Add empty arrays for missing columns to maintain compatibility
                    sample['smell'] = pd.Series(dtype=str)
                    sample['notes'] = pd.Series(dtype=str)
                    
                    data_found = True
                    print(f"DEBUG: Sample {sample_idx + 1} successfully processed")
                else:
                    print(f"DEBUG: Sample {sample_idx + 1} has insufficient data - Puffs: {len(puffs_clean)}, TPM: {len(tpm_clean)}")
                
            except Exception as e:
                print(f"ERROR: Failed to extract data for sample {sample_idx + 1}: {e}")
                import traceback
                traceback.print_exc()
            
            if data_found:
                samples.append(sample)
                print(f"DEBUG: Successfully added sample {sample_idx + 1}")
        
        print(f"DEBUG: Extracted {len(samples)} samples from cart format")
        return samples
        
    except Exception as e:
        print(f"ERROR: Failed to extract samples from cart format: {e}")
        import traceback
        traceback.print_exc()
        return []

def convert_cart_format_to_template(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Convert cart format legacy files to standardized template.
    """
    print(f"DEBUG: Converting cart format file: {legacy_file_path}")
    
    # Determine template path
    if template_path is None:
        template_path = os.path.join(os.path.abspath("."), "resources", 
                                     "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Extract samples using cart format function
    all_legacy_samples = extract_samples_from_cart_format(legacy_file_path)
    if not all_legacy_samples:
        raise ValueError("No valid legacy sample data found in cart format.")

    print(f"DEBUG: Successfully extracted {len(all_legacy_samples)} samples")
    
    # Filter samples to only include those with meaningful data
    legacy_samples = filter_legacy_samples(all_legacy_samples)
    
    if not legacy_samples:
        raise ValueError("No samples with meaningful data found after filtering.")
    
    print(f"DEBUG: Processing {len(legacy_samples)} filtered samples for template conversion")

    # Metadata mapping for cart format (same positions as v2)
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"], 
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"], 
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping (same as existing)
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True), 
        "after_weight": (2, True),
        "draw_pressure": (3, True),  # This will use PV1 data
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each sample (using same logic as v2)
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        print(f"DEBUG: Processing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # Write project name to first row
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        # Prepare metadata values
        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        # Write metadata to template
        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            if value:
                print(f"DEBUG: Set {template_key} to '{value}' at row {tpl_row}, col {col_offset + tpl_col_offset}")
                ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # Write data columns (using same logic as v2)
        for data_key, (data_col_offset, is_numeric) in DATA_COL_MAPPING.items():
            if data_key in sample:
                data_series = sample[data_key]
                target_col = col_offset + data_col_offset
                
                # Convert pandas Series to list for consistent handling
                if hasattr(data_series, 'tolist'):
                    data_values = data_series.tolist()
                elif hasattr(data_series, '__iter__'):
                    data_values = list(data_series)
                else:
                    data_values = [data_series] if data_series else []
                
                # Remove NaN and empty values
                clean_values = []
                for val in data_values:
                    if pd.notna(val) and str(val).strip() != '':
                        clean_values.append(val)
                
                print(f"DEBUG: Writing {data_key} data to column {target_col} ({len(clean_values)} values)")
                
                # Write the data starting from row 4 (data_start_row)
                for row_idx, value in enumerate(clean_values, start=4):
                    if is_numeric:
                        try:
                            numeric_value = float(value)
                            ws.cell(row=row_idx, column=target_col, value=numeric_value)
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=target_col, value=value)
                    else:
                        ws.cell(row=row_idx, column=target_col, value=str(value))

    # Keep only the Intense Test sheet
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"Legacy_{base_name}"[:31]
    ws.title = new_sheet_name
    
    # Ensure legacy data directory exists
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    new_file_name = f"{base_name} Legacy Cart.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)
    
    print(f"DEBUG: Saved processed cart format file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]

def calculate_normalized_tpm_for_sample(sample_data, tpm_data):
    """
    Calculate normalized TPM by dividing TPM by puff time.
    Reuses the logic from the plotting functions.
    
    Args:
        sample_data: DataFrame containing sample data
        tpm_data: Series of TPM values
        
    Returns:
        str: Formatted normalized TPM value or empty string
    """
    debug_print("DEBUG: Calculating Normalized TPM for data extraction")
    
    try:
        # Convert TPM data to numeric
        tpm_numeric = pd.to_numeric(tpm_data, errors='coerce').dropna()
        if tpm_numeric.empty:
            debug_print("DEBUG: No valid TPM data for normalization")
            return ""
        
        # Extract puffing regime from row 1, column 8 (index 0, 7)
        puff_time = None
        puffing_regime = None
        
        if sample_data.shape[0] > 0 and sample_data.shape[1] > 7:
            puffing_regime_cell = sample_data.iloc[0, 7]  # Row 1, Column 8 (H)
            if pd.notna(puffing_regime_cell):
                puffing_regime = str(puffing_regime_cell).strip()
                debug_print(f"DEBUG: Found puffing regime: '{puffing_regime}'")
                
                # Extract puff time using regex pattern
                import re
                pattern = r'mL/(\d+(?:\.\d+)?)s/'
                match = re.search(pattern, puffing_regime, re.IGNORECASE)
                if match:
                    puff_time = float(match.group(1))
                    debug_print(f"DEBUG: Extracted puff time: {puff_time}s")
                else:
                    debug_print(f"DEBUG: Could not extract puff time from: '{puffing_regime}'")
        
        # Apply normalization
        if puff_time is not None and puff_time > 0:
            avg_normalized_tpm = (tpm_numeric / puff_time).mean()
            result = f"{round_values(avg_normalized_tpm):.2f}"
            debug_print(f"DEBUG: Calculated normalized TPM: {result} mg/s")
            return result
        else:
            debug_print("DEBUG: Using default puff time of 3.0s for normalization")
            default_puff_time = 3.0
            avg_normalized_tpm = (tmp_numeric / default_puff_time).mean()
            result = f"{round_values(avg_normalized_tpm):.2f}"
            debug_print(f"DEBUG: Calculated normalized TPM with default: {result} mg/s")
            return result
            
    except Exception as e:
        debug_print(f"DEBUG: Error calculating normalized TPM: {e}")
        return ""

def calculate_usage_efficiency_for_sample(sample_data):
    """
    Calculate usage efficiency using the Excel formula logic.
    Formula: ((first_tpm * first_puffs + sum(tpm * incremental_puffs)) / 1000) / initial_oil_mass * 100
    
    Args:
        sample_data: DataFrame containing sample data
        
    Returns:
        str: Formatted usage efficiency percentage or empty string
    """
    debug_print("DEBUG: Calculating usage efficiency for sample")
    
    try:
        if sample_data.shape[0] < 4 or sample_data.shape[1] < 9:
            debug_print(f"DEBUG: Insufficient data shape {sample_data.shape} for usage efficiency calculation")
            return ""
        
        # Get initial oil mass from H3 (column 7, row 1 with -1 indexing = row 2)
        initial_oil_mass_val = sample_data.iloc[1, 7]
        if pd.isna(initial_oil_mass_val) or initial_oil_mass_val == 0:
            debug_print(f"DEBUG: Invalid initial oil mass: {initial_oil_mass_val}")
            return ""
        
        # Get puffs values from column A (column 0) starting from row 4 (row 3 with -1 indexing)
        puffs_values = pd.to_numeric(sample_data.iloc[3:, 0], errors='coerce').dropna()
        
        # Get TPM values from column I (column 8) starting from row 4 (row 3 with -1 indexing)
        tpm_values = pd.to_numeric(sample_data.iloc[3:, 8], errors='coerce').dropna()
        
        if len(puffs_values) == 0 or len(tpm_values) == 0:
            debug_print(f"DEBUG: No valid puffs ({len(puffs_values)}) or TPM ({len(tpm_values)}) data")
            return ""
        
        initial_oil_mass_mg = float(initial_oil_mass_val) * 1000  # Convert g to mg
        debug_print(f"DEBUG: Initial oil mass: {initial_oil_mass_val}g ({initial_oil_mass_mg}mg)")
        
        # Calculate total aerosol mass following Excel formula logic
        total_aerosol_mass_mg = 0
        
        # Align the arrays to same length
        min_length = min(len(puffs_values), len(tpm_values))
        puffs_aligned = puffs_values.iloc[:min_length]
        tpm_aligned = tpm_values.iloc[:min_length]
        
        debug_print(f"DEBUG: Processing {min_length} data points for efficiency calculation")
        
        for i in range(min_length):
            tpm_val = tpm_aligned.iloc[i]
            puffs_val = puffs_aligned.iloc[i]
            
            if not pd.isna(tpm_val) and not pd.isna(puffs_val):
                if i == 0:
                    # First measurement: TPM * total puffs
                    aerosol_mass = tpm_val * puffs_val
                    debug_print(f"DEBUG: Row {i}: First measurement - {tpm_val} * {puffs_val} = {aerosol_mass}")
                else:
                    # Subsequent measurements: TPM * (current_puffs - previous_puffs)
                    previous_puffs = puffs_aligned.iloc[i-1]
                    if not pd.isna(previous_puffs):
                        incremental_puffs = puffs_val - previous_puffs
                        aerosol_mass = tpm_val * incremental_puffs
                        debug_print(f"DEBUG: Row {i}: {tpm_val} * ({puffs_val} - {previous_puffs}) = {aerosol_mass}")
                    else:
                        aerosol_mass = tpm_val * puffs_val
                        debug_print(f"DEBUG: Row {i}: Previous puffs NaN, using {tpm_val} * {puffs_val} = {aerosol_mass}")
                
                total_aerosol_mass_mg += aerosol_mass
        
        # Calculate usage efficiency: (total aerosol mass / initial oil mass) * 100
        if initial_oil_mass_mg > 0:
            calculated_efficiency = (total_aerosol_mass_mg / initial_oil_mass_mg) * 100
            usage_efficiency = f"{round_values(calculated_efficiency):.1f}%"
            
            debug_print(f"DEBUG: Usage efficiency calculation complete:")
            debug_print(f"  - Total aerosol mass: {round_values(total_aerosol_mass_mg):.2f}mg")
            debug_print(f"  - Initial oil mass: {initial_oil_mass_mg}mg")
            debug_print(f"  - Calculated usage efficiency: {usage_efficiency}")
            
            return usage_efficiency
        else:
            debug_print(f"DEBUG: Invalid initial oil mass for efficiency calculation: {initial_oil_mass_mg}")
            return ""
            
    except Exception as e:
        debug_print(f"DEBUG: Error calculating usage efficiency: {e}")
        import traceback
        traceback.print_exc()
        return ""

def extract_initial_oil_mass(sample_data):
    """
    Extract initial oil mass from sample data.
    
    Args:
        sample_data: DataFrame containing sample data
        
    Returns:
        str: Initial oil mass value or empty string
    """
    try:
        if sample_data.shape[0] > 1 and sample_data.shape[1] > 7:
            initial_oil_mass_val = sample_data.iloc[1, 7]  # Row 2, Column 8 (H)
            if pd.notna(initial_oil_mass_val):
                debug_print(f"DEBUG: Extracted initial oil mass: {initial_oil_mass_val}")
                return str(initial_oil_mass_val)
        debug_print("DEBUG: Could not extract initial oil mass")
        return ""
    except Exception as e:
        debug_print(f"DEBUG: Error extracting initial oil mass: {e}")
        return ""
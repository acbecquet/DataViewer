"""
services/file_service.py
Consolidated file I/O operations and calculation service.
This consolidates all file operations from file_manager.py, vap_file_manager.py,
and calculation functionality from processing.py and data_collection_window.py.
"""

# Standard library imports
import os
import re
import copy
import shutil
import time
import uuid
import tempfile
import threading
import subprocess
import traceback
import json
import zipfile
import math
import statistics
from typing import Optional, Dict, Any, List, Tuple, Union
from pathlib import Path
from datetime import datetime

# Third party imports
import pandas as pd
import numpy as np
import psutil
import openpyxl
from openpyxl import Workbook, load_workbook


def debug_print(message: str):
    """Debug print function for file operations."""
    print(f"DEBUG: FileService - {message}")


def round_values(value: float, decimal_places: int = 3) -> float:
    """Round values for display consistency."""
    try:
        return round(float(value), decimal_places)
    except (ValueError, TypeError):
        return 0.0


class FileService:
    """
    Consolidated service for file I/O operations and calculations.
    Handles Excel files, VAP3 files, database operations, and mathematical calculations.
    """
    
    def __init__(self, database_service=None):
        """Initialize the file service."""
        debug_print("Initializing FileService")
        
        # File format support
        self.supported_formats = ['.xlsx', '.xls', '.csv', '.vap3']
        
        # Database integration
        self.database_service = database_service
        
        # File operation caches
        self.loaded_files_cache = {}
        self.stored_files_cache = set()
        
        # TPM calculation parameters
        self.default_puff_time = 3.0
        self.default_puffs = 10
        
        # Calculation settings
        self.precision_digits = 3
        self.statistical_threshold = 2
        
        # Formulation database for viscosity calculations
        self.formulation_database = {}
        self.viscosity_models = {}
        
        debug_print("FileService initialized successfully")
        debug_print(f"Supported formats: {', '.join(self.supported_formats)}")
    
    # ===================== EXCEL FILE OPERATIONS =====================
    
    def load_excel_file(self, file_path: str, legacy_mode: str = None, 
                       skip_database_storage: bool = False, force_reload: bool = False) -> Tuple[bool, Dict[str, Any], str]:
        """
        Load an Excel file and return processed sheet data.
        Enhanced with caching, legacy mode support, and database integration.
        """
        debug_print(f"Loading Excel file: {file_path}")
        debug_print(f"Legacy mode: {legacy_mode}, Skip DB: {skip_database_storage}, Force reload: {force_reload}")
        
        try:
            # Validate file path
            if not Path(file_path).exists():
                return False, {}, f"File not found: {file_path}"
            
            # Check cache if not forcing reload
            cache_key = f"{file_path}_{legacy_mode}"
            if not force_reload and cache_key in self.loaded_files_cache:
                debug_print("Using cached file data")
                return True, self.loaded_files_cache[cache_key], "Loaded from cache"
            
            # Load Excel file with error handling
            try:
                debug_print("Reading Excel file with openpyxl engine")
                sheets_dict = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            except Exception as e:
                debug_print(f"openpyxl failed, trying xlrd: {e}")
                try:
                    sheets_dict = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
                except Exception as e2:
                    return False, {}, f"Failed to read Excel file with both engines: {e2}"
            
            debug_print(f"Loaded {len(sheets_dict)} sheets: {list(sheets_dict.keys())}")
            
            # Process sheets based on legacy mode
            filtered_sheets = {}
            
            for sheet_name, raw_data in sheets_dict.items():
                debug_print(f"Processing sheet: {sheet_name}")
                
                # Apply legacy mode processing if specified
                if legacy_mode == "legacy":
                    processed_data = self._process_legacy_sheet(raw_data, sheet_name)
                elif legacy_mode == "cart":
                    processed_data = self._process_cart_sheet(raw_data, sheet_name)
                else:
                    processed_data = self._process_standard_sheet(raw_data, sheet_name)
                
                if processed_data is not None:
                    filtered_sheets[sheet_name] = {
                        'data': processed_data,
                        'original_data': raw_data.copy(),
                        'sheet_name': sheet_name,
                        'file_path': file_path,
                        'processing_mode': legacy_mode or 'standard',
                        'header_data': self._extract_header_data(raw_data)
                    }
            
            # Cache the results
            result_data = {
                'filtered_sheets': filtered_sheets,
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'load_time': datetime.now().isoformat()
            }
            self.loaded_files_cache[cache_key] = result_data
            
            # Store in database if requested and not already stored
            if not skip_database_storage and self.database_service and file_path not in self.stored_files_cache:
                self._store_file_in_database(file_path, result_data)
            
            debug_print(f"Successfully loaded Excel file with {len(filtered_sheets)} sheets")
            return True, result_data, "Success"
            
        except Exception as e:
            error_msg = f"Failed to load Excel file: {e}"
            debug_print(f"ERROR: {error_msg}")
            traceback.print_exc()
            return False, {}, error_msg
    
    def save_excel_file(self, file_path: str, sheets_data: Dict[str, pd.DataFrame], 
                       preserve_formatting: bool = True) -> Tuple[bool, str]:
        """Save data to an Excel file with optional formatting preservation."""
        debug_print(f"Saving Excel file: {file_path}")
        
        try:
            # Create directory if it doesn't exist
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            if preserve_formatting and Path(file_path).exists():
                # Load existing workbook to preserve formatting
                debug_print("Preserving existing workbook formatting")
                wb = load_workbook(file_path)
                
                for sheet_name, data in sheets_data.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # Update only data cells, preserve formatting
                        for r_idx, row in data.iterrows():
                            for c_idx, value in enumerate(row):
                                # Adjust for Excel 1-based indexing and potential header offset
                                cell_row = r_idx + 4  # Assuming data starts at row 4
                                cell_col = c_idx + 1
                                if cell_row <= ws.max_row and cell_col <= ws.max_column:
                                    ws.cell(row=cell_row, column=cell_col, value=value)
                    else:
                        # Add new sheet if it doesn't exist
                        ws = wb.create_sheet(sheet_name)
                        for r_idx, row in data.iterrows():
                            for c_idx, value in enumerate(row):
                                ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                
                wb.save(file_path)
            else:
                # Create new workbook
                debug_print("Creating new Excel workbook")
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, data in sheets_data.items():
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            debug_print(f"Successfully saved {len(sheets_data)} sheets to {file_path}")
            return True, "Success"
            
        except Exception as e:
            error_msg = f"Failed to save Excel file: {e}"
            debug_print(f"ERROR: {error_msg}")
            return False, error_msg
    
    def _process_standard_sheet(self, raw_data: pd.DataFrame, sheet_name: str) -> Optional[pd.DataFrame]:
        """Process a standard format sheet."""
        debug_print(f"Processing standard sheet: {sheet_name}")
        
        try:
            # Skip non-data sheets
            skip_sheets = ['Summary', 'Charts', 'Graphs', 'Overview']
            if any(skip in sheet_name for skip in skip_sheets):
                debug_print(f"Skipping summary sheet: {sheet_name}")
                return None
            
            # Validate minimum data requirements
            if raw_data.shape[0] < 4 or raw_data.shape[1] < 3:
                debug_print(f"Sheet {sheet_name} has insufficient data")
                return None
            
            # Extract data starting from row 4 (index 3)
            data_rows = raw_data.iloc[3:].copy()
            
            # Clean and process data
            data_rows = self._clean_sheet_data(data_rows)
            
            # Validate required columns exist
            if data_rows.empty:
                debug_print(f"No valid data found in sheet {sheet_name}")
                return None
            
            debug_print(f"Processed standard sheet {sheet_name}: {data_rows.shape}")
            return data_rows
            
        except Exception as e:
            debug_print(f"Error processing standard sheet {sheet_name}: {e}")
            return None
    
    def _process_legacy_sheet(self, raw_data: pd.DataFrame, sheet_name: str) -> Optional[pd.DataFrame]:
        """Process a legacy format sheet with different column mappings."""
        debug_print(f"Processing legacy sheet: {sheet_name}")
        
        try:
            # Legacy format typically has different column structure
            # Map legacy columns to standard format
            if raw_data.shape[1] >= 12:  # Legacy 12-column format
                # Map columns: A->A, K->B, L->C, etc.
                legacy_data = raw_data.iloc[3:].copy()
                
                # Create standard format mapping
                standard_data = pd.DataFrame()
                standard_data.iloc[:, 0] = legacy_data.iloc[:, 0]  # Puffs
                standard_data.iloc[:, 1] = legacy_data.iloc[:, 10]  # Before weight (K->B)
                standard_data.iloc[:, 2] = legacy_data.iloc[:, 11]  # After weight (L->C)
                
                debug_print(f"Converted legacy format for {sheet_name}")
                return self._clean_sheet_data(standard_data)
            else:
                debug_print(f"Legacy sheet {sheet_name} doesn't have expected 12 columns")
                return self._process_standard_sheet(raw_data, sheet_name)
                
        except Exception as e:
            debug_print(f"Error processing legacy sheet {sheet_name}: {e}")
            return self._process_standard_sheet(raw_data, sheet_name)
    
    def _process_cart_sheet(self, raw_data: pd.DataFrame, sheet_name: str) -> Optional[pd.DataFrame]:
        """Process a cart format sheet."""
        debug_print(f"Processing cart sheet: {sheet_name}")
        
        try:
            # Cart format processing - convert to standard format
            cart_data = raw_data.iloc[3:].copy()
            
            # Apply cart-specific transformations
            if 'Cart' not in sheet_name:
                new_sheet_name = f"{sheet_name} Cart"
                debug_print(f"Renaming cart sheet to: {new_sheet_name}")
            
            return self._clean_sheet_data(cart_data)
            
        except Exception as e:
            debug_print(f"Error processing cart sheet {sheet_name}: {e}")
            return self._process_standard_sheet(raw_data, sheet_name)
    
    def _clean_sheet_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate sheet data."""
        if data.empty:
            return data
        
        # Remove completely empty rows
        data = data.dropna(how='all')
        
        # Convert numeric columns
        for col in data.columns:
            if data[col].dtype == 'object':
                data[col] = pd.to_numeric(data[col], errors='coerce')
        
        return data
    
    def _extract_header_data(self, raw_data: pd.DataFrame) -> Dict[str, Any]:
        """Extract header information from the first few rows."""
        header_data = {}
        
        try:
            if raw_data.shape[0] >= 3:
                # Extract sample information
                header_data['samples'] = []
                
                # Look for sample data in first 3 rows
                for i in range(min(3, raw_data.shape[0])):
                    row_data = {}
                    for j in range(min(10, raw_data.shape[1])):  # First 10 columns
                        value = raw_data.iloc[i, j]
                        if pd.notna(value):
                            row_data[f'col_{j}'] = value
                    
                    if row_data:
                        header_data['samples'].append(row_data)
                
                debug_print(f"Extracted header data with {len(header_data['samples'])} sample entries")
        
        except Exception as e:
            debug_print(f"Error extracting header data: {e}")
        
        return header_data
    
    # ===================== VAP3 FILE OPERATIONS =====================
    
    def load_vap3_file(self, file_path: str) -> Tuple[bool, Dict[str, Any], str]:
        """Load a VAP3 file and return its contents."""
        debug_print(f"Loading VAP3 file: {file_path}")
        
        try:
            if not Path(file_path).exists():
                return False, {}, f"VAP3 file not found: {file_path}"
            
            vap3_data = {}
            
            with zipfile.ZipFile(file_path, 'r') as archive:
                # Load main data components
                if 'filtered_sheets.json' in archive.namelist():
                    sheets_data = json.loads(archive.read('filtered_sheets.json').decode('utf-8'))
                    
                    # Convert JSON data back to DataFrames
                    filtered_sheets = {}
                    for sheet_name, sheet_info in sheets_data.items():
                        if 'data' in sheet_info:
                            df = pd.DataFrame(sheet_info['data'])
                            filtered_sheets[sheet_name] = {
                                'data': df,
                                'sheet_name': sheet_name,
                                'header_data': sheet_info.get('header_data', {}),
                                'processing_mode': sheet_info.get('processing_mode', 'standard')
                            }
                    
                    vap3_data['filtered_sheets'] = filtered_sheets
                    debug_print(f"Loaded {len(filtered_sheets)} sheets from VAP3")
                
                # Load plot options
                if 'plot_options.json' in archive.namelist():
                    vap3_data['plot_options'] = json.loads(archive.read('plot_options.json').decode('utf-8'))
                
                # Load plot settings
                if 'plot_settings.json' in archive.namelist():
                    vap3_data['plot_settings'] = json.loads(archive.read('plot_settings.json').decode('utf-8'))
                
                # Load image crop states
                if 'image_crop_states.json' in archive.namelist():
                    vap3_data['image_crop_states'] = json.loads(archive.read('image_crop_states.json').decode('utf-8'))
                
                # Load sample images
                sample_images = {}
                sample_image_crop_states = {}
                
                if 'sample_images/metadata.json' in archive.namelist():
                    metadata = json.loads(archive.read('sample_images/metadata.json').decode('utf-8'))
                    debug_print(f"Found sample images metadata: {metadata.get('sample_count', 0)} samples")
                    
                    # Extract sample images
                    for file_info in archive.infolist():
                        if file_info.filename.startswith('sample_images/') and not file_info.filename.endswith('.json'):
                            # Parse sample ID from path
                            path_parts = file_info.filename.split('/')
                            if len(path_parts) >= 3:
                                sample_id = path_parts[1]
                                
                                if sample_id not in sample_images:
                                    sample_images[sample_id] = []
                                
                                # Extract image to temporary file
                                temp_dir = tempfile.mkdtemp()
                                temp_path = os.path.join(temp_dir, os.path.basename(file_info.filename))
                                
                                with archive.open(file_info.filename) as src, open(temp_path, 'wb') as dst:
                                    dst.write(src.read())
                                
                                sample_images[sample_id].append(temp_path)
                    
                    vap3_data['sample_images'] = sample_images
                    debug_print(f"Loaded sample images for {len(sample_images)} samples")
                
                # Load sample image crop states
                if 'sample_images/crop_states.json' in archive.namelist():
                    vap3_data['sample_image_crop_states'] = json.loads(
                        archive.read('sample_images/crop_states.json').decode('utf-8'))
                
                # Load sample header data
                if 'sample_images/metadata.json' in archive.namelist():
                    metadata = json.loads(archive.read('sample_images/metadata.json').decode('utf-8'))
                    vap3_data['sample_header_data'] = metadata.get('header_data', {})
            
            vap3_data['file_path'] = file_path
            vap3_data['file_name'] = Path(file_path).name
            vap3_data['load_time'] = datetime.now().isoformat()
            
            debug_print(f"Successfully loaded VAP3 file with {len(vap3_data)} components")
            return True, vap3_data, "Success"
            
        except Exception as e:
            error_msg = f"Failed to load VAP3 file: {e}"
            debug_print(f"ERROR: {error_msg}")
            traceback.print_exc()
            return False, {}, error_msg
    
    def save_vap3_file(self, file_path: str, filtered_sheets: Dict, sheet_images: Dict = None,
                      plot_options: List = None, image_crop_states: Dict = None,
                      plot_settings: Dict = None, sample_images: Dict = None,
                      sample_image_crop_states: Dict = None, sample_header_data: Dict = None) -> Tuple[bool, str]:
        """Save data as a comprehensive VAP3 file with sample images."""
        debug_print(f"Saving VAP3 file: {file_path}")
        
        try:
            # Create directory if needed
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as archive:
                # Save filtered sheets
                if filtered_sheets:
                    sheets_data = {}
                    for sheet_name, sheet_info in filtered_sheets.items():
                        if 'data' in sheet_info and isinstance(sheet_info['data'], pd.DataFrame):
                            sheets_data[sheet_name] = {
                                'data': sheet_info['data'].to_dict('records'),
                                'sheet_name': sheet_name,
                                'header_data': sheet_info.get('header_data', {}),
                                'processing_mode': sheet_info.get('processing_mode', 'standard')
                            }
                    
                    archive.writestr('filtered_sheets.json', json.dumps(sheets_data, indent=2))
                    debug_print(f"Saved {len(sheets_data)} sheets to VAP3")
                
                # Save plot options
                if plot_options:
                    archive.writestr('plot_options.json', json.dumps(plot_options, indent=2))
                
                # Save plot settings
                if plot_settings:
                    archive.writestr('plot_settings.json', json.dumps(plot_settings, indent=2))
                
                # Save image crop states
                if image_crop_states:
                    archive.writestr('image_crop_states.json', json.dumps(image_crop_states, indent=2))
                
                # Save sample images with metadata
                if sample_images:
                    debug_print(f"Saving sample images for {len(sample_images)} samples")
                    
                    # Create sample images metadata
                    metadata = {
                        'timestamp': datetime.now().isoformat(),
                        'sample_count': len(sample_images),
                        'header_data': sample_header_data or {}
                    }
                    archive.writestr('sample_images/metadata.json', json.dumps(metadata, indent=2))
                    
                    # Save each sample's images
                    for sample_id, image_paths in sample_images.items():
                        debug_print(f"Saving {len(image_paths)} images for {sample_id}")
                        
                        for i, img_path in enumerate(image_paths):
                            if os.path.exists(img_path):
                                img_ext = os.path.splitext(img_path)[1]
                                sample_img_path = f'sample_images/{sample_id}/image_{i}{img_ext}'
                                archive.write(img_path, sample_img_path)
                    
                    # Save sample image crop states
                    if sample_image_crop_states:
                        archive.writestr('sample_images/crop_states.json',
                                       json.dumps(sample_image_crop_states, indent=2))
            
            debug_print(f"Successfully saved VAP3 file: {file_path}")
            return True, "Success"
            
        except Exception as e:
            error_msg = f"Failed to save VAP3 file: {e}"
            debug_print(f"ERROR: {error_msg}")
            # Clean up corrupted file
            if Path(file_path).exists():
                try:
                    os.remove(file_path)
                except:
                    pass
            return False, error_msg
    
    # ===================== DATABASE INTEGRATION =====================
    
    def _store_file_in_database(self, file_path: str, file_data: Dict[str, Any]) -> bool:
        """Store file data in database."""
        if not self.database_service:
            debug_print("No database service available")
            return False
        
        try:
            debug_print(f"Storing file in database: {file_path}")
            
            # Create temporary VAP3 for database storage
            with tempfile.NamedTemporaryFile(suffix='.vap3', delete=False) as temp_file:
                temp_vap3_path = temp_file.name
            
            # Save current state as VAP3
            success, message = self.save_vap3_file(
                temp_vap3_path,
                file_data.get('filtered_sheets', {}),
                plot_options=file_data.get('plot_options', [])
            )
            
            if success:
                # Store in database
                metadata = {
                    'display_filename': Path(file_path).name,
                    'original_filename': Path(file_path).name,
                    'original_path': file_path,
                    'creation_date': time.strftime('%Y-%m-%d %H:%M:%S'),
                    'sheet_count': len(file_data.get('filtered_sheets', {}))
                }
                
                file_id = self.database_service.store_file(temp_vap3_path, metadata)
                
                if file_id:
                    self.stored_files_cache.add(file_path)
                    debug_print(f"File stored in database with ID: {file_id}")
                    return True
            
            # Clean up temporary file
            try:
                os.remove(temp_vap3_path)
            except:
                pass
            
            return False
            
        except Exception as e:
            debug_print(f"Error storing file in database: {e}")
            return False
    
    # ===================== CALCULATION OPERATIONS =====================
    
    def calculate_tpm_from_weights(self, puffs: pd.Series, before_weights: pd.Series, 
                                   after_weights: pd.Series, test_type: str = "standard") -> pd.Series:
        """Calculate TPM from weight differences and puffing intervals."""
        debug_print(f"Calculating TPM for {test_type} test with {len(puffs)} data points")
        
        try:
            # Calculate weight differences
            weight_diff = before_weights - after_weights
            
            # Calculate puffing intervals
            puffing_intervals = pd.Series(index=puffs.index, dtype=float)
            
            for i, idx in enumerate(puffs.index):
                if i == 0:
                    # First row: use current puffs value
                    current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else self.default_puffs
                    puffing_intervals.loc[idx] = current_puffs
                else:
                    # Subsequent rows: current_puffs - previous_puffs
                    prev_idx = puffs.index[i-1]
                    current_puffs = puffs.loc[idx] if pd.notna(puffs.loc[idx]) else 0
                    prev_puffs = puffs.loc[prev_idx] if pd.notna(puffs.loc[prev_idx]) else 0
                    puff_interval = current_puffs - prev_puffs
                    
                    # Validate interval
                    if puff_interval <= 0 or pd.isna(current_puffs):
                        if pd.isna(current_puffs) or current_puffs == 0:
                            puffing_intervals.loc[idx] = self.default_puffs
                        else:
                            puffing_intervals.loc[idx] = current_puffs
                    else:
                        puffing_intervals.loc[idx] = puff_interval
            
            # Calculate TPM
            tpm_values = pd.Series(index=weight_diff.index, dtype=float)
            
            for idx in weight_diff.index:
                weight = weight_diff.loc[idx]
                interval = puffing_intervals.loc[idx]
                
                if pd.notna(weight) and pd.notna(interval) and interval > 0:
                    tpm = (weight * 1000) / interval  # Convert to mg and divide by puffs
                    tpm_values.loc[idx] = round_values(tpm)
                else:
                    tpm_values.loc[idx] = None
            
            debug_print(f"Calculated TPM for {tpm_values.notna().sum()} valid data points")
            return tpm_values
            
        except Exception as e:
            debug_print(f"Error calculating TPM: {e}")
            return pd.Series(dtype=float)
    
    def calculate_normalized_tpm(self, tpm_data: pd.Series, sample_data: pd.DataFrame) -> str:
        """Calculate normalized TPM by dividing TPM by puff time."""
        debug_print("Calculating normalized TPM")
        
        try:
            # Convert TPM data to numeric
            tpm_numeric = pd.to_numeric(tpm_data, errors='coerce').dropna()
            if tpm_numeric.empty:
                return ""
            
            # Extract puff time from sample data
            puff_time = self.default_puff_time
            
            if sample_data.shape[0] > 0 and sample_data.shape[1] > 7:
                puffing_regime_cell = sample_data.iloc[0, 7]
                if pd.notna(puffing_regime_cell):
                    puffing_regime = str(puffing_regime_cell).strip()
                    
                    # Extract puff time using regex
                    import re
                    pattern = r'mL/(\d+(?:\.\d+)?)'
                    match = re.search(pattern, puffing_regime)
                    if match:
                        puff_time = float(match.group(1))
                        debug_print(f"Extracted puff time: {puff_time}s")
            
            # Calculate normalized TPM
            avg_tpm = tpm_numeric.mean()
            normalized_tpm = avg_tpm / puff_time
            
            result = f"{round_values(normalized_tpm, 2)}"
            debug_print(f"Normalized TPM: {result}")
            return result
            
        except Exception as e:
            debug_print(f"Error calculating normalized TPM: {e}")
            return ""
    
    def calculate_usage_efficiency(self, sample_data: pd.DataFrame, tpm_data: pd.Series) -> str:
        """Calculate usage efficiency from sample data and TPM values."""
        debug_print("Calculating usage efficiency")
        
        try:
            if sample_data.shape[0] < 4 or sample_data.shape[1] < 9:
                return ""
            
            # Get initial oil mass from H3 (column 7, row 1)
            initial_oil_mass_val = sample_data.iloc[1, 7]
            if pd.isna(initial_oil_mass_val):
                return ""
            
            initial_oil_mass = float(initial_oil_mass_val)
            
            # Get puffs and TPM values
            puffs_values = pd.to_numeric(sample_data.iloc[3:, 0], errors='coerce').dropna()
            tpm_values = pd.to_numeric(tpm_data, errors='coerce').dropna()
            
            if puffs_values.empty or tpm_values.empty:
                return ""
            
            # Calculate total mass vaporized
            total_mass_vaporized = tpm_values.sum() / 1000  # Convert mg to g
            
            # Calculate efficiency
            efficiency = (total_mass_vaporized / initial_oil_mass) * 100
            
            result = f"{round_values(efficiency, 1)}%"
            debug_print(f"Usage efficiency: {result}")
            return result
            
        except Exception as e:
            debug_print(f"Error calculating usage efficiency: {e}")
            return ""
    
    def calculate_statistics(self, data: Union[List, pd.Series, np.ndarray]) -> Dict[str, float]:
        """Calculate comprehensive statistics for a dataset."""
        debug_print("Calculating dataset statistics")
        
        try:
            # Convert to numeric array
            if isinstance(data, pd.Series):
                numeric_data = pd.to_numeric(data, errors='coerce').dropna()
            elif isinstance(data, list):
                numeric_data = pd.Series(data, dtype=float).dropna()
            else:
                numeric_data = pd.Series(data, dtype=float).dropna()
            
            if len(numeric_data) < self.statistical_threshold:
                debug_print(f"Insufficient data points: {len(numeric_data)}")
                return {}
            
            stats = {
                'count': len(numeric_data),
                'mean': round_values(numeric_data.mean()),
                'median': round_values(numeric_data.median()),
                'std': round_values(numeric_data.std()),
                'min': round_values(numeric_data.min()),
                'max': round_values(numeric_data.max()),
                'range': round_values(numeric_data.max() - numeric_data.min())
            }
            
            # Add percentiles
            stats['q25'] = round_values(numeric_data.quantile(0.25))
            stats['q75'] = round_values(numeric_data.quantile(0.75))
            stats['iqr'] = round_values(stats['q75'] - stats['q25'])
            
            # Add coefficient of variation
            if stats['mean'] != 0:
                stats['cv'] = round_values((stats['std'] / stats['mean']) * 100)
            else:
                stats['cv'] = 0.0
            
            debug_print(f"Calculated statistics for {stats['count']} data points")
            return stats
            
        except Exception as e:
            debug_print(f"Error calculating statistics: {e}")
            return {}
    
    # ===================== FILE VALIDATION AND UTILITIES =====================
    
    def validate_file_path(self, file_path: str) -> Tuple[bool, str]:
        """Validate if a file path is accessible and supported."""
        if not Path(file_path).exists():
            return False, f"File does not exist: {file_path}"
        
        file_ext = Path(file_path).suffix.lower()
        if file_ext not in self.supported_formats:
            return False, f"Unsupported file format: {file_ext}"
        
        try:
            # Check if file is readable
            Path(file_path).stat()
            return True, "Valid file"
        except PermissionError:
            return False, "Permission denied"
        except Exception as e:
            return False, f"File access error: {e}"
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get comprehensive information about a file."""
        try:
            path = Path(file_path)
            stat = path.stat()
            
            info = {
                'filename': path.name,
                'size_bytes': stat.st_size,
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'modified_time': stat.st_mtime,
                'modified_date': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'extension': path.suffix.lower(),
                'exists': path.exists(),
                'is_supported': path.suffix.lower() in self.supported_formats
            }
            
            # Add format-specific information
            if path.suffix.lower() in ['.xlsx', '.xls']:
                try:
                    sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                    info['sheet_count'] = len(sheets)
                    info['sheet_names'] = list(sheets.keys())
                except:
                    info['sheet_count'] = 0
                    info['sheet_names'] = []
            
            debug_print(f"Retrieved file info for {path.name}")
            return info
            
        except Exception as e:
            debug_print(f"Failed to get file info: {e}")
            return {
                'filename': Path(file_path).name if file_path else 'Unknown',
                'exists': False,
                'error': str(e)
            }
    
    def create_new_template(self, template_type: str = "standard", 
                           test_selections: List[str] = None) -> Tuple[bool, str, str]:
        """Create a new Excel template file."""
        debug_print(f"Creating new template: {template_type}")
        
        try:
            # Create new workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Add sheets based on test selections
            if test_selections:
                for test_name in test_selections:
                    ws = wb.create_sheet(title=test_name)
                    self._setup_template_sheet(ws, template_type)
            else:
                # Default single sheet
                ws = wb.create_sheet(title="Test 1")
                self._setup_template_sheet(ws, template_type)
            
            # Save to temporary file
            temp_dir = tempfile.mkdtemp()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Template_{template_type}_{timestamp}.xlsx"
            file_path = os.path.join(temp_dir, filename)
            
            wb.save(file_path)
            
            debug_print(f"Created template: {file_path}")
            return True, file_path, f"Template created successfully: {filename}"
            
        except Exception as e:
            error_msg = f"Failed to create template: {e}"
            debug_print(f"ERROR: {error_msg}")
            return False, "", error_msg
    
    def _setup_template_sheet(self, worksheet, template_type: str):
        """Set up a template sheet with proper headers and formatting."""
        try:
            # Add standard headers
            headers = [
                ["Sample Information", "", "", "", "", "", "", "", ""],
                ["Media:", "", "Voltage:", "", "", "Power:", "", "", ""],
                ["Viscosity:", "", "", "", "", "", "", "", ""],
                ["Puffs", "Before Weight", "After Weight", "TPM", "", "", "", "", ""]
            ]
            
            for row_idx, row_data in enumerate(headers, 1):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Add sample data rows
            for row in range(5, 25):  # 20 data rows
                worksheet.cell(row=row, column=1, value="")  # Puffs
                worksheet.cell(row=row, column=2, value="")  # Before Weight
                worksheet.cell(row=row, column=3, value="")  # After Weight
                worksheet.cell(row=row, column=4, value="")  # TPM
            
            debug_print(f"Set up {template_type} template sheet")
            
        except Exception as e:
            debug_print(f"Error setting up template sheet: {e}")
    
    def monitor_file_changes(self, file_path: str, callback_func) -> bool:
        """Monitor a file for changes and call callback when modified."""
        debug_print(f"Starting file monitoring for: {file_path}")
        
        try:
            if not Path(file_path).exists():
                return False
            
            initial_mtime = Path(file_path).stat().st_mtime
            
            def monitor_loop():
                while True:
                    try:
                        current_mtime = Path(file_path).stat().st_mtime
                        if current_mtime != initial_mtime:
                            debug_print("File change detected")
                            callback_func(file_path)
                            break
                        time.sleep(1)
                    except FileNotFoundError:
                        debug_print("File was deleted during monitoring")
                        break
                    except Exception as e:
                        debug_print(f"Error during file monitoring: {e}")
                        break
            
            # Start monitoring in separate thread
            monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
            monitor_thread.start()
            
            return True
            
        except Exception as e:
            debug_print(f"Error starting file monitoring: {e}")
            return False
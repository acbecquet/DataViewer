"""
Header Data Processor Module for DataViewer Application

This module handles extraction, validation, and application of header data
from Excel files and loaded sheet data, supporting both old and new formats.
"""

# Standard library imports
import os
import re
import traceback

# Third party imports
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Local imports
from utils import debug_print


class HeaderDataProcessor:
    """Handles header data extraction, validation, and application operations."""
    
    def __init__(self, file_manager):
        """Initialize with reference to parent FileManager."""
        self.file_manager = file_manager
        self.gui = file_manager.gui
        self.root = file_manager.root
        
    def extract_existing_header_data(self, file_path, selected_test):
        """Extract existing header data from Excel files or loaded .vap3 data."""
        debug_print(f"DEBUG: Extracting header data from {file_path} for test {selected_test}")

        try:
            # Check if this is a .vap3 file or temporary file that doesn't exist
            if file_path.endswith('.vap3') or not os.path.exists(file_path):
                debug_print("DEBUG: Detected .vap3 file or non-existent file, extracting from loaded data")
                return self.extract_header_data_from_loaded_sheets(selected_test)
            else:
                debug_print("DEBUG: Detected Excel file, extracting from file using openpyxl")
                return self.extract_header_data_from_excel_file(file_path, selected_test)

        except Exception as e:
            debug_print(f"ERROR: Failed to extract header data: {e}")
            import traceback
            traceback.print_exc()
            return None

    def extract_header_data_from_loaded_sheets(self, selected_test):
        """Extract header data from already-loaded sheet data (for .vap3 files)."""
        debug_print(f"DEBUG: Extracting header data from loaded sheets for test: {selected_test}")

        try:
            # Check if the sheet is loaded
            if not hasattr(self.gui, 'filtered_sheets') or selected_test not in self.gui.filtered_sheets:
                print(f"ERROR: Sheet {selected_test} not found in loaded data")
                return None

            sheet_info = self.gui.filtered_sheets[selected_test]
            debug_print(f"DEBUG: sheet_info keys: {list(sheet_info.keys())}")

            # FOR .VAP3 FILES: Use the stored header_data JSON directly
            if 'header_data' in sheet_info:
                debug_print(f"DEBUG: header_data key exists, value type: {type(sheet_info['header_data'])}")
                debug_print(f"DEBUG: header_data value: {sheet_info['header_data']}")

                if sheet_info['header_data'] is not None:
                    debug_print("DEBUG: Found stored header_data in loaded sheets - using directly")
                    header_data = sheet_info['header_data']

                    # Ensure backwards compatibility for device_type
                    if 'common' in header_data:
                        if 'device_type' not in header_data['common']:
                            header_data['common']['device_type'] = 'T58G'
                            debug_print("DEBUG: Added default device_type T58G for backwards compatibility")

                    # Validate structure
                    debug_print(f"DEBUG: About to validate header_data: {header_data}")
                    if self.validate_header_data(header_data):
                        debug_print(f"DEBUG: Successfully extracted header data from .vap3 for {selected_test}")
                        return header_data
                    else:
                        debug_print("DEBUG: Stored header data failed validation, falling back to extraction")
                else:
                    debug_print("DEBUG: header_data key exists but value is None")
            else:
                debug_print("DEBUG: No header_data key found in sheet_info")

            # FALLBACK: Try to extract from sheet data (for backwards compatibility or corrupted header data)
            print("DEBUG: No valid stored header_data found, attempting extraction from sheet data")
            sheet_data = sheet_info['data']

            if sheet_data.empty:
                debug_print("DEBUG: Sheet data is empty")
                return None

            debug_print(f"DEBUG: Sheet data shape: {sheet_data.shape}")
            debug_print(f"DEBUG: Column headers: {list(sheet_data.columns[:20])}")  # Show first 20 column headers

            # Determine sample count from data structure
            sample_count = self.determine_sample_count_from_data(sheet_data, selected_test)
            debug_print(f"DEBUG: Determined sample count: {sample_count}")

            # Extract header data using simplified structure - only tester and device_type are common
            header_data = {
                'common': {
                    'tester': '',
                    'device_type': 'T58G'
                },
                'samples': [],
                'test': selected_test,
                'num_samples': sample_count
            }

            try:
                # Extract tester from a common location (row 1, column 3)
                if len(sheet_data) > 2 and len(sheet_data.columns) > 3:
                    tester_value = sheet_data.iloc[1, 3] if not pd.isna(sheet_data.iloc[1, 3]) else ""
                    header_data['common']['tester'] = str(tester_value).strip()
                    debug_print(f"DEBUG: Extracted tester: '{header_data['common']['tester']}'")

                # Extract sample-specific data for each sample - each sample gets its own values
                for i in range(sample_count):

                    sample_data = {
                        'id': f'Sample {i+1}',
                        'resistance': '',
                        'media': '',
                        'viscosity': '',
                        'voltage': '',
                        'puffing_regime': '60mL/3s/30s',
                        'oil_mass': ''
                    }

                    debug_print(f"DEBUG: Processing sample {i+1} with column offset {col_offset}")

                    try:
                        # Sample ID - look in the header row (row 0) at the sample's column offset
                        if col_offset < len(sheet_data.columns):
                            sample_id_value = sheet_data.columns[col_offset] if col_offset < len(sheet_data.columns) else ""
                            if pd.notna(sample_id_value) and str(sample_id_value).strip() and not str(sample_id_value).startswith('Unnamed'):
                                sample_data['id'] = str(sample_id_value).strip()
                            else:
                                # Try looking in the actual data rows for sample ID
                                for row_idx in range(min(3, len(sheet_data))):
                                    cell_value = sheet_data.iloc[row_idx, col_offset] if col_offset < len(sheet_data.columns) else ""
                                    if pd.notna(cell_value) and str(cell_value).strip():
                                        sample_data['id'] = str(cell_value).strip()
                                        break

                        # Extract sample-specific data from this sample's column block
                        sample_start_col = col_offset - 5  # Go back to start of this sample's block

                        # Resistance - typically at sample_start_col + 3 (row 0)
                        if sample_start_col + 3 < len(sheet_data.columns) and len(sheet_data) > 0:
                            resistance = sheet_data.iloc[0, sample_start_col + 3]
                            if pd.notna(resistance) and str(resistance).strip():
                                sample_data['resistance'] = str(resistance).strip()
                                debug_print(f"DEBUG: Sample {i+1} resistance: '{sample_data['resistance']}'")

                        # Media - typically at sample_start_col + 1 (row 0)
                        if sample_start_col + 1 < len(sheet_data.columns) and len(sheet_data) > 0:
                            media = sheet_data.iloc[0, sample_start_col + 1]
                            if pd.notna(media) and str(media).strip():
                                sample_data['media'] = str(media).strip()
                                debug_print(f"DEBUG: Sample {i+1} media: '{sample_data['media']}'")

                        # Viscosity - typically at sample_start_col + 1 (row 1)
                        if sample_start_col + 1 < len(sheet_data.columns) and len(sheet_data) > 1:
                            viscosity = sheet_data.iloc[1, sample_start_col + 1]
                            if pd.notna(viscosity) and str(viscosity).strip():
                                sample_data['viscosity'] = str(viscosity).strip()
                                debug_print(f"DEBUG: Sample {i+1} viscosity: '{sample_data['viscosity']}'")

                        # Voltage - typically at sample_start_col + 5 (row 1)
                        if sample_start_col + 5 < len(sheet_data.columns) and len(sheet_data) > 1:
                            voltage = sheet_data.iloc[1, sample_start_col + 5]
                            if pd.notna(voltage) and str(voltage).strip():
                                sample_data['voltage'] = str(voltage).strip()
                                debug_print(f"DEBUG: Sample {i+1} voltage: '{sample_data['voltage']}'")

                        # Puffing regime - typically at sample_start_col + 7 (row 0)
                        if sample_start_col + 7 < len(sheet_data.columns) and len(sheet_data) > 0:
                            puffing = sheet_data.iloc[0, sample_start_col + 7]
                            if pd.notna(puffing) and str(puffing).strip():
                                sample_data['puffing_regime'] = str(puffing).strip()
                                debug_print(f"DEBUG: Sample {i+1} puffing regime: '{sample_data['puffing_regime']}'")

                        # Oil mass - typically at sample_start_col + 7 (row 1)
                        if sample_start_col + 7 < len(sheet_data.columns) and len(sheet_data) > 1:
                            oil_mass = sheet_data.iloc[1, sample_start_col + 7]
                            if pd.notna(oil_mass) and str(oil_mass).strip():
                                sample_data['oil_mass'] = str(oil_mass).strip()
                                debug_print(f"DEBUG: Sample {i+1} oil mass: '{sample_data['oil_mass']}'")

                        debug_print(f"DEBUG: Sample {i+1} final data: {sample_data}")

                    except Exception as e:
                        debug_print(f"DEBUG: Error extracting data for sample {i+1}: {e}")

                    # Add the sample data to the header_data
                    header_data['samples'].append(sample_data)

            except Exception as e:
                debug_print(f"DEBUG: Error extracting header data: {e}")

            debug_print(f"DEBUG: Final extracted header data: {sample_count} samples")
            debug_print(f"DEBUG: Final samples: {header_data['samples']}")
            return header_data

        except Exception as e:
            debug_print(f"ERROR: Failed to extract header data from loaded sheets: {e}")
            import traceback
            traceback.print_exc()
            return None

    def extract_header_data_from_excel_file(self, file_path, selected_test):
        """Extract header data from Excel file with enhanced old format support."""
        try:
            debug_print(f"DEBUG: Extracting header data from Excel file for test: {selected_test}")
            wb = load_workbook(file_path, read_only=True)

            if selected_test not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet '{selected_test}' not found in workbook")
                return None

            ws = wb[selected_test]

            # First, detect if this sheet uses old or new format
            format_type = self.detect_sheet_format(ws)
            debug_print(f"DEBUG: Detected sheet format: {format_type}")

            if format_type == "old":
                return self.extract_old_format_header_data(ws, selected_test)
            else:
                return self.extract_new_format_header_data(ws, selected_test)

        except Exception as e:
            debug_print(f"ERROR: Exception extracting header data from Excel file: {e}")
            traceback.print_exc()
            return None

    def detect_sheet_format(self, ws):
        """Detect if a sheet uses old or new template format."""
        try:
            # Look for old format indicators in first few rows
            old_format_indicators = 0
            new_format_indicators = 0

            for row in range(1, 6):  # Check first 5 rows
                for col in range(1, 11):  # Check first 10 columns
                    try:
                        cell_val = str(ws.cell(row=row, column=col).value or "").lower().strip()

                        # Old format indicators
                        if re.search(r"project\s*:", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Project:' at row {row}, col {col}")
                        if re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Ri (Ohms)' at row {row}, col {col}")
                        if re.search(r"rf\s*\(\s*ohms?\s*\)", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Rf (Ohms)' at row {row}, col {col}")

                        # New format indicators
                        if re.search(r"sample\s*(id|name)\s*:", cell_val):
                            new_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Sample ID/Name:' at row {row}, col {col}")
                        if re.search(r"resistance\s*\(\s*ohms?\s*\)\s*:", cell_val) and "ri" not in cell_val and "rf" not in cell_val:
                            new_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Resistance (Ohms):' at row {row}, col {col}")

                    except Exception:
                        continue

            debug_print(f"DEBUG: Format detection - Old indicators: {old_format_indicators}, New indicators: {new_format_indicators}")

            if old_format_indicators > new_format_indicators:
                return "old"
            elif new_format_indicators > old_format_indicators:
                return "new"
            else:
                return "new"  # Default to new format

        except Exception as e:
            debug_print(f"DEBUG: Error detecting sheet format: {e}")
            return "new"

    def extract_old_format_header_data(self, ws, selected_test):
        """Extract header data from old format sheet - each sample gets individual data."""
        debug_print("DEBUG: Extracting header data using old format logic")

        # Only extract tester and device_type as common data
        common_data = {
            'tester': str(ws.cell(row=1, column=5).value or ""),
            'device_type': 'T58G'
        }

        debug_print(f"DEBUG: Extracted old format common data: {common_data}")

        # Extract sample data by scanning for Project and Sample pairs
        samples = []
        sample_count = 0

        # Keep checking until we don't find any more samples
        i = 0
        while True:
            # Calculate base column for this sample (12 columns per sample)
            base_col = 6 + (i * 12)

            # Check if we've gone beyond the worksheet columns
            if base_col > ws.max_column:
                debug_print(f"DEBUG: Reached end of worksheet at column {base_col}")
                break

            # Look for Project and Sample in the header area
            project_value = None
            sample_value = None

            # Sample-specific data extraction from this sample's column block
            sample_data = {
                'id': f'Sample {i+1}',  # Default, will be overwritten if found
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }

            # Search in the first few rows for Project and Sample within this sample's block
            sample_found = False
            for row in range(1, 4):
                for col_offset in range(min(12, ws.max_column - base_col + 1)):  # Don't go beyond max column
                    col = base_col + col_offset
                    if col > ws.max_column:
                        break

                    try:
                        cell_val = str(ws.cell(row=row, column=col).value or "").lower().strip()
                        next_cell_val = str(ws.cell(row=row, column=col+1).value or "").strip() if col+1 <= ws.max_column else ""

                        if re.search(r"project\s*:", cell_val) and next_cell_val:
                            project_value = next_cell_val
                            sample_found = True
                            debug_print(f"DEBUG: Found project '{project_value}' at row {row}, col {col}")
                        elif re.search(r"sample\s*:", cell_val) and next_cell_val:
                            sample_value = next_cell_val
                            sample_found = True
                            debug_print(f"DEBUG: Found sample '{sample_value}' at row {row}, col {col}")
                        elif re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val) and next_cell_val:
                            sample_data['resistance'] = next_cell_val
                            debug_print(f"DEBUG: Found resistance '{next_cell_val}' at row {row}, col {col}")
                    except Exception:
                        continue

            # If we found a sample, extract all the individual data for this sample
            if sample_found:
                # Create sample ID from project and/or sample values
                if project_value and sample_value:
                    sample_data['id'] = f"{project_value} {sample_value}".strip()
                elif project_value:
                    sample_data['id'] = project_value.strip()
                elif sample_value:
                    sample_data['id'] = sample_value.strip()

                # Extract individual sample data from this sample's column block
                sample_base = i * 12  # 0, 12, 24, 36, etc.

                # Media - typically at column 2 + sample_base
                if 2 + sample_base <= ws.max_column:
                    media_val = str(ws.cell(row=2, column=2 + sample_base).value or "").strip()
                    if media_val:
                        sample_data['media'] = media_val
                        debug_print(f"DEBUG: Sample {i+1} media: '{media_val}'")

                # Viscosity - typically at column 2 + sample_base
                if 2 + sample_base <= ws.max_column:
                    viscosity_val = str(ws.cell(row=3, column=2 + sample_base).value or "").strip()
                    if viscosity_val:
                        sample_data['viscosity'] = viscosity_val
                        debug_print(f"DEBUG: Sample {i+1} viscosity: '{viscosity_val}'")

                # Voltage - look for voltage in this sample's area
                if 5 + sample_base <= ws.max_column:
                    voltage_val = str(ws.cell(row=3, column=6 + sample_base).value or "").strip()
                    if voltage_val:
                        sample_data['voltage'] = voltage_val
                        debug_print(f"DEBUG: Sample {i+1} voltage: '{voltage_val}'")

                # Puffing regime - typically at column 8 + sample_base
                if 8 + sample_base <= ws.max_column:
                    puffing_val = str(ws.cell(row=2, column=8 + sample_base).value or "").strip()
                    if puffing_val:
                        sample_data['puffing_regime'] = puffing_val
                        debug_print(f"DEBUG: Sample {i+1} puffing regime: '{puffing_val}'")

                # Oil mass - look for oil mass in this sample's area
                if 8 + sample_base <= ws.max_column:
                    oil_mass_val = str(ws.cell(row=3, column=8 + sample_base).value or "").strip()
                    if oil_mass_val:
                        sample_data['oil_mass'] = oil_mass_val
                        debug_print(f"DEBUG: Sample {i+1} oil mass: '{oil_mass_val}'")

                samples.append(sample_data)
                sample_count += 1
                debug_print(f"DEBUG: Created old format sample {sample_count} with individual data: {sample_data}")

                # Move to next sample
                i += 1
            else:
                # No more samples found
                debug_print(f"DEBUG: No more old format samples found after checking {i+1} positions")
                break

        debug_print(f"DEBUG: Total samples found: {sample_count}")

        if sample_count == 0:
            debug_print("DEBUG: No old format samples found, using default single sample")
            samples = [{
                'id': 'Sample 1',
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }]
            sample_count = 1

        header_data = {
            'common': common_data,
            'samples': samples,
            'test': selected_test,
            'num_samples': sample_count
        }

        debug_print(f"DEBUG: Final old format header data: {sample_count} samples")
        debug_print(f"DEBUG: Old format samples with individual data: {samples}")
        debug_print(f"DEBUG: Old format common: {common_data}")

        return header_data

    def extract_new_format_header_data(self, ws, selected_test):
        """Extract header data from new format sheet - each sample gets individual data."""
        debug_print("DEBUG: Extracting header data using new format logic")

        # Only extract tester and device_type as common data
        common_data = {
            'tester': str(ws.cell(row=2, column=4).value or ""),
            'device_type': 'T58G'
        }

        debug_print(f"DEBUG: Extracted new format common data: {common_data}")

        # Extract sample data by scanning for sample blocks
        samples = []
        sample_count = 0

        # Keep checking until we don't find any more samples
        i = 0
        while True:
            # Sample ID position: row 1, columns 6, 18, 30, 42, 54, 66, etc.
            sample_id_col = 6 + (i * 12)

            # Check if we've gone beyond the worksheet columns
            if sample_id_col > ws.max_column:
                debug_print(f"DEBUG: Reached end of worksheet at column {sample_id_col}")
                break

            # Check if there's a sample ID at this position
            sample_id_cell = ws.cell(row=1, column=sample_id_col)

            if sample_id_cell.value:
                sample_id = str(sample_id_cell.value).strip()

                # Calculate the base column for this sample's block
                base_col = i * 12

                # Extract ALL sample-specific data from this sample's column block
                sample_data = {
                    'id': sample_id,
                    'resistance': str(ws.cell(row=2, column=4 + base_col).value or "").strip(),
                    'media': str(ws.cell(row=2, column=2 + base_col).value or "").strip(),
                    'viscosity': str(ws.cell(row=3, column=2 + base_col).value or "").strip(),
                    'voltage': str(ws.cell(row=3, column=6 + base_col).value or "").strip(),
                    'puffing_regime': str(ws.cell(row=2, column=8 + base_col).value or "").strip(),
                    'oil_mass': str(ws.cell(row=3, column=8 + base_col).value or "").strip()
                }

                # Clean up the puffing regime value if it has label prefix
                if 'puffing regime:' in sample_data['puffing_regime'].lower():
                    sample_data['puffing_regime'] = sample_data['puffing_regime'].replace('Puffing Regime:', '').strip()

                if not sample_data['puffing_regime']:
                    sample_data['puffing_regime'] = '60mL/3s/30s'  # Default

                samples.append(sample_data)
                sample_count += 1
                debug_print(f"DEBUG: Found new format sample {sample_count} with individual data: {sample_data}")

                # Move to next sample
                i += 1
            else:
                # If no sample ID, we've reached the end of samples
                debug_print(f"DEBUG: No more new format samples found after checking {i+1} positions")
                break

        debug_print(f"DEBUG: Total samples found: {sample_count}")

        if sample_count == 0:
            debug_print("DEBUG: No new format samples found, using default single sample")
            samples = [{
                'id': 'Sample 1',
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }]
            sample_count = 1

        header_data = {
            'common': common_data,
            'samples': samples,
            'test': selected_test,
            'num_samples': sample_count
        }

        debug_print(f"DEBUG: Final new format header data: {sample_count} samples")
        debug_print(f"DEBUG: New format samples with individual data: {samples}")
        debug_print(f"DEBUG: New format common: {common_data}")

        return header_data

    def extract_header_data_from_excel_file_old(self, file_path, selected_test):
        """Extract header data from Excel file using openpyxl (existing method)."""
        debug_print(f"DEBUG: Extracting header data from Excel file: {file_path} for test {selected_test}")

        try:
            wb = openpyxl.load_workbook(file_path)

            if selected_test not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet {selected_test} not found in file. Available sheets: {wb.sheetnames}")
                return None

            ws = wb[selected_test]
            debug_print(f"DEBUG: Successfully opened sheet '{selected_test}'")

            # Extract tester name from corrected position: row 3, column 4
            tester_cell = ws.cell(row=3, column=4)
            tester = ""
            if tester_cell.value:
                tester_value = str(tester_cell.value)
                # Remove any label prefix if present
                if "tester:" in tester_value.lower():
                    tester = tester_value.split(":", 1)[1].strip() if ":" in tester_value else ""
                else:
                    tester = tester_value.strip()

            debug_print(f"DEBUG: Extracted tester: '{tester}' from cell D3: '{tester_cell.value}'")

            # Extract common data from corrected positions
            common_data = {
                'tester': tester,
                'media': str(ws.cell(row=2, column=2).value or ""),          # Row 2, Col B
                'viscosity': str(ws.cell(row=3, column=2).value or ""),      # Row 3, Col B
                'voltage': str(ws.cell(row=3, column=6).value or ""),        # Row 3, Col F (corrected)
                'oil_mass': str(ws.cell(row=3, column=8).value or ""),       # Row 3, Col H (corrected)
                'puffing_regime': str(ws.cell(row=2, column=8).value or "Standard")  # Row 2, Col H
            }

            debug_print(f"DEBUG: Extracted common data with corrected positions: {common_data}")

            # Extract sample data by scanning the first row for sample IDs
            samples = []
            sample_count = 0

            # Scan the first row looking for sample blocks (starting from column 6, then every 12 columns)
            for i in range(6):  # Check up to 6 samples maximum
                # Sample ID position: row 1, columns 6, 18, 30, 42, 54, 66
                sample_id_col = 6 + (i * 12)
                resistance_col = 4 + (i * 12)  # Row 2, columns 4, 16, 28, 40, 52, 64

                # Check if there's a sample ID at this position
                sample_id_cell = ws.cell(row=1, column=sample_id_col)
                resistance_cell = ws.cell(row=2, column=resistance_col)

                if sample_id_cell.value:
                    sample_id = str(sample_id_cell.value).strip()
                    resistance = str(resistance_cell.value or "").strip()

                    samples.append({
                        'id': str(sample_id).strip(),
                        'resistance': str(resistance).strip() if resistance else ""
                    })
                    sample_count += 1
                    debug_print(f"DEBUG: Found valid sample {sample_count}: ID='{sample_id}', Resistance='{resistance}'")
                else:
                    # If no sample ID, we've reached the end of samples
                    debug_print(f"DEBUG: No more samples found after checking {i+1} positions")
                    break

            if sample_count == 0:
                debug_print("DEBUG: No samples found, using default single sample")
                samples = [{'id': 'Sample 1', 'resistance': ''}]
                sample_count = 1

            header_data = {
                'common': common_data,
                'samples': samples,
                'test': selected_test,
                'num_samples': sample_count
            }

            debug_print(f"DEBUG: Final extracted header data from Excel file: {sample_count} samples")
            debug_print(f"DEBUG: Samples: {samples}")
            debug_print(f"DEBUG: Common: {common_data}")

            wb.close()
            return header_data

        except Exception as e:
            debug_print(f"ERROR: Exception extracting header data from Excel file: {e}")
            traceback.print_exc()
            return None

    def migrate_header_data_for_backwards_compatibility(self, header_data):
        """Ensure header data has device type field for backwards compatibility."""
        if header_data and 'common' in header_data:
            if 'device_type' not in header_data['common']:
                header_data['common']['device_type'] = None
                debug_print("DEBUG: Added device_type: None for backwards compatibility")
        return header_data

    def validate_header_data(self, header_data):
        """Validate that header data has sufficient content for data collection."""
        debug_print("DEBUG: Validating extracted header data")

        if not header_data:
            debug_print("DEBUG: Header data is None")
            return False

        # Check for samples - this is the only requirement to proceed
        samples = header_data.get('samples', [])
        if not samples:
            debug_print("DEBUG: No samples found")
            return False

        # Check that at least one sample has an ID
        has_valid_sample = False
        for sample in samples:
            if sample.get('id', '').strip():
                has_valid_sample = True
                break

        if not has_valid_sample:
            debug_print("DEBUG: No samples with valid IDs found")
            return False

        debug_print(f"DEBUG: Found {len(samples)} samples with valid IDs - proceeding to data collection")
        debug_print("DEBUG: Header data validation passed")
        return True

    def apply_header_data_to_file(self, file_path, header_data):
        """
        Apply the header data to the Excel file.
        Enhanced to correctly apply headers to all sample blocks with proper column mapping.
        """
        try:
            debug_print(f"DEBUG: Applying header data to {file_path} for {header_data['num_samples']} samples")

            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Get the sheet for the selected test
            if header_data["test"] in wb.sheetnames:
                ws = wb[header_data["test"]]

                debug_print(f"DEBUG: Successfully opened sheet '{header_data['test']}'")

                # Set the test name at row 1, column 1 (this should be done once)
                ws.cell(row=1, column=1, value=header_data["test"])
                debug_print(f"DEBUG: Set test name '{header_data['test']}' at row 1, column 1")

                # Get common data once
                common_data = header_data["common"]
                samples_data = header_data.get("samples", [])

                # Apply sample-specific data for each sample block
                num_samples = header_data["num_samples"]
                for i in range(num_samples):
                    # Calculate column offset (12 columns per sample)
                    # Sample blocks start at column 1, so offsets are 0, 12, 24, 36, etc.
                    col_offset = i * 12
                    sample_data = samples_data[i] if i < len(samples_data) else {}

                    debug_print(f"DEBUG: Processing sample {i+1} with column offset {col_offset}")
                    debug_print(f"DEBUG: Sample {i+1}: ID='{sample_data.get('id', '')}', Resistance='{sample_data.get('resistance', '')}'")

                    # Row 1, Column F (6) + offset: Sample ID
                    sample_id = sample_data.get('id', f'Sample {i+1}')
                    ws.cell(row=1, column=6 + col_offset, value=sample_id)
                    debug_print(f"DEBUG: Set sample ID '{sample_id}' at row 1, column {6 + col_offset}")

                    # Row 2, Column D (4) + offset: Resistance
                    resistance = sample_data.get("resistance", "")
                    if resistance:
                        try:
                            resistance_value = float(resistance)
                            ws.cell(row=2, column=4 + col_offset, value=resistance_value)
                        except ValueError:
                            ws.cell(row=2, column=4 + col_offset, value=resistance)
                        debug_print(f"DEBUG: Set resistance '{resistance}' at row 2, column {4 + col_offset}")

                    # Row 3, Column D (4) + offset: Tester name (from common data)
                    tester_name = common_data.get("tester", "")
                    if tester_name:
                        tester_col = 4 + col_offset
                        ws.cell(row=3, column=tester_col, value=tester_name)
                        debug_print(f"DEBUG: Set tester '{tester_name}' at row 3, column {tester_col} for sample {i+1}")

                    # Sample-specific data from the sample record
                    # Row 2, Column B (2) + offset: Media
                    media = sample_data.get("media", "")
                    if media:
                        media_col = 2 + col_offset
                        ws.cell(row=2, column=media_col, value=media)
                        debug_print(f"DEBUG: Set media '{media}' at row 2, column {media_col} for sample {i+1}")

                    # Row 3, Column B (2) + offset: Viscosity
                    viscosity = sample_data.get("viscosity", "")
                    if viscosity:
                        viscosity_col = 2 + col_offset
                        try:
                            viscosity_value = float(viscosity)
                            ws.cell(row=3, column=viscosity_col, value=viscosity_value)
                        except ValueError:
                            ws.cell(row=3, column=viscosity_col, value=viscosity)
                        debug_print(f"DEBUG: Set viscosity '{viscosity}' at row 3, column {viscosity_col} for sample {i+1}")

                    # Row 3, Column F (6) + offset: Voltage
                    voltage = sample_data.get("voltage", "")
                    if voltage:
                        try:
                            voltage_value = float(voltage)
                            ws.cell(row=3, column=6 + col_offset, value=voltage_value)
                        except ValueError:
                            ws.cell(row=3, column=6 + col_offset, value=voltage)

                    # Row 2, Column F (6) + offset: Calculated Power
                    calculated_power = sample_data.get("calculated_power", "")
                    if calculated_power:
                        try:
                            power_value = float(calculated_power)
                            ws.cell(row=2, column=6 + col_offset, value=power_value)
                            debug_print(f"DEBUG: Set calculated power '{power_value}' at row 2, column {6 + col_offset} for sample {i+1}")
                        except ValueError:
                            ws.cell(row=2, column=6 + col_offset, value=calculated_power)
                            debug_print(f"DEBUG: Set calculated power (string) '{calculated_power}' at row 2, column {6 + col_offset} for sample {i+1}")
                    else:
                        debug_print(f"DEBUG: No calculated power available for sample {i+1}")

                    # Row 3, Column H (8) + offset: Oil Mass
                    oil_mass = sample_data.get("oil_mass", "")
                    if oil_mass:
                        oil_mass_col = 8 + col_offset
                        try:
                            oil_mass_value = float(oil_mass)
                            ws.cell(row=3, column=oil_mass_col, value=oil_mass_value)
                        except ValueError:
                            ws.cell(row=3, column=oil_mass_col, value=oil_mass)
                        debug_print(f"DEBUG: Set oil mass '{oil_mass}' at row 3, column {oil_mass_col} for sample {i+1}")

                    # Row 2, Column H + offset: Puffing Regime
                    puffing_regime = sample_data.get("puffing_regime", "60mL/3s/30s")
                    if puffing_regime:
                        puffing_regime_col = 8 + col_offset
                        ws.cell(row=2, column=puffing_regime_col, value=puffing_regime)
                        debug_print(f"DEBUG: Set puffing regime '{puffing_regime}' at row 2, column {puffing_regime_col} for sample {i+1}")

                # Calculate the last column used
                last_sample_column = ((num_samples - 1) * 12) + 12
                max_column = ws.max_column
                debug_print(f"DEBUG: Last sample column: {last_sample_column}, Max column: {max_column}")

                # Save the workbook
                wb.save(file_path)
                debug_print(f"DEBUG: Successfully saved workbook to {file_path}")

                debug_print(f"SUCCESS: Applied header data for {num_samples} samples to {file_path}")

            else:
                error_msg = f"Sheet '{header_data['test']}' not found in the file."
                debug_print(f"ERROR: {error_msg}")
                raise Exception(error_msg)

        except Exception as e:
            debug_print(f"ERROR: Error applying header data: {e}")
            debug_print("DEBUG: Full traceback:")
            import traceback
            traceback.print_exc()
            raise

    def determine_sample_count_from_data(self, sheet_data, test_name):
        """Determine the number of samples based on existing data structure."""
        debug_print(f"DEBUG: Determining sample count from data for test: {test_name}")

        try:
            total_columns = len(sheet_data.columns)
            debug_print(f"DEBUG: Total columns in data: {total_columns}")
                
            columns_per_sample = 12  # Standard format

            debug_print(f"DEBUG: Using {columns_per_sample} columns per sample for test type")

            # Look for actual sample data by checking each potential sample position
            actual_samples = 0

            # Start checking from the first sample position and keep going until no more data found
            i = 0
            while True:
                sample_id_col = 5 + (i * columns_per_sample)  # Sample ID column position

                if sample_id_col >= total_columns:
                    debug_print(f"DEBUG: Reached end of data at column {sample_id_col}")
                    break

                # Check if there's meaningful data in this sample's area
                sample_has_data = False
                for row_idx in range(min(5, len(sheet_data))):
                    for col_offset in range(min(columns_per_sample, total_columns - sample_id_col)):
                        cell_value = sheet_data.iloc[row_idx, sample_id_col + col_offset]
                        if pd.notna(cell_value) and str(cell_value).strip():
                            sample_has_data = True
                            break
                    if sample_has_data:
                        break

                if sample_has_data:
                    actual_samples = i + 1
                    debug_print(f"DEBUG: Found data for sample {actual_samples}")
                    i += 1
                else:
                    debug_print(f"DEBUG: No data found for sample {i + 1}, stopping count")
                    break

            final_sample_count = max(1, actual_samples)
            debug_print(f"DEBUG: Final determined sample count: {final_sample_count}")

            return final_sample_count

        except Exception as e:
            debug_print(f"ERROR: Error determining sample count: {e}")
            return 1  # Default to 1 sample if we can't determine
"""
Core File Operations Module for DataViewer Application

This module handles basic file loading, selection, and state management operations.
"""

# Standard library imports
import os
import re
import copy
import shutil
import time
import uuid
import tempfile
import threading
import subprocess
import traceback
from typing import Optional

# Third party imports
import pandas as pd
import psutil
import openpyxl
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label, Button, ttk, Frame, Listbox, Scrollbar

# Local imports
import processing
from resource_utils import get_resource_path
from utils import (
    is_valid_excel_file,
    load_excel_file,
    get_save_path,
    is_standard_file,
    FONT,
    APP_BACKGROUND_COLOR,
    load_excel_file_with_formulas,
    debug_print,
    show_success_message,
    plotting_sheet_test
)
from excel_image_extractor import extract_and_load_excel_images

class CoreFileOperations:
    """Handles core file operations like loading, reloading, and file state management."""
    
    def __init__(self, file_manager):
        """Initialize with reference to parent FileManager."""
        self.file_manager = file_manager
        self.gui = file_manager.gui
        self.root = file_manager.root
        
    def load_excel_file(self, file_path, legacy_mode: str = None, skip_database_storage: bool = False, force_reload: bool = False) -> None:
        """
        Load the selected Excel file and process its sheets.
        Enhanced with caching and optional database storage skip.

        Args:
            file_path (str): Path to the Excel file
            legacy_mode (str, optional): Legacy processing mode
            skip_database_storage (bool): If True, skip storing in database
            force_reload (bool): If True, bypass cache and reload from file
        """
        debug_print(f"DEBUG: load_excel_file called for {file_path}, skip_db_storage={skip_database_storage}, force_reload={force_reload}")

        # Check cache first (unless force_reload is True)
        cache_key = f"{file_path}_{legacy_mode}"
        if not force_reload and cache_key in self.file_manager.loaded_files_cache:
            debug_print("DEBUG: Using cached file data instead of reprocessing")
            cached_data = self.file_manager.loaded_files_cache[cache_key]
            self.gui.filtered_sheets = cached_data['filtered_sheets']
            self.gui.sheets = cached_data.get('sheets', {})
            # Fix: Make sure full_sample_data exists
            if hasattr(self.gui, 'full_sample_data'):
                self.gui.full_sample_data = cached_data.get('full_sample_data', pd.DataFrame())

            # Set the selected sheet without triggering full UI update
            if cached_data['filtered_sheets']:
                first_sheet = list(cached_data['filtered_sheets'].keys())[0]
                self.gui.selected_sheet.set(first_sheet)
            return

        # Clear cache entry if force_reload is True
        if force_reload and cache_key in self.file_manager.loaded_files_cache:
            debug_print(f"DEBUG: Force reload requested - clearing cache entry for {file_path}")
            del self.file_manager.loaded_files_cache[cache_key]

        try:
            # Ensure the file is a valid Excel file.
            if not is_valid_excel_file(os.path.basename(file_path)):
                raise ValueError(f"Invalid Excel file selected: {file_path}")

            debug_print(f"DEBUG: {'Force reloading' if force_reload else 'Loading'} file from disk: {file_path}")

            # extract embedded images from Excel file
            debug_print("DEBUG: Checking for embedded images in Excel file")
            try:
                num_images = extract_and_load_excel_images(self.gui, file_path,current_sheet=None)
                if num_images > 0:
                    debug_print(f"DEBUG: Successfully extracted {num_images} embedded images from Excel")
            except Exception as img_error:
                debug_print(f"DEBUG: Failed to extract images from Excel: {img_error}")
                # don't fail the entire load if image extraction fails
                pass

            debug_print(f"DEBUG: Checking if file is standard format: {file_path}")

            if not is_standard_file(file_path):
                debug_print("DEBUG: File is legacy format, processing accordingly")
                # Legacy file processing
                legacy_dir = os.path.join(os.path.abspath("."), "legacy data")
                if not os.path.exists(legacy_dir):
                    os.makedirs(legacy_dir)
                    debug_print(f"DEBUG: Created legacy data directory: {legacy_dir}")

                legacy_wb = load_workbook(file_path)
                legacy_sheetnames = legacy_wb.sheetnames
                debug_print(f"DEBUG: Legacy file sheets: {legacy_sheetnames}")

                template_path_default = os.path.join(os.path.abspath("."), "resources",
                                         "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")

                if not os.path.exists(template_path_default):
                    raise FileNotFoundError(f"Template file not found: {template_path_default}")

                wb_template = load_workbook(template_path_default)
                template_sheet_names = wb_template.sheetnames
                debug_print(f"DEBUG: Template sheets: {template_sheet_names}")

                if legacy_mode is None:
                    if len(legacy_sheetnames) == 1 and legacy_sheetnames[0] not in template_sheet_names:
                        legacy_mode = "file"
                        debug_print("DEBUG: Auto-detected legacy mode: file")
                    else:
                        legacy_mode = "standards"
                        debug_print("DEBUG: Auto-detected legacy mode: standards")

                final_sheets = {}
                full_sample_data = pd.DataFrame()

                if legacy_mode == "file":
                    debug_print("DEBUG: Processing as legacy file")
                    converted = processing.process_legacy_file_auto_detect(file_path)
                    key = f"Legacy_{os.path.basename(file_path)}"
                    final_sheets = {key: {"data": converted, "is_empty": converted.empty}}
                    full_sample_data = converted

                    self.gui.filtered_sheets = final_sheets
                    self.gui.full_sample_data = full_sample_data
                    default_key = list(final_sheets.keys())[0]
                    self.gui.selected_sheet.set(default_key)
                    debug_print(f"DEBUG: Legacy file processed, key: {default_key}")

                elif legacy_mode == "standards":
                    debug_print("DEBUG: Processing as legacy standards file")
                    converted_dict = processing.convert_legacy_standards_using_template(file_path)

                    self.gui.sheets = converted_dict
                    self.gui.filtered_sheets = {
                        name: {"data": data, "is_empty": data.empty}
                        for name, data in self.gui.sheets.items()
                    }
                    # Fix: Safely concatenate sheets
                    sheet_data_list = []
                    for sheet_info in self.gui.filtered_sheets.values():
                        if not sheet_info["data"].empty:
                            sheet_data_list.append(sheet_info["data"])

                    if sheet_data_list:
                        self.gui.full_sample_data = pd.concat(sheet_data_list, axis=1)
                    else:
                        self.gui.full_sample_data = pd.DataFrame()

                    first_sheet = list(self.gui.filtered_sheets.keys())[0]
                    self.gui.selected_sheet.set(first_sheet)
                    debug_print(f"DEBUG: Legacy standards processed, first sheet: {first_sheet}")
                else:
                    raise ValueError(f"Unknown legacy mode: {legacy_mode}")

                # Store in database only if not skipping and not already stored
                if not skip_database_storage and file_path not in self.file_manager.stored_files_cache:
                    debug_print("DEBUG: Storing legacy file in database")
                    self.file_manager._store_file_in_database(file_path)
                    self.file_manager.stored_files_cache.add(file_path)
            else:
                # Standard file processing
                debug_print("DEBUG: Processing as standard file")
                self.gui.sheets = load_excel_file(file_path)
                self.gui.filtered_sheets = {
                    name: {"data": data, "is_empty": data.empty}
                    for name, data in self.gui.sheets.items()
                }

                # Fix: Safely concatenate sheets and handle empty sheets
                sheet_data_list = []
                for sheet_info in self.gui.filtered_sheets.values():
                    if not sheet_info["data"].empty:
                        sheet_data_list.append(sheet_info["data"])

                if sheet_data_list:
                    self.gui.full_sample_data = pd.concat(sheet_data_list, axis=1)
                else:
                    self.gui.full_sample_data = pd.DataFrame()

                first_sheet = list(self.gui.filtered_sheets.keys())[0]
                self.gui.selected_sheet.set(first_sheet)
                debug_print(f"DEBUG: Standard file processed, first sheet: {first_sheet}")

                # Store in database only if not skipping and not already stored
                if not skip_database_storage and file_path not in self.file_manager.stored_files_cache:
                    debug_print("DEBUG: Storing standard file in database")
                    self.file_manager._store_file_in_database(file_path)
                    self.file_manager.stored_files_cache.add(file_path)

            # Cache the processed data
            cache_data = {
                'filtered_sheets': copy.deepcopy(self.gui.filtered_sheets),
                'sheets': copy.deepcopy(getattr(self.gui, 'sheets', {})),
                'full_sample_data': self.gui.full_sample_data.copy() if hasattr(self.gui, 'full_sample_data') and not self.gui.full_sample_data.empty else pd.DataFrame()
            }
            self.file_manager.loaded_files_cache[cache_key] = cache_data
            debug_print(f"DEBUG: Cached processed data for {file_path}")

        except Exception as e:
            error_msg = f"Error occurred while loading file: {e}"
            debug_print(f"ERROR: {error_msg}")
            traceback.print_exc()
            messagebox.showerror("Error", error_msg)

    def load_initial_file(self) -> None:
        """Handle file loading directly on the main thread."""
        file_paths = filedialog.askopenfilenames(
            title="Select Excel File(s)",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_paths:
            show_success_message("Info", "No files were selected", self.gui.root)
            return

        self.gui.progress_dialog.show_progress_bar("Loading files...")
        total_files = len(file_paths)

        try:
            for index, file_path in enumerate(file_paths, 1):
                # Update progress
                progress = int((index / total_files) * 100)
                self.gui.progress_dialog.update_progress_bar(progress)

                # Load and process file
                self.load_excel_file(file_path)

                # Store results
                self.gui.all_filtered_sheets.append({
                    "file_name": os.path.basename(file_path),
                    "file_path": file_path,
                    "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
                })

                # Allow GUI to refresh
                self.root.update_idletasks()

            # Final setup after loading all files
            if self.gui.all_filtered_sheets:
                last_file = self.gui.all_filtered_sheets[-1]
                self.set_active_file(last_file["file_name"])
                self.file_manager.update_file_dropdown()
                self.file_manager.update_ui_for_current_file()

        except Exception as e:
            messagebox.showerror("Loading Error", f"Failed to load files: {str(e)}")
        finally:
            self.gui.progress_dialog.hide_progress_bar()
            self.root.update_idletasks()

    def reload_excel_file(self) -> None:
        """Reload the Excel file into the program, preserving the state of the UI."""
        try:
            self.load_excel_file(self.gui.file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reload the Excel file: {e}")

    def load_file(self, file_path):
        """Load a file without showing dialogs - used for data collection flow."""
        debug_print(f"DEBUG: Loading file for data collection: {file_path}")
        try:
            # Use the optimized ensure method instead of full reload
            self.ensure_file_is_loaded_in_ui(file_path)
            debug_print(f"DEBUG: File loaded successfully: {file_path}")
            return True

        except Exception as e:
            debug_print(f"DEBUG: Error loading file: {e}")
            messagebox.showerror("Error", f"Failed to load file: {e}")
            return False

    def ask_open_file(self) -> str:
        """Prompt the user to select an Excel file."""
        return filedialog.askopenfilename(
            title="Select Standardized Testing File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

    def add_data(self) -> None:
        """Handle adding a new data file directly and update UI accordingly."""
        file_paths = filedialog.askopenfilenames(
            title="Select Excel File(s)", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_paths:
            show_success_message("Info", "No files were selected", self.gui.root)
            return

        for file_path in file_paths:
            self.load_excel_file(file_path)
            self.gui.all_filtered_sheets.append({
                "file_name": os.path.basename(file_path),
                "file_path": file_path,
                "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
            })
        self.file_manager.update_file_dropdown()
        last_file = self.gui.all_filtered_sheets[-1]
        self.set_active_file(last_file["file_name"])
        self.file_manager.update_ui_for_current_file()
        show_success_message("Success", f"Data from {len(file_paths)} file(s) added successfully.", self.gui.root)

    def set_active_file(self, file_name: str) -> None:
        """Set the active file based on the given file name."""
        for file_data in self.gui.all_filtered_sheets:
            if file_data["file_name"] == file_name:
                self.gui.current_file = file_name
                self.gui.file_path = file_data.get("file_path", None)
                self.gui.filtered_sheets = file_data["filtered_sheets"]
                if self.gui.file_path is None:
                    raise ValueError(f"No file path associated with the file '{file_name}'.")
                break
        else:
            raise ValueError(f"File '{file_name}' not found.")

    def ensure_file_is_loaded_in_ui(self, file_path):
        """Ensure the file is properly loaded in the UI without redundant processing."""
        debug_print(f"DEBUG: Ensuring file {file_path} is loaded in UI")

        # Special handling for .vap3 files - they're already loaded in memory
        if file_path.endswith('.vap3'):
            debug_print("DEBUG: .vap3 file detected - file should already be loaded in memory")

            # For .vap3 files, just verify the UI state is correct
            if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                debug_print("DEBUG: .vap3 file data is already loaded, ensuring UI is updated")

                # Check if file is in all_filtered_sheets
                file_name = None
                for file_data in self.gui.all_filtered_sheets:
                    if file_data.get("file_path") == file_path:
                        file_name = file_data["file_name"]
                        break

                if file_name:
                    debug_print(f"DEBUG: Found .vap3 file in loaded files: {file_name}")
                    try:
                        self.set_active_file(file_name)
                        self.file_manager.update_ui_for_current_file()
                        return True
                    except Exception as e:
                        debug_print(f"ERROR: Failed to set active .vap3 file: {e}")
                        return False
                else:
                    debug_print("DEBUG: .vap3 file not found in loaded files list")
                    # The file data is loaded but not in the list - this is OK for database files
                    # Just update the UI with current data
                    if hasattr(self.gui, 'update_displayed_sheet') and hasattr(self.gui, 'selected_sheet'):
                        current_sheet = self.gui.selected_sheet.get()
                        if current_sheet in self.gui.filtered_sheets:
                            self.gui.update_displayed_sheet(current_sheet)
                            debug_print(f"DEBUG: Updated display for current sheet: {current_sheet}")
                            return True
                    debug_print("DEBUG: Could not update .vap3 file UI")
                    return False
            else:
                debug_print("ERROR: .vap3 file should be loaded but no data found")
                return False

        # Regular Excel file handling
        # Validate file path
        if not file_path or not os.path.exists(file_path):
            debug_print(f"ERROR: Invalid file path: {file_path}")
            return False

        # Check if file is already in the UI state
        file_name = os.path.basename(file_path)

        # Look for existing entry in all_filtered_sheets
        existing_entry = None
        for file_data in self.gui.all_filtered_sheets:
            if file_data["file_path"] == file_path:
                existing_entry = file_data
                break

        if existing_entry:
            debug_print("DEBUG: File already in UI state, just updating active file")
            try:
                self.set_active_file(existing_entry["file_name"])
                self.file_manager.update_ui_for_current_file()
                return True
            except Exception as e:
                debug_print(f"ERROR: Failed to set active file: {e}")
                return False
        else:
            debug_print("DEBUG: File not in UI state, adding it")
            try:
                # Load file without database storage (skip_database_storage=True)
                self.load_excel_file(file_path, skip_database_storage=True)

                # Add to all_filtered_sheets
                self.gui.all_filtered_sheets.append({
                    "file_name": file_name,
                    "file_path": file_path,
                    "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
                })

                self.file_manager.update_file_dropdown()
                self.set_active_file(file_name)
                self.file_manager.update_ui_for_current_file()
                debug_print(f"DEBUG: Successfully added file {file_name} to UI")
                return True

            except Exception as e:
                debug_print(f"ERROR: Failed to load file into UI: {e}")
                traceback.print_exc()
                return False

    def close_current_file(self):
        """Close the currently active file and remove it from the session."""
        if not self.gui.current_file:
            messagebox.showinfo("No File", "No file is currently loaded.")
            return
    
        current_file = self.gui.current_file
    
        # Confirm before closing
        confirm = messagebox.askyesno(
            "Confirm Close",
            f"Close '{current_file}'?\n\nAny unsaved changes will be lost."
        )
    
        if not confirm:
            return
    
        debug_print(f"DEBUG: Closing file: {current_file}")
    
        # Remove from all_filtered_sheets
        self.gui.all_filtered_sheets = [
            file_data for file_data in self.gui.all_filtered_sheets 
            if file_data["file_name"] != current_file
        ]
    
        # Remove from sheet_images if present
        if hasattr(self.gui, 'sheet_images') and current_file in self.gui.sheet_images:
            del self.gui.sheet_images[current_file]
    
        # Clear current state
        self.gui.filtered_sheets = {}
        self.gui.sheets = {}
        self.gui.current_file = None
        self.gui.file_path = None
        self.gui.selected_sheet.set("")  # Clear selected sheet
    
        # Clear UI
        self.gui.ui_manager.clear_dynamic_frame()
    
        # Clear sheet dropdown - ADD THIS BACK
        if hasattr(self.gui, 'drop_down_menu') and self.gui.drop_down_menu:
            self.gui.drop_down_menu['values'] = []
            self.gui.drop_down_menu.set('')
            debug_print("DEBUG: Cleared sheet dropdown")
    
        # Update file dropdown
        self.file_manager.ui_manager.update_file_dropdown()
    
        # If other files exist, switch to the last one
        if self.gui.all_filtered_sheets:
            last_file = self.gui.all_filtered_sheets[-1]
            self.file_manager.set_active_file(last_file["file_name"])
            self.file_manager.ui_manager.update_ui_for_current_file()
            debug_print(f"DEBUG: Switched to file: {last_file['file_name']}")
        else:
            # No files left - clear file dropdown selection too
            if hasattr(self.gui, 'file_dropdown_var'):
                self.gui.file_dropdown_var.set('')
            debug_print("DEBUG: No files remaining, showing startup menu")
            self.gui.show_startup_menu()
    
        show_success_message("File Closed", f"'{current_file}' has been closed.", self.gui.root)
"""
Test Workflow Module for DataViewer Application

This module handles test creation, template management, and data collection workflow
including test selection, header data entry, and workflow coordination.
"""

# Standard library imports
import os
import shutil
import traceback

# Third party imports
import openpyxl
from tkinter import messagebox

# Local imports
from resource_utils import get_resource_path
from utils import get_save_path, debug_print
from test_selection_dialog import TestSelectionDialog
from test_start_menu import TestStartMenu
from header_data_dialog import HeaderDataDialog


class TestWorkflow:
    """Handles test creation workflow and data collection coordination."""
    
    def __init__(self, file_manager):
        """Initialize with reference to parent FileManager."""
        self.file_manager = file_manager
        self.gui = file_manager.gui
        self.root = file_manager.root
        
    def create_new_template(self, parent_window=None):
        """Create a new file from the template."""
        return self.create_new_file_with_tests(parent_window)

    def create_new_file_with_tests(self, parent_window=None):
        """Create a new file with only the selected tests."""
        # Get save path before showing test selection dialog
        file_path = get_save_path(".xlsx")
        if not file_path:
            return None

        # Get template path
        template_path = get_resource_path("resources/Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")

        # Load template to get available tests
        wb = openpyxl.load_workbook(template_path)
        available_tests = [sheet for sheet in wb.sheetnames if sheet != "Sheet1"]

        # Show test selection dialog
        test_dialog = TestSelectionDialog(self.gui.root, available_tests)
        ok_clicked, selected_tests = test_dialog.show()

        if not ok_clicked or not selected_tests:
            return None

        # Create a new file with selected tests
        try:
            shutil.copy(template_path, file_path)
            new_wb = openpyxl.load_workbook(file_path)

            # Remove unselected tests
            for sheet_name in list(new_wb.sheetnames):
                if sheet_name not in selected_tests and sheet_name != "Sheet1":
                    del new_wb[sheet_name]

            new_wb.save(file_path)

            # Close parent window if provided
            if parent_window and parent_window.winfo_exists():
                parent_window.destroy()

            # Show test start menu
            self.show_test_start_menu(file_path)

            return file_path
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while creating the file: {str(e)}")
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
            return None

    def show_test_start_menu(self, file_path, original_filename = None):
        """
        Show the test start menu for a file.
        If file is already loaded, try to extract existing header data and skip the dialog.
        """
        debug_print(f"DEBUG: Showing test start menu for file: {file_path}")

        # Store the original filename for later use
        if original_filename:
            debug_print(f"DEBUG: Storing original filename for data collection: {original_filename}")
            # You can store this in a class attribute or pass it through the flow
            self.current_original_filename = original_filename
        else:
            self.current_original_filename = None

        # For .vap3 files or files that don't exist, use loaded sheet data instead
        available_tests = None
        if file_path.endswith('.vap3') or not os.path.exists(file_path):
            # Use the currently loaded sheets instead of trying to read the file
            if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                available_tests = list(self.gui.filtered_sheets.keys())
                debug_print(f"DEBUG: Using loaded sheet names for .vap3 file: {available_tests}")
            else:
                debug_print("ERROR: No loaded sheets available for .vap3 file")
                messagebox.showerror("Error", "No sheet data is currently loaded. Please load a file first.")
                return

        # Show the test start menu with available tests
        start_menu = TestStartMenu(self.gui.root, file_path, available_tests)
        result, selected_test = start_menu.show()

        if result == "start_test" and selected_test:
            debug_print(f"DEBUG: Starting test: {selected_test}")

            # For existing files, always try to extract header data first
            debug_print("DEBUG: Attempting to extract existing header data from file/loaded data")

            # Try to extract existing header data
            existing_header_data = self.extract_existing_header_data(file_path, selected_test)

            if existing_header_data and self.validate_header_data(existing_header_data):
                debug_print("DEBUG: Found complete header data - skipping header dialog and going directly to data collection")
                # Go directly to data collection with existing header data
                self.start_data_collection_with_header_data(file_path, selected_test, existing_header_data)
            else:
                debug_print("DEBUG: Header data incomplete or missing - showing header dialog")
                # Only show header data dialog if we couldn't find valid header data
                self.show_header_data_dialog(file_path, selected_test)

        elif result == "view_raw_file":
            debug_print("DEBUG: Viewing raw file requested")
            # Load the file for viewing if not already loaded
            if not hasattr(self.gui, 'file_path') or self.gui.file_path != file_path:
                self.load_file(file_path)
        else:
            debug_print("DEBUG: Test start menu was cancelled or closed")

    def show_header_data_dialog(self, file_path, selected_test):
        """Show the header data dialog for a selected test."""
        debug_print(f"DEBUG: Showing header data dialog for {selected_test}")

        # Try to determine sample count from existing data first
        initial_sample_count = 1
        try:
            if hasattr(self.gui, 'filtered_sheets') and selected_test in self.gui.filtered_sheets:
                sheet_data = self.gui.filtered_sheets[selected_test]['data']
                initial_sample_count = self.determine_sample_count_from_data(sheet_data, selected_test)
                debug_print(f"DEBUG: Determined {initial_sample_count} samples from existing data")
        except Exception as e:
            debug_print(f"DEBUG: Could not determine sample count from data: {e}")

        # Create initial header data structure with detected sample count
        initial_header_data = {
            'common': {'tester': ''},
            'samples': [
                {
                    'id': f'Sample {i+1}',
                    'resistance': '',
                    'media': '',
                    'viscosity': '',
                    'voltage': '',
                    'puffing_regime': '60mL/3s/30s',
                    'oil_mass': ''
                } for i in range(initial_sample_count)
            ],
            'test': selected_test,
            'num_samples': initial_sample_count
        }

        # Show the header data dialog
        header_dialog = HeaderDataDialog(self.gui.root, file_path, selected_test,
                                       edit_mode=False, current_data=initial_header_data)
        result, header_data = header_dialog.show()

        if result:
            debug_print("DEBUG: Header data dialog completed successfully")
            # Apply header data to the file
            self.apply_header_data_to_file(file_path, header_data)

            debug_print("DEBUG: Proceeding to data collection window")
            # Show the data collection window
            from data_collection import DataCollectionWindow
            data_collection_window = DataCollectionWindow(self.gui, file_path, selected_test, header_data)
            data_result = data_collection_window.show()

            if data_result == "load_file":
                # Load the file for viewing
                self.load_file(file_path)
            elif data_result == "cancel":
                # User cancelled data collection, just load the file
                self.load_file(file_path)
        else:
            debug_print("DEBUG: Header data dialog was cancelled")

    def start_data_collection_with_header_data(self, file_path, selected_test, header_data):
        """Start data collection directly with existing header data."""
        debug_print("DEBUG: Starting data collection with existing header data")

        # Ensure header data has all required fields for each sample
        if 'samples' in header_data:
            for i, sample in enumerate(header_data['samples']):
                # Ensure each sample has all required fields
                required_fields = ['id', 'resistance', 'media', 'viscosity', 'voltage', 'puffing_regime', 'oil_mass']
                for field in required_fields:
                    if field not in sample:
                        if field == 'id':
                            sample[field] = f"Sample {i+1}"
                        elif field == 'puffing_regime':
                            sample[field] = '60mL/3s/30s'
                        else:
                            sample[field] = ''

        # Ensure common data has device_type
        if 'common' not in header_data:
            header_data['common'] = {}
        if 'device_type' not in header_data['common']:
            header_data['common']['device_type'] = 'T58G'  # Changed default to T58G

        debug_print(f"DEBUG: Validated header data structure with all fields: {header_data}")

        try:
            # Show the data collection window directly
            from data_collection import DataCollectionWindow

            # Pass the original filename to the data collection window
            original_filename = getattr(self, 'current_original_filename', None)
            data_collection_window = DataCollectionWindow(self.gui, file_path, selected_test, header_data, original_filename=original_filename)
            data_result = data_collection_window.show()

            if data_result in ["load_file", "cancel"]:
                debug_print("DEBUG: Data collection completed - data should already be updated in main GUI")
                # The data collection window should have already updated the main GUI
                # Just refresh the currently displayed sheet if needed
                if hasattr(self.gui, 'selected_sheet') and self.gui.selected_sheet.get() == selected_test:
                    self.gui.root.after(100, lambda: self.gui.update_displayed_sheet(selected_test))

        except Exception as e:
            debug_print(f"DEBUG: Error starting data collection: {e}")
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to start data collection: {e}")
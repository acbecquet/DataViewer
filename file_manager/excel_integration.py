"""
Excel Integration Module for DataViewer Application

This module handles opening Excel files for external editing and processing
changes made in Excel back into the application.
"""

# Standard library imports
import os
import time
import uuid
import tempfile
import threading
import subprocess
import traceback

# Third party imports
import pandas as pd
from openpyxl import Workbook
from tkinter import messagebox, ttk

# Local imports
from utils import debug_print


class ExcelIntegration:
    """Handles Excel external editing integration."""
    
    def __init__(self, file_manager):
        """Initialize with reference to parent FileManager."""
        self.file_manager = file_manager
        self.gui = file_manager.gui
        self.root = file_manager.root
        
    def open_raw_data_in_excel(self, sheet_name=None) -> None:
        """Opens an Excel file for a specific sheet, allowing edits, and then updates the application when Excel closes."""
        try:
            # Validate the current state
            if not hasattr(self.gui, 'filtered_sheets') or not self.gui.filtered_sheets:
                messagebox.showerror("Error", "No data is currently loaded.")
                return

            # If no sheet specified, use the currently selected sheet
            if not sheet_name and hasattr(self.gui, 'selected_sheet'):
                sheet_name = self.gui.selected_sheet.get()

            if not sheet_name or sheet_name not in self.gui.filtered_sheets:
                messagebox.showerror("Error", "Please select a valid sheet first.")
                return

            # Get the sheet data
            sheet_data = self.gui.filtered_sheets[sheet_name]['data']

            # Create a temporary Excel file
            import tempfile
            import uuid
            from openpyxl import Workbook

            # Generate a unique identifier
            unique_id = str(uuid.uuid4()).split('-')[0]

            # Ensure the sheet name only contains valid characters
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c == ' ')
            safe_sheet_name = safe_sheet_name.replace(' ', '_')[:15]

            # Create a temporary file
            temp_dir = os.path.abspath(tempfile.gettempdir())
            temp_file = os.path.join(temp_dir, f"dataviewer_{safe_sheet_name}_{unique_id}.xlsx")

            # Create a new workbook with just this sheet
            wb = Workbook()
            ws = wb.active
            ws.title = safe_sheet_name[:31]

            # Write the headers
            for col_idx, column_name in enumerate(sheet_data.columns, 1):
                ws.cell(row=1, column=col_idx, value=str(column_name))

            # Write the data
            for row_idx, row in enumerate(sheet_data.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Save the workbook
            wb.save(temp_file)

            if not os.path.exists(temp_file):
                raise FileNotFoundError(f"Failed to create temporary file at {temp_file}")

            # Record the file modification time before opening
            original_mod_time = os.path.getmtime(temp_file)

            # Create status notification
            status_text = f"Opening {sheet_name} in Excel. Changes will be imported when Excel closes."
            status_label = ttk.Label(self.gui.root, text=status_text, relief="sunken", anchor="w")
            status_label.pack(side="bottom", fill="x")
            self.gui.root.update_idletasks()

            # File monitor state
            class FileMonitorState:
                def __init__(self, initial_mod_time):
                    self.last_mod_time = initial_mod_time
                    self.has_changed = False

            monitor_state = FileMonitorState(original_mod_time)

            # Force a new Excel instance
            import subprocess
            cmd = f'start /wait "" "excel.exe" /x "{os.path.abspath(temp_file)}"'

            try:
                subprocess.Popen(cmd, shell=True)
            except Exception as e:
                debug_print(f"Error launching Excel with command: {e}")
                os.startfile(os.path.abspath(temp_file))

            time.sleep(2.0)

            # Monitor file in background thread
            def monitor_file_lock():
                file_open = True
                while file_open:
                    file_locked = False
                    try:
                        with open(temp_file, 'r+b') as test_lock:
                            pass
                    except PermissionError:
                        file_locked = True
                    except FileNotFoundError:
                        file_locked = False
                        file_open = False
                    except Exception:
                        file_locked = False

                    file_open = file_locked

                    if os.path.exists(temp_file):
                        try:
                            current_mod_time = os.path.getmtime(temp_file)
                            if current_mod_time > monitor_state.last_mod_time:
                                monitor_state.has_changed = True
                                monitor_state.last_mod_time = current_mod_time
                        except Exception:
                            pass

                    if not file_open:
                        self.gui.root.after(500, lambda: self._process_excel_changes(temp_file, sheet_name, monitor_state.has_changed, status_label))
                        break

                    time.sleep(1.0)

            # Start monitoring thread
            monitor_thread = threading.Thread(target=monitor_file_lock, daemon=True)
            self.gui.threads.append(monitor_thread)
            monitor_thread.start()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {e}")
            traceback.print_exc()

    def _process_excel_changes(self, temp_file, sheet_name, file_changed, status_label=None):
        """Process changes made in Excel after it closes."""
        try:
            if status_label and status_label.winfo_exists():
                status_label.destroy()

            if file_changed and os.path.exists(temp_file):
                try:
                    max_retries = 3
                    retry_count = 0
                    read_success = False
                    modified_data = None

                    while not read_success and retry_count < max_retries:
                        try:
                            modified_data = pd.read_excel(temp_file)
                            read_success = True
                        except Exception:
                            retry_count += 1
                            time.sleep(1.0)

                    if not read_success:
                        raise Exception(f"Failed to read Excel file after {max_retries} attempts")

                    # Update the filtered sheets with new data
                    self.gui.filtered_sheets[sheet_name]['data'] = modified_data

                    # If using VAP3 file, update it
                    if hasattr(self.gui, 'file_path') and self.gui.file_path.endswith('.vap3'):
                        from vap_file_manager import VapFileManager
                        vap_manager = VapFileManager()
                        vap_manager.save_to_vap3(
                            self.gui.file_path,
                            self.gui.filtered_sheets,
                            self.gui.sheet_images,
                            self.gui.plot_options,
                            getattr(self.gui, 'image_crop_states', {})
                        )

                    # Refresh the GUI display
                    self.gui.update_displayed_sheet(sheet_name)

                    # Show success message
                    success_label = ttk.Label(
                        self.gui.root,
                        text="Changes from Excel have been imported successfully.",
                        relief="sunken",
                        anchor="w"
                    )
                    success_label.pack(side="bottom", fill="x")
                    self.gui.root.after(3000, lambda: success_label.destroy() if success_label.winfo_exists() else None)

                except Exception as e:
                    traceback.print_exc()
                    messagebox.showerror("Error", f"Failed to read modified Excel file: {e}")
            elif file_changed and not os.path.exists(temp_file):
                show_success_message("Information", "Excel file was modified but appears to have been moved or renamed. Changes could not be imported.", self.gui.root)
            else:
                info_label = ttk.Label(
                    self.gui.root,
                    text="Excel file was closed without changes.",
                    relief="sunken",
                    anchor="w"
                )
                info_label.pack(side="bottom", fill="x")
                self.gui.root.after(3000, lambda: info_label.destroy() if info_label.winfo_exists() else None)

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to process Excel changes: {e}")
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass
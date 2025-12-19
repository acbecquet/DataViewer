"""
legacy_processing.py
Developed by Charlie Becquet
Legacy Data Processing module for the DataViewer application.
"""

import re
import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Any
from utils import (
    debug_print,
    read_sheet_with_values,
    extract_meta_data,
    map_meta_data_to_template,
    load_excel_file,
    plotting_sheet_test
)

# Module constants for legacy processing
CART_FORMAT_INDICATORS = ['cart #', 'cart#', 'cartridge #']
OLD_FORMAT_INDICATORS = ['project:', 'sample:']
NEW_FORMAT_INDICATORS = ['sample id:', 'resistance (ohms):']


def extract_samples_from_old_file(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Extract samples from an old Excel file by robustly locating the headers "puffs" and "TPM (mg/puff)" anywhere in the sheet.
    For each sample header found, a new set of meta_data (from up to 3 rows above) and column data is extracted.

    Returns a list of dictionaries, one per sample, each containing:
        - 'sample_name': The value found near the puffs header.
        - 'puffs': A pandas Series of the puffs column.
        - 'tpm': A pandas Series of the TPM column.
        - 'header_row': The index of the header row.
        - Additional meta_data if detected.
    """
    # Read the sheet without assuming a header row.
    df = read_sheet_with_values(file_path, sheet_name)
    samples = []
    nrows, ncols = df.shape

    # Regex patterns for meta_data and data headers.
    meta_data_patterns = {
        "sample_name": [
            r"(cart(ridge)?\s*#|sample\s*(name|id))",  # e.g. "Cart #", "Sample ID"
            r"puffing\s*data\s*for\s*:?\s*" # e.g puffing data for:
        ],
        "resistance": [
            r"\bri\s*\(?\s*ohms?\s*\)?\s*:?\s*",  # e.g. "Ri (Ohms)"
            r"resistance\s*\(?ohms?\)?\s*:?\s*"
        ],
        "voltage": [r"voltage\s*:?\s*"],
        "viscosity": [r"viscosity\b\s*:?\s*"],
        "puffing_regime": [r"\b(puff(ing)?\s*regime|puff\s*settings?)\s*:?\s*"],
        "initial_oil_mass": [r"initial\s*oil\s*mass\b\s*:?\s*"],
        "date": [r"date\s*:?\s*"],
        "media": [r"media\s*:?\s*"]
    }

    data_header_patterns = {
        "puffs": r"puffs",
        "tpm": r"tpm\s*\(mg\s*/\s*puff\)",
        "before_weight": r"before\s*weight/g",
        "after_weight": r"after\s*weight/g",
        "draw_pressure": r"pv1|draw\s*pressure\s*\(kpa\)",
        "smell": r"smell",
        "notes": r"notes"
    }

    numeric_columns = {"puffs", "tpm", "before_weight", "after_weight", "draw_pressure"}

    # Track processed columns for sample data and meta_data separately.
    processed_cols = {row: [] for row in range(nrows)}
    processed_meta_data = {row: [] for row in range(nrows)}
    # Set a threshold: cells within this number of columns are considered too close.
    proximity_threshold = 1

    # Loop over every cell in the sheet.
    for row in range(nrows):
        for col in range(ncols):
            # Skip this cell if it's near a previously processed sample header in the same row.
            if any(abs(col - proc_col) < proximity_threshold for proc_col in processed_cols[row]):
                continue

            cell_val = df.iat[row, col]
            if header_matches(cell_val, data_header_patterns["puffs"]):
                # New sample header found.
                sample = {"sample_name": str(cell_val).strip(), "header_row": row}

                # Mark this column as processed for sample data.
                processed_cols[row].append(col)

                # -----------------------------
                # Extract meta_data (up to 3 rows above)
                # -----------------------------
                start_search_row = max(0, row - 3)
                meta_data_found = {}
                for r in range(row - 1, start_search_row - 1, -1):
                    for c in range(ncols):
                        # Skip if this cell in meta_data row was already used.
                        if any(abs(c - pm) < proximity_threshold for pm in processed_meta_data[r]):
                            continue
                        cell_val_above = df.iat[r, c]
                        for key, patterns in meta_data_patterns.items():
                            if key in meta_data_found:
                                continue  # Already found this key for the current sample.
                            for pattern in patterns:
                                if header_matches(cell_val_above, pattern):
                                    value = df.iat[r, c + 1] if (c + 1 < ncols) else None
                                    meta_data_found[key] = value
                                    # Mark this column as used for meta_data in row r.
                                    processed_meta_data[r].append(c)
                                    break
                            if key in meta_data_found:
                                break
                sample.update(meta_data_found)

                # -----------------------------
                # Extract column data for this sample.
                # -----------------------------
                data_cols = {}
                for key, pattern in data_header_patterns.items():
                    # Look to the right from the current header cell.
                    for j in range(col, ncols):
                        if header_matches(df.iat[row, j], pattern):
                            raw_data = df.iloc[row + 1:, j]
                            if key in numeric_columns:
                                data_cols[key] = pd.to_numeric(raw_data, errors='coerce').dropna()
                            else:
                                data_cols[key] = raw_data.astype(str).replace("nan", "")
                            break
                # Only add the sample if both "puffs" and "tpm" data were found.
                if "puffs" in data_cols and "tpm" in data_cols:
                    sample.update(data_cols)
                    samples.append(sample)
    return samples
        
def extract_samples_from_cart_format(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Extract samples from the cart format legacy files.

    This format has:
    - Cart # in A2, value in B2
    - Media in A3, value in B3
    - Ri in C2, value in D2
    - Viscosity in G2, value in H2
    - Voltage in G3, value in H3
    - PV1-PV5 instead of draw pressure
    - TPM in last column (no avg/std dev after)
    """
    print(f"DEBUG: Extracting samples from cart format: {file_path}")

    try:
        if sheet_name is None:
            wb = load_workbook(file_path)
            sheet_name = wb.sheetnames[0]
            wb.close()

        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        print(f"DEBUG: Loaded sheet with shape: {df.shape}")

        samples = []
        nrows, ncols = df.shape

        # Find the data start row (look for "Puffs" or numbers in first column)
        data_start_row = None
        header_row = None
        for row in range(min(15, nrows)):
            cell_val = str(df.iloc[row, 0]).lower() if pd.notna(df.iloc[row, 0]) else ""
            if "puff" in cell_val:
                data_start_row = row + 1  # Data starts one row after headers
                header_row = row  # Headers are at this row
                print(f"DEBUG: Found headers at row {row}, data starts at row {row + 1}")
                break

        if data_start_row is None or header_row is None:
            print("ERROR: Could not find header/data start row")
            return []

        # Print all headers for debugging
        print("DEBUG: Headers found:")
        for col in range(min(ncols, 20)):  # Check first 20 columns
            header_val = str(df.iloc[header_row, col]) if pd.notna(df.iloc[header_row, col]) else ""
            if header_val.strip():
                print(f"  Col {col}: '{header_val}'")

        # Find TPM columns by scanning the header row
        tpm_columns = []
        for col in range(ncols):
            header_val = str(df.iloc[header_row, col]).lower() if pd.notna(df.iloc[header_row, col]) else ""
            if "tpm" in header_val:
                tpm_columns.append(col)
                print(f"DEBUG: Found TPM at column {col}: '{df.iloc[header_row, col]}'")

        if not tpm_columns:
            print("ERROR: No TPM columns found")
            return []

        # For cart format, we expect one main sample. TPM should be the last data column
        # Let's find the sample boundaries more systematically

        # Find sample blocks by looking for the standard pattern: Puffs, Before weight, After weight, etc.
        sample_blocks = []

        # Look for "Puffs" columns which indicate sample starts
        puffs_columns = []
        for col in range(ncols):
            header_val = str(df.iloc[header_row, col]).lower() if pd.notna(df.iloc[header_row, col]) else ""
            if header_val.strip() == "puffs" or header_val.strip() == "puff":
                puffs_columns.append(col)
                print(f"DEBUG: Found Puffs column at {col}")

        if not puffs_columns:
            # If no explicit "Puffs" column, assume first column is puffs for cart format
            puffs_columns = [0]
            print("DEBUG: No 'Puffs' header found, assuming column 0 is puffs")

        # For each puffs column, find the corresponding TPM column
        for puffs_col in puffs_columns:
            # Find the nearest TPM column to the right
            nearest_tpm = None
            for tpm_col in tpm_columns:
                if tpm_col > puffs_col:
                    if nearest_tpm is None or tpm_col < nearest_tpm:
                        nearest_tpm = tpm_col

            if nearest_tpm is not None:
                sample_blocks.append((puffs_col, nearest_tpm))
                print(f"DEBUG: Sample block from column {puffs_col} to {nearest_tpm}")

        # If no sample blocks found, create one assuming standard cart format layout
        if not sample_blocks and tpm_columns:
            sample_blocks = [(0, tpm_columns[0])]
            print(f"DEBUG: Created default sample block from 0 to {tpm_columns[0]}")

        print(f"DEBUG: Found {len(sample_blocks)} sample blocks: {sample_blocks}")

        # Extract shared metadata first
        shared_metadata = {}

        # Cart # / Sample ID from B2 (row 1, col 1)
        sample_id = str(df.iloc[1, 1]) if pd.notna(df.iloc[1, 1]) else ""
        if sample_id and sample_id != 'nan':
            shared_metadata['sample_name'] = sample_id

        # Media from B3 (row 2, col 1)
        media = str(df.iloc[2, 1]) if pd.notna(df.iloc[2, 1]) else ""
        shared_metadata['media'] = media

        # Ri (resistance) from D2 (row 1, col 3)
        resistance = str(df.iloc[1, 3]) if pd.notna(df.iloc[1, 3]) else ""
        shared_metadata['resistance'] = resistance

        # Viscosity from H2 (row 1, col 7)
        viscosity = str(df.iloc[1, 7]) if (ncols > 7 and pd.notna(df.iloc[1, 7])) else ""
        shared_metadata['viscosity'] = viscosity

        # Voltage from H3 (row 2, col 7)
        voltage = str(df.iloc[2, 7]) if (ncols > 7 and pd.notna(df.iloc[2, 7])) else ""
        shared_metadata['voltage'] = voltage

        print(f"DEBUG: Extracted shared metadata - Sample: '{shared_metadata.get('sample_name', '')}', Media: '{shared_metadata.get('media', '')}', Resistance: '{shared_metadata.get('resistance', '')}', Viscosity: '{shared_metadata.get('viscosity', '')}', Voltage: '{shared_metadata.get('voltage', '')}'")

        # Extract each sample
        for sample_idx, (start_col, tpm_col) in enumerate(sample_blocks):
            print(f"DEBUG: Processing sample {sample_idx + 1} from col {start_col} to {tpm_col}")

            sample = shared_metadata.copy()
            if sample_idx > 0:  # For multiple samples, add index to name
                base_name = shared_metadata.get('sample_name', 'Sample')
                sample['sample_name'] = f"{base_name}_{sample_idx + 1}"

            # Set default values for missing metadata
            sample.setdefault('date', '')
            sample.setdefault('puffing_regime', '')
            sample.setdefault('initial_oil_mass', '')

            # Define expected column positions based on standard format
            puffs_col = start_col
            before_weight_col = start_col + 1
            after_weight_col = start_col + 2

            # Find PV1 column (look for pressure column between after_weight and TPM)
            pv1_col = None
            for check_col in range(start_col + 3, tpm_col):
                header_val = str(df.iloc[header_row, check_col]).lower() if pd.notna(df.iloc[header_row, check_col]) else ""
                if "pv1" in header_val or "pressure" in header_val:
                    pv1_col = check_col
                    print(f"DEBUG: Found pressure column at {check_col}: '{df.iloc[header_row, check_col]}'")
                    break

            # If no specific PV1 found, use the first column after after_weight
            if pv1_col is None and start_col + 3 < tpm_col:
                pv1_col = start_col + 3
                print(f"DEBUG: Using default pressure column at {pv1_col}")

            # Extract data arrays
            data_found = False
            try:
                # Puffs data
                puffs_data = df.iloc[data_start_row:, puffs_col]
                puffs_clean = pd.to_numeric(puffs_data, errors='coerce').dropna()

                # Before weight data
                before_weight_data = df.iloc[data_start_row:, before_weight_col] if before_weight_col < ncols else pd.Series(dtype=float)
                before_weight_clean = pd.to_numeric(before_weight_data, errors='coerce').dropna()

                # After weight data
                after_weight_data = df.iloc[data_start_row:, after_weight_col] if after_weight_col < ncols else pd.Series(dtype=float)
                after_weight_clean = pd.to_numeric(after_weight_data, errors='coerce').dropna()

                # PV1 data (for draw pressure)
                if pv1_col is not None and pv1_col < ncols:
                    pv1_data = df.iloc[data_start_row:, pv1_col]
                    pv1_clean = pd.to_numeric(pv1_data, errors='coerce').dropna()
                else:
                    pv1_clean = pd.Series(dtype=float)

                # TPM data
                tpm_data = df.iloc[data_start_row:, tpm_col]
                tpm_clean = pd.to_numeric(tpm_data, errors='coerce').dropna()

                print(f"DEBUG: Data extraction results:")
                print(f"  Puffs: {len(puffs_clean)} values")
                print(f"  Before weight: {len(before_weight_clean)} values")
                print(f"  After weight: {len(after_weight_clean)} values")
                print(f"  PV1/Pressure: {len(pv1_clean)} values")
                print(f"  TPM: {len(tpm_clean)} values")

                # Only add sample if we have meaningful data
                if len(puffs_clean) > 0 and len(tpm_clean) > 0:
                    sample['puffs'] = puffs_clean
                    sample['before_weight'] = before_weight_clean
                    sample['after_weight'] = after_weight_clean
                    sample['draw_pressure'] = pv1_clean  # Use PV1 for draw pressure
                    sample['tpm'] = tpm_clean

                    # Add empty arrays for missing columns to maintain compatibility
                    sample['smell'] = pd.Series(dtype=str)
                    sample['notes'] = pd.Series(dtype=str)

                    data_found = True
                    print(f"DEBUG: Sample {sample_idx + 1} successfully processed")
                else:
                    print(f"DEBUG: Sample {sample_idx + 1} has insufficient data - Puffs: {len(puffs_clean)}, TPM: {len(tpm_clean)}")

            except Exception as e:
                print(f"ERROR: Failed to extract data for sample {sample_idx + 1}: {e}")
                import traceback
                traceback.print_exc()

            if data_found:
                samples.append(sample)
                print(f"DEBUG: Successfully added sample {sample_idx + 1}")

        print(f"DEBUG: Extracted {len(samples)} samples from cart format")
        return samples

    except Exception as e:
        print(f"ERROR: Failed to extract samples from cart format: {e}")
        import traceback
        traceback.print_exc()
        return []

def process_legacy_file_auto_detect(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Automatically detect template format and use appropriate processing function.
    Updated to handle cart format.
    """
    print(f"DEBUG: Auto-detecting format for {legacy_file_path}")

    format_type = detect_template_format(legacy_file_path)
    print(f"DEBUG: Detected format: {format_type}")

    if format_type == "cart_format":
        print("DEBUG: Using cart format processing")
        return convert_cart_format_to_template(legacy_file_path, template_path)
    elif format_type == "old":
        print("DEBUG: Using enhanced processing for old format")
        return convert_legacy_file_using_template_v2(legacy_file_path, template_path)
    else:
        print("DEBUG: Using standard processing for new/unknown format")
        return convert_legacy_file_using_template(legacy_file_path, template_path)

def detect_template_format(file_path: str, sheet_name: Optional[str] = None) -> str:
    """
    Detect whether this is an old, new, or cart template format.

    Returns:
        "cart_format": Cart format with Cart # in A2, Media in A3, Ri in C2
        "old": Old format with Project/Sample fields and Ri/Rf
        "new": New format with Sample ID and single Resistance
        "unknown": Could not determine format
    """
    try:
        print(f"DEBUG: Detecting template format for {file_path}")
        df = read_sheet_with_values(file_path, sheet_name)
        nrows, ncols = df.shape

        # First check for cart format (most specific)
        if nrows > 3 and ncols > 8:
            # Check specific positions for cart format indicators
            a2_val = str(df.iloc[1, 0]).lower() if pd.notna(df.iloc[1, 0]) else ""
            a3_val = str(df.iloc[2, 0]).lower() if pd.notna(df.iloc[2, 0]) else ""
            c2_val = str(df.iloc[1, 2]).lower() if pd.notna(df.iloc[1, 2]) else ""

            print(f"DEBUG: Checking cart format - A2: '{a2_val}', A3: '{a3_val}', C2: '{c2_val}'")

            # Cart format has "Cart #" in A2, "Media" in A3, and "Ri" in C2
            if ("cart" in a2_val and "#" in a2_val and
                "media" in a3_val and
                "ri" in c2_val):
                print("DEBUG: Detected cart format")
                return "cart_format"

        # Look for old vs new format indicators in first few rows
        old_format_indicators = 0
        new_format_indicators = 0

        for row in range(min(5, nrows)):
            for col in range(min(10, ncols)):
                cell_val = str(df.iat[row, col]).lower().strip()

                # Old format indicators
                if re.search(r"project\s*:", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Project:' at row {row}, col {col}")
                if re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Ri (Ohms)' at row {row}, col {col}")
                if re.search(r"rf\s*\(\s*ohms?\s*\)", cell_val):
                    old_format_indicators += 1
                    print(f"DEBUG: Found 'Rf (Ohms)' at row {row}, col {col}")

                # New format indicators
                if re.search(r"sample\s*(id|name)\s*:", cell_val):
                    new_format_indicators += 1
                    print(f"DEBUG: Found 'Sample ID/Name:' at row {row}, col {col}")
                if re.search(r"resistance\s*\(\s*ohms?\s*\)\s*:", cell_val) and "ri" not in cell_val and "rf" not in cell_val:
                    new_format_indicators += 1
                    print(f"DEBUG: Found 'Resistance (Ohms):' at row {row}, col {col}")

        print(f"DEBUG: Format detection - Old indicators: {old_format_indicators}, New indicators: {new_format_indicators}")

        if old_format_indicators > new_format_indicators:
            return "old"
        elif new_format_indicators > old_format_indicators:
            return "new"
        else:
            return "unknown"

    except Exception as e:
        print(f"DEBUG: Error detecting template format: {e}")
        return "unknown"

def convert_legacy_file_using_template(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Converts a legacy Excel file to the standardized template format.
    - Uses 12-column blocks per sample.
    - Maps meta_data into exact template positions.
    - For each sample block, only rows up to the first empty 'After weight/g'
        are written and all cells below that row (within that block) are cleared.
    """
    # Determine the template path.
    if template_path is None:
        # Try new template first, fall back to old
        new_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - December 2025.xlsx")
        old_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
        if os.path.exists(new_template_path):
            template_path = new_template_path
            print("DEBUG: Using new template for legacy conversion")
        elif os.path.exists(old_template_path):
            template_path = old_template_path
            print("DEBUG: Using old template for legacy conversion")
        else:
            raise FileNotFoundError("No template file found")

    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Load legacy samples using our extraction function.
    legacy_samples = extract_samples_from_old_file(legacy_file_path)
    if not legacy_samples:
        raise ValueError("No valid legacy sample data found.")

    # meta_data mapping: each key maps to a list of regex patterns and the target (row, col offset).
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"],
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"],
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping relative to the sample block start.
    # Note: "after_weight" is at column offset 2.
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True),
        "after_weight": (2, True),
        "draw_pressure": (3, True),
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each legacy sample.
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        #debug_print(f"\nProcessing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # --- 1. meta_data Handling ---
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # --- 2. Data Column Handling with Row Slicing ---
        # Determine the cutoff row for this sample block based on "after_weight".
        raw_after_weight = sample.get("after_weight", pd.Series(dtype=object))
        raw_after_weight = raw_after_weight.reset_index(drop=True)
        cutoff = len(raw_after_weight)
        for i, val in enumerate(raw_after_weight):
            if pd.isna(val) or str(val).strip() == "":
                cutoff = i
                break
        #debug_print(f"Sample {sample_idx + 1}: Writing {cutoff} data rows based on 'After weight/g' column.")

        # Now loop over each key in DATA_COL_MAPPING and write only rows up to the cutoff.
        for key, (rel_col, is_numeric) in DATA_COL_MAPPING.items():
            raw_data = sample.get(key, pd.Series(dtype=object))
            raw_data = raw_data.reset_index(drop=True)
            sliced_data = raw_data.iloc[:cutoff]
            if is_numeric:
                data = pd.to_numeric(sliced_data, errors="coerce")
            else:
                data = sliced_data.astype(str).replace("nan", "").replace("None", "")
            for row_offset, value in enumerate(data):
                target_row = 5 + row_offset  # Data starts at row 5.
                target_col = col_offset + rel_col
                try:
                    ws.cell(row=target_row, column=target_col,
                            value=float(value) if is_numeric else str(value))
                except Exception as e:
                    #debug_print(f"Error writing {key} at ({target_row},{target_col}): {e}")
                    ws.cell(row=target_row, column=target_col, value="ERROR")

        # --- 3. Clearing Extra Rows in the Block ---
        # For rows below the cutoff (in the sample block), clear all cells in the block's columns.
        # We assume the block occupies 12 columns starting at col_offset.
        start_clear_row = 5 + cutoff  # first row to clear in the block.
        # Use the maximum row in the worksheet as the ending point.
        for row_clear in range(start_clear_row, ws.max_row + 1):
            for col_clear in range(col_offset, col_offset + 12):
                ws.cell(row=row_clear, column=col_clear).value = None

    # --- 4. Final Cleanup ---
    # Delete any extra columns beyond the last sample block.
    last_sample_col = len(legacy_samples) * 12
    if ws.max_column > last_sample_col:
        ws.delete_cols(last_sample_col + 1, ws.max_column - last_sample_col)

    # Delete all sheets except the template sheet.
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters).
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"{base_name} Data"[:31]
    ws.title = new_sheet_name
    folder_path = os.path.join(os.path.abspath("."), "legacy data")

    new_file_name = f"{base_name} Legacy.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)
    #debug_print(f"\nSaved processed file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]

def convert_legacy_standards_using_template(legacy_file_path: str, template_path: str = None) -> dict:
    """
    Converts a legacy standards Excel file to the standardized template format.
    """
    from openpyxl.cell.cell import MergedCell

    if template_path is None:
        # Try new template first, fall back to old
        new_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - December 2025.xlsx")
        old_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
        if os.path.exists(new_template_path):
            template_path = new_template_path
            print("DEBUG: Using new template for legacy conversion")
        elif os.path.exists(old_template_path):
            template_path = old_template_path
            print("DEBUG: Using old template for legacy conversion")
        else:
            raise FileNotFoundError("No template file found")

    wb_template = load_workbook(template_path, read_only=False)
    legacy_wb = load_workbook(legacy_file_path, read_only=False)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    new_file_name = f"{base_name} Legacy Standards.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)

    SIMPLE_meta_data_MAPPING = {
        "sample_name": r"(cart(ridge)?\s*#|sample\s*(name|id))",
        "voltage": r"voltage",
        "viscosity": r"viscosity",
        "resistance": r"(ri\s*\(?\s*ohms?\s*\)?|resistance)",
        "puff_regime": r"puff(ing)?\s*regime",
        "initial_oil": r"initial\s*oil\s*mass",
        "date": r"date",
        "media": r"media"
    }

    for sheet_name in wb_template.sheetnames:
        # Create a new sheet to avoid merged cell issues
        new_sheet_name = f"{sheet_name}_new"
        new_ws = wb_template.create_sheet(title=new_sheet_name)

        if sheet_name in legacy_wb.sheetnames:
            legacy_ws = legacy_wb[sheet_name]
            legacy_df = read_sheet_with_values(legacy_file_path, sheet_name)

            # Check if this is a plotting sheet
            is_plotting = plotting_sheet_test(sheet_name, legacy_df)

            # Process based on sheet type
            if is_plotting:
                # For plotting sheets: copy the entire legacy sheet
                processed_df = legacy_df.copy()
                meta_data = extract_meta_data(legacy_ws, SIMPLE_meta_data_MAPPING)

                # Write to the new sheet
                for i, row in processed_df.iterrows():
                    for j, value in enumerate(row):
                        new_ws.cell(row=i + 1, column=j + 1, value=value)
            else:
                # For non-plotting sheets: copy the legacy sheet
                processed_df = legacy_df.copy()
                for i, row in processed_df.iterrows():
                    for j, value in enumerate(row):
                        new_ws.cell(row=i + 1, column=j + 1, value=value)
        else:
            # If sheet not in legacy file, copy from template sheet
            ws = wb_template[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                for col_idx, value in enumerate(row, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)

    # Remove original sheets and rename new ones
    for sheet_name in list(wb_template.sheetnames):
        if not sheet_name.endswith('_new'):
            del wb_template[sheet_name]

    for sheet_name in list(wb_template.sheetnames):
        if sheet_name.endswith('_new'):
            original_name = sheet_name[:-4]  # Remove _new suffix
            wb_template[sheet_name].title = original_name

    wb_template.save(new_file_path)
    return load_excel_file(new_file_path)

def extract_samples_from_old_file_v2(file_path: str, sheet_name: Optional[str] = None) -> list:
    """
    Enhanced version of extract_samples_from_old_file that handles both old and new template formats.
    Old format differences:
    - Has separate "Project:" and "Sample:" fields instead of "Sample ID:"
    - Has "Ri (Ohms)" and "Rf (Ohms)" instead of just "Resistance (Ohms)"
    - Test type is in sheet name, not top-left corner

    Returns a list of dictionaries, one per sample, each containing:
        - 'sample_name': Combined from Project + Sample for old format
        - 'puffs': A pandas Series of the puffs column
        - 'tpm': A pandas Series of the TPM column
        - 'header_row': The index of the header row
        - Additional metadata if detected
    """
    print(f"DEBUG: Starting enhanced extraction from file: {file_path}")
    print(f"DEBUG: Target sheet: {sheet_name}")

    # Read the sheet without assuming a header row
    df = read_sheet_with_values(file_path, sheet_name)
    samples = []
    nrows, ncols = df.shape

    print(f"DEBUG: Sheet dimensions: {nrows} rows x {ncols} columns")

    # Enhanced regex patterns for metadata - now includes both old and new format patterns
    meta_data_patterns = {
        "sample_name": [
            r"(cart(ridge)?\s*#|sample\s*(name|id))",  # New format: "Cart #", "Sample ID"
            r"puffing\s*data\s*for\s*:?\s*",  # New format: "puffing data for:"
            r"^sample\s*:?\s*(?:\.\d+)?$"  # Old format: "Sample:", "Sample:.1", "Sample:.2", etc.
        ],
        "project": [
            r"^project\s*:?\s*(?:\.\d+)?$"  # Old format: "Project:", "Project:.1", "Project:.2", etc.
        ],
        "resistance": [
            r"\bri\s*\(?\s*ohms?\s*\)?\s*:?\s*(?:\.\d+)?",  # Old format: "Ri (Ohms)", "Ri (Ohms).1", etc.
            r"resistance\s*\(?ohms?\)?\s*:?\s*(?:\.\d+)?"   # New format: "Resistance (Ohms)", "Resistance (Ohms).1", etc.
        ],
        "voltage": [r"voltage\s*:?\s*(?:\.\d+)?"],  # "voltage:", "voltage:.1", etc.
        "viscosity": [r"viscosity\b\s*:?\s*(?:\.\d+)?"],  # "viscosity:", "viscosity:.1", etc.
        "puffing_regime": [r"\b(puff(ing)?\s*regime|puff\s*settings?)\s*:?\s*(?:\.\d+)?"]  # With suffixes
    }

    # FIXED: More specific data header patterns to avoid false matches
    data_header_patterns = {
        "puffs": r"^puffs?$",  # EXACT match for "puff" or "puffs" - not part of another phrase
        "tpm": r"\btpm\b",
        "before_weight": r"before.{0,10}weight",
        "after_weight": r"after.{0,10}weight",
        "draw_pressure": r"draw.{0,10}pressure",
        "smell": r"\bsmell\b",
        "notes": r"\bnotes?\b"
    }

    numeric_columns = {"puffs", "tpm", "before_weight", "after_weight", "draw_pressure"}

    # Track which rows and columns have been used for metadata to avoid duplication
    processed_meta_data = {r: [] for r in range(nrows)}
    processed_cols = {r: [] for r in range(nrows)}
    proximity_threshold = 8  # Increased to 8 to account for 12-column sample blocks

    print(f"DEBUG: Starting cell-by-cell scanning...")

    # Cell-by-cell scanning for "puffs" headers (indicates sample start)
    for row in range(nrows):
        for col in range(ncols):
            # Skip this cell if it's near a previously processed sample header in the same row
            if any(abs(col - proc_col) < proximity_threshold for proc_col in processed_cols[row]):
                continue

            cell_val = df.iat[row, col]
            # More strict matching - check if it's exactly "puffs" and not part of another phrase
            if header_matches(cell_val, data_header_patterns["puffs"]):
                print(f"DEBUG: Found puffs header at row {row}, col {col}")

                # Additional validation: ensure this is actually the start of a sample block
                # Check if this column is at expected sample positions (0, 12, 24, 36, ...)
                if col % 12 != 0:
                    print(f"DEBUG: Skipping col {col} - not at expected sample start position (should be multiple of 12)")
                    continue

                # New sample header found
                sample = {"sample_name": str(cell_val).strip(), "header_row": row}

                # Mark this column as processed for sample data
                processed_cols[row].append(col)

                # Calculate the expected column range for this sample (12 columns per sample)
                sample_start_col = col
                sample_end_col = col + 12

                print(f"DEBUG: Searching for metadata from row {max(0, row - 3)} to {row - 1} within columns {sample_start_col} to {sample_end_col}")

                # Extract metadata (up to 3 rows above) within this sample's column range
                start_search_row = max(0, row - 3)
                meta_data_found = {}
                project_value = None
                sample_value = None

                # FIXED: More thorough metadata search with better debugging
                for r in range(row - 1, start_search_row - 1, -1):
                    for c in range(sample_start_col, min(sample_end_col, ncols)):
                        # Skip if this cell in metadata row was already used
                        if any(abs(c - pm) < 2 for pm in processed_meta_data[r]):  # Reduced threshold for metadata
                            continue

                        cell_val_above = df.iat[r, c]
                        cell_str = str(cell_val_above).strip().lower()

                        # ENHANCED: More explicit project pattern matching
                        if header_matches(cell_val_above, meta_data_patterns["project"][0]):
                            # Look for the value in the next cell (to the right)
                            if c + 1 < ncols:
                                value = df.iat[r, c + 1]
                                if value and str(value).strip().lower() not in ['nan', 'none', '']:
                                    project_value = str(value).strip()
                                    print(f"DEBUG: Found project value: {project_value}")
                                    processed_meta_data[r].append(c)

                        # ENHANCED: More explicit sample pattern matching
                        elif header_matches(cell_val_above, meta_data_patterns["sample_name"][2]):  # Use the old format pattern
                            # Look for the value in the next cell (to the right)
                            if c + 1 < ncols:
                                value = df.iat[r, c + 1]
                                if value and str(value).strip().lower() not in ['nan', 'none', '']:
                                    sample_value = str(value).strip()
                                    print(f"DEBUG: Found sample value: {sample_value}")
                                    processed_meta_data[r].append(c)

                        # Check for other metadata patterns
                        else:
                            for key, patterns in meta_data_patterns.items():
                                if key in meta_data_found or key in ["project", "sample_name"]:
                                    continue  # Already found this key or handled above
                                for pattern in patterns:
                                    if header_matches(cell_val_above, pattern):
                                        value = df.iat[r, c + 1] if (c + 1 < ncols) else None
                                        meta_data_found[key] = value
                                        processed_meta_data[r].append(c)
                                        break
                                if key in meta_data_found:
                                    break

                # Combine project and sample for old format sample names
                if project_value and sample_value:
                    combined_sample_name = f"{project_value} {sample_value}"
                    print(f"DEBUG: Combined old format sample name: {combined_sample_name}")
                    meta_data_found["sample_name"] = combined_sample_name
                elif project_value:
                    meta_data_found["sample_name"] = project_value
                    print(f"DEBUG: Using project as sample name: {project_value}")
                elif sample_value:
                    meta_data_found["sample_name"] = sample_value
                    print(f"DEBUG: Using sample as sample name: {sample_value}")
                else:
                    # Use fallback sample name based on position
                    fallback_name = f"Sample {len(samples) + 1}"
                    print(f"DEBUG: Using fallback sample name: {fallback_name}")
                    meta_data_found["sample_name"] = fallback_name

                print(f"DEBUG: Final metadata found for sample: {meta_data_found}")
                sample.update(meta_data_found)

                # Extract column data for this sample within the sample's column range
                data_cols = {}
                print(f"DEBUG: Extracting column data starting from col {col}")

                for key, pattern in data_header_patterns.items():
                    # Look to the right from the current header cell within the sample range
                    for j in range(col, min(sample_end_col, ncols)):
                        if header_matches(df.iat[row, j], pattern):
                            print(f"DEBUG: Found {key} column at position {j}")
                            raw_data = df.iloc[row + 1:, j]
                            if key in numeric_columns:
                                data_cols[key] = pd.to_numeric(raw_data, errors='coerce').dropna()
                            else:
                                data_cols[key] = raw_data.astype(str).replace("nan", "")
                            break

                # Only add the sample if both "puffs" and "tpm" data were found
                if "puffs" in data_cols and "tpm" in data_cols:
                    sample.update(data_cols)
                    samples.append(sample)
                    print(f"DEBUG: Successfully added sample {len(samples)}: {meta_data_found.get('sample_name', 'Unknown')}")
                else:
                    print(f"DEBUG: Skipping sample - missing required data. Found: {list(data_cols.keys())}")

    print(f"DEBUG: Extraction complete. Found {len(samples)} valid samples")
    for i, sample in enumerate(samples):
        print(f"DEBUG: Sample {i+1}: {sample.get('sample_name', 'Unknown')} - {len(sample.get('puffs', []))} puffs")

    return samples

def is_legacy_sample_empty(sample):
    """
    Check if a legacy sample has meaningful data.
    Enhanced to be more graceful with different data formats.

    Args:
        sample (dict): Legacy sample data with fields like 'puffs', 'tpm', etc.

    Returns:
        bool: True if sample is empty/invalid, False if it has meaningful data
    """
    try:
        # Check for TPM data (primary indicator of valid sample)
        tpm_data = sample.get('tpm', [])
        if hasattr(tpm_data, '__len__') and len(tpm_data) > 0:
            # Convert to list if it's a pandas Series
            if hasattr(tpm_data, 'tolist'):
                tpm_values = tpm_data.tolist()
            else:
                tpm_values = list(tpm_data)

            # Check if there are any non-zero, non-NaN TPM values
            valid_tpm_count = 0
            for val in tpm_values:
                try:
                    numeric_val = float(val)
                    if numeric_val > 0 and not math.isnan(numeric_val):
                        valid_tpm_count += 1
                except (ValueError, TypeError):
                    continue

            if valid_tpm_count > 0:
                print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has {valid_tpm_count} valid TPM values")
                return False  # Has meaningful data

        # Check puffs data as secondary indicator
        puffs_data = sample.get('puffs', [])
        if hasattr(puffs_data, '__len__') and len(puffs_data) > 2:  # Need at least 3 data points
            # If we have substantial puffs data, consider it valid even without TPM
            print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has {len(puffs_data)} puffs data points")
            return False

        # Check if this is just metadata without data
        metadata_fields = ['sample_name', 'voltage', 'viscosity', 'resistance', 'media', 'date', 'puffing_regime', 'initial_oil_mass']
        has_metadata = any(sample.get(field, '') for field in metadata_fields)

        if has_metadata and not any(sample.get(field, []) for field in ['puffs', 'tpm', 'before_weight', 'after_weight']):
            print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' has metadata but no data arrays - treating as empty")
            return True

        print(f"DEBUG: Sample '{sample.get('sample_name', 'Unknown')}' appears empty - no meaningful data")
        return True  # No meaningful data found

    except Exception as e:
        print(f"DEBUG: Error checking legacy sample: {e}")
        return False  # If error, assume not empty to be safe

def filter_legacy_samples(legacy_samples):
    """
    Filter legacy samples to only include those with meaningful data.

    Args:
        legacy_samples (list): List of legacy sample dictionaries

    Returns:
        list: Filtered list containing only samples with meaningful data
    """
    import math

    filtered_samples = []

    for i, sample in enumerate(legacy_samples):
        sample_name = sample.get('sample_name', f'Sample {i+1}')

        if not is_legacy_sample_empty(sample):
            filtered_samples.append(sample)
            print(f"DEBUG: Keeping sample {len(filtered_samples)}: {sample_name}")
        else:
            print(f"DEBUG: Filtering out empty sample: {sample_name}")

    print(f"DEBUG: Filtered from {len(legacy_samples)} to {len(filtered_samples)} meaningful samples")
    return filtered_samples

def convert_legacy_file_using_template_v2(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Enhanced version that handles both old and new template formats.
    Uses the enhanced extraction function and properly handles:
    - Combined Project + Sample fields from old format
    - Ri vs Rf resistance selection
    - Test type detection from sheet name for old format
    - Filters out empty samples before writing to template
    """
    import math  # Add this import for the filtering functions
    print(f"DEBUG: Starting enhanced template conversion for: {legacy_file_path}")

    # Determine the template path
    if template_path is None:
        # Try new template first, fall back to old
        new_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - December 2025.xlsx")
        old_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
        if os.path.exists(new_template_path):
            template_path = new_template_path
            print("DEBUG: Using new template for legacy conversion")
        elif os.path.exists(old_template_path):
            template_path = old_template_path
            print("DEBUG: Using old template for legacy conversion")
        else:
            raise FileNotFoundError("No template file found")

    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Load legacy samples using our enhanced extraction function
    all_legacy_samples = extract_samples_from_old_file_v2(legacy_file_path)
    if not all_legacy_samples:
        raise ValueError("No valid legacy sample data found.")

    print(f"DEBUG: Successfully extracted {len(all_legacy_samples)} samples")

    # Filter samples to only include those with meaningful data
    legacy_samples = filter_legacy_samples(all_legacy_samples)

    if not legacy_samples:
        raise ValueError("No samples with meaningful data found after filtering.")

    print(f"DEBUG: Processing {len(legacy_samples)} filtered samples for template conversion")

    # Enhanced metadata mapping with old format support
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"],
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"],
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping relative to the sample block start
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True),
        "after_weight": (2, True),
        "draw_pressure": (3, True),
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each FILTERED legacy sample
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        print(f"DEBUG: Processing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # Metadata Handling
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        # Write metadata to template
        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            if value:
                print(f"DEBUG: Set {template_key} to '{value}' at row {tpl_row}, col {col_offset + tpl_col_offset}")
                ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # Data Handling
        for data_key, (data_col_offset, is_numeric) in DATA_COL_MAPPING.items():
            if data_key in sample:
                data_series = sample[data_key]
                target_col = col_offset + data_col_offset

                # Convert pandas Series to list for consistent handling
                if hasattr(data_series, 'tolist'):
                    data_values = data_series.tolist()
                elif hasattr(data_series, '__iter__'):
                    data_values = list(data_series)
                else:
                    data_values = [data_series] if data_series else []

                # Remove NaN and empty values
                clean_values = []
                for val in data_values:
                    if pd.notna(val) and str(val).strip() != '':
                        clean_values.append(val)

                print(f"DEBUG: Writing {data_key} data to column {target_col} ({len(clean_values)} values)")

                # Write the data starting from row 4 (data_start_row)
                for row_idx, value in enumerate(clean_values, start=4):
                    if is_numeric:
                        try:
                            numeric_value = float(value)
                            ws.cell(row=row_idx, column=target_col, value=numeric_value)
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=target_col, value=value)
                    else:
                        ws.cell(row=row_idx, column=target_col, value=str(value))

    # Keep only the Intense Test sheet
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"Legacy_{base_name}"[:31]
    ws.title = new_sheet_name

    # Ensure legacy data directory exists
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    new_file_name = f"{base_name} Legacy.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)

    print(f"DEBUG: Saved processed file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]

def convert_cart_format_to_template(legacy_file_path: str, template_path: str = None) -> pd.DataFrame:
    """
    Convert cart format legacy files to standardized template.
    """
    print(f"DEBUG: Converting cart format file: {legacy_file_path}")

    # Determine template path
    if template_path is None:
        # Try new template first, fall back to old
        new_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - December 2025.xlsx")
        old_template_path = os.path.join(os.path.abspath("."), "resources",
                                            "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
        if os.path.exists(new_template_path):
            template_path = new_template_path
            print("DEBUG: Using new template for legacy conversion")
        elif os.path.exists(old_template_path):
            template_path = old_template_path
            print("DEBUG: Using old template for legacy conversion")
        else:
            raise FileNotFoundError("No template file found")

    template_sheet = "Intense Test"
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template file.")
    ws = wb[template_sheet]

    # Extract samples using cart format function
    all_legacy_samples = extract_samples_from_cart_format(legacy_file_path)
    if not all_legacy_samples:
        raise ValueError("No valid legacy sample data found in cart format.")

    print(f"DEBUG: Successfully extracted {len(all_legacy_samples)} samples")

    # Filter samples to only include those with meaningful data
    legacy_samples = filter_legacy_samples(all_legacy_samples)

    if not legacy_samples:
        raise ValueError("No samples with meaningful data found after filtering.")

    print(f"DEBUG: Processing {len(legacy_samples)} filtered samples for template conversion")

    # Metadata mapping for cart format (same positions as v2)
    meta_data_MAPPING = {
        "Sample ID:": (
            [r"cart\s*#:?", r"sample\s*(name|id):?"],
            (1, 5)
        ),
        "Voltage:": ([r"voltage:?"], (3, 5)),
        "Viscosity:": ([r"viscosity:?"], (3, 1)),
        "Resistance (Ohms):": (
            [r"ri\s*\(\s*ohms?\s*\)", r"resistance\s*\(?ohms?\)?\s*:?\s*"],
            (2, 3)
        ),
        "Puffing Regime:": ([r"puffing\s*regime:?", r"puff\s*regime:?"], (2, 7)),
        "Initial Oil Mass:": ([r"initial\s*oil\s*mass:?"], (3, 7)),
        "Date:": ([r"date:?"], (1, 3)),
        "Media:": ([r"media:?"], (2, 1))
    }

    # Data column mapping (same as existing)
    DATA_COL_MAPPING = {
        "puffs": (0, True),
        "before_weight": (1, True),
        "after_weight": (2, True),
        "draw_pressure": (3, True),  # This will use PV1 data
        "smell": (5, False),
        "notes": (7, False),
        "tpm": (8, True)
    }

    # Process each sample (using same logic as v2)
    for sample_idx, sample in enumerate(legacy_samples):
        col_offset = 1 + (sample_idx * 12)
        print(f"DEBUG: Processing sample {sample_idx + 1} at columns {col_offset} to {col_offset + 11}")

        # Write project name to first row
        sample_name = sample.get("sample_name", f"Sample {sample_idx + 1}")
        ws.cell(row=1, column=col_offset, value=os.path.splitext(os.path.basename(legacy_file_path))[0])

        # Prepare metadata values
        meta_data_values = {
            "Sample ID:": sample.get("sample_name", ""),
            "Voltage:": sample.get("voltage", ""),
            "Viscosity:": sample.get("viscosity", ""),
            "Resistance (Ohms):": sample.get("resistance", ""),
            "Puffing Regime:": sample.get("puffing_regime", ""),
            "Initial Oil Mass:": sample.get("initial_oil_mass", ""),
            "Date:": sample.get("date", ""),
            "Media:": sample.get("media", "")
        }

        # Write metadata to template
        for template_key, (patterns, (tpl_row, tpl_col_offset)) in meta_data_MAPPING.items():
            value = meta_data_values.get(template_key, "")
            if value:
                print(f"DEBUG: Set {template_key} to '{value}' at row {tpl_row}, col {col_offset + tpl_col_offset}")
                ws.cell(row=tpl_row, column=col_offset + tpl_col_offset, value=value)

        # Write data columns (using same logic as v2)
        for data_key, (data_col_offset, is_numeric) in DATA_COL_MAPPING.items():
            if data_key in sample:
                data_series = sample[data_key]
                target_col = col_offset + data_col_offset

                # Convert pandas Series to list for consistent handling
                if hasattr(data_series, 'tolist'):
                    data_values = data_series.tolist()
                elif hasattr(data_series, '__iter__'):
                    data_values = list(data_series)
                else:
                    data_values = [data_series] if data_series else []

                # Remove NaN and empty values
                clean_values = []
                for val in data_values:
                    if pd.notna(val) and str(val).strip() != '':
                        clean_values.append(val)

                print(f"DEBUG: Writing {data_key} data to column {target_col} ({len(clean_values)} values)")

                # Write the data starting from row 4 (data_start_row)
                for row_idx, value in enumerate(clean_values, start=4):
                    if is_numeric:
                        try:
                            numeric_value = float(value)
                            ws.cell(row=row_idx, column=target_col, value=numeric_value)
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=target_col, value=value)
                    else:
                        ws.cell(row=row_idx, column=target_col, value=str(value))

    # Keep only the Intense Test sheet
    sheets_to_keep = [template_sheet]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]

    # Rename the sheet based on the legacy file name (limited to 31 characters)
    base_name = os.path.splitext(os.path.basename(legacy_file_path))[0]
    new_sheet_name = f"Legacy_{base_name}"[:31]
    ws.title = new_sheet_name

    # Ensure legacy data directory exists
    folder_path = os.path.join(os.path.abspath("."), "legacy data")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    new_file_name = f"{base_name} Legacy Cart.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)
    wb.save(new_file_path)

    print(f"DEBUG: Saved processed cart format file to: {new_file_path}")
    return load_excel_file(new_file_path)[new_sheet_name]
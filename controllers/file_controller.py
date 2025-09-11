# controllers/file_controller.py
"""
controllers/file_controller.py
File operations controller that coordinates between file service and models.
This replaces the coordination logic currently in file_manager.py.
"""

import os
import re
import copy
import shutil
import time
import uuid
import tempfile
import threading
import subprocess
import traceback
from typing import Optional, Dict, Any, List, Tuple

import pandas as pd
import psutil
import openpyxl
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label, Button, ttk, Frame, Listbox, Scrollbar

# Local imports from refactor_files that need to be consolidated
import processing
from resource_utils import get_resource_path
from utils import (
    is_valid_excel_file,
    load_excel_file,
    get_save_path,
    is_standard_file,
    FONT,
    APP_BACKGROUND_COLOR,
    load_excel_file_with_formulas,
    debug_print,
    show_success_message
)
from database_manager import DatabaseManager

# Model imports
from models.file_model import FileModel, FileState
from models.data_model import DataModel


class FileController:
    """Controller for file operations and data loading."""
    
    def __init__(self, file_model: FileModel, data_model: DataModel, 
                 file_service: Any, database_service: Any):
        """Initialize the file controller."""
        self.file_model = file_model
        self.data_model = data_model
        self.file_service = file_service
        self.database_service = database_service
        
        # Initialize database manager
        self.db_manager = DatabaseManager()
        
        # Cross-controller references (set later)
        self.plot_controller: Optional['PlotController'] = None
        
        # Add cache to prevent redundant operations
        self.loaded_files_cache = {}  # Cache for loaded file data
        self.stored_files_cache = set()  # Track files already stored in database
        
        # GUI reference will be set when connected
        self.gui = None
        
        print("DEBUG: FileController initialized")
        print(f"DEBUG: Connected to FileModel and DataModel")
        print("DEBUG: Database manager initialized")
        print("DEBUG: Cache systems initialized")
    
    def set_gui_reference(self, gui):
        """Set reference to main GUI for UI updates."""
        self.gui = gui
        print("DEBUG: FileController connected to GUI")
    
    def set_plot_controller(self, plot_controller: 'PlotController'):
        """Set reference to plot controller for notifications."""
        self.plot_controller = plot_controller
        print("DEBUG: FileController connected to PlotController")
    
    def load_initial_file(self) -> None:
        """Load the initial Excel file with file dialog selection."""
        debug_print("DEBUG: FileController starting initial file load")
        
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel file",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                debug_print("DEBUG: No file selected")
                return
            
            debug_print(f"DEBUG: Selected file: {file_path}")
            
            # Load the file
            success = self.load_excel_file(file_path)
            
            if success:
                debug_print("DEBUG: Initial file load completed successfully")
                # Update UI if GUI reference is available
                if self.gui:
                    self.gui.populate_sheet_dropdown()
                    self.gui.populate_file_dropdown()
            else:
                debug_print("ERROR: Initial file load failed")
                
        except Exception as e:
            debug_print(f"ERROR: Exception during initial file load: {e}")
            messagebox.showerror("Error", f"Failed to load file: {e}")
    
    def load_excel_file(self, file_path: str, legacy_mode: str = None, 
                       skip_database_storage: bool = False, force_reload: bool = False) -> bool:
        """Load the selected Excel file and process its sheets."""
        debug_print(f"DEBUG: FileController load_excel_file called for {file_path}")
        debug_print(f"DEBUG: Parameters - legacy_mode={legacy_mode}, skip_db_storage={skip_database_storage}, force_reload={force_reload}")

        # Check cache first (unless force_reload is True)
        cache_key = f"{file_path}_{legacy_mode}"
        if not force_reload and cache_key in self.loaded_files_cache:
            debug_print("DEBUG: Using cached file data instead of reprocessing")
            cached_data = self.loaded_files_cache[cache_key]
            
            if self.gui:
                self.gui.filtered_sheets = cached_data['filtered_sheets']
                self.gui.sheets = cached_data.get('sheets', {})
                
                # Set the selected sheet without triggering full UI update
                if cached_data['filtered_sheets']:
                    first_sheet = list(cached_data['filtered_sheets'].keys())[0]
                    self.gui.selected_sheet.set(first_sheet)
            return True

        # Clear cache entry if force_reload is True
        if force_reload and cache_key in self.loaded_files_cache:
            debug_print(f"DEBUG: Force reload requested - clearing cache entry for {file_path}")
            del self.loaded_files_cache[cache_key]

        try:
            # Ensure the file is a valid Excel file
            if not is_valid_excel_file(os.path.basename(file_path)):
                raise ValueError(f"Invalid Excel file selected: {file_path}")

            debug_print(f"DEBUG: {'Force reloading' if force_reload else 'Loading'} file from disk: {file_path}")
            debug_print(f"DEBUG: Checking if file is standard format: {file_path}")
        
            if not is_standard_file(file_path):
                debug_print("DEBUG: File is legacy format, processing accordingly")
                return self._process_legacy_file(file_path, legacy_mode, skip_database_storage)
            else:
                debug_print("DEBUG: File is standard format, processing accordingly")
                return self._process_standard_file(file_path, skip_database_storage, force_reload, cache_key)
                
        except Exception as e:
            debug_print(f"ERROR: FileController failed to load {file_path}: {e}")
            traceback.print_exc()
            if self.gui:
                messagebox.showerror("Error", f"Error loading file: {e}")
            return False
    
    def _process_legacy_file(self, file_path: str, legacy_mode: str, skip_database_storage: bool) -> bool:
        """Process legacy format Excel file."""
        debug_print("DEBUG: Processing legacy file")
        
        try:
            # Legacy file processing
            legacy_dir = os.path.join(os.path.abspath("."), "legacy data")
            if not os.path.exists(legacy_dir):
                os.makedirs(legacy_dir)
                debug_print(f"DEBUG: Created legacy data directory: {legacy_dir}")
        
            legacy_wb = load_workbook(file_path)
            legacy_sheetnames = legacy_wb.sheetnames
            debug_print(f"DEBUG: Legacy file sheets: {legacy_sheetnames}")

            template_path_default = os.path.join(os.path.abspath("."), "resources", 
                                               "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
            
            if not os.path.exists(template_path_default):
                raise FileNotFoundError(f"Template file not found: {template_path_default}")

            # Process legacy file according to mode
            if legacy_mode == 'viscosity':
                debug_print("DEBUG: Processing legacy viscosity file")
                processed_data = processing.convert_legacy_viscosity_using_template(file_path, template_path_default)
            elif legacy_mode == 'standards':
                debug_print("DEBUG: Processing legacy standards file")
                processed_data = processing.convert_legacy_standards_using_template(file_path, template_path_default)
            else:
                # Default legacy processing
                debug_print("DEBUG: Processing general legacy file")
                processed_data = processing.convert_legacy_using_template(file_path, template_path_default)
                
            # Update GUI with processed data
            if self.gui and processed_data:
                self.gui.filtered_sheets = processed_data
                debug_print(f"DEBUG: Loaded {len(processed_data)} sheets from legacy file")
                
                # Store in database if not skipping
                if not skip_database_storage:
                    self._store_in_database(file_path, processed_data)
                
                # Cache the result
                cache_key = f"{file_path}_{legacy_mode}"
                self.loaded_files_cache[cache_key] = {
                    'filtered_sheets': processed_data,
                    'sheets': processed_data
                }
                
                return True
            else:
                debug_print("ERROR: No processed data returned from legacy file")
                return False
                
        except Exception as e:
            debug_print(f"ERROR: Failed to process legacy file: {e}")
            traceback.print_exc()
            return False
    
    def _process_standard_file(self, file_path: str, skip_database_storage: bool, 
                              force_reload: bool, cache_key: str) -> bool:
        """Process standard format Excel file."""
        debug_print("DEBUG: Processing standard file")
        
        try:
            # Load Excel file using the utility function
            sheets = load_excel_file(file_path)
            debug_print(f"DEBUG: Loaded {len(sheets)} sheets from Excel file")
            
            # Filter sheets to get only data sheets
            filtered_sheets = {}
            for sheet_name, sheet_data in sheets.items():
                if processing.is_data_sheet(sheet_data):
                    filtered_sheets[sheet_name] = {
                        'data': sheet_data,
                        'source_file': os.path.basename(file_path)
                    }
            
            debug_print(f"DEBUG: Filtered to {len(filtered_sheets)} data sheets")
            
            # Update GUI with loaded data
            if self.gui:
                self.gui.filtered_sheets = filtered_sheets
                self.gui.sheets = sheets
                self.gui.file_path = file_path
                
                # Add to file list
                file_name = os.path.basename(file_path)
                file_data = {
                    "file_name": file_name,
                    "file_path": file_path,
                    "filtered_sheets": filtered_sheets
                }
                
                # Check if already in all_filtered_sheets
                existing_index = None
                for i, existing_file in enumerate(self.gui.all_filtered_sheets):
                    if existing_file["file_path"] == file_path:
                        existing_index = i
                        break
                
                if existing_index is not None:
                    # Update existing entry
                    self.gui.all_filtered_sheets[existing_index] = file_data
                else:
                    # Add new entry
                    self.gui.all_filtered_sheets.append(file_data)
            
            # Store in database if not skipping
            if not skip_database_storage:
                self._store_in_database(file_path, filtered_sheets)
            
            # Cache the result
            self.loaded_files_cache[cache_key] = {
                'filtered_sheets': filtered_sheets,
                'sheets': sheets
            }
            
            debug_print("DEBUG: Standard file processing completed successfully")
            return True
            
        except Exception as e:
            debug_print(f"ERROR: Failed to process standard file: {e}")
            traceback.print_exc()
            return False
    
    def _store_in_database(self, file_path: str, filtered_sheets: Dict) -> None:
        """Store file data in database."""
        try:
            debug_print("DEBUG: Storing file data in database")
            
            # Check if already stored to avoid duplicates
            if file_path in self.stored_files_cache:
                debug_print(f"DEBUG: File {file_path} already stored in database")
                return
            
            # Store using database manager
            file_id = self.db_manager.store_file(file_path, filtered_sheets)
            
            if file_id:
                self.stored_files_cache.add(file_path)
                debug_print(f"DEBUG: Successfully stored file in database with ID: {file_id}")
            else:
                debug_print("WARNING: Failed to store file in database")
                
        except Exception as e:
            debug_print(f"ERROR: Database storage failed: {e}")
    
    def batch_load_folder(self) -> None:
        """Batch load Excel files from a selected folder and all subfolders."""
        debug_print("DEBUG: Starting batch folder loading process")
    
        # Get folder selection from user
        folder_path = filedialog.askdirectory(
            title="Select Folder for Batch Loading (will scan all subfolders)"
        )
    
        if not folder_path:
            debug_print("DEBUG: No folder selected, canceling batch load")
            return
    
        try:
            # Get test names from processing dictionary for filename matching
            test_names = self._get_test_names_for_matching()
            debug_print(f"DEBUG: Got {len(test_names)} test names for matching: {test_names[:5]}...")
        
            # Scan folder recursively for Excel files
            debug_print(f"DEBUG: Scanning folder: {folder_path}")
            excel_files = self._scan_folder_for_excel_files(folder_path)
            debug_print(f"DEBUG: Found {len(excel_files)} Excel files total")
        
            # Filter files based on test name matching
            matching_files = self._filter_files_by_test_names(excel_files, test_names)
            debug_print(f"DEBUG: Found {len(matching_files)} files matching test names")
        
            if not matching_files:
                messagebox.showinfo("No Matching Files", 
                                   f"No Excel files found in '{folder_path}' with test names in their filename.\n\n"
                                   f"Looking for files containing any of these test names: {', '.join(test_names[:10])}{'...' if len(test_names) > 10 else ''}")
                return
        
            # Confirm with user before proceeding
            proceed = messagebox.askyesno("Batch Load Confirmation", 
                                        f"Found {len(matching_files)} Excel files to load.\n\n"
                                        f"This may take several minutes. Proceed with batch loading?")
        
            if proceed:
                self._perform_batch_loading(matching_files)
                
        except Exception as e:
            debug_print(f"ERROR: Batch loading failed: {e}")
            messagebox.showerror("Error", f"Batch loading failed: {e}")
    
    def _get_test_names_for_matching(self) -> List[str]:
        """Get test names from processing dictionary for filename matching."""
        try:
            test_names = []
            if hasattr(processing, 'processing_dict'):
                for test_type, test_data in processing.processing_dict.items():
                    test_names.append(test_type.lower())
                    if isinstance(test_data, dict) and 'aliases' in test_data:
                        test_names.extend([alias.lower() for alias in test_data['aliases']])
            
            # Add common test name patterns
            common_patterns = ['test', 'sample', 'data', 'analysis', 'measurement']
            test_names.extend(common_patterns)
            
            # Remove duplicates and return
            return list(set(test_names))
            
        except Exception as e:
            debug_print(f"ERROR: Failed to get test names: {e}")
            return ['test', 'sample', 'data']  # Fallback list
    
    def _scan_folder_for_excel_files(self, folder_path: str) -> List[str]:
        """Scan folder recursively for Excel files."""
        debug_print(f"DEBUG: Starting recursive scan of folder: {folder_path}")
    
        excel_files = []
        excel_extensions = ('.xlsx', '.xls')
    
        try:
            for root, dirs, files in os.walk(folder_path):
                debug_print(f"DEBUG: Scanning directory: {root}")
            
                for file in files:
                    if file.lower().endswith(excel_extensions):
                        full_path = os.path.join(root, file)
                    
                        # Skip temporary Excel files
                        if file.startswith('~$'):
                            debug_print(f"DEBUG: Skipping temporary file: {file}")
                            continue
                    
                        # Check if file is accessible
                        try:
                            if os.access(full_path, os.R_OK):
                                excel_files.append(full_path)
                                debug_print(f"DEBUG: Added Excel file: {file}")
                            else:
                                debug_print(f"DEBUG: Skipping inaccessible file: {file}")
                        except Exception as e:
                            debug_print(f"DEBUG: Error checking file access for {file}: {e}")
                            continue
        
            debug_print(f"DEBUG: Completed folder scan, found {len(excel_files)} Excel files")
            return excel_files
        
        except Exception as e:
            debug_print(f"ERROR: Failed to scan folder: {e}")
            raise
    
    def _filter_files_by_test_names(self, excel_files: List[str], test_names: List[str]) -> List[str]:
        """Filter Excel files by checking if filename contains any test names."""
        debug_print(f"DEBUG: Filtering {len(excel_files)} files against {len(test_names)} test names")
    
        matching_files = []
    
        for file_path in excel_files:
            filename = os.path.basename(file_path).lower()
            debug_print(f"DEBUG: Checking file: {filename}")
        
            # Check if any test name is contained in the filename
            for test_name in test_names:
                if test_name in filename:
                    matching_files.append(file_path)
                    debug_print(f"DEBUG: MATCH - '{filename}' contains '{test_name}'")
                    break
    
        debug_print(f"DEBUG: Filtered to {len(matching_files)} matching files")
        return matching_files
    
    def _perform_batch_loading(self, file_paths: List[str]) -> None:
        """Perform the actual batch loading with accurate progress tracking."""
        debug_print(f"DEBUG: Starting batch loading of {len(file_paths)} files")
        
        # Create progress dialog
        progress_dialog = Toplevel(self.gui.root if self.gui else None)
        progress_dialog.title("Batch Loading Progress")
        progress_dialog.geometry("400x150")
        progress_dialog.transient(self.gui.root if self.gui else None)
        progress_dialog.grab_set()
        
        # Progress label
        progress_label = Label(progress_dialog, text="Initializing batch load...")
        progress_label.pack(pady=10)
        
        # Progress bar
        progress_bar = ttk.Progressbar(progress_dialog, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        progress_bar['maximum'] = len(file_paths)
        
        # Status label
        status_label = Label(progress_dialog, text="")
        status_label.pack(pady=5)
        
        # Center the dialog
        if self.gui:
            self.gui.center_window(progress_dialog)
        
        # Track progress
        loaded_count = 0
        failed_count = 0
        failed_files = []
        
        try:
            for i, file_path in enumerate(file_paths):
                try:
                    # Update progress display
                    filename = os.path.basename(file_path)
                    progress_label.config(text=f"Loading file {i+1} of {len(file_paths)}")
                    status_label.config(text=f"Processing: {filename}")
                    progress_bar['value'] = i
                    progress_dialog.update()
                    
                    debug_print(f"DEBUG: Batch loading file {i+1}/{len(file_paths)}: {filename}")
                    
                    # Load the file
                    success = self.load_excel_file(file_path, skip_database_storage=False)
                    
                    if success:
                        loaded_count += 1
                        debug_print(f"DEBUG: Successfully loaded {filename}")
                    else:
                        failed_count += 1
                        failed_files.append(filename)
                        debug_print(f"WARNING: Failed to load {filename}")
                        
                except Exception as e:
                    failed_count += 1
                    failed_files.append(os.path.basename(file_path))
                    debug_print(f"ERROR: Exception loading {file_path}: {e}")
            
            # Final progress update
            progress_bar['value'] = len(file_paths)
            progress_label.config(text="Batch loading completed!")
            status_label.config(text=f"Loaded: {loaded_count}, Failed: {failed_count}")
            progress_dialog.update()
            
            # Show completion message
            completion_msg = f"Batch loading completed!\n\nSuccessfully loaded: {loaded_count} files\nFailed: {failed_count} files"
            if failed_files:
                completion_msg += f"\n\nFailed files:\n" + "\n".join(failed_files[:10])
                if len(failed_files) > 10:
                    completion_msg += f"\n... and {len(failed_files) - 10} more"
            
            messagebox.showinfo("Batch Load Complete", completion_msg)
            
            # Update GUI
            if self.gui:
                self.gui.populate_file_dropdown()
                self.gui.populate_sheet_dropdown()
            
        except Exception as e:
            debug_print(f"ERROR: Critical error during batch loading: {e}")
            messagebox.showerror("Error", f"Critical error during batch loading: {e}")
            
        finally:
            progress_dialog.destroy()
            debug_print(f"DEBUG: Batch loading completed - {loaded_count} success, {failed_count} failed")
    
    def load_vap3_file(self) -> None:
        """Load a .vap3 file using the file dialog."""
        debug_print("DEBUG: FileController loading VAP3 file")
        
        try:
            file_path = filedialog.askopenfilename(
                title="Open VAP3 file",
                filetypes=[("VAP3 files", "*.vap3"), ("All files", "*.*")]
            )
            
            if not file_path:
                debug_print("DEBUG: No VAP3 file selected")
                return
            
            # Load VAP3 file
            success = self._load_vap3_file_internal(file_path)
            
            if success:
                debug_print("DEBUG: VAP3 file loaded successfully")
                # Update UI
                if self.gui:
                    self.gui.populate_sheet_dropdown()
                    self.gui.populate_file_dropdown()
            else:
                debug_print("ERROR: VAP3 file loading failed")
                
        except Exception as e:
            debug_print(f"ERROR: Exception during VAP3 file load: {e}")
            messagebox.showerror("Error", f"Failed to load VAP3 file: {e}")
    
    def _load_vap3_file_internal(self, file_path: str, display_name: str = None, 
                                append_to_existing: bool = False) -> bool:
        """Internal method to load VAP3 file data."""
        debug_print(f"DEBUG: Loading VAP3 file internally: {file_path}")
        
        try:
            from vap_file_manager import VapFileManager
            vap_manager = VapFileManager()
            
            # Load VAP3 data
            vap_data = vap_manager.load_from_vap3(file_path)
            debug_print(f"DEBUG: VAP3 data keys: {list(vap_data.keys())}")
            
            if not vap_data or 'filtered_sheets' not in vap_data:
                debug_print("ERROR: Invalid VAP3 data structure")
                return False
            
            # Extract data
            filtered_sheets = vap_data['filtered_sheets']
            sheet_images = vap_data.get('sheet_images', {})
            plot_options = vap_data.get('plot_options', {})
            image_crop_states = vap_data.get('image_crop_states', {})
            
            # Update GUI
            if self.gui:
                if append_to_existing:
                    # Append to existing data
                    for sheet_name, sheet_data in filtered_sheets.items():
                        self.gui.filtered_sheets[sheet_name] = sheet_data
                    
                    # Merge other data
                    if hasattr(self.gui, 'sheet_images'):
                        self.gui.sheet_images.update(sheet_images)
                    if hasattr(self.gui, 'plot_options'):
                        self.gui.plot_options.update(plot_options)
                    if hasattr(self.gui, 'image_crop_states'):
                        self.gui.image_crop_states.update(image_crop_states)
                else:
                    # Replace existing data
                    self.gui.filtered_sheets = filtered_sheets
                    self.gui.sheet_images = sheet_images
                    self.gui.plot_options = plot_options
                    self.gui.image_crop_states = image_crop_states
                
                # Set file path
                self.gui.file_path = file_path
                
                # Add to file list
                file_name = display_name or os.path.basename(file_path)
                file_data = {
                    "file_name": file_name,
                    "file_path": file_path,
                    "filtered_sheets": filtered_sheets
                }
                
                if not append_to_existing:
                    self.gui.all_filtered_sheets = [file_data]
                else:
                    self.gui.all_filtered_sheets.append(file_data)
                
                # Load sample images if they exist
                if 'sample_images' in vap_data and vap_data['sample_images']:
                    debug_print("DEBUG: Loading sample images from VAP3")
                    self.gui.load_sample_images_from_vap3(vap_data)
            
            debug_print("DEBUG: VAP3 file loaded successfully")
            return True
            
        except Exception as e:
            debug_print(f"ERROR: Failed to load VAP3 file: {e}")
            traceback.print_exc()
            return False
    
    def save_as_vap3(self) -> None:
        """Save the current session as a .vap3 file."""
        debug_print("DEBUG: FileController saving as VAP3")
        
        try:
            if not self.gui or not hasattr(self.gui, 'filtered_sheets') or not self.gui.filtered_sheets:
                messagebox.showerror("Error", "No data loaded to save")
                return
            
            # Get save path
            file_path = filedialog.asksaveasfilename(
                title="Save as VAP3 file",
                defaultextension=".vap3",
                filetypes=[("VAP3 files", "*.vap3"), ("All files", "*.*")]
            )
            
            if not file_path:
                debug_print("DEBUG: No save path selected")
                return
            
            # Prepare data for saving
            filtered_sheets = self.gui.filtered_sheets
            sheet_images = getattr(self.gui, 'sheet_images', {})
            plot_options = getattr(self.gui, 'plot_options', {})
            image_crop_states = getattr(self.gui, 'image_crop_states', {})
            
            # Prepare sample images data
            sample_images = {}
            sample_image_crop_states = {}
            sample_header_data = {}
            
            # Check for pending sample images
            if hasattr(self.gui, 'pending_sample_images'):
                sample_images = getattr(self.gui, 'pending_sample_images', {})
                sample_image_crop_states = getattr(self.gui, 'pending_sample_image_crop_states', {})
                sample_header_data = getattr(self.gui, 'pending_sample_header_data', {})
                debug_print(f"DEBUG: Found pending sample images: {len(sample_images)} samples")
            
            # Save as VAP3
            from vap_file_manager import VapFileManager
            vap_manager = VapFileManager()
            
            success = vap_manager.save_to_vap3(
                file_path,
                filtered_sheets,
                sheet_images,
                plot_options,
                image_crop_states,
                sample_images,
                sample_image_crop_states,
                sample_header_data
            )
            
            if success:
                debug_print(f"DEBUG: Successfully saved VAP3 file: {file_path}")
                self.gui.file_path = file_path
                messagebox.showinfo("Success", f"File saved successfully as {os.path.basename(file_path)}")
            else:
                debug_print("ERROR: Failed to save VAP3 file")
                messagebox.showerror("Error", "Failed to save VAP3 file")
                
        except Exception as e:
            debug_print(f"ERROR: Exception during VAP3 save: {e}")
            messagebox.showerror("Error", f"Failed to save VAP3 file: {e}")
    
    def create_new_template(self, startup_menu) -> None:
        """Create a new template file."""
        debug_print("DEBUG: FileController creating new template")
        
        try:
            # Close the startup menu
            startup_menu.destroy()
            
            # Launch test start menu for new template creation
            if self.gui:
                test_start_menu = TestStartMenu(self.gui.root, self)
            else:
                debug_print("ERROR: No GUI reference available for new template creation")
                
        except Exception as e:
            debug_print(f"ERROR: Failed to create new template: {e}")
            messagebox.showerror("Error", f"Failed to create new template: {e}")
    
    def show_database_browser(self, comparison_mode: bool = False) -> None:
        """Show database browser for file selection."""
        debug_print(f"DEBUG: FileController showing database browser, comparison_mode={comparison_mode}")
        
        try:
            # Get files from database
            files = self.db_manager.get_files_with_sheet_info()
            
            if not files:
                messagebox.showinfo("No Files", "No files found in database")
                return
            
            # Create database browser dialog
            self._create_database_browser_dialog(files, comparison_mode)
            
        except Exception as e:
            debug_print(f"ERROR: Failed to show database browser: {e}")
            messagebox.showerror("Error", f"Failed to show database browser: {e}")
    
    def _create_database_browser_dialog(self, files: List[Dict], comparison_mode: bool) -> None:
        """Create and show database browser dialog."""
        debug_print(f"DEBUG: Creating database browser dialog with {len(files)} files")
        
        # Create dialog window
        dialog = Toplevel(self.gui.root if self.gui else None)
        dialog.title("Database Browser" + (" - Comparison Mode" if comparison_mode else ""))
        dialog.geometry("800x600")
        if self.gui:
            dialog.transient(self.gui.root)
            dialog.grab_set()
        
        # Create file list
        file_frame = Frame(dialog)
        file_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # File listbox with scrollbar
        listbox_frame = Frame(file_frame)
        listbox_frame.pack(fill=tk.BOTH, expand=True)
        
        file_listbox = Listbox(listbox_frame, selectmode=tk.MULTIPLE if comparison_mode else tk.SINGLE)
        scrollbar = Scrollbar(listbox_frame, orient=tk.VERTICAL, command=file_listbox.yview)
        file_listbox.config(yscrollcommand=scrollbar.set)
        
        file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate file list
        listbox_to_file_mapping = {}
        for i, file_record in enumerate(files):
            filename = file_record['filename']
            created_str = file_record['created_at'].strftime('%Y-%m-%d %H:%M')
            display_text = f"{filename} ({created_str})"
            
            file_listbox.insert(tk.END, display_text)
            listbox_to_file_mapping[i] = file_record
        
        # Button frame
        button_frame = Frame(dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=5)
        
        if comparison_mode:
            # Comparison mode buttons
            def on_compare():
                selected_items = file_listbox.curselection()
                if len(selected_items) < 2:
                    messagebox.showwarning("Warning", "Please select at least 2 files for comparison")
                    return
                
                # Load selected files for comparison
                selected_files = [listbox_to_file_mapping[i] for i in selected_items]
                dialog.destroy()
                
                if self.gui:
                    self.gui.load_files_from_database_for_comparison(selected_files)
            
            Button(button_frame, text="Compare Selected", command=on_compare).pack(side=tk.LEFT, padx=5)
        else:
            # Regular mode buttons
            def on_load():
                selected_items = file_listbox.curselection()
                if not selected_items:
                    messagebox.showwarning("Warning", "Please select a file to load")
                    return
                
                # Load selected file
                selected_file = listbox_to_file_mapping[selected_items[0]]
                dialog.destroy()
                
                self.load_from_database(selected_file)
            
            Button(button_frame, text="Load Selected", command=on_load).pack(side=tk.LEFT, padx=5)
        
        # Close button
        Button(button_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Center the dialog
        if self.gui:
            self.gui.center_window(dialog)
    
    def load_from_database(self, file_record: Dict, append_to_existing: bool = False) -> bool:
        """Load a specific file from the database."""
        debug_print(f"DEBUG: Loading file from database: {file_record['filename']}")
        
        try:
            # Get file data from database
            file_id = file_record['id']
            filename = file_record['filename']
            
            # Create temporary VAP3 file
            temp_vap3_path = os.path.join(tempfile.gettempdir(), f"db_file_{file_id}_{uuid.uuid4().hex[:8]}.vap3")
            display_filename = f"{filename} (from database)"
            
            # Export from database to VAP3
            success = self.db_manager.export_file_to_vap3(file_id, temp_vap3_path)
            
            if not success:
                debug_print("ERROR: Failed to export file from database")
                return False
            
            # Load the temporary VAP3 file
            from vap_file_manager import VapFileManager
            vap_manager = VapFileManager()
            vap_data = vap_manager.load_from_vap3(temp_vap3_path)
            
            if self.gui:
                # Store VAP3 data in GUI for sample image loading
                self.gui.current_vap_data = vap_data
                debug_print(f"DEBUG: Loaded VAP3 data with keys: {list(vap_data.keys())}")
            
            # Use the enhanced VAP3 loading that handles sample images
            success = self._load_vap3_file_internal(temp_vap3_path, display_name=display_filename, 
                                                 append_to_existing=append_to_existing)
            
            if success and self.gui:
                # Load sample images if they exist
                if 'sample_images' in vap_data and vap_data['sample_images']:
                    debug_print("DEBUG: Loading sample images from database VAP3")
                    self.gui.load_sample_images_from_vap3(vap_data)
                
                total_files = len(self.gui.all_filtered_sheets)
                debug_print(f"DEBUG: Successfully loaded from database. Total files: {total_files}")
                
                # Update UI
                self.gui.populate_file_dropdown()
                self.gui.populate_sheet_dropdown()
            
            # Clean up temporary file
            try:
                os.remove(temp_vap3_path)
                debug_print("DEBUG: Cleaned up temporary VAP3 file")
            except Exception as e:
                debug_print(f"WARNING: Failed to clean up temporary file: {e}")
            
            return success
            
        except Exception as e:
            debug_print(f"ERROR: Failed to load file from database: {e}")
            traceback.print_exc()
            return False
    
    def open_raw_data_in_excel(self, sheet_name: str) -> None:
        """Open raw data in Excel for editing."""
        debug_print(f"DEBUG: Opening raw data in Excel for sheet: {sheet_name}")
        
        try:
            if not self.gui or not hasattr(self.gui, 'filtered_sheets'):
                debug_print("ERROR: No data available to open")
                return
            
            if sheet_name not in self.gui.filtered_sheets:
                debug_print(f"ERROR: Sheet {sheet_name} not found in filtered sheets")
                return
            
            # Get sheet data
            sheet_data = self.gui.filtered_sheets[sheet_name]['data']
            
            # Create temporary Excel file
            temp_file = os.path.join(tempfile.gettempdir(), f"temp_data_{uuid.uuid4().hex[:8]}.xlsx")
            
            # Save data to temporary file
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            debug_print(f"DEBUG: Created temporary Excel file: {temp_file}")
            
            # Open in Excel
            if os.name == 'nt':  # Windows
                os.startfile(temp_file)
            elif os.name == 'posix':  # macOS and Linux
                subprocess.Popen(['open', temp_file])
            
            # Monitor file for changes
            self._monitor_excel_file_changes(temp_file, sheet_name)
            
        except Exception as e:
            debug_print(f"ERROR: Failed to open raw data in Excel: {e}")
            messagebox.showerror("Error", f"Failed to open raw data in Excel: {e}")
    
    def _monitor_excel_file_changes(self, temp_file: str, sheet_name: str) -> None:
        """Monitor Excel file for changes and import them back."""
        debug_print(f"DEBUG: Starting file monitoring for: {temp_file}")
        
        if not self.gui:
            debug_print("ERROR: No GUI reference for file monitoring")
            return
        
        # Create status label
        status_label = ttk.Label(
            self.gui.root,
            text=f"Excel file opened for editing. Changes will be imported when file is closed.",
            relief="sunken",
            anchor="w"
        )
        status_label.pack(side="bottom", fill="x")
        
        # Start monitoring thread
        def monitor_file():
            debug_print("DEBUG: File monitoring thread started")
            original_mtime = os.path.getmtime(temp_file)
            file_changed = False
            
            while os.path.exists(temp_file):
                try:
                    # Check if file is still open in Excel
                    if self._is_file_open_in_excel(temp_file):
                        time.sleep(2)  # Check every 2 seconds
                        
                        # Check for modifications
                        current_mtime = os.path.getmtime(temp_file)
                        if current_mtime > original_mtime:
                            file_changed = True
                            original_mtime = current_mtime
                            debug_print("DEBUG: File modification detected")
                        
                        continue
                    else:
                        # File is closed
                        debug_print("DEBUG: Excel file has been closed")
                        break
                        
                except Exception as e:
                    debug_print(f"ERROR: Exception in file monitoring: {e}")
                    break
            
            # Import changes if file was modified
            try:
                if status_label and status_label.winfo_exists():
                    status_label.destroy()
                
                if file_changed and os.path.exists(temp_file):
                    try:
                        # Read modified data with retries
                        max_retries = 3
                        retry_count = 0
                        read_success = False
                        modified_data = None
                        
                        while not read_success and retry_count < max_retries:
                            try:
                                modified_data = pd.read_excel(temp_file)
                                read_success = True
                            except Exception:
                                retry_count += 1
                                time.sleep(1.0)
                        
                        if not read_success:
                            raise Exception(f"Failed to read Excel file after {max_retries} attempts")
                        
                        # Update the filtered sheets with new data
                        self.gui.filtered_sheets[sheet_name]['data'] = modified_data
                        
                        # If using VAP3 file, update it
                        if hasattr(self.gui, 'file_path') and self.gui.file_path.endswith('.vap3'):
                            from vap_file_manager import VapFileManager
                            vap_manager = VapFileManager()
                            vap_manager.save_to_vap3(
                                self.gui.file_path,
                                self.gui.filtered_sheets,
                                getattr(self.gui, 'sheet_images', {}),
                                getattr(self.gui, 'plot_options', {}),
                                getattr(self.gui, 'image_crop_states', {})
                            )
                        
                        # Refresh the GUI display
                        self.gui.update_displayed_sheet(sheet_name)
                        
                        # Show success message
                        show_success_message("Success", "Changes from Excel have been imported successfully.", self.gui.root)
                        
                    except Exception as e:
                        debug_print(f"ERROR: Failed to import changes: {e}")
                        messagebox.showerror("Error", f"Failed to read modified Excel file: {e}")
                        
                elif file_changed and not os.path.exists(temp_file):
                    show_success_message("Information", "Excel file was modified but appears to have been moved or renamed. Changes could not be imported.", self.gui.root)
                else:
                    info_label = ttk.Label(
                        self.gui.root,
                        text="Excel file was closed without changes.",
                        relief="sunken",
                        anchor="w"
                    )
                    info_label.pack(side="bottom", fill="x")
                    self.gui.root.after(3000, lambda: info_label.destroy() if info_label.winfo_exists() else None)
                
                # Clean up temporary file
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        debug_print("DEBUG: Cleaned up temporary Excel file")
                except Exception as e:
                    debug_print(f"WARNING: Failed to clean up temporary file: {e}")
                    
            except Exception as e:
                debug_print(f"ERROR: Exception in file monitoring cleanup: {e}")
        
        # Start monitoring in separate thread
        monitor_thread = threading.Thread(target=monitor_file, daemon=True)
        monitor_thread.start()
        
        if self.gui and hasattr(self.gui, 'threads'):
            self.gui.threads.append(monitor_thread)
    
    def _is_file_open_in_excel(self, file_path: str) -> bool:
        """Check if file is currently open in Excel."""
        try:
            # Try to open file for writing - if it fails, it's likely open in Excel
            with open(file_path, 'r+b'):
                pass
            return False  # File is not locked, so not open in Excel
        except (IOError, OSError):
            return True  # File is locked, likely open in Excel
    
    def ensure_file_is_loaded_in_ui(self, file_path: str) -> bool:
        """Ensure file is loaded in UI for data collection."""
        debug_print(f"DEBUG: Ensuring file {file_path} is loaded in UI")
        
        try:
            # Load file for data collection
            success = self.load_excel_file(file_path)
            
            if success:
                debug_print(f"DEBUG: File loaded successfully: {file_path}")
                return True
            else:
                debug_print(f"DEBUG: Error loading file: {file_path}")
                if self.gui:
                    messagebox.showerror("Error", f"Failed to load file: {file_path}")
                return False
                
        except Exception as e:
            debug_print(f"DEBUG: Exception loading file: {e}")
            if self.gui:
                messagebox.showerror("Error", f"Failed to load file: {e}")
            return False
    
    def set_active_file(self, file_name: str) -> None:
        """Set the active file from the loaded files list."""
        debug_print(f"DEBUG: Setting active file: {file_name}")
        
        try:
            if not self.gui or not hasattr(self.gui, 'all_filtered_sheets'):
                debug_print("ERROR: No GUI or file data available")
                return
            
            # Find the file in all_filtered_sheets
            target_file = None
            for file_data in self.gui.all_filtered_sheets:
                if file_data["file_name"] == file_name:
                    target_file = file_data
                    break
            
            if not target_file:
                debug_print(f"ERROR: File {file_name} not found in loaded files")
                raise ValueError(f"File '{file_name}' not found in loaded files")
            
            # Set as active
            self.gui.filtered_sheets = target_file["filtered_sheets"]
            self.gui.file_path = target_file["file_path"]
            
            debug_print(f"DEBUG: Successfully set active file: {file_name}")
            
        except Exception as e:
            debug_print(f"ERROR: Failed to set active file: {e}")
            raise
    
    def update_ui_for_current_file(self) -> None:
        """Update UI components for the current active file."""
        debug_print("DEBUG: Updating UI for current file")
        
        try:
            if not self.gui:
                debug_print("ERROR: No GUI reference available")
                return
            
            # Update sheet dropdown
            self.gui.populate_sheet_dropdown()
            
            # Update display if there are sheets
            if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                first_sheet = list(self.gui.filtered_sheets.keys())[0]
                self.gui.selected_sheet.set(first_sheet)
                self.gui.update_displayed_sheet(first_sheet)
            
            debug_print("DEBUG: UI updated successfully for current file")
            
        except Exception as e:
            debug_print(f"ERROR: Failed to update UI for current file: {e}")
    
    # Notify plot controller when data changes
    def on_data_changed(self):
        """Notify plot controller of data changes."""
        debug_print("DEBUG: FileController notifying of data changes")
        
        if self.plot_controller:
            self.plot_controller.on_data_changed()
            debug_print("DEBUG: Plot controller notified of data changes")
        else:
            debug_print("WARNING: No plot controller available to notify")
    
    # Getter methods for UI access
    def get_loaded_files(self) -> List[Dict]:
        """Get list of currently loaded files."""
        if self.gui and hasattr(self.gui, 'all_filtered_sheets'):
            return self.gui.all_filtered_sheets
        return []
    
    def get_current_file_path(self) -> Optional[str]:
        """Get current file path."""
        if self.gui and hasattr(self.gui, 'file_path'):
            return self.gui.file_path
        return None
    
    def get_filtered_sheets(self) -> Dict:
        """Get current filtered sheets."""
        if self.gui and hasattr(self.gui, 'filtered_sheets'):
            return self.gui.filtered_sheets
        return {}
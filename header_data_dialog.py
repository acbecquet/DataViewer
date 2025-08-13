"""
header_data_dialog.py
Developed by Charlie Becquet.
Dialog for entering test header data.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from utils import APP_BACKGROUND_COLOR, FONT, debug_print

class HeaderDataDialog:
    def __init__(self, parent, file_path, selected_test, edit_mode=False, current_data=None):
        """
        Initialize the header data dialog.
    
        Args:
            parent (tk.Tk): The parent window.
            file_path (str): Path to the created file.
            selected_test (str): The selected test name.
            edit_mode (bool): Whether this is edit mode or new entry mode.
            current_data (dict): Existing header data if in edit mode.
        """
        self.parent = parent
        self.file_path = file_path
        self.selected_test = selected_test
        self.edit_mode = edit_mode
        self.current_data = current_data
        self.result = None
        self.header_data = {}
    
        debug_print(f"DEBUG: HeaderDataDialog initialized - edit_mode: {edit_mode}, has_current_data: {current_data is not None}")
    
        # Initialize references for cleanup
        self.canvas = None
        self.mousewheel_binding_id = None
    
        self.dialog = tk.Toplevel(parent)
        title_text = f"Edit Header Data - {selected_test}" if edit_mode else f"Header Data - {selected_test}"
        self.dialog.title(title_text)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.configure(bg=APP_BACKGROUND_COLOR)
        self.dialog.geometry("700x600")  # Increased size for sample-specific fields
    
        # Set up cleanup when dialog is destroyed
        self.dialog.protocol("WM_DELETE_WINDOW", self.cleanup_and_close)

        # Load existing data if in edit mode
        if self.edit_mode:
            if self.current_data:
                # Use provided current data (from data collection window)
                self.existing_data = self.current_data
                debug_print("DEBUG: Using provided current data for header editing")
            else:
                # Load from file (from main GUI)
                self.load_existing_header_data()
    
        self.create_widgets()
        self.center_window()
    
    def center_window(self):
        """Center the dialog on the screen."""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_widgets(self):
        """Create the dialog widgets."""
        debug_print("DEBUG: Creating widgets for HeaderDataDialog")
        
        # Main container with scrolling capability
        main_container = ttk.Frame(self.dialog)
        main_container.pack(fill="both", expand=True)
        
        # Create a canvas with scrollbar
        self.canvas = tk.Canvas(main_container, bg=APP_BACKGROUND_COLOR)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        
        # Pack scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # Frame to hold the form
        self.form_frame = ttk.Frame(self.canvas, padding=20)
        
        # Add the form frame to the canvas
        canvas_frame = self.canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        
        # Configure canvas scrolling
        def configure_canvas(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            self.canvas.itemconfig(canvas_frame, width=event.width)
        
        self.form_frame.bind("<Configure>", configure_canvas)
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(canvas_frame, width=e.width))
        
        # Add mouse wheel scrolling with proper error handling
        def _on_mousewheel(event):
            # Check if canvas still exists and is valid
            if self.canvas and self.canvas.winfo_exists():
                try:
                    self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                except tk.TclError as e:
                    debug_print(f"DEBUG: Canvas scrolling error (canvas may be destroyed): {e}")
                    # Unbind the event if canvas is invalid
                    self.cleanup_mousewheel_binding()
            else:
                debug_print("DEBUG: Canvas no longer exists, cleaning up mousewheel binding")
                self.cleanup_mousewheel_binding()
        
        # Use bind instead of bind_all to limit scope to this canvas
        self.canvas.bind("<MouseWheel>", _on_mousewheel)
        
        # Also bind to the dialog to capture mousewheel when over the dialog
        self.dialog.bind("<MouseWheel>", _on_mousewheel)
        
        debug_print("DEBUG: Canvas and mousewheel binding created successfully")
        
        # Header
        header_label = ttk.Label(
            self.form_frame,
            text=f"Enter header data for {self.selected_test}",
            font=("Arial", 14),
            foreground="black",
            background=APP_BACKGROUND_COLOR
        )
        header_label.grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 20))
        
        # === COMMON FIELDS SECTION ===
        common_frame = ttk.LabelFrame(self.form_frame, text="Common Information", padding=10)
        common_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(0, 20))
        
        # Tester
        ttk.Label(
            common_frame, 
            text="Tester:", 
            foreground="black",
            background=APP_BACKGROUND_COLOR
        ).grid(row=0, column=0, sticky="w", pady=5)
        
        self.tester_var = tk.StringVar()
        ttk.Entry(common_frame, textvariable=self.tester_var, width=30).grid(
            row=0, column=1, sticky="ew", padx=(10, 0), pady=5
        )
        
        # Number of Samples
        ttk.Label(
            common_frame, 
            text="Number of Samples:", 
            foreground="black",
            background=APP_BACKGROUND_COLOR
        ).grid(row=1, column=0, sticky="w", pady=5)
        
        self.num_samples_var = tk.IntVar(value=1)
        samples_spinbox = ttk.Spinbox(
            common_frame, 
            from_=1, 
            to=10, 
            textvariable=self.num_samples_var, 
            width=5,
            command=self.update_sample_fields
        )
        samples_spinbox.grid(row=1, column=1, sticky="w", padx=(10, 0), pady=5)
        
        # Configure grid for common frame
        common_frame.columnconfigure(1, weight=1)
        
        # === SAMPLE-SPECIFIC FIELDS SECTION ===
        self.samples_frame = ttk.Frame(self.form_frame)
        self.samples_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=10)
        
        # Configure grid
        self.form_frame.columnconfigure(0, weight=1)
        
        # Initialize sample field variables
        self.sample_id_vars = []
        self.sample_media_vars = []
        self.sample_viscosity_vars = []
        self.sample_voltage_vars = []
        self.sample_puffing_regime_vars = []
        self.sample_oil_mass_vars = []
        self.resistance_vars = []
        self.update_sample_fields()
        
        # Populate existing data if in edit mode
        if self.edit_mode and hasattr(self, 'existing_data'):
            # Use after_idle to ensure all widgets are created first
            self.dialog.after_idle(self.populate_existing_data)
        
        # Buttons
        button_frame = ttk.Frame(self.form_frame)
        button_frame.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(20, 0))
        
        ttk.Button(button_frame, text="Cancel", command=self.on_cancel).pack(side="right")
        
        button_text = "Update" if self.edit_mode else "Continue"
        ttk.Button(button_frame, text=button_text, command=self.on_continue).pack(side="right", padx=(0, 5))
        
        debug_print("DEBUG: All widgets created successfully")
    
    def populate_existing_data(self):
        """Populate form fields with existing data."""
        if not hasattr(self, 'existing_data') or not self.existing_data:
            return
        
        debug_print("DEBUG: Populating form with existing data")
    
        # Handle both file format and current_data format
        if 'common' in self.existing_data:
            # Data collection format
            common_data = self.existing_data.get('common', {})
            self.tester_var.set(common_data.get('tester', ''))
            samples_data = self.existing_data.get('samples', [])
        else:
            # File format
            self.tester_var.set(self.existing_data.get('tester', ''))
            samples_data = self.existing_data.get('samples', [])
        
        # Set number of samples and populate sample data
        num_existing_samples = len(samples_data)
        if num_existing_samples > 0:
            self.num_samples_var.set(num_existing_samples)
            self.update_sample_fields()
            
            # Populate sample-specific data
            for i, sample_data in enumerate(samples_data):
                if i < len(self.sample_id_vars):
                    self.sample_id_vars[i].set(sample_data.get('id', ''))
                    self.resistance_vars[i].set(sample_data.get('resistance', ''))
                    self.sample_media_vars[i].set(sample_data.get('media', ''))
                    self.sample_viscosity_vars[i].set(sample_data.get('viscosity', ''))
                    self.sample_voltage_vars[i].set(sample_data.get('voltage', ''))
                    self.sample_puffing_regime_vars[i].set(sample_data.get('puffing_regime', '60mL/3s/30s'))
                    self.sample_oil_mass_vars[i].set(sample_data.get('oil_mass', ''))
        
        debug_print("DEBUG: Form populated with existing data")

    def load_existing_header_data(self):
        """Load existing header data from the Excel file."""
        debug_print(f"DEBUG: Loading existing header data from {self.file_path}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path)
            
            if self.selected_test not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet {self.selected_test} not found in file")
                return
                
            ws = wb[self.selected_test]
            
            # Load existing data from known positions
            self.existing_data = {
                'tester': ws.cell(row=1, column=1).value or "",
                'samples': []
            }
            
            # Load sample data (assuming max 10 samples)
            for i in range(10):
                col_offset = i * 12
                sample_id = ws.cell(row=1, column=6 + col_offset).value
                resistance = ws.cell(row=3, column=4 + col_offset).value
                media = ws.cell(row=2, column=2 + col_offset).value
                viscosity = ws.cell(row=3, column=2 + col_offset).value
                voltage = ws.cell(row=2, column=6 + col_offset).value
                oil_mass = ws.cell(row=2, column=8 + col_offset).value
                
                if sample_id:  # Only add if sample ID exists
                    self.existing_data['samples'].append({
                        'id': str(sample_id),
                        'resistance': str(resistance) if resistance else "",
                        'media': str(media) if media else "",
                        'viscosity': str(viscosity) if viscosity else "",
                        'voltage': str(voltage) if voltage else "",
                        'puffing_regime': "60mL/3s/30s",  # Default value
                        'oil_mass': str(oil_mass) if oil_mass else ""
                    })
            
            debug_print(f"DEBUG: Loaded existing data: {len(self.existing_data['samples'])} samples")
            
        except Exception as e:
            debug_print(f"DEBUG: Error loading existing header data: {e}")
            self.existing_data = None

    def cleanup_mousewheel_binding(self):
        """Clean up mousewheel event bindings."""
        try:
            if self.canvas and self.canvas.winfo_exists():
                self.canvas.unbind("<MouseWheel>")
                debug_print("DEBUG: Canvas mousewheel binding cleaned up")
        except tk.TclError:
            debug_print("DEBUG: Canvas already destroyed, binding cleanup not needed")
        
        try:
            if self.dialog and self.dialog.winfo_exists():
                self.dialog.unbind("<MouseWheel>")
                debug_print("DEBUG: Dialog mousewheel binding cleaned up")
        except tk.TclError:
            debug_print("DEBUG: Dialog already destroyed, binding cleanup not needed")
    
    def cleanup_and_close(self):
        """Clean up resources and close the dialog when window is closed."""
        debug_print("DEBUG: Window close button clicked - cleaning up HeaderDataDialog resources")
        self.cleanup_mousewheel_binding()
        self.on_cancel()
    
    def update_sample_fields(self):
        """Update the sample-specific fields with all required fields per sample."""
        debug_print(f"DEBUG: Updating sample fields for {self.num_samples_var.get()} samples")

        # Clear existing widgets
        for widget in self.samples_frame.winfo_children():
            widget.destroy()

        # Get the number of samples
        num_samples = self.num_samples_var.get()

        # Ensure we have enough variables for all sample-specific fields
        while len(self.sample_id_vars) < num_samples:
            self.sample_id_vars.append(tk.StringVar())
            self.resistance_vars.append(tk.StringVar())
            self.sample_media_vars.append(tk.StringVar())
            self.sample_viscosity_vars.append(tk.StringVar())
            self.sample_voltage_vars.append(tk.StringVar())
            self.sample_puffing_regime_vars.append(tk.StringVar(value="60mL/3s/30s"))
            self.sample_oil_mass_vars.append(tk.StringVar())

        # Create UI for each sample with all required fields
        for i in range(num_samples):
            sample_frame = ttk.LabelFrame(self.samples_frame, text=f"Sample {i+1}", padding=10)
            sample_frame.pack(fill='x', pady=5)

            # Create a grid layout within each sample frame
            # Row 0: Sample ID and Resistance
            ttk.Label(sample_frame, text="Sample ID:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_id_vars[i], width=20).grid(row=0, column=1, padx=5, pady=2)
            
            ttk.Label(sample_frame, text="Resistance (Ω):").grid(row=0, column=2, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.resistance_vars[i], width=20).grid(row=0, column=3, padx=5, pady=2)

            # Row 1: Media and Viscosity
            ttk.Label(sample_frame, text="Media:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_media_vars[i], width=20).grid(row=1, column=1, padx=5, pady=2)
            
            ttk.Label(sample_frame, text="Viscosity (cP):").grid(row=1, column=2, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_viscosity_vars[i], width=20).grid(row=1, column=3, padx=5, pady=2)

            # Row 2: Voltage and Puffing Regime
            ttk.Label(sample_frame, text="Voltage (V):").grid(row=2, column=0, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_voltage_vars[i], width=20).grid(row=2, column=1, padx=5, pady=2)
            
            ttk.Label(sample_frame, text="Puffing Regime:").grid(row=2, column=2, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_puffing_regime_vars[i], width=20).grid(row=2, column=3, padx=5, pady=2)

            # Row 3: Oil Mass
            ttk.Label(sample_frame, text="Initial Oil Mass (g):").grid(row=3, column=0, sticky="w", padx=5, pady=2)
            ttk.Entry(sample_frame, textvariable=self.sample_oil_mass_vars[i], width=20).grid(row=3, column=1, padx=5, pady=2)

        debug_print(f"DEBUG: Created sample-specific fields for {num_samples} samples")
    
    def validate_data(self):
        """Validate the header data."""
        debug_print("DEBUG: Validating header data")
        
        # Check for empty required fields
        if not self.tester_var.get().strip():
            messagebox.showerror("Validation Error", "Tester name is required.")
            return False
            
        # Check sample-specific fields
        num_samples = self.num_samples_var.get()
        for i in range(num_samples):
            if not self.sample_id_vars[i].get().strip():
                messagebox.showerror("Validation Error", f"Sample {i+1} ID is required.")
                return False
                
            # Validate numeric fields for each sample
            try:
                if self.sample_viscosity_vars[i].get().strip():
                    float(self.sample_viscosity_vars[i].get().strip())
                if self.sample_voltage_vars[i].get().strip():
                    float(self.sample_voltage_vars[i].get().strip())
                if self.sample_oil_mass_vars[i].get().strip():
                    float(self.sample_oil_mass_vars[i].get().strip())
                if self.resistance_vars[i].get().strip():
                    float(self.resistance_vars[i].get().strip())
            except ValueError:
                messagebox.showerror("Validation Error", f"Numeric fields for Sample {i+1} must be valid numbers.")
                return False
        
        debug_print("DEBUG: Header data validation successful")
        return True
    
    def collect_header_data(self):
        """Collect and return the header data with proper structure."""
        debug_print("DEBUG: Collecting header data")

        header_data = {
            "test": self.selected_test,
            "common": {
                "tester": self.tester_var.get().strip()
            },
            "num_samples": self.num_samples_var.get(),
            "samples": []
        }

        # Collect individual sample data with all fields
        for i in range(self.num_samples_var.get()):
            sample_data = {
                "id": self.sample_id_vars[i].get().strip(),
                "resistance": self.resistance_vars[i].get().strip(),
                "media": self.sample_media_vars[i].get().strip(),
                "viscosity": self.sample_viscosity_vars[i].get().strip(),
                "voltage": self.sample_voltage_vars[i].get().strip(),
                "puffing_regime": self.sample_puffing_regime_vars[i].get().strip(),
                "oil_mass": self.sample_oil_mass_vars[i].get().strip()
            }
            header_data["samples"].append(sample_data)

        debug_print(f"DEBUG: Collected header data: {header_data}")
        return header_data
    
    def on_continue(self):
        """Handle Continue button click."""
        debug_print("DEBUG: Continue button clicked")
        
        if self.validate_data():
            self.header_data = self.collect_header_data()
            self.result = True
            self.cleanup_mousewheel_binding()
            self.dialog.destroy()
            debug_print("DEBUG: Header data collection completed successfully")
    
    def on_cancel(self):
        """Handle Cancel button click."""
        debug_print("DEBUG: Dialog cancelled - setting result to False")
        self.result = False
        self.dialog.destroy()
    
    def show(self):
        """
        Show the dialog and wait for user input.
        
        Returns:
            tuple: (result, header_data) where result is True if Continue was clicked,
                   and header_data is a dictionary of header data.
        """
        debug_print("DEBUG: Showing HeaderDataDialog")
        self.dialog.wait_window()
        debug_print(f"DEBUG: HeaderDataDialog closed with result: {self.result}")
        if self.result:
            debug_print("DEBUG: Dialog succeeded - continuing to data collection window")
        else:
            debug_print("DEBUG: Dialog was cancelled - not proceeding to data collection")
        return self.result, self.header_data
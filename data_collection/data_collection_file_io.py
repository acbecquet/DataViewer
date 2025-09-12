"""
data_collection_file_io.py
Developed by Charlie Becquet
File IO class for data collection window
"""

# pylint: disable=no-member
# This module is part of a multiple inheritance structure where attributes
# are defined in other parent classes (DataCollectionData, DataCollectionHandlers, etc.)

import os
import copy
import time
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import messagebox
from utils import debug_print, load_excel_file_with_formulas, show_success_message

class DataCollectionFileIO:

    def apply_header_changes_to_file(self):
        """Apply header changes to the Excel file using the new sample-specific structure."""
        print("DEBUG: Applying header changes to Excel file")

        # For .vap3 files, we don't update the physical file directly
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            print("DEBUG: .vap3 file detected, updating header data in memory only")
            # The header data is already updated in self.header_data
            # Just mark that we have changes
            self.mark_unsaved_changes()
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path)

            if self.test_name not in wb.sheetnames:
                print(f"DEBUG: Sheet {self.test_name} not found")
                return

            ws = wb[self.test_name]

            # Get common data and samples data
            common_data = self.header_data.get('common', {})
            samples_data = self.header_data.get('samples', [])

            # Apply sample-specific data for each sample block
            num_samples = self.header_data.get('num_samples', 0)
            for i in range(num_samples):
                col_offset = i * 12
                sample_data = samples_data[i] if i < len(samples_data) else {}

                print(f"DEBUG: Applying header changes for sample {i+1} with offset {col_offset}")

                # Row 1, Column F (6) + offset: Sample ID
                sample_id = sample_data.get('id', f'Sample {i+1}')
                ws.cell(row=1, column=6 + col_offset, value=sample_id)

                # Row 2, Column D (4) + offset: Resistance
                resistance = sample_data.get("resistance", "")
                if resistance:
                    try:
                        resistance_value = float(resistance)
                        ws.cell(row=2, column=4 + col_offset, value=resistance_value)
                    except ValueError:
                        ws.cell(row=2, column=4 + col_offset, value=resistance)

                # Row 3, Column D (4) + offset: Tester name
                tester_name = common_data.get("tester", "")
                if tester_name:
                    ws.cell(row=3, column=4 + col_offset, value=tester_name)

                # Row 2, Column B (2) + offset: Media
                media = sample_data.get("media", "")
                if media:
                    ws.cell(row=2, column=2 + col_offset, value=media)

                # Row 3, Column B (2) + offset: Viscosity
                viscosity = sample_data.get("viscosity", "")
                if viscosity:
                    try:
                        viscosity_value = float(viscosity)
                        ws.cell(row=3, column=2 + col_offset, value=viscosity_value)
                    except ValueError:
                        ws.cell(row=3, column=2 + col_offset, value=viscosity)

                # Row 3, Column F (6) + offset: Voltage
                voltage = sample_data.get("voltage", "")
                if voltage:
                    try:
                        voltage_value = float(voltage)
                        ws.cell(row=3, column=6 + col_offset, value=voltage_value)
                    except ValueError:
                        ws.cell(row=3, column=6 + col_offset, value=voltage)

                # Calculate and apply power to Row 2, Column F (6) + offset
                resistance = sample_data.get("resistance", "")
                device_type = self.header_data.get('common', {}).get('device_type', 'EVO')

                # Device type mapping for dR values
                device_dr_mapping = {
                    'T58G': 0.9,
                    'EVO': 0.15,
                    'EVOMAX': 0.15,
                    'T28': 0.1,
                    'T51': 0.8,
                    'other': 0.15,
                    None: 0.0
                }

                try:
                    if voltage and resistance:
                        voltage_val = float(voltage)
                        resistance_val = float(resistance)
                        dr_value = device_dr_mapping.get(device_type, 0.0)

                        print(f"DEBUG: Voltage: {voltage_val}V, Resistance: {resistance_val}Ω")
                        print(f"DEBUG: Device type lookup: '{device_type}' -> dR = {dr_value}")


                        # Explicit handling for None device type (backwards compatibility)
                        if device_type is None:
                            dr_value = 0.0
                            print(f"DEBUG: Using dR = 0 for backwards compatibility (device_type is None)")

                        calculated_power = (voltage_val ** 2) / (resistance_val + dr_value)
                        ws.cell(row=2, column=6 + col_offset, value=calculated_power)
                        print(f"DEBUG: Calculated and applied power {calculated_power:.3f}W for sample {i+1} (V={voltage_val}, R={resistance_val}, dR={dr_value}, device_type={device_type})")
                    else:
                        print(f"DEBUG: Cannot calculate power for sample {i+1} - missing voltage or resistance")
                except (ValueError, TypeError) as e:
                    print(f"DEBUG: Error calculating power for sample {i+1}: {e}")

                # Row 3, Column H (8) + offset: Oil Mass
                oil_mass = sample_data.get("oil_mass", "")
                if oil_mass:
                    try:
                        oil_mass_value = float(oil_mass)
                        ws.cell(row=3, column=8 + col_offset, value=oil_mass_value)
                    except ValueError:
                        ws.cell(row=3, column=8 + col_offset, value=oil_mass)

                # Row 2, Column H (8) + offset: Puffing Regime
                puffing_regime = sample_data.get("puffing_regime", "60mL/3s/30s")
                if puffing_regime:
                    ws.cell(row=2, column=8 + col_offset, value=puffing_regime)

            # Save the workbook
            wb.save(self.file_path)
            debug_print("DEBUG: Header changes applied successfully to Excel file")

        except Exception as e:
            debug_print(f"DEBUG: Error applying header changes to file: {e}")
            # For .vap3 files, this is expected - just update in memory
            if self.file_path.endswith('.vap3'):
                debug_print("DEBUG: .vap3 file format detected, header changes stored in memory only")
            else:
                # For actual Excel files, show the error
                from tkinter import messagebox
                messagebox.showerror("Error", f"Could not update header data in file: {e}")

    def _save_to_excel(self):
        """Save data to the appropriate file format."""
        debug_print(f"DEBUG: _save_to_excel() starting - file: {self.file_path}")

        # Check if this is a .vap3 file or temporary file
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: Detected .vap3 file or non-existent file, saving to loaded sheets")
            self._save_to_loaded_sheets()
        else:
            debug_print("DEBUG: Detected Excel file, saving using openpyxl")
            self._save_to_excel_file()

        debug_print("DEBUG: Excel save completed, updating main GUI...")

        # For .vap3 files, the data is already updated in memory via _save_to_loaded_sheets
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: .vap3 file detected, skipping file-based update")
            # The _save_to_loaded_sheets method should have already updated the main GUI data
        else:
            debug_print("DEBUG: Excel file detected, updating from file")
            self._update_excel_data_in_main_gui()

    def _save_to_excel_file(self):
        """Save data to the Excel file."""
        debug_print(f"DEBUG: _save_to_excel() starting - file: {self.file_path}")

        # Load the workbook
        wb = openpyxl.load_workbook(self.file_path)
        debug_print(f"DEBUG: Loaded workbook, sheets: {wb.sheetnames}")

        # Get the sheet for this test
        if self.test_name not in wb.sheetnames:
            raise Exception(f"Sheet '{self.test_name}' not found in the file.")

        ws = wb[self.test_name]
        debug_print(f"DEBUG: Opened sheet '{self.test_name}'")

        # Define green fill for TPM cells
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Determine column layout based on test type
        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
            columns_per_sample = 8  # Including chronography
            debug_print(f"DEBUG: Using User Test Simulation format with 8 columns per sample")
        else:
            columns_per_sample = 12  # Standard format
            debug_print(f"DEBUG: Using standard format with 12 columns per sample")

        # Track how much data we're actually writing
        total_data_written = 0

        # For each sample, write the data
        for sample_idx in range(self.num_samples):
            sample_id = f"Sample {sample_idx+1}"

            # Calculate column offset based on test type
            col_offset = sample_idx * columns_per_sample

            debug_print(f"DEBUG: Writing data for {sample_id} at column offset {col_offset}")

            sample_data_written = 0

            # Write the data starting at row 5
            for i, puff in enumerate(self.data[sample_id]["puffs"]):
                row = i + 5  # Row 5 is the first data row

                # Only write if we have actual data (not just empty rows)
                has_data = (
                    self.data[sample_id]["before_weight"][i] or
                    self.data[sample_id]["after_weight"][i] or
                    self.data[sample_id]["draw_pressure"][i] or
                    self.data[sample_id]["smell"][i] or
                    self.data[sample_id]["notes"][i] or
                    self.data[sample_id]["tpm"][i] is not None
                )

                # For User Test Simulation, also check chronography
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    has_data = has_data or (i < len(self.data[sample_id]["chronography"]) and self.data[sample_id]["chronography"][i])

                if not has_data:
                    continue  # Skip empty rows

                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    # User Test Simulation column layout (8 columns)
                    # Chronography column (A + offset)
                    if i < len(self.data[sample_id]["chronography"]) and self.data[sample_id]["chronography"][i]:
                        ws.cell(row=row, column=1 + col_offset, value=str(self.data[sample_id]["chronography"][i]))

                    # Puffs column (B + offset)
                    ws.cell(row=row, column=2 + col_offset, value=puff)

                    # Before weight column (C + offset)
                    if self.data[sample_id]["before_weight"][i]:
                        try:
                            ws.cell(row=row, column=3 + col_offset, value=float(self.data[sample_id]["before_weight"][i]))
                        except:
                            ws.cell(row=row, column=3 + col_offset, value=self.data[sample_id]["before_weight"][i])

                    # After weight column (D + offset)
                    if self.data[sample_id]["after_weight"][i]:
                        try:
                            ws.cell(row=row, column=4 + col_offset, value=float(self.data[sample_id]["after_weight"][i]))
                        except:
                            ws.cell(row=row, column=4 + col_offset, value=self.data[sample_id]["after_weight"][i])

                    # Draw pressure column (E + offset)
                    if self.data[sample_id]["draw_pressure"][i]:
                        try:
                            ws.cell(row=row, column=5 + col_offset, value=float(self.data[sample_id]["draw_pressure"][i]))
                        except:
                            ws.cell(row=row, column=5 + col_offset, value=self.data[sample_id]["draw_pressure"][i])

                    # Skip resistance column (F + offset) - not used in User Test Simulation

                    # Failure column (G + offset)
                    if self.data[sample_id]["smell"][i]:
                        try:
                            ws.cell(row=row, column=6 + col_offset, value=float(self.data[sample_id]["smell"][i]))
                        except:
                            ws.cell(row=row, column=6 + col_offset, value=self.data[sample_id]["smell"][i])

                    # Notes column (H + offset)
                    if self.data[sample_id]["notes"][i]:
                        ws.cell(row=row, column=7 + col_offset, value=str(self.data[sample_id]["notes"][i]))

                    debug_print(f"DEBUG: Saved User Test Simulation row {i} for {sample_id}")

                else:
                    # Standard column layout (12 columns)
                    # Puffs column (A + offset)
                    ws.cell(row=row, column=1 + col_offset, value=puff)

                    # Before weight column (B + offset)
                    if self.data[sample_id]["before_weight"][i]:
                        try:
                            ws.cell(row=row, column=2 + col_offset, value=float(self.data[sample_id]["before_weight"][i]))
                        except:
                            ws.cell(row=row, column=2 + col_offset, value=self.data[sample_id]["before_weight"][i])

                    # After weight column (C + offset)
                    if self.data[sample_id]["after_weight"][i]:
                        try:
                            ws.cell(row=row, column=3 + col_offset, value=float(self.data[sample_id]["after_weight"][i]))
                        except:
                            ws.cell(row=row, column=3 + col_offset, value=self.data[sample_id]["after_weight"][i])

                    # Draw pressure column (D + offset)
                    if self.data[sample_id]["draw_pressure"][i]:
                        try:
                            ws.cell(row=row, column=4 + col_offset, value=float(self.data[sample_id]["draw_pressure"][i]))
                        except:
                            ws.cell(row=row, column=4 + col_offset, value=self.data[sample_id]["draw_pressure"][i])

                    if self.data[sample_id]["resistance"][i]:
                        try:
                            ws.cell(row=row, column=5 + col_offset, value=float(self.data[sample_id]["resistance"][i]))
                        except:
                            ws.cell(row=row, column=5 + col_offset, value=self.data[sample_id]["resistance"][i])

                    # Smell column (F + offset)
                    if self.data[sample_id]["smell"][i]:
                        try:
                            ws.cell(row=row, column=6 + col_offset, value=float(self.data[sample_id]["smell"][i]))
                        except:
                            ws.cell(row=row, column=6 + col_offset, value=self.data[sample_id]["smell"][i])

                    if self.data[sample_id]["clog"][i]:
                        try:
                            ws.cell(row=row, column=7 + col_offset, value=float(self.data[sample_id]["clog"][i]))
                        except:
                            ws.cell(row=row, column=7 + col_offset, value=self.data[sample_id]["clog"][i])

                    # Notes column (H + offset)
                    if self.data[sample_id]["notes"][i]:
                        ws.cell(row=row, column=8 + col_offset, value=str(self.data[sample_id]["notes"][i]))

                    # TPM column (I + offset) - if calculated
                    if i < len(self.data[sample_id]["tpm"]) and self.data[sample_id]["tpm"][i] is not None:
                        tpm_cell = ws.cell(row=row, column=9 + col_offset, value=float(self.data[sample_id]["tpm"][i]))
                        tpm_cell.fill = green_fill

                sample_data_written += 1

            total_data_written += sample_data_written
            debug_print(f"DEBUG: Wrote {sample_data_written} data rows for {sample_id}")

        # Save the workbook
        debug_print(f"DEBUG: Saving workbook with {total_data_written} total data rows written...")

        debug_print("DEBUG: About to save Excel file - verifying data to be saved:")
        for sample_idx in range(min(2, self.num_samples)):  # Check first 2 samples
            sample_id = f"Sample {sample_idx+1}"
            col_offset = sample_idx * columns_per_sample

            debug_print(f"DEBUG: Sample {sample_idx+1} data preview:")
            for i in range(min(3, len(self.data[sample_id]["puffs"]))):  # First 3 rows
                row = i + 5
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    chrono_val = ws.cell(row=row, column=1 + col_offset).value
                    puff_val = ws.cell(row=row, column=2 + col_offset).value
                    before_val = ws.cell(row=row, column=3 + col_offset).value
                    after_val = ws.cell(row=row, column=4 + col_offset).value
                    debug_print(f"DEBUG:   Row {i}: Chrono={chrono_val}, Puff={puff_val}, Before={before_val}, After={after_val}")
                else:
                    puff_val = ws.cell(row=row, column=1 + col_offset).value
                    before_val = ws.cell(row=row, column=2 + col_offset).value
                    after_val = ws.cell(row=row, column=3 + col_offset).value
                    tpm_val = ws.cell(row=row, column=9 + col_offset).value
                    debug_print(f"DEBUG:   Row {i}: Puff={puff_val}, Before={before_val}, After={after_val}, TPM={tpm_val}")

        wb.save(self.file_path)
        debug_print(f"DEBUG: Excel file saved successfully to {self.file_path}")
        self._refresh_main_gui_notes_display

    def _save_to_loaded_sheets(self):
        """Save data directly to the loaded sheets in memory."""
        debug_print("DEBUG: _save_to_loaded_sheets() starting")

        # Use original filename if available, otherwise fall back to file_path
        display_filename = self.original_filename if self.original_filename else os.path.basename(self.file_path)

        debug_print(f"DEBUG: Saving to loaded sheets with display filename: {display_filename}")

        try:
            # Find the correct file data in all_filtered_sheets
            current_file_data = None

            # For .vap3 files, the matching logic needs to be more flexible
            if self.file_path.endswith('.vap3'):
                debug_print("DEBUG: .vap3 file detected, using flexible matching")

                # Try multiple matching strategies for .vap3 files
                for file_data in self.parent.all_filtered_sheets:
                    # First try original filename match
                    if (self.original_filename and
                        (file_data.get("original_filename") == self.original_filename or
                         file_data.get("database_filename") == self.original_filename or
                         file_data.get("file_name") == self.original_filename)):
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching .vap3 file by original filename: {self.original_filename}")
                        break

                    # Then try display filename match
                    if (display_filename and
                        (file_data.get("file_name") == display_filename or
                         file_data.get("display_filename") == display_filename)):
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching .vap3 file by display filename: {display_filename}")
                        break

                # If still no match, just use the first loaded file for .vap3 files
                if not current_file_data and self.parent.all_filtered_sheets:
                    current_file_data = self.parent.all_filtered_sheets[0]
                    debug_print("DEBUG: Using first loaded file as fallback for .vap3 file")
            else:
                # Regular Excel file matching
                for file_data in self.parent.all_filtered_sheets:
                    if file_data.get("file_path") == self.file_path:
                        current_file_data = file_data
                        debug_print(f"DEBUG: Found matching Excel file: {self.file_path}")
                        break

            if not current_file_data:
                debug_print(f"ERROR: Could not find file data for {display_filename}")
                return

            debug_print(f"DEBUG: Saving data to loaded sheets for test: {self.test_name}")

            # Check if the sheet exists in loaded data
            if not hasattr(self.parent, 'filtered_sheets') or self.test_name not in self.parent.filtered_sheets:
                debug_print(f"ERROR: Sheet {self.test_name} not found in loaded data")
                raise Exception(f"Sheet '{self.test_name}' not found in loaded data")

            sheet_data = self.parent.filtered_sheets[self.test_name]['data'].copy()
            debug_print(f"DEBUG: Found loaded sheet data with shape: {sheet_data.shape}")

            # Determine format based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8  # [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                debug_print("DEBUG: Saving in User Test Simulation format")
            else:
                columns_per_sample = 12  # [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM, etc.]
                debug_print("DEBUG: Saving in standard format")

            total_data_written = 0

            # Save data for each sample
            for sample_idx in range(self.num_samples):
                sample_name = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample

                debug_print(f"DEBUG: Saving data for {sample_name} with column offset {col_offset}")

                sample_data = self.data[sample_name]
                sample_data_written = 0

                # Get the length of data to write
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    data_length = len(sample_data.get("chronography", []))
                else:
                    data_length = len(sample_data.get("puffs", []))

                debug_print(f"DEBUG: {sample_name} has {data_length} rows of data to write")

                # Write data starting from row 3 (index 3)
                for i in range(data_length):
                    data_row_idx = 3 + i  # Changed from 4 + i to 3 + i

                    try:
                        if data_row_idx >= len(sheet_data):
                            debug_print(f"DEBUG: Extending sheet data to accommodate row {data_row_idx}")
                            # Extend the DataFrame with empty rows
                            new_rows = data_row_idx - len(sheet_data) + 1
                            empty_df = pd.DataFrame([[""] * len(sheet_data.columns)] * new_rows,
                                                  columns=sheet_data.columns)
                            sheet_data = pd.concat([sheet_data, empty_df], ignore_index=True)

                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User simulation format
                            values_to_write = [
                                sample_data["chronography"][i] if i < len(sample_data["chronography"]) else "",
                                sample_data["puffs"][i] if i < len(sample_data["puffs"]) else "",
                                sample_data["before_weight"][i] if i < len(sample_data["before_weight"]) else "",
                                sample_data["after_weight"][i] if i < len(sample_data["after_weight"]) else "",
                                sample_data["draw_pressure"][i] if i < len(sample_data["draw_pressure"]) else "",
                                sample_data["smell"][i] if i < len(sample_data["smell"]) else "",  # failure
                                sample_data["notes"][i] if i < len(sample_data["notes"]) else "",
                                sample_data["tpm"][i] if i < len(sample_data["tpm"]) and sample_data["tpm"][i] is not None else ""
                            ]
                        else:
                            # Standard format
                            values_to_write = [
                                sample_data["puffs"][i] if i < len(sample_data["puffs"]) else "",
                                sample_data["before_weight"][i] if i < len(sample_data["before_weight"]) else "",
                                sample_data["after_weight"][i] if i < len(sample_data["after_weight"]) else "",
                                sample_data["draw_pressure"][i] if i < len(sample_data["draw_pressure"]) else "",
                                sample_data["resistance"][i] if i < len(sample_data["resistance"]) else "",
                                sample_data["smell"][i] if i < len(sample_data["smell"]) else "",
                                sample_data["clog"][i] if i < len(sample_data["clog"]) else "",
                                sample_data["notes"][i] if i < len(sample_data["notes"]) else "",
                                sample_data["tpm"][i] if i < len(sample_data["tpm"]) and sample_data["tpm"][i] is not None else ""
                            ]

                        # Write values to the appropriate columns
                        for col_idx, value in enumerate(values_to_write):
                            target_col = col_offset + col_idx
                            if target_col < len(sheet_data.columns):
                                sheet_data.iloc[data_row_idx, target_col] = value

                        sample_data_written += 1

                    except Exception as e:
                        debug_print(f"DEBUG: Error writing row {i} for {sample_name}: {e}")
                        continue

                total_data_written += sample_data_written
                debug_print(f"DEBUG: Wrote {sample_data_written} data rows for {sample_name}")

            # Update the loaded sheet data in memory
            self.parent.filtered_sheets[self.test_name]['data'] = sheet_data
            debug_print(f"DEBUG: Updated loaded sheet data with {total_data_written} total data rows")

            self._refresh_main_gui_notes_display()

            debug_print("DEBUG: _save_to_loaded_sheets completed successfully")

            # FIX: Save sample notes to header data
            debug_print("DEBUG: Saving sample notes to header data")
            if not hasattr(self, 'header_data') or not self.header_data:
                self.header_data = {'samples': []}

            # Ensure header_data has enough sample entries
            while len(self.header_data['samples']) < self.num_samples:
                self.header_data['samples'].append({})

            # Save sample notes to header data
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                if sample_id in self.data and 'sample_notes' in self.data[sample_id]:
                    self.header_data['samples'][i]['sample_notes'] = self.data[sample_id]['sample_notes']
                    debug_print(f"DEBUG: Saved sample notes for {sample_id}: {self.data[sample_id]['sample_notes'][:50]}...")

            # Also update the UI state in all_filtered_sheets
            if hasattr(self.parent, 'all_filtered_sheets'):
                for file_data in self.parent.all_filtered_sheets:
                    if self.test_name in file_data.get('filtered_sheets', {}):
                        file_data['filtered_sheets'][self.test_name]['data'] = sheet_data.copy()
                        # CRITICAL: Save header data to the filtered_sheets structure
                        file_data['filtered_sheets'][self.test_name]['header_data'] = self.header_data.copy()
                        debug_print("DEBUG: Updated all_filtered_sheets with new data and header data")
                        break

            # CRITICAL: Ensure header data is stored in the main filtered_sheets
            if self.test_name in self.parent.filtered_sheets:
                self.parent.filtered_sheets[self.test_name]['header_data'] = self.header_data.copy()
                debug_print("DEBUG: Stored header data in filtered_sheets for .vap3 persistence")

            debug_print("DEBUG: Successfully saved data to loaded sheets")

            debug_print("DEBUG: Loaded sheets save completed - main GUI data should now be current")

            # Ensure the main GUI's current filtered_sheets reflects the updated data
            if hasattr(self.parent, 'filtered_sheets') and self.test_name in self.parent.filtered_sheets:
                # The data should already be updated, but let's ensure it's current
                debug_print(f"DEBUG: Main GUI filtered_sheets for {self.test_name} is current")

        except Exception as e:
            debug_print(f"ERROR: Failed to save data to loaded sheets: {e}")
            import traceback
            traceback.print_exc()
            raise e

    def load_existing_data_from_loaded_sheets(self):
        """Load existing data from already-loaded sheet data (for .vap3 files)."""
        debug_print(f"DEBUG: Loading existing data from loaded sheets for test: {self.test_name}")

        try:
            # Get the loaded sheet data
            if not hasattr(self.parent, 'filtered_sheets') or self.test_name not in self.parent.filtered_sheets:
                debug_print(f"ERROR: Sheet {self.test_name} not found in loaded data")
                return

            sheet_data = self.parent.filtered_sheets[self.test_name]['data']
            debug_print(f"DEBUG: Found loaded sheet data with shape: {sheet_data.shape}")

            debug_print("DEBUG: First 10 rows and 16 columns of actual DataFrame data:")
            for i in range(min(10, len(sheet_data))):
                row_preview = []
                for j in range(min(16, len(sheet_data.columns))):
                    val = sheet_data.iloc[i, j]
                    if pd.isna(val):
                        row_preview.append("NaN")
                    else:
                        val_str = str(val)[:10]  # Truncate long values
                        row_preview.append(f"'{val_str}'")
                debug_print(f"  Row {i}: {row_preview}")

            # Determine the data format based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8
                debug_print("DEBUG: Loading User Test Simulation format")
            else:
                columns_per_sample = 12
                debug_print("DEBUG: Loading standard format")

            # Load data for each sample
            loaded_data_count = 0

            for sample_idx in range(self.num_samples):
                sample_id = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample

                debug_print(f"DEBUG: Loading data for {sample_id} with column offset {col_offset}")

                debug_print(f"DEBUG: Clearing existing template data for {sample_id}")
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    self.data[sample_id]["chronography"] = []
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["smell"] = []
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []
                else:
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["resistance"] = []
                    self.data[sample_id]["smell"] = []
                    self.data[sample_id]["clog"] = []
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []

                # Load data starting from row 4 (DataFrame index 3)
                sample_had_data = False
                row_count = 0

                data_start_row = 3  # Data starts at row 4 (0-indexed row 3)

                for data_row_idx in range(data_start_row, min(len(sheet_data), 100)):  # Limit to reasonable number of rows
                    try:
                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User Test Simulation format: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                            chronography_col = 0 + col_offset
                            puffs_col = 1 + col_offset
                            before_weight_col = 2 + col_offset
                            after_weight_col = 3 + col_offset
                            draw_pressure_col = 4 + col_offset
                            failure_col = 5 + col_offset  # Stored in smell field
                            notes_col = 6 + col_offset
                            tpm_col = 7 + col_offset

                            # Extract values if columns exist
                            values = {}

                            # Helper function to clean values and strip quotes
                            def clean_value(raw_val):
                                if pd.isna(raw_val):
                                    return ""
                                val_str = str(raw_val).strip()
                                # Remove surrounding quotes if present
                                if val_str.startswith("'") and val_str.endswith("'"):
                                    val_str = val_str[1:-1]
                                elif val_str.startswith('"') and val_str.endswith('"'):
                                    val_str = val_str[1:-1]
                                return val_str

                            debug_print(f"DEBUG: Processing row {data_row_idx} for {sample_id}")

                            if chronography_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, chronography_col]
                                values['chronography'] = clean_value(raw_val)
                                debug_print(f"DEBUG: chronography raw: {raw_val}, cleaned: {values['chronography']}")
                            else:
                                values['chronography'] = ""

                            if puffs_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, puffs_col]
                                values['puffs'] = clean_value(raw_val)
                                debug_print(f"DEBUG: puffs raw: {raw_val}, cleaned: {values['puffs']}")
                            else:
                                values['puffs'] = ""

                            if before_weight_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, before_weight_col]
                                values['before_weight'] = clean_value(raw_val)
                                debug_print(f"DEBUG: before_weight raw: {raw_val}, cleaned: {values['before_weight']}")
                            else:
                                values['before_weight'] = ""

                            if after_weight_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, after_weight_col]
                                values['after_weight'] = clean_value(raw_val)
                                debug_print(f"DEBUG: after_weight raw: {raw_val}, cleaned: {values['after_weight']}")
                            else:
                                values['after_weight'] = ""

                            if draw_pressure_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, draw_pressure_col]
                                values['draw_pressure'] = clean_value(raw_val)
                            else:
                                values['draw_pressure'] = ""

                            if failure_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, failure_col]
                                values['failure'] = clean_value(raw_val)
                            else:
                                values['failure'] = ""

                            if notes_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, notes_col]
                                values['notes'] = clean_value(raw_val)
                            else:
                                values['notes'] = ""

                            if tpm_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, tpm_col]
                                cleaned_val = clean_value(raw_val)
                                if cleaned_val and cleaned_val.strip():
                                    try:
                                        values['tpm'] = float(cleaned_val)
                                        debug_print(f"DEBUG: tpm raw: {raw_val}, cleaned: {cleaned_val}, float: {values['tpm']}")
                                    except (ValueError, TypeError):
                                        values['tpm'] = None
                                        debug_print(f"DEBUG: tpm conversion failed for: {cleaned_val}")
                                else:
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data (not all empty)
                            if any([values['chronography'], values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Storing data for row {data_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                                # Append the data (building from scratch)
                                self.data[sample_id]["chronography"].append(values['chronography'])
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["smell"].append(values['failure'])
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Row {data_row_idx} has no meaningful data, skipping")

                        else:
                            # Standard format: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
                            puffs_col = 0 + col_offset
                            before_weight_col = 1 + col_offset
                            after_weight_col = 2 + col_offset
                            draw_pressure_col = 3 + col_offset
                            resistance_col = 4 + col_offset
                            smell_col = 5 + col_offset
                            clog_col = 6 + col_offset
                            notes_col = 7 + col_offset
                            tpm_col = 8 + col_offset  # TPM is typically in column 9 for standard format

                            # Extract values if columns exist
                            values = {}

                            # Helper function to clean values and strip quotes
                            def clean_value(raw_val):
                                if pd.isna(raw_val):
                                    return ""
                                val_str = str(raw_val).strip()
                                # Remove surrounding quotes if present
                                if val_str.startswith("'") and val_str.endswith("'"):
                                    val_str = val_str[1:-1]
                                elif val_str.startswith('"') and val_str.endswith('"'):
                                    val_str = val_str[1:-1]
                                return val_str

                            debug_print(f"DEBUG: Processing row {data_row_idx} for {sample_id}")

                            for field, col_idx in [
                                ('puffs', puffs_col), ('before_weight', before_weight_col),
                                ('after_weight', after_weight_col), ('draw_pressure', draw_pressure_col),
                                ('resistance', resistance_col), ('smell', smell_col),
                                ('clog', clog_col), ('notes', notes_col)
                            ]:
                                if col_idx < len(sheet_data.columns):
                                    raw_val = sheet_data.iloc[data_row_idx, col_idx]
                                    values[field] = clean_value(raw_val)
                                    debug_print(f"DEBUG: {field} raw: {raw_val}, cleaned: {values[field]}")
                                else:
                                    values[field] = ""

                            # Handle TPM separately
                            if tpm_col < len(sheet_data.columns):
                                raw_val = sheet_data.iloc[data_row_idx, tpm_col]
                                cleaned_val = clean_value(raw_val)
                                if cleaned_val and cleaned_val.strip():
                                    try:
                                        values['tpm'] = float(cleaned_val)
                                        debug_print(f"DEBUG: tpm raw: {raw_val}, cleaned: {cleaned_val}, float: {values['tpm']}")
                                    except (ValueError, TypeError):
                                        values['tpm'] = None
                                        debug_print(f"DEBUG: tpm conversion failed for: {cleaned_val}")
                                else:
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data (not all empty)
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Storing data for row {data_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")
                                # Append the data (building from scratch)
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["resistance"].append(values['resistance'])
                                self.data[sample_id]["smell"].append(values['smell'])
                                self.data[sample_id]["clog"].append(values['clog'])
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Row {data_row_idx} has no meaningful data, skipping")

                    except Exception as e:
                        debug_print(f"DEBUG: Error processing row {data_row_idx} for {sample_id}: {e}")
                        continue

                if sample_had_data:
                    loaded_data_count += 1
                    debug_print(f"DEBUG: {sample_id} - Loaded {row_count} rows of existing data (cleared template)")
                else:
                    debug_print(f"DEBUG: {sample_id} - No existing data found")
                    # If no data was found, ensure we still have empty arrays (not template data)
                    if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                        self.data[sample_id]["chronography"] = []
                        self.data[sample_id]["puffs"] = []
                        self.data[sample_id]["before_weight"] = []
                        self.data[sample_id]["after_weight"] = []
                        self.data[sample_id]["draw_pressure"] = []
                        self.data[sample_id]["smell"] = []
                        self.data[sample_id]["notes"] = []
                        self.data[sample_id]["tpm"] = []
                    else:
                        self.data[sample_id]["puffs"] = []
                        self.data[sample_id]["before_weight"] = []
                        self.data[sample_id]["after_weight"] = []
                        self.data[sample_id]["draw_pressure"] = []
                        self.data[sample_id]["resistance"] = []
                        self.data[sample_id]["smell"] = []
                        self.data[sample_id]["clog"] = []
                        self.data[sample_id]["notes"] = []
                        self.data[sample_id]["tpm"] = []

            debug_print(f"DEBUG: Successfully loaded existing data for {loaded_data_count} samples from loaded sheets")

            # Recalculate TPM values for all samples
            for i in range(self.num_samples):
                sample_id = f"Sample {i + 1}"
                self.calculate_tpm(sample_id)
                debug_print(f"DEBUG: Calculated TPM for {sample_id}")

            # update tksheets to show loaded data
            if hasattr(self, 'sample_sheets') and self.sample_sheets:
                for i, sample_sheet in enumerate(self.sample_sheets):
                    if i < self.num_samples:
                        sample_id = f"Sample {i + 1}"
                        self.update_tksheet(sample_sheet, sample_id)
                        debug_print(f"DEBUG: Updated tksheet for {sample_id}")
            else:
                debug_print("DEBUG: sample_sheets not available yet, data loading completed")

            # Update the stats panel
            self.update_stats_panel()

            debug_print("DEBUG: Existing data loading from loaded sheets completed successfully")

        except Exception as e:
            debug_print(f"ERROR: Failed to load existing data from loaded sheets: {e}")
            import traceback
            traceback.print_exc()

    def load_existing_data_from_file(self):
        """Load existing data from file or loaded sheets depending on file type."""
        debug_print(f"DEBUG: Loading existing data for file: {self.file_path}")

        # Check if this is a .vap3 file or if the file doesn't exist (temporary file)
        if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
            debug_print("DEBUG: Detected .vap3 file or non-existent file, loading from loaded sheets")
            self.load_existing_data_from_loaded_sheets()
        else:
            debug_print("DEBUG: Detected Excel file, loading from file using openpyxl")
            self.load_existing_data_from_excel_file()

    def load_existing_data_from_excel_file(self):
        """Load existing data from the Excel file into the data collection interface."""
        debug_print(f"DEBUG: Loading existing data from file: {self.file_path}")

        try:
            # Load the workbook and calculate formulas
            wb = openpyxl.load_workbook(self.file_path, data_only=True)  # data_only=True evaluates formulas

            # Check if the test sheet exists
            if self.test_name not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet '{self.test_name}' not found in file")
                return

            ws = wb[self.test_name]
            debug_print(f"DEBUG: Successfully opened sheet '{self.test_name}'")

            # Determine column layout based on test type
            if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8  # Including chronography
                debug_print(f"DEBUG: Loading User Test Simulation format with 8 columns per sample")
            else:
                columns_per_sample = 12  # Standard format
                debug_print(f"DEBUG: Loading standard format with 12 columns per sample")

            # Load data for each sample
            loaded_data_count = 0
            for sample_idx in range(self.num_samples):
                sample_id = f"Sample {sample_idx + 1}"
                col_offset = sample_idx * columns_per_sample

                debug_print(f"DEBUG: Loading data for {sample_id} with column offset {col_offset}")

                # CLEAR ALL EXISTING TEMPLATE DATA for this sample (like in load_existing_data_from_loaded_sheets)
                debug_print(f"DEBUG: Clearing existing template data for {sample_id}")
                if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                    self.data[sample_id]["chronography"] = []
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["smell"] = []  # Used for failure in user simulation
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []
                else:
                    self.data[sample_id]["puffs"] = []
                    self.data[sample_id]["before_weight"] = []
                    self.data[sample_id]["after_weight"] = []
                    self.data[sample_id]["draw_pressure"] = []
                    self.data[sample_id]["resistance"] = []
                    self.data[sample_id]["smell"] = []
                    self.data[sample_id]["clog"] = []
                    self.data[sample_id]["notes"] = []
                    self.data[sample_id]["tpm"] = []

                # Read data starting from row 5 (Excel row 5 = index 4)
                row_count = 0
                sample_had_data = False

                for excel_row_idx in range(5, min(ws.max_row + 1, 100)):  # Start from row 5, limit to 100 rows max
                    try:
                        if self.test_name in ["User Test Simulation", "User Simulation Test"]:
                            # User Test Simulation format: [Chronography, Puffs, Before Weight, After Weight, Draw Pressure, Failure, Notes, TPM]
                            chronography_col = 1 + col_offset  # Column A + offset
                            puffs_col = 2 + col_offset          # Column B + offset
                            before_weight_col = 3 + col_offset  # Column C + offset
                            after_weight_col = 4 + col_offset   # Column D + offset
                            draw_pressure_col = 5 + col_offset  # Column E + offset
                            failure_col = 6 + col_offset        # Column F + offset (stored in smell field)
                            notes_col = 7 + col_offset          # Column G + offset
                            tpm_col = 8 + col_offset            # Column H + offset

                            # Extract values and clean them
                            def clean_excel_value(cell_value):
                                if cell_value is None:
                                    return ""
                                return str(cell_value).strip().strip('"').strip("'")

                            values = {}
                            values['chronography'] = clean_excel_value(ws.cell(row=excel_row_idx, column=chronography_col).value)
                            values['puffs'] = clean_excel_value(ws.cell(row=excel_row_idx, column=puffs_col).value)
                            values['before_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=before_weight_col).value)
                            values['after_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=after_weight_col).value)
                            values['draw_pressure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=draw_pressure_col).value)
                            values['failure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=failure_col).value)
                            values['notes'] = clean_excel_value(ws.cell(row=excel_row_idx, column=notes_col).value)

                            # Handle TPM calculation
                            tpm_cell_value = ws.cell(row=excel_row_idx, column=tpm_col).value
                            if tpm_cell_value is not None:
                                try:
                                    values['tpm'] = float(tpm_cell_value)
                                except (ValueError, TypeError):
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Appending User Test Simulation data for Excel row {excel_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")


                                self.data[sample_id]["chronography"].append(values['chronography'])
                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["smell"].append(values['failure'])  # failure stored in smell field
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Excel row {excel_row_idx} has no meaningful data, skipping")

                        else:
                            # Standard format: [Puffs, Before Weight, After Weight, Draw Pressure, Resistance, Smell, Clog, Notes, TPM]
                            puffs_col = 1 + col_offset          # Column A + offset
                            before_weight_col = 2 + col_offset  # Column B + offset
                            after_weight_col = 3 + col_offset   # Column C + offset
                            draw_pressure_col = 4 + col_offset  # Column D + offset
                            resistance_col = 5 + col_offset     # Column E + offset
                            smell_col = 6 + col_offset          # Column F + offset
                            clog_col = 7 + col_offset           # Column G + offset
                            notes_col = 8 + col_offset          # Column H + offset
                            tpm_col = 9 + col_offset            # Column I + offset

                            # Extract values and clean them
                            def clean_excel_value(cell_value):
                                if cell_value is None:
                                    return ""
                                return str(cell_value).strip().strip('"').strip("'")

                            values = {}
                            values['puffs'] = clean_excel_value(ws.cell(row=excel_row_idx, column=puffs_col).value)
                            values['before_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=before_weight_col).value)
                            values['after_weight'] = clean_excel_value(ws.cell(row=excel_row_idx, column=after_weight_col).value)
                            values['draw_pressure'] = clean_excel_value(ws.cell(row=excel_row_idx, column=draw_pressure_col).value)
                            values['resistance'] = clean_excel_value(ws.cell(row=excel_row_idx, column=resistance_col).value)
                            values['smell'] = clean_excel_value(ws.cell(row=excel_row_idx, column=smell_col).value)
                            values['clog'] = clean_excel_value(ws.cell(row=excel_row_idx, column=clog_col).value)
                            values['notes'] = clean_excel_value(ws.cell(row=excel_row_idx, column=notes_col).value)

                            # Handle TPM calculation
                            tpm_cell_value = ws.cell(row=excel_row_idx, column=tpm_col).value
                            if tpm_cell_value is not None:
                                try:
                                    values['tpm'] = float(tpm_cell_value)
                                except (ValueError, TypeError):
                                    values['tpm'] = None
                            else:
                                values['tpm'] = None

                            # Check if this row has any meaningful data
                            if any([values['puffs'], values['before_weight'], values['after_weight']]):
                                debug_print(f"DEBUG: Appending standard data for Excel row {excel_row_idx}: puffs={values['puffs']}, before={values['before_weight']}, after={values['after_weight']}")


                                self.data[sample_id]["puffs"].append(values['puffs'])
                                self.data[sample_id]["before_weight"].append(values['before_weight'])
                                self.data[sample_id]["after_weight"].append(values['after_weight'])
                                self.data[sample_id]["draw_pressure"].append(values['draw_pressure'])
                                self.data[sample_id]["resistance"].append(values['resistance'])
                                self.data[sample_id]["smell"].append(values['smell'])
                                self.data[sample_id]["clog"].append(values['clog'])
                                self.data[sample_id]["notes"].append(values['notes'])
                                self.data[sample_id]["tpm"].append(values['tpm'])

                                sample_had_data = True
                                row_count += 1
                            else:
                                debug_print(f"DEBUG: Excel row {excel_row_idx} has no meaningful data, skipping")

                    except Exception as e:
                        debug_print(f"DEBUG: Error processing Excel row {excel_row_idx} for {sample_id}: {e}")
                        continue

                if sample_had_data:
                    loaded_data_count += 1
                    debug_print(f"DEBUG: {sample_id} - Loaded {row_count} rows of existing data from Excel")
                else:
                    debug_print(f"DEBUG: {sample_id} - No existing data found in Excel")

            wb.close()

            if loaded_data_count > 0:
                debug_print(f"DEBUG: Successfully loaded existing data from Excel file for {loaded_data_count} samples")

                # Recalculate TPM values for all samples FIRST
                for i in range(self.num_samples):
                    sample_id = f"Sample {i + 1}"
                    self.calculate_tpm(sample_id)
                    debug_print(f"DEBUG: Calculated TPM for {sample_id}")

                # Update all treeviews to show the loaded data WITH calculated TPM
                if hasattr(self, 'sample_sheets') and self.sample_sheets:
                    for i, sample_sheet in enumerate(self.sample_sheets):
                        if i < self.num_samples:
                            sample_id = f"Sample {i + 1}"
                            self.update_tksheet(sample_sheet, sample_id)
                            debug_print(f"DEBUG: Updated tksheet for {sample_id} with TPM values")
                else:
                    debug_print("DEBUG: sample_sheets not available yet, TPM calculation completed")

                # Update the stats panel if available
                if hasattr(self, 'update_stats_panel'):
                    self.update_stats_panel()
            else:
                debug_print("DEBUG: No existing data found in Excel file")

        except Exception as e:
            debug_print(f"ERROR: Failed to load existing data from Excel file: {e}")
            import traceback
            traceback.print_exc()

    def refresh_main_gui_after_save(self):
        """Update the main GUI data structures after saving data."""
        try:
            debug_print("DEBUG: Refreshing main GUI after data collection save")

            # For .vap3 files or database files, the data is already updated in memory
            # We just need to ensure the main GUI reflects the current state
            if self.file_path.endswith('.vap3') or not os.path.exists(self.file_path):
                self._update_vap3_data_in_main_gui()
            else:
                # For regular Excel files, reload from file
                self._update_excel_data_in_main_gui()

            # Mark this file as modified in the staging area
            self._mark_file_as_modified()

            debug_print("DEBUG: Main GUI refresh completed")

        except Exception as e:
            debug_print(f"ERROR: Failed to refresh main GUI: {e}")
            import traceback
            traceback.print_exc()

    def _update_excel_data_in_main_gui(self):
        """Update main GUI data for Excel files."""
        try:
            # Only do this for actual Excel files that exist
            if not os.path.exists(self.file_path) or not self.file_path.endswith(('.xlsx', '.xls')):
                debug_print("DEBUG: File doesn't exist or isn't Excel, skipping Excel update")
                return

            from utils import load_excel_file

            debug_print(f"DEBUG: Reloading Excel file with formula evaluation: {self.file_path}")
            excel_data = load_excel_file(self.file_path)

            if self.test_name not in excel_data:
                debug_print(f"DEBUG: Sheet {self.test_name} not found in reloaded data")
                return

            new_sheet_data = excel_data[self.test_name]
            debug_print(f"DEBUG: Reloaded sheet {self.test_name} with shape: {new_sheet_data.shape}")

            # Update the main GUI's filtered_sheets
            if hasattr(self.parent, 'filtered_sheets') and self.test_name in self.parent.filtered_sheets:
                self.parent.filtered_sheets[self.test_name]['data'] = new_sheet_data
                self.parent.filtered_sheets[self.test_name]['is_empty'] = new_sheet_data.empty
                debug_print(f"DEBUG: Updated sheet {self.test_name} in main GUI filtered_sheets")

            # Update all_filtered_sheets for the current file
            if hasattr(self.parent, 'all_filtered_sheets') and hasattr(self.parent, 'current_file'):
                for file_data in self.parent.all_filtered_sheets:
                    if file_data["file_name"] == self.parent.current_file:
                        if self.test_name in file_data["filtered_sheets"]:
                            file_data["filtered_sheets"][self.test_name]['data'] = new_sheet_data
                            file_data["filtered_sheets"][self.test_name]['is_empty'] = new_sheet_data.empty
                            debug_print(f"DEBUG: Updated sheet {self.test_name} in all_filtered_sheets for file {self.parent.current_file}")
                        break

        except Exception as e:
            debug_print(f"ERROR: Failed to update Excel data in main GUI: {e}")
            raise

    def _update_vap3_data_in_main_gui(self):
        """Update main GUI data for .vap3 files - data is already in memory."""
        try:
            debug_print("DEBUG: Updating .vap3 data in main GUI - data should already be current")

            # For .vap3 files, the data should have already been updated in the save process
            # The _save_to_loaded_sheets method should have updated the main GUI data structures

            # Just ensure the all_filtered_sheets stays synchronized
            if hasattr(self.parent, 'filtered_sheets') and hasattr(self.parent, 'all_filtered_sheets'):
                current_file = getattr(self.parent, 'current_file', None)
                if current_file:
                    for file_data in self.parent.all_filtered_sheets:
                        if file_data["file_name"] == current_file:
                            # Ensure the all_filtered_sheets data matches current filtered_sheets
                            file_data["filtered_sheets"] = copy.deepcopy(self.parent.filtered_sheets)
                            debug_print(f"DEBUG: Synchronized .vap3 data for file {current_file}")
                            break

            debug_print("DEBUG: .vap3 data update completed")

        except Exception as e:
            debug_print(f"ERROR: Failed to update .vap3 data in main GUI: {e}")
            raise

    def _mark_file_as_modified(self):
        """Mark the current file as modified in the staging area."""
        try:
            if hasattr(self.parent, 'all_filtered_sheets') and hasattr(self.parent, 'current_file'):
                for file_data in self.parent.all_filtered_sheets:
                    if file_data["file_name"] == self.parent.current_file:
                        file_data["is_modified"] = True
                        file_data["last_modified"] = time.time()
                        debug_print(f"DEBUG: Marked file {self.parent.current_file} as modified")
                        break

            # Update the main GUI window title to show modified status
            if hasattr(self.parent, 'root') and hasattr(self.parent, 'current_file'):
                current_title = self.parent.root.title()
                if not current_title.endswith(" *"):
                    self.parent.root.title(current_title + " *")
                    debug_print("DEBUG: Updated window title to show modified status")

        except Exception as e:
            debug_print(f"ERROR: Failed to mark file as modified: {e}")

    def _update_application_state(self):
        """Update the main application's state if this is a VAP3 file."""
        # Check if the parent has methods to update state
        if hasattr(self.parent, 'filtered_sheets') and hasattr(self.parent, 'file_path'):
            if self.parent.file_path and self.parent.file_path.endswith('.vap3'):
                debug_print("DEBUG: Updating VAP3 file and application state")
                try:
                    # Import here to avoid circular imports
                    from vap_file_manager import VapFileManager

                    vap_manager = VapFileManager()

                    # Get current application state
                    image_crop_states = getattr(self.parent, 'image_crop_states', {})
                    plot_settings = {
                        'selected_plot_type': getattr(self.parent, 'selected_plot_type', tk.StringVar()).get()
                    }

                    # Save to VAP3 file
                    success = vap_manager.save_to_vap3(
                        self.parent.file_path,
                        self.parent.filtered_sheets,
                        getattr(self.parent, 'sheet_images', {}),
                        getattr(self.parent, 'plot_options', []),
                        image_crop_states,
                        plot_settings
                    )

                    if success:
                        debug_print("DEBUG: VAP3 file updated successfully")
                    else:
                        debug_print("DEBUG: Failed to update VAP3 file")

                except Exception as e:
                    debug_print(f"DEBUG: Error updating VAP3 file: {e}")

    def _refresh_main_gui_notes_display(self):
        """Refresh the main GUI notes display with updated sample notes."""
        try:
            debug_print("DEBUG: Refreshing main GUI notes display")

            # Ensure the main GUI has the updated header data with sample notes
            if hasattr(self.parent, 'filtered_sheets') and self.test_name in self.parent.filtered_sheets:
                # Update header data in the main GUI's filtered sheets
                debug_print(f"DEBUG: self.header_data before update: {self.header_data}")

                self.parent.filtered_sheets[self.test_name]['header_data'] = self.header_data
                debug_print(f"DEBUG: Updated header_data in main GUI for {self.test_name}")


                # Trigger notes display update in main GUI
                if hasattr(self.parent, 'update_notes_display'):
                    self.parent.update_notes_display(self.test_name)
                    debug_print(f"DEBUG: Triggered main GUI notes display update for {self.test_name}")

        except Exception as e:
            debug_print(f"ERROR: Failed to refresh main GUI notes display: {e}")

    def export_csv(self):
        """Export current data to CSV files."""

        from tkinter import filedialog

        # Ask for directory to save CSV files
        directory = filedialog.askdirectory(title="Select directory to save CSV files")
        if not directory:
            return

        try:
            for i in range(self.num_samples):
                sample_id = f"Sample {i+1}"
                sample_name = self.header_data["samples"][i]["id"]

                # Create DataFrame for this sample
                df_data = {
                    "Puffs": self.data[sample_id]["puffs"],
                    "Before Weight (g)": self.data[sample_id]["before_weight"],
                    "After Weight (g)": self.data[sample_id]["after_weight"],
                    "Draw Pressure (kPa)": self.data[sample_id]["draw_pressure"],
                    "Smell": self.data[sample_id]["smell"],
                    "Notes": self.data[sample_id]["notes"],
                    "TPM (mg/puff)": self.data[sample_id]["tpm"]
                }

                df = pd.DataFrame(df_data)

                # Remove empty rows
                df = df.dropna(how='all', subset=["Before Weight (g)", "After Weight (g)"])

                csv_filename = f"{self.test_name}_{sample_name}_data.csv"
                csv_path = os.path.join(directory, csv_filename)
                df.to_csv(csv_path, index=False)

            show_success_message("Export Complete", f"Exported {self.num_samples} CSV files to {directory}", self.window)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export CSV files: {e}")

    def save_data(self, exit_after=False, show_confirmation=True, auto_save=False):
        """Unified method for saving data."""
        # End any active editing
        self.end_editing()

        # For auto-save, we'll set a flag to avoid certain behaviors
        if auto_save:
            self._auto_save_in_progress = True

        # Confirm save if requested and not auto-saving
        if show_confirmation and not auto_save:
            if not messagebox.askyesno("Confirm Save", "Save the collected data to the file?"):
                return False

        try:
            # Calculate TPM values for all samples
            for i in range(self.num_samples):
                sample_id = f"Sample {i+1}"
                self.calculate_tpm(sample_id)

            # Save to Excel file
            self._save_to_excel()

            self._save_sample_images()

            # Update application state if needed
            if hasattr(self.parent, 'filtered_sheets'):
                self._update_application_state()

            # Refresh the main GUI if not auto-saving
            if not auto_save:
                self.refresh_main_gui_after_save()

            # Mark as saved
            self.has_unsaved_changes = False
            self.update_save_status(False)

            # Show confirmation if requested (not for auto-save)
            if show_confirmation and not auto_save and not exit_after:
                show_success_message("Save Complete", "Data saved successfully.", self.window)

            # Clean up auto-save flag
            if auto_save:
                self._auto_save_in_progress = False

            # Exit if requested
            if exit_after:
                debug_print("DEBUG: Save and exit requested - calling on_window_close()")
                self.result = "load_file"
                # Call on_window_close() to properly restore main window before destroying
                self.on_window_close()
                return True

            return True

        except Exception as e:
            # Clean up auto-save flag
            if auto_save:
                self._auto_save_in_progress = False

            error_msg = f"Failed to save data: {e}"
            self.log(error_msg, "error")

            if not auto_save:  # Don't show errors for auto-save
                messagebox.showerror("Save Error", error_msg)

            return False

    def _save_sample_images(self):
        """Save sample images to the appropriate file format."""
        try:
            debug_print("DEBUG: _save_sample_images() starting")

            # Check if we have sample images to save
            if not hasattr(self, 'sample_images') or not self.sample_images:
                debug_print("DEBUG: No sample images to save")
                return

            debug_print(f"DEBUG: Saving sample images for {len(self.sample_images)} samples")

            # For VAP3 files or if the parent supports it, store in memory for later saving
            if self.file_path.endswith('.vap3') or hasattr(self.parent, 'filtered_sheets'):
                debug_print("DEBUG: Storing sample images in parent for VAP3 save")

                # Store sample images in parent for later VAP3 save
                self.parent.pending_sample_images = self.sample_images.copy()
                self.parent.pending_sample_image_crop_states = getattr(self, 'sample_image_crop_states', {}).copy()
                self.parent.pending_sample_header_data = self.header_data.copy()

                debug_print(f"DEBUG: Stored {len(self.sample_images)} sample groups in parent")

                # If this is already a VAP3 file, save it now
                if self.file_path.endswith('.vap3'):
                    self._save_vap3_with_sample_images()

            debug_print("DEBUG: Sample images saved successfully")

        except Exception as e:
            debug_print(f"ERROR: Failed to save sample images: {e}")
            import traceback
            traceback.print_exc()
            # Don't fail the entire save process for image save issues

    def _save_vap3_with_sample_images(self):
        """Save the VAP3 file with sample images included."""
        try:
            debug_print("DEBUG: Saving VAP3 file with sample images")

            # Import the enhanced VAP file manager
            from vap_file_manager import VapFileManager
            vap_manager = VapFileManager()

            # Get current application state from parent
            filtered_sheets = getattr(self.parent, 'filtered_sheets', {})
            sheet_images = getattr(self.parent, 'sheet_images', {})
            plot_options = getattr(self.parent, 'plot_options', [])
            image_crop_states = getattr(self.parent, 'image_crop_states', {})

            # Plot settings
            plot_settings = {}
            if hasattr(self.parent, 'selected_plot_type'):
                plot_settings['selected_plot_type'] = self.parent.selected_plot_type.get()

            # Sample images
            sample_images = getattr(self.parent, 'pending_sample_images', {})
            sample_image_crop_states = getattr(self.parent, 'pending_sample_image_crop_states', {})
            sample_header_data = getattr(self.parent, 'pending_sample_header_data', {})

            # Save to VAP3 file with sample images
            success = vap_manager.save_to_vap3(
                self.file_path,
                filtered_sheets,
                sheet_images,
                plot_options,
                image_crop_states,
                plot_settings,
                sample_images,
                sample_image_crop_states,
                sample_header_data
            )

            if success:
                debug_print("DEBUG: VAP3 file with sample images saved successfully")
            else:
                debug_print("ERROR: Failed to save VAP3 file with sample images")

        except Exception as e:
            debug_print(f"ERROR: Failed to save VAP3 with sample images: {e}")
            import traceback
            traceback.print_exc()
            # Don't fail the entire save for this

    def save_sample_images_to_vap3(self):
        """Save sample-specific images to the VAP3 file."""
        try:
            debug_print("DEBUG: Saving sample images to VAP3 file")

            if not self.sample_images:
                debug_print("DEBUG: No sample images to save")
                return

            # We'll modify the VAP file manager to handle this
            # For now, store the sample images in the parent for later saving
            if hasattr(self.parent, 'pending_sample_images'):
                self.parent.pending_sample_images = self.sample_images.copy()
                self.parent.pending_sample_image_crop_states = self.sample_image_crop_states.copy()
                self.parent.pending_sample_header_data = self.header_data.copy()
                debug_print(f"DEBUG: Stored {len(self.sample_images)} sample image groups for later saving")

        except Exception as e:
            debug_print(f"ERROR: Failed to save sample images: {e}")
            import traceback
            traceback.print_exc()
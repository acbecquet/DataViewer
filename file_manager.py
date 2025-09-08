"""
File Management Module for DataViewer Application

This module handles all file-related operations including:
- Loading Excel files (standard and legacy formats)
- Saving and loading .vap3 files (custom format)
- Opening Excel files for editing and monitoring changes
- Managing file selections and UI updates related to files
"""

# Standard library imports
import os
import re
import copy
import shutil
import time
import uuid
import tempfile
import threading
import subprocess
import traceback
from typing import Optional

# Third party imports
import pandas as pd
import psutil
import openpyxl
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label, Button, ttk, Frame, Listbox, Scrollbar
from test_selection_dialog import TestSelectionDialog
from test_start_menu import TestStartMenu
from header_data_dialog import HeaderDataDialog

# Local imports
import processing
from resource_utils import get_resource_path
from utils import (
    is_valid_excel_file,
    load_excel_file,
    get_save_path,
    is_standard_file,
    FONT,
    APP_BACKGROUND_COLOR,
    load_excel_file_with_formulas,
    debug_print,
    show_success_message
)
from database_manager import DatabaseManager

class FileManager:
    """File Management Module for DataViewer.
    
    This class is initialized with a reference to the main TestingGUI instance
    so that it can update its state (sheets, selected_sheet, etc.).
    """
    def __init__(self, gui):
        self.gui = gui
        self.root = gui.root
        self.db_manager = DatabaseManager()
        
        # Add cache to prevent redundant operations
        self.loaded_files_cache = {}  # Cache for loaded file data
        self.stored_files_cache = set()  # Track files already stored in database
        
    def load_excel_file(self, file_path, legacy_mode: str = None, skip_database_storage: bool = False, force_reload: bool = False) -> None:
        """
        Load the selected Excel file and process its sheets.
        Enhanced with caching and optional database storage skip.

        Args:
            file_path (str): Path to the Excel file
            legacy_mode (str, optional): Legacy processing mode
            skip_database_storage (bool): If True, skip storing in database
            force_reload (bool): If True, bypass cache and reload from file
        """
        debug_print(f"DEBUG: load_excel_file called for {file_path}, skip_db_storage={skip_database_storage}, force_reload={force_reload}")

        # Check cache first (unless force_reload is True)
        cache_key = f"{file_path}_{legacy_mode}"
        if not force_reload and cache_key in self.loaded_files_cache:
            debug_print("DEBUG: Using cached file data instead of reprocessing")
            cached_data = self.loaded_files_cache[cache_key]
            self.gui.filtered_sheets = cached_data['filtered_sheets']
            self.gui.sheets = cached_data.get('sheets', {})
            # Fix: Make sure full_sample_data exists
            if hasattr(self.gui, 'full_sample_data'):
                self.gui.full_sample_data = cached_data.get('full_sample_data', pd.DataFrame())
    
            # Set the selected sheet without triggering full UI update
            if cached_data['filtered_sheets']:
                first_sheet = list(cached_data['filtered_sheets'].keys())[0]
                self.gui.selected_sheet.set(first_sheet)
            return

        # Clear cache entry if force_reload is True
        if force_reload and cache_key in self.loaded_files_cache:
            debug_print(f"DEBUG: Force reload requested - clearing cache entry for {file_path}")
            del self.loaded_files_cache[cache_key]

        try:
            # Ensure the file is a valid Excel file.
            if not is_valid_excel_file(os.path.basename(file_path)):
                raise ValueError(f"Invalid Excel file selected: {file_path}")

            debug_print(f"DEBUG: {'Force reloading' if force_reload else 'Loading'} file from disk: {file_path}")
            debug_print(f"DEBUG: Checking if file is standard format: {file_path}")
        
            if not is_standard_file(file_path):
                debug_print("DEBUG: File is legacy format, processing accordingly")
                # Legacy file processing
                legacy_dir = os.path.join(os.path.abspath("."), "legacy data")
                if not os.path.exists(legacy_dir):
                    os.makedirs(legacy_dir)
                    debug_print(f"DEBUG: Created legacy data directory: {legacy_dir}")
            
                legacy_wb = load_workbook(file_path)
                legacy_sheetnames = legacy_wb.sheetnames
                debug_print(f"DEBUG: Legacy file sheets: {legacy_sheetnames}")

                template_path_default = os.path.join(os.path.abspath("."), "resources", 
                                         "Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
                                     
                if not os.path.exists(template_path_default):
                    raise FileNotFoundError(f"Template file not found: {template_path_default}")
                
                wb_template = load_workbook(template_path_default)
                template_sheet_names = wb_template.sheetnames
                debug_print(f"DEBUG: Template sheets: {template_sheet_names}")

                if legacy_mode is None:
                    if len(legacy_sheetnames) == 1 and legacy_sheetnames[0] not in template_sheet_names:
                        legacy_mode = "file"
                        debug_print("DEBUG: Auto-detected legacy mode: file")
                    else:
                        legacy_mode = "standards"
                        debug_print("DEBUG: Auto-detected legacy mode: standards")

                final_sheets = {}
                full_sample_data = pd.DataFrame()

                if legacy_mode == "file":
                    debug_print("DEBUG: Processing as legacy file")
                    converted = processing.process_legacy_file_auto_detect(file_path)
                    key = f"Legacy_{os.path.basename(file_path)}"
                    final_sheets = {key: {"data": converted, "is_empty": converted.empty}}
                    full_sample_data = converted

                    self.gui.filtered_sheets = final_sheets
                    self.gui.full_sample_data = full_sample_data
                    default_key = list(final_sheets.keys())[0]
                    self.gui.selected_sheet.set(default_key)
                    debug_print(f"DEBUG: Legacy file processed, key: {default_key}")

                elif legacy_mode == "standards":
                    debug_print("DEBUG: Processing as legacy standards file")
                    converted_dict = processing.convert_legacy_standards_using_template(file_path)
                
                    self.gui.sheets = converted_dict
                    self.gui.filtered_sheets = {
                        name: {"data": data, "is_empty": data.empty}
                        for name, data in self.gui.sheets.items()
                    }
                    # Fix: Safely concatenate sheets
                    sheet_data_list = []
                    for sheet_info in self.gui.filtered_sheets.values():
                        if not sheet_info["data"].empty:
                            sheet_data_list.append(sheet_info["data"])
                
                    if sheet_data_list:
                        self.gui.full_sample_data = pd.concat(sheet_data_list, axis=1)
                    else:
                        self.gui.full_sample_data = pd.DataFrame()
                    
                    first_sheet = list(self.gui.filtered_sheets.keys())[0]
                    self.gui.selected_sheet.set(first_sheet)
                    debug_print(f"DEBUG: Legacy standards processed, first sheet: {first_sheet}")
                else:
                    raise ValueError(f"Unknown legacy mode: {legacy_mode}")

                # Store in database only if not skipping and not already stored
                if not skip_database_storage and file_path not in self.stored_files_cache:
                    debug_print("DEBUG: Storing legacy file in database")
                    self._store_file_in_database(file_path)
                    self.stored_files_cache.add(file_path)
            else:
                # Standard file processing
                debug_print("DEBUG: Processing as standard file")
                self.gui.sheets = load_excel_file(file_path)
                self.gui.filtered_sheets = {
                    name: {"data": data, "is_empty": data.empty}
                    for name, data in self.gui.sheets.items()
                }
            
                # Fix: Safely concatenate sheets and handle empty sheets
                sheet_data_list = []
                for sheet_info in self.gui.filtered_sheets.values():
                    if not sheet_info["data"].empty:
                        sheet_data_list.append(sheet_info["data"])
            
                if sheet_data_list:
                    self.gui.full_sample_data = pd.concat(sheet_data_list, axis=1)
                else:
                    self.gui.full_sample_data = pd.DataFrame()
                
                first_sheet = list(self.gui.filtered_sheets.keys())[0]
                self.gui.selected_sheet.set(first_sheet)
                debug_print(f"DEBUG: Standard file processed, first sheet: {first_sheet}")
            
                # Store in database only if not skipping and not already stored
                if not skip_database_storage and file_path not in self.stored_files_cache:
                    debug_print("DEBUG: Storing standard file in database")
                    self._store_file_in_database(file_path)
                    self.stored_files_cache.add(file_path)

            # Cache the processed data
            cache_data = {
                'filtered_sheets': copy.deepcopy(self.gui.filtered_sheets),
                'sheets': copy.deepcopy(getattr(self.gui, 'sheets', {})),
                'full_sample_data': self.gui.full_sample_data.copy() if hasattr(self.gui, 'full_sample_data') and not self.gui.full_sample_data.empty else pd.DataFrame()
            }
            self.loaded_files_cache[cache_key] = cache_data
            debug_print(f"DEBUG: Cached processed data for {file_path}")

        except Exception as e:
            error_msg = f"Error occurred while loading file: {e}"
            debug_print(f"ERROR: {error_msg}")
            traceback.print_exc()
            messagebox.showerror("Error", error_msg)

    def load_initial_file(self) -> None:
        """Handle file loading directly on the main thread."""
        file_paths = filedialog.askopenfilenames(
            title="Select Excel File(s)", 
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_paths:
            show_success_message("Info", "No files were selected", self.gui.root)
            return

        self.gui.progress_dialog.show_progress_bar("Loading files...")
        total_files = len(file_paths)
    
        try:
            for index, file_path in enumerate(file_paths, 1):
                # Update progress
                progress = int((index / total_files) * 100)
                self.gui.progress_dialog.update_progress_bar(progress)
            
                # Load and process file
                self.load_excel_file(file_path)
            
                # Store results
                self.gui.all_filtered_sheets.append({
                    "file_name": os.path.basename(file_path),
                    "file_path": file_path,
                    "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
                })
            
                # Allow GUI to refresh
                self.root.update_idletasks()

            # Final setup after loading all files
            if self.gui.all_filtered_sheets:
                last_file = self.gui.all_filtered_sheets[-1]
                self.set_active_file(last_file["file_name"])
                self.update_file_dropdown()
                self.update_ui_for_current_file()

        except Exception as e:
            messagebox.showerror("Loading Error", f"Failed to load files: {str(e)}")
        finally:
            self.gui.progress_dialog.hide_progress_bar()
            self.root.update_idletasks()

    def _store_file_in_database(self, original_file_path, display_filename=None):
        """
        Store the current file in the database.
        Enhanced with duplicate checking, better error handling, and sample image support.
        """
        debug_print(f"DEBUG: Checking if file {original_file_path} needs database storage")

        # Check if already stored
        if original_file_path in self.stored_files_cache:
            debug_print("DEBUG: File already stored in database, skipping")
            return

        try:
            debug_print("DEBUG: Storing new file in database...")
            # Show progress dialog
            self.gui.progress_dialog.show_progress_bar("Storing file in database...")
            self.gui.root.update_idletasks()

            # CREATE: Collect sample notes from current filtered_sheets (ADD THIS SECTION)
            sample_notes_data = {}
            for sheet_name, sheet_info in self.gui.filtered_sheets.items():
                header_data = sheet_info.get('header_data')
                if header_data and 'samples' in header_data:
                    sheet_notes = {}
                    for i, sample_data in enumerate(header_data['samples']):
                        sample_notes = sample_data.get('sample_notes', '')
                        if sample_notes.strip():
                            sheet_notes[f"Sample {i+1}"] = sample_notes
                
                    if sheet_notes:
                        sample_notes_data[sheet_name] = sheet_notes
                        debug_print(f"DEBUG: Collected notes for sheet {sheet_name}: {len(sheet_notes)} samples")

            # Create a temporary VAP3 file
            with tempfile.NamedTemporaryFile(suffix='.vap3', delete=False) as temp_file:
                temp_vap3_path = temp_file.name
                debug_print(f"DEBUG: Created temporary VAP3 file: {temp_vap3_path}")

            # Save current state as VAP3
            from vap_file_manager import VapFileManager
            vap_manager = VapFileManager()

            # Collect plot settings safely
            plot_settings = {}
            if hasattr(self.gui, 'selected_plot_type'):
                plot_settings['selected_plot_type'] = self.gui.selected_plot_type.get()

            debug_print(f"DEBUG: Plot settings: {plot_settings}")

            # Get image crop states safely
            image_crop_states = getattr(self.gui, 'image_crop_states', {})
            if hasattr(self.gui, 'image_loader') and self.gui.image_loader:
                image_crop_states.update(getattr(self.gui.image_loader, 'image_crop_states', {}))

            debug_print(f"DEBUG: Image crop states: {len(image_crop_states)} items")

            # Collect sample images for database storage
            sample_images = {}
            sample_image_crop_states = {}
            sample_header_data = {}
        
            # Check if we have pending sample images
            if hasattr(self.gui, 'pending_sample_images'):
                sample_images = getattr(self.gui, 'pending_sample_images', {})
                sample_image_crop_states = getattr(self.gui, 'pending_sample_image_crop_states', {})
                sample_header_data = getattr(self.gui, 'pending_sample_header_data', {})
                debug_print(f"DEBUG: Found pending sample images: {len(sample_images)} samples")
        
            # Also check sample_image_metadata structure for all files/sheets
            if hasattr(self.gui, 'sample_image_metadata'):
                current_file = getattr(self.gui, 'current_file', None)
                if current_file and current_file in self.gui.sample_image_metadata:
                    for sheet_name, metadata in self.gui.sample_image_metadata[current_file].items():
                        sheet_sample_images = metadata.get('sample_images', {})
                        if sheet_sample_images:
                            # Merge with existing sample images
                            sample_images.update(sheet_sample_images)
                            sample_image_crop_states.update(metadata.get('sample_image_crop_states', {}))
                            if not sample_header_data:  # Use first header data found
                                sample_header_data = metadata.get('header_data', {})
                            debug_print(f"DEBUG: Found sample images in metadata for {sheet_name}: {len(sheet_sample_images)} samples")

            debug_print(f"DEBUG: Total sample images for database storage: {len(sample_images)} samples")

            # Extract and construct the display filename
            if display_filename is None:
                if original_file_path.endswith('.vap3'):
                    display_filename = os.path.basename(original_file_path)
                else:
                    # For Excel files, create .vap3 extension
                    base_name = os.path.splitext(os.path.basename(original_file_path))[0]
                    display_filename = f"{base_name}.vap3"
        
            debug_print(f"DEBUG: Using display filename: {display_filename}")

            # Save the VAP3 file with sample images included
            success = vap_manager.save_to_vap3(
                temp_vap3_path,
                self.gui.filtered_sheets,
                getattr(self.gui, 'sheet_images', {}),
                getattr(self.gui, 'plot_options', []),
                image_crop_states,
                plot_settings,
                sample_images,
                sample_image_crop_states,
                sample_header_data
            )
        
            if not success:
                raise Exception("Failed to create temporary VAP3 file")
        
            debug_print("DEBUG: VAP3 file created successfully with sample images")

            # MODIFY: Store metadata about the original file (ADD sample_notes to metadata)
            meta_data = {
                'display_filename': display_filename,
                'original_filename': os.path.basename(original_file_path),
                'original_path': original_file_path,
                'creation_date': time.strftime('%Y-%m-%d %H:%M:%S'),
                'sheet_count': len(self.gui.filtered_sheets),
                'plot_options': getattr(self.gui, 'plot_options', []),
                'plot_settings': plot_settings,
                'has_sample_images': bool(sample_images),
                'sample_count': len(sample_images),
                'sample_notes': sample_notes_data  # ADD THIS LINE
            }

            debug_print(f"DEBUG: Metadata to store: {meta_data}")

            # Store the VAP3 file in the database with the proper display filename
            file_id = self.db_manager.store_vap3_file(temp_vap3_path, meta_data)
            debug_print(f"DEBUG: File stored with ID: {file_id}")

            # Store sheet metadata
            for sheet_name, sheet_info in self.gui.filtered_sheets.items():
                is_plotting = processing.plotting_sheet_test(sheet_name, sheet_info["data"])
                is_empty = sheet_info.get("is_empty", False)

                sheet_id = self.db_manager.store_sheet_info(
                    file_id, 
                    sheet_name, 
                    is_plotting, 
                    is_empty
                )
                debug_print(f"DEBUG: Stored sheet '{sheet_name}' with ID: {sheet_id}")

            # Store associated images
            if hasattr(self.gui, 'sheet_images') and hasattr(self.gui, 'current_file') and self.gui.current_file in self.gui.sheet_images:
                current_file = self.gui.current_file
                for sheet_name, image_paths in self.gui.sheet_images[current_file].items():
                    for image_path in image_paths:
                        if os.path.exists(image_path):
                            crop_enabled = image_crop_states.get(image_path, False)
                            self.db_manager.store_image(file_id, sheet_name, image_path, crop_enabled)
                            debug_print(f"DEBUG: Stored image for sheet {sheet_name}: {os.path.basename(image_path)}")

            # Clean up temporary file
            try:
                os.unlink(temp_vap3_path)
            except:
                pass

            # Add to cache to prevent re-storing
            self.stored_files_cache.add(original_file_path)
        
            debug_print("DEBUG: File stored in database successfully")

        except Exception as e:
            debug_print(f"ERROR: Failed to store file in database: {e}")
            import traceback
            traceback.print_exc()
            raise e
        finally:
            self.gui.progress_dialog.hide_progress_bar()
    

    def handle_data_collection_close(self, data_collection_window, result):
        """Handle closing of data collection window with sample image support."""
        try:
            if result == "load_file":
                # Check for sample images before loading
                if hasattr(data_collection_window, 'sample_images') and data_collection_window.sample_images:
                    debug_print("DEBUG: Data collection closed with sample images")
                
                    # Store sample images in main GUI for saving
                    self.gui.pending_sample_images = data_collection_window.sample_images.copy()
                    self.gui.pending_sample_image_crop_states = data_collection_window.sample_image_crop_states.copy()
                    self.gui.pending_sample_header_data = data_collection_window.header_data.copy()
                
                    # Transfer formatted images
                    data_collection_window.transfer_images_to_main_gui()
                
                    # Save with sample images
                    if self.gui.file_path and self.gui.file_path.endswith('.vap3'):
                        self.gui.save_with_sample_images()
            
                # Load the file for viewing
                self.load_file(data_collection_window.file_path)
            
            # Restore main window visibility
            if hasattr(data_collection_window, 'main_window_was_visible') and data_collection_window.main_window_was_visible:
                self.gui.root.deiconify()
        
        except Exception as e:
            debug_print(f"ERROR: Failed to handle data collection close: {e}")
            import traceback
            traceback.print_exc()

    def load_from_database(self, file_id=None, show_success_msg=True, batch_operation=False):
        """Load a file from the database."""
        debug_print(f"DEBUG: load_from_database called with file_id={file_id}, show_success_message={show_success_msg}, batch_operation={batch_operation}")
        try:
            # Only show progress dialog if not part of a batch operation
            if not batch_operation:
                self.gui.progress_dialog.show_progress_bar("Loading from database...")
                self.gui.root.update_idletasks()

            if file_id is None:
                # Show a dialog to select from available files
                file_list = self.db_manager.list_files()
                if not file_list:
                    show_success_message("Info", "No files found in the database.", self.gui.root)
                    return False

                # Create file selection dialog
                selection_dialog = FileSelectionDialog(self.gui.root, file_list)
                file_id = selection_dialog.show()
            
                if file_id is None:
                    debug_print("DEBUG: No file selected from dialog")
                    return False

            # Get file data from database - CORRECTED METHOD NAME
            file_data = self.db_manager.get_file_by_id(file_id)
            if not file_data:
                if show_success_msg:
                    messagebox.showerror("Error", "File not found in database.")
                return False
        
            raw_database_filename = file_data['filename']
            debug_print(f"DEBUG: Raw database filename: {raw_database_filename}")
            created_at = file_data.get('created_at')

            # Get the proper display filename - prioritize metadata, then fallback to database filename
            display_filename = None

            # First, try to get display_filename from metadata
            if 'meta_data' in file_data and file_data['meta_data']:
                display_filename = file_data['meta_data'].get('display_filename')
                debug_print(f"DEBUG: display_filename from metadata: '{display_filename}'")
    
                # If not found, try original_filename from metadata and construct .vap3 name
                if not display_filename:
                    original_filename = file_data['meta_data'].get('original_filename')
                    if original_filename:
                        # Remove extension and add .vap3
                        display_filename = os.path.splitext(original_filename)[0] + '.vap3'
                        debug_print(f"DEBUG: Constructed filename from original_filename: '{display_filename}'")

            # Final fallback to database filename field
            if not display_filename:
                display_filename = file_data['filename']
                debug_print(f"DEBUG: Using database filename as fallback: '{display_filename}'")

            # Check if we already have files loaded to determine if we should append
            append_to_existing = len(self.gui.all_filtered_sheets) > 0

            # Save the VAP3 file to a temporary location
            with tempfile.NamedTemporaryFile(suffix='.vap3', delete=False) as temp_file:
                temp_file.write(file_data['file_content'])
                temp_vap3_path = temp_file.name

            try:
                # CRITICAL: Load and process VAP3 data BEFORE loading the file
                from vap_file_manager import VapFileManager
                vap_manager = VapFileManager()
                vap_data = vap_manager.load_from_vap3(temp_vap3_path)
                debug_print(f"DEBUG: VAP3 data keys: {list(vap_data.keys())}")
                debug_print(f"DEBUG: Sample images in vap_data: {list(vap_data.get('sample_images', {}).keys())}")
                debug_print(f"DEBUG: Sample image counts: {[(k, len(v)) for k, v in vap_data.get('sample_images', {}).items()]}")
                # Store VAP3 data in GUI for sample image loading
                self.gui.current_vap_data = vap_data
                debug_print(f"DEBUG: Loaded VAP3 data with keys: {list(vap_data.keys())}")
            
                # Use the enhanced VAP3 loading that handles sample images
                success = self.load_vap3_file(temp_vap3_path, display_name=display_filename, append_to_existing=append_to_existing)
            
                if success:
                    total_files = len(self.gui.all_filtered_sheets)
                    debug_print(f"DEBUG: Successfully loaded from database. Total files: {total_files}")
                
                    # CRITICAL FIX: Load sample images from the VAP3 data
                    # Load sample images if they exist
                    if 'sample_images' in vap_data and vap_data['sample_images']:
                        debug_print(f"DEBUG: Loading sample images from database VAP3")
                    
                        # Load sample images and populate main GUI
                        self.gui.load_sample_images_from_vap3(vap_data)
                    
                        # CRITICAL: Also populate the sample_image_metadata structure
                        sample_images = vap_data.get('sample_images', {})
                        sample_crop_states = vap_data.get('sample_image_crop_states', {})
                        sample_header_data = vap_data.get('sample_images_metadata', {}).get('header_data', {})
                    
                        debug_print(f"DEBUG: Sample images content: {sample_images}")
                        debug_print(f"DEBUG: Sample images metadata content: {sample_header_data}")
                        debug_print(f"DEBUG: Sample crop states content: {sample_crop_states}")
                        debug_print(f"DEBUG: Sample images keys: {list(sample_images.keys()) if sample_images else 'Empty'}")
                        debug_print(f"DEBUG: Total sample image files: {sum(len(imgs) for imgs in sample_images.values()) if sample_images else 0}")

                        # Also check regular sheet images for comparison
                        sheet_images = vap_data.get('sheet_images', {})
                        debug_print(f"DEBUG: Sheet images content: {sheet_images}")
                        if sheet_images:
                            for file_key, sheets in sheet_images.items():
                                debug_print(f"DEBUG: File '{file_key}' sheet images:")
                                for sheet_name, images in sheets.items():
                                    debug_print(f"DEBUG:   Sheet '{sheet_name}': {len(images)} images - {images[:2] if images else 'None'}...")

                        if sample_images:
                            if not hasattr(self.gui, 'sample_image_metadata'):
                                self.gui.sample_image_metadata = {}
                            if display_filename not in self.gui.sample_image_metadata:
                                self.gui.sample_image_metadata[display_filename] = {}
                        
                            # Determine which sheet this belongs to
                            test_name = sample_header_data.get('test', 'Unknown Test')
                            if test_name in self.gui.filtered_sheets:
                                self.gui.sample_image_metadata[display_filename][test_name] = {
                                    'sample_images': sample_images,
                                    'sample_image_crop_states': sample_crop_states,
                                    'header_data': sample_header_data,
                                    'test_name': test_name
                                }
                                debug_print(f"DEBUG: Populated sample_image_metadata for {test_name} in file {display_filename}")
                            else:
                                # If test_name not found, try to find a matching sheet
                                for sheet_name in self.gui.filtered_sheets.keys():
                                    if sheet_name.lower() == test_name.lower() or test_name.lower() in sheet_name.lower():
                                        self.gui.sample_image_metadata[display_filename][sheet_name] = {
                                            'sample_images': sample_images,
                                            'sample_image_crop_states': sample_crop_states,
                                            'header_data': sample_header_data,
                                            'test_name': sheet_name
                                        }
                                        debug_print(f"DEBUG: Populated sample_image_metadata for matched sheet {sheet_name} in file {display_filename}")
                                        break
                    else:
                        debug_print("DEBUG: No sample images found in VAP3 data")
                
                    # Store database-specific metadata in the latest file entry
                    if self.gui.all_filtered_sheets:
                        latest_file = self.gui.all_filtered_sheets[-1]
                        latest_file['database_filename'] = raw_database_filename
                        latest_file['database_created_at'] = created_at
            
                        # Also store in the original_filename if not already set
                        if 'original_filename' not in latest_file:
                            latest_file['original_filename'] = raw_database_filename
            
                        debug_print(f"DEBUG: Stored database filename in metadata: {raw_database_filename}")

                    # Show the success message only if requested and not in batch mode
                    if show_success_msg and not batch_operation:
                        if total_files > 1:
                            show_success_message("Success", f"VAP3 file loaded successfully: {display_filename}\nTotal files loaded: {total_files}", self.gui.root)
                        else:
                            show_success_message("Success", f"VAP3 file loaded successfully: {display_filename}", self.gui.root)
        
                    return True
                else:
                    if show_success_msg:
                        messagebox.showerror("Error", f"Failed to load file: {display_filename}")
                    return False
                
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_vap3_path)
                except:
                    pass

        except Exception as e:
            if show_success_msg:  # Only show error dialog if not in batch mode
                messagebox.showerror("Error", f"Error loading file from database: {e}")
            debug_print(f"ERROR: Error loading file from database: {e}")
            traceback.print_exc()
            return False
        finally:
            # Hide progress dialog only if not in batch mode
            if not batch_operation:
                self.gui.progress_dialog.hide_progress_bar()

    def load_multiple_from_database(self, file_ids):
        """Load multiple files from the database with a single progress dialog and success message."""
        if not file_ids:
            return

        loaded_files = []
        failed_files = []
        total_sample_images_loaded = 0

        try:
            # Show progress dialog for the entire batch operation
            self.gui.progress_dialog.show_progress_bar("Loading files from database...")
            self.gui.root.update_idletasks()
    
            total_files = len(file_ids)
            debug_print(f"DEBUG: Starting batch load of {total_files} files")
    
            for i, file_id in enumerate(file_ids):
                try:
                    # Update progress
                    progress = int(((i + 1) / total_files) * 100)
                    self.gui.progress_dialog.update_progress_bar(progress)
                    self.gui.root.update_idletasks()
            
                    debug_print(f"DEBUG: Loading file {i + 1}/{total_files} (ID: {file_id})")
            
                    # Load this file (suppress individual success messages and progress dialogs)
                    success = self.load_from_database(file_id, show_success_msg=False, batch_operation=True)
            
                    if success:
                        # Get the file info for the success message - CORRECTED METHOD NAME
                        file_data = self.db_manager.get_file_by_id(file_id)
                        if file_data:
                            display_filename = None
                            if 'meta_data' in file_data and file_data['meta_data']:
                                display_filename = file_data['meta_data'].get('display_filename')
                                if not display_filename:
                                    original_filename = file_data['meta_data'].get('original_filename')
                                    if original_filename:
                                        display_filename = os.path.splitext(original_filename)[0] + '.vap3'
                            if not display_filename:
                                display_filename = file_data['filename']
                    
                            loaded_files.append(display_filename)
                            debug_print(f"DEBUG: Successfully loaded: {display_filename}")
                        
                            # Count sample images for this file
                            if (hasattr(self.gui, 'sample_image_metadata') and 
                                display_filename in self.gui.sample_image_metadata):
                                for sheet_metadata in self.gui.sample_image_metadata[display_filename].values():
                                    sample_images = sheet_metadata.get('sample_images', {})
                                    file_sample_count = sum(len(images) for images in sample_images.values())
                                    total_sample_images_loaded += file_sample_count
                                    debug_print(f"DEBUG: Loaded {file_sample_count} sample images for {display_filename}")
                    else:
                        failed_files.append(f"File ID {file_id}")
                        debug_print(f"DEBUG: Failed to load file ID: {file_id}")
                
                except Exception as e:
                    failed_files.append(f"File ID {file_id}")
                    debug_print(f"DEBUG: Exception loading file ID {file_id}: {e}")
    
            # Update final progress
            self.gui.progress_dialog.update_progress_bar(100)
            self.gui.root.update_idletasks()
    
            # Update window title with final count
            total_loaded = len(self.gui.all_filtered_sheets)
            if total_loaded > 1:
                self.gui.root.title(f"DataViewer - {total_loaded} files loaded")
            elif total_loaded == 1:
                self.gui.root.title("DataViewer - 1 file loaded")
    
            # Show single summary message
            if failed_files:
                if loaded_files:
                    # Partial success
                    success_count = len(loaded_files)
                    failed_count = len(failed_files)
                    message = f"Batch load completed:\n\n"
                    message += f"✓ Successfully loaded: {success_count} files\n"
                    message += f"✗ Failed to load: {failed_count} files\n\n"
                    if total_sample_images_loaded > 0:
                        message += f"📷 Sample images loaded: {total_sample_images_loaded}\n\n"
                    message += f"Total files now loaded: {len(self.gui.all_filtered_sheets)}"
                    messagebox.showwarning("Partial Success", message)
                else:
                    # Complete failure
                    messagebox.showerror("Error", f"Failed to load all {len(failed_files)} selected files.")
            else:
                # Complete success
                if len(loaded_files) == 1:
                    message = f"Successfully loaded 1 file:\n{loaded_files[0]}"
                    if total_sample_images_loaded > 0:
                        message += f"\n\n📷 Sample images loaded: {total_sample_images_loaded}"
                else:
                    message = f"Successfully loaded {len(loaded_files)} files:\n\n"
                    # Show first few filenames, then "and X more" if too many
                    if len(loaded_files) <= 5:
                        message += "\n".join([f"• {name}" for name in loaded_files])
                    else:
                        message += "\n".join([f"• {name}" for name in loaded_files[:3]])
                        message += f"\n• ... and {len(loaded_files) - 3} more files"
            
                    if total_sample_images_loaded > 0:
                        message += f"\n\n📷 Total sample images loaded: {total_sample_images_loaded}"
                    message += f"\n\nTotal files now loaded: {len(self.gui.all_filtered_sheets)}"
        
                show_success_message("Success", message, self.gui.root)
    
        except Exception as e:
            messagebox.showerror("Error", f"Error during batch loading: {e}")
            debug_print(f"ERROR: Batch loading error: {e}")
            traceback.print_exc()
        finally:
            # Hide progress dialog
            self.gui.progress_dialog.hide_progress_bar()

    def load_from_database_by_id(self, file_id):
        """Load a file from database by its ID."""
        debug_print(f"DEBUG: Loading file from database with ID: {file_id}")
    
        try:
            # Get file data from database
            file_data = self.db_manager.get_file_by_id(file_id)
        
            if file_data and 'filepath' in file_data:
                # Load the file
                self.load_excel_file(file_data['filepath'], skip_database_storage=True)
                return True
            else:
                debug_print(f"DEBUG: Could not find file data for ID: {file_id}")
                return False
            
        except Exception as e:
            debug_print(f"DEBUG: Error loading file from database ID {file_id}: {e}")
            return False

    def center_window(self, window, width, height):
        """Center a window on the screen."""
        try:
            # Get screen dimensions
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()
        
            # Calculate center position
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
        
            # Set window geometry with center position
            window.geometry(f"{width}x{height}+{x}+{y}")
        
            debug_print(f"DEBUG: Centered window {width}x{height} at position ({x}, {y}) on {screen_width}x{screen_height} screen")
        
        except Exception as e:
            debug_print(f"DEBUG: Error centering window: {e}")
            # Fallback to basic geometry if centering fails
            window.geometry(f"{width}x{height}")

    def show_database_browser(self, comparison_mode=False):
        """Show a dialog to browse files stored in the database with multi-select support."""
        # Create a dialog to browse the database
        dialog = Toplevel(self.gui.root)

        # Update title based on mode
        if comparison_mode:
            dialog.title("Select Files for Comparison")
        else:
            dialog.title("Database Browser")
    
        dialog.transient(self.gui.root)
        self.center_window(dialog, 900, 700)

        # Create frames for UI elements
        top_frame = Frame(dialog)
        top_frame.pack(fill="x", padx=10, pady=10)

        list_frame = Frame(dialog)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)

        bottom_frame = Frame(dialog)
        bottom_frame.pack(fill="x", padx=10, pady=10)

        # Store grouped files data
        file_groups = {}  # Key: base filename, Value: list of file records
        listbox_to_file_mapping = {}  # Key: listbox index, Value: file record

        def extract_base_filename(filename):
            """Extract base filename without timestamp and extension."""
            
            # Remove common timestamp patterns and extensions
            base = re.sub(r'\s+\d{4}-\d{2}-\d{2}.*', '', filename)  # Remove date patterns
            base = re.sub(r'\s+copy.*', '', base, flags=re.IGNORECASE)  # Remove "copy" variants
            base = re.sub(r'\.[^.]*$', '', base)  # Remove extension
            return base.strip()


        def sort_file_groups(groups, sort_method):
            """Sort file groups based on the selected method."""
            debug_print(f"DEBUG: Sorting files by {sort_method}")
            
            if sort_method == "newest_first":
                # Sort by most recent update (within each group, then by group)
                sorted_groups = {}
                for base_name, versions in groups.items():
                    # Each group is already sorted by creation date (newest first)
                    sorted_groups[base_name] = versions
                # Sort group keys by the most recent file in each group
                sorted_keys = sorted(sorted_groups.keys(), 
                                   key=lambda name: sorted_groups[name][0]['created_at'], reverse=True)
            elif sort_method == "oldest_first":
                # Sort by oldest update
                sorted_groups = {}
                for base_name, versions in groups.items():
                    sorted_groups[base_name] = versions
                # Sort group keys by the most recent file in each group (oldest first)
                sorted_keys = sorted(sorted_groups.keys(), 
                                   key=lambda name: sorted_groups[name][0]['created_at'], reverse=False)
            elif sort_method == "alphabetical_asc":
                # Sort alphabetically A to Z
                sorted_groups = groups
                sorted_keys = sorted(groups.keys())
            elif sort_method == "alphabetical_desc":
                # Sort alphabetically Z to A  
                sorted_groups = groups
                sorted_keys = sorted(groups.keys(), reverse=True)
            else:
                # Default to newest first
                sorted_groups = groups
                sorted_keys = sorted(groups.keys(), 
                                   key=lambda name: groups[name][0]['created_at'], reverse=True)
            
            return sorted_groups, sorted_keys

        def populate_listbox(filter_keyword = None):
            """Populate the listbox with files according to current sort method."""
            try:
                debug_print("DEBUG: Populating listbox with sorted files")
                
                # Clear current listbox
                file_listbox.delete(0, tk.END)
                listbox_to_file_mapping.clear()
                
                # Get current sort method
                current_sort = sort_var.get()
                if current_sort == "Newest First":
                    sort_method = "newest_first"
                elif current_sort == "Oldest First":
                    sort_method = "oldest_first"
                elif current_sort == "A to Z":
                    sort_method = "alphabetical_asc"
                elif current_sort == "Z to A":
                    sort_method = "alphabetical_desc"
                else:
                    sort_method = "newest_first"

                # Apply filter if provided
                filtered_groups = file_groups.copy()
                if filter_keyword and filter_keyword.strip():
                    keyword_lower = filter_keyword.strip().lower()
                    filtered_groups = {
                        base_name: versions 
                        for base_name, versions in file_groups.items() 
                        if keyword_lower in base_name.lower()
                    }
                    debug_print(f"DEBUG: Applied filter '{filter_keyword}' - {len(filtered_groups)} groups remaining")
              
                
                # Sort the file groups
                sorted_groups, sorted_keys = sort_file_groups(filtered_groups, sort_method)
                
                # Populate listbox with sorted groups
                listbox_index = 0
                for base_name in sorted_keys:
                    latest_file = sorted_groups[base_name][0]  # First item is newest
                    count = len(sorted_groups[base_name])

                    if count > 1:
                        display_text = f"{base_name} ({count} versions, latest: {latest_file['created_at'].strftime('%Y-%m-%d %H:%M')})"
                    else:
                        display_text = f"{base_name} ({latest_file['created_at'].strftime('%Y-%m-%d %H:%M')})"

                    file_listbox.insert(tk.END, display_text)
                    listbox_to_file_mapping[listbox_index] = latest_file
                    debug_print(f"DEBUG: Mapped listbox index {listbox_index} to file ID {latest_file['id']} for base name '{base_name}'")
                    listbox_index += 1
                    
                update_selection_info()
                debug_print(f"DEBUG: Listbox populated with {listbox_index} items")
                
            except Exception as e:
                debug_print(f"DEBUG: Error populating listbox: {e}")
                messagebox.showerror("Error", f"Error displaying files: {e}")

        try:
            # Get files from database
            files = self.db_manager.list_files()
            debug_print(f"DEBUG: Found {len(files)} files in database")

            # Group files by base filename
            for file_record in files:
                base_name = extract_base_filename(file_record['filename'])
                if base_name not in file_groups:
                    file_groups[base_name] = []
                file_groups[base_name].append(file_record)

            # Sort each group by creation date (newest first)
            for group in file_groups.values():
                group.sort(key=lambda x: x['created_at'], reverse=True)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load files from database: {e}")
            dialog.destroy()
            return

        # Create header
        if comparison_mode:
            header_text = "Select files for comparison analysis (minimum 2 required)"
        else:
            header_text = "Database Files (grouped by base name, showing latest version)"

        Label(top_frame, text=header_text, font=("Arial", 12, "bold")).pack()

        # Add sorting controls
        sort_frame = Frame(top_frame)
        sort_frame.pack(fill="x", pady=5)
        
        Label(sort_frame, text="Sort by:", font=("Arial", 10)).pack(side="left", padx=(0, 5))
        
        # Sort options variable
        sort_var = tk.StringVar(value="newest_first")
        
        sort_options = [
            ("Newest First", "newest_first"),
            ("Oldest First", "oldest_first"), 
            ("A to Z", "alphabetical_asc"),
            ("Z to A", "alphabetical_desc")
        ]
        
        sort_dropdown = ttk.Combobox(sort_frame, textvariable=sort_var, values=[option[1] for option in sort_options], 
                                   state="readonly", width=15)
        sort_dropdown.pack(side="left", padx=5)
        
        # Set display values
        sort_dropdown['values'] = [option[0] for option in sort_options]
        sort_dropdown.set("Newest First")

        # Right side - Filter controls
        filter_var = tk.StringVar()
        Label(sort_frame, text="Filter:", font=("Arial", 10)).pack(side="right", padx=(5, 0))
        filter_entry = tk.Entry(sort_frame, textvariable=filter_var, width=20, font=("Arial", 10))
        filter_entry.pack(side="right", padx=(0, 5))

        # Create listbox with scrollbar
        listbox_frame = Frame(list_frame)

        # Create listbox with scrollbar
        listbox_frame = Frame(list_frame)
        listbox_frame.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side="right", fill="y")

        # Use extended selection mode for multi-select
        file_listbox = tk.Listbox(listbox_frame, yscrollcommand=scrollbar.set, 
                                 selectmode=tk.EXTENDED, font=FONT)
        file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=file_listbox.yview)

        # Selection info label (create before functions that use it)
        selection_info = Label(list_frame, text="", font=("Arial", 10))
        selection_info.pack(pady=5)

        # DEFINE update_selection_info FIRST since populate_listbox needs it
        def update_selection_info():
            """Update the selection information label."""
            debug_print("DEBUG: Updating selection info")
            try:
                selected_count = len(file_listbox.curselection())
                if comparison_mode:
                    if selected_count < 2:
                        selection_info.config(text=f"Selected: {selected_count} files (need at least 2 for comparison)", 
                                            fg="red")
                    else:
                        selection_info.config(text=f"Selected: {selected_count} files (ready for comparison)", 
                                            fg="green")
                else:
                    selection_info.config(text=f"Selected: {selected_count} files", fg="black")
                debug_print(f"DEBUG: Selection info updated - {selected_count} files selected")
            except Exception as e:
                debug_print(f"DEBUG: Error updating selection info: {e}")

        # Initial population with default sort
        populate_listbox()

        # Bind sort dropdown change event
        def on_sort_change(event=None):
            """Handle sort method change."""
            debug_print(f"DEBUG: Sort method changed to {sort_var.get()}")
            current_filter = filter_var.get()
            populate_listbox(current_filter if current_filter else None)
            
        sort_dropdown.bind('<<ComboboxSelected>>', on_sort_change)

        # Bind filter entry change event
        def on_filter_change(event=None):
            """Handle filter keyword change."""
            filter_keyword = filter_var.get()
            debug_print(f"DEBUG: Filter keyword changed to '{filter_keyword}'")
            populate_listbox(filter_keyword if filter_keyword else None)
          
        filter_entry.bind('<KeyRelease>', on_filter_change)
        filter_entry.bind('<FocusOut>', on_filter_change)

        def show_version_history_dialog(base_name, versions, main_listbox_index):
            """Show dialog with version history and allow version selection."""
            history_dialog = Toplevel(dialog)
            history_dialog.title(f"Version History - {base_name}")

            # Hide the window immediately to prevent stutter during positioning
            history_dialog.withdraw()

            history_dialog.transient(dialog)
            history_dialog.grab_set()

            # Create frames
            top_frame_hist = Frame(history_dialog)
            top_frame_hist.pack(fill="x", padx=10, pady=10)

            list_frame_hist = Frame(history_dialog)
            list_frame_hist.pack(fill="both", expand=True, padx=10, pady=10)

            bottom_frame_hist = Frame(history_dialog)
            bottom_frame_hist.pack(fill="x", padx=10, pady=10)

            # Header
            Label(top_frame_hist, text=f"Version History: {base_name}", 
                  font=("Arial", 14, "bold")).pack()
            Label(top_frame_hist, text="Double-click a version to select it for loading. Use Ctrl+click for multiple selection.", 
                  font=("Arial", 10)).pack(pady=5)

            # Version listbox with scrollbar
            hist_listbox_frame = Frame(list_frame_hist)
            hist_listbox_frame.pack(fill="both", expand=True)

            hist_scrollbar = tk.Scrollbar(hist_listbox_frame)
            hist_scrollbar.pack(side="right", fill="y")

            version_listbox = tk.Listbox(hist_listbox_frame, yscrollcommand=hist_scrollbar.set, selectmode=tk.EXTENDED, font=FONT)
            version_listbox.pack(side="left", fill="both", expand=True)
            hist_scrollbar.config(command=version_listbox.yview)

            # Populate with versions (newest first)
            for i, version in enumerate(versions):
                created_str = version['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                is_current = i == 0
                status = " [CURRENT]" if is_current else f" [v{len(versions)-i}]"
                display_text = f"{version['filename']}{status} - {created_str}"
                version_listbox.insert(tk.END, display_text)

            def on_version_double_click(event):
                """Handle double-click on version to replace in main browser."""
                selection = version_listbox.curselection()
                if not selection:
                    return
    
                selected_version_idx = selection[0]
                selected_version = versions[selected_version_idx]
    
                # Update the main listbox display
                created_str = selected_version['created_at'].strftime('%Y-%m-%d %H:%M')
                if selected_version_idx == 0:
                    # Current version selected - restore normal display
                    count = len(versions)
                    if count > 1:
                        display_text = f"{base_name} ({count} versions, latest: {created_str})"
                    else:
                        display_text = f"{base_name} ({created_str})"
                    # Remove from restoration tracking if it exists
                    if hasattr(dialog, 'original_mappings') and main_listbox_index in dialog.original_mappings:
                        del dialog.original_mappings[main_listbox_index]
                else:
                    # Older version selected - show with indicator
                    version_num = len(versions) - selected_version_idx
                    display_text = f"{base_name} [v{version_num}: {created_str}] ★"
        
                    # Store original mapping for restoration if not already stored
                    if not hasattr(dialog, 'original_mappings'):
                        dialog.original_mappings = {}
                    if main_listbox_index not in dialog.original_mappings:
                        dialog.original_mappings[main_listbox_index] = versions[0]  # Store original (newest)
    
                # Update main listbox
                file_listbox.delete(main_listbox_index)
                file_listbox.insert(main_listbox_index, display_text)
    
                # Update mapping to point to selected version
                listbox_to_file_mapping[main_listbox_index] = selected_version
    
                # Keep the item selected in main listbox
                file_listbox.selection_clear(0, tk.END)
                file_listbox.selection_set(main_listbox_index)
    
                debug_print(f"DEBUG: Temporarily replaced {base_name} with version {selected_version_idx + 1}")
    
                history_dialog.destroy()

            # Bind double-click to version listbox
            version_listbox.bind("<Double-Button-1>", on_version_double_click)

            # Close button
            def on_close():
                history_dialog.destroy()

            def on_delete_versions():
                """Delete selected versions from the version history."""
                selected_items = version_listbox.curselection()
                if not selected_items:
                    messagebox.showwarning("Warning", "Please select at least one version to delete.")
                    return

                version_ids = []
                version_names = []
                for idx in selected_items:
                    if idx < len(versions):
                        version_record = versions[idx]
                        version_ids.append(version_record['id'])
                        version_names.append(version_record['filename'])

                if not version_ids:
                    messagebox.showwarning("Warning", "No valid versions selected for deletion.")
                    return

                # Confirm deletion
                if len(version_ids) == 1:
                    confirm_msg = f"Are you sure you want to delete this version:\n\n{version_names[0]}\n\nThis action cannot be undone."
                else:
                    confirm_msg = f"Are you sure you want to delete {len(version_ids)} versions?\n\nThis action cannot be undone."

                if not messagebox.askyesno("Confirm Deletion", confirm_msg):
                    return

                try:
                    debug_print(f"DEBUG: Starting deletion of {len(version_ids)} versions")
                    success_count, error_count = self.db_manager.delete_multiple_files(version_ids)
                
                    if success_count > 0:
                        if error_count == 0:
                            messagebox.showinfo("Success", f"Successfully deleted {success_count} version(s).")
                        else:
                            messagebox.showwarning("Partial Success", f"Deleted {success_count} version(s), but {error_count} failed.")
                    
                        # Check if we deleted the most recent version
                        most_recent_deleted = 0 in selected_items
                    
                        if most_recent_deleted:
                            # Find the new most recent version and update main listbox
                            remaining_versions = [v for i, v in enumerate(versions) if i not in selected_items]
                            if remaining_versions:
                                # Sort remaining versions by date (newest first)
                                remaining_versions.sort(key=lambda x: x['created_at'], reverse=True)
                                new_most_recent = remaining_versions[0]
                            
                                # Update the main listbox display
                                base_name = extract_base_filename(new_most_recent['filename'])
                                count = len(remaining_versions)
                                created_str = new_most_recent['created_at'].strftime('%Y-%m-%d %H:%M')
                            
                                if count > 1:
                                    display_text = f"{base_name} ({count} versions, latest: {created_str})"
                                else:
                                    display_text = f"{base_name} ({created_str})"
                            
                                # Update the item in main listbox
                                file_listbox.delete(main_listbox_index)
                                file_listbox.insert(main_listbox_index, display_text)
                                listbox_to_file_mapping[main_listbox_index] = new_most_recent
                            
                                debug_print(f"DEBUG: Updated main listbox after deleting most recent version")
                            else:
                                # All versions deleted, remove from main listbox
                                file_listbox.delete(main_listbox_index)
                                if main_listbox_index in listbox_to_file_mapping:
                                    del listbox_to_file_mapping[main_listbox_index]
                                debug_print(f"DEBUG: Removed entry from main listbox - all versions deleted")
                    
                        # Close the version history dialog and refresh
                        history_dialog.destroy()
                        update_selection_info()
                    else:
                        messagebox.showerror("Error", "Failed to delete any versions.")
            
                except Exception as e:
                    debug_print(f"DEBUG: Error during version deletion: {e}")
                    messagebox.showerror("Error", f"Error during deletion: {e}")

            # Add delete button on the left, close on the right
            Button(bottom_frame_hist, text="Delete Selected", command=on_delete_versions, 
                   bg="#f44336", fg="white", font=FONT).pack(side="left", padx=5)
        
            Button(bottom_frame_hist, text="Close", command=on_close, 
                   bg="#f44336", fg="black", font=FONT).pack(side="right", padx=5)

            # Center the window and then show it (prevents stutter)
            self.center_window(history_dialog, 600, 400)
            history_dialog.deiconify()

        def on_delete():
            selected_items = file_listbox.curselection()
            if not selected_items:
                messagebox.showwarning("Warning", "Please select at least one file to delete.")
                return

            file_ids = []
            filenames = []
            for idx in selected_items:
                if idx in listbox_to_file_mapping:
                    file_record = listbox_to_file_mapping[idx]
                    file_ids.append(file_record['id'])
                    filenames.append(file_record['filename'])

            if not file_ids:
                messagebox.showwarning("Warning", "No valid files selected for deletion.")
                return

            # Confirm deletion
            if len(file_ids) == 1:
                confirm_msg = f"Are you sure you want to delete the file:\n\n{filenames[0]}\n\nThis action cannot be undone."
            else:
                confirm_msg = f"Are you sure you want to delete {len(file_ids)} files?\n\nThis action cannot be undone."

            if not messagebox.askyesno("Confirm Deletion", confirm_msg):
                return

            try:
                debug_print(f"DEBUG: Starting deletion of {len(file_ids)} files")
                success_count, error_count = self.db_manager.delete_multiple_files(file_ids)
                
                if success_count > 0:
                    if error_count == 0:
                        messagebox.showinfo("Success", f"Successfully deleted {success_count} file(s).")
                    else:
                        messagebox.showwarning("Partial Success", f"Deleted {success_count} file(s), but {error_count} failed.")
                    
                    # Refresh the file list after deletion
                    refresh_file_list()
                else:
                    messagebox.showerror("Error", "Failed to delete any files.")
            
            except Exception as e:
                debug_print(f"DEBUG: Error during file deletion: {e}")
                messagebox.showerror("Error", f"Error during deletion: {e}")

        def refresh_file_list():
            """Refresh the file list display after deletion."""
            try:
                debug_print("DEBUG: Refreshing file list after deletion")
                
                # Clear current listbox
                file_listbox.delete(0, tk.END)
                listbox_to_file_mapping.clear()
                file_groups.clear()
                
                # Reload files from database
                files = self.db_manager.list_files()
                debug_print(f"DEBUG: Reloaded {len(files)} files from database")
                
                # Regroup files
                for file_record in files:
                    base_name = extract_base_filename(file_record['filename'])
                    if base_name not in file_groups:
                        file_groups[base_name] = []
                    file_groups[base_name].append(file_record)
                
                # Sort each group by creation date (newest first)
                for base_name in file_groups:
                    file_groups[base_name].sort(key=lambda x: x['created_at'], reverse=True)
                
                # Repopulate listbox
                populate_listbox()
                
                update_selection_info()
                debug_print("DEBUG: File list refresh completed")
                
            except Exception as e:
                debug_print(f"DEBUG: Error refreshing file list: {e}")
                messagebox.showerror("Error", f"Error refreshing file list: {e}")

        def on_double_click(event):
            """Handle double-click to show version history for selected files."""
            selected_items = list(file_listbox.curselection())
            if not selected_items:
                return
    
            # Process each selected file one by one
            def process_next_file(file_indices):
                if not file_indices:
                    return  # All done
        
                idx = file_indices[0]
                remaining = file_indices[1:]
        
                if idx not in listbox_to_file_mapping:
                    # Skip this one and process next
                    if remaining:
                        dialog.after(100, lambda: process_next_file(remaining))
                    return
            
                selected_file = listbox_to_file_mapping[idx]
                base_name = extract_base_filename(selected_file['filename'])
        
                # Check if this file has multiple versions
                if base_name not in file_groups or len(file_groups[base_name]) <= 1:
                    # Single version - skip to next file
                    if remaining:
                        dialog.after(100, lambda: process_next_file(remaining))
                    return
        
                # Show version history dialog
                # We need to ensure the dialog is modal and wait for it to close before proceeding
                def show_dialog_and_continue():
                    show_version_history_dialog(base_name, file_groups[base_name], idx)
                    # After dialog closes, process next file
                    if remaining:
                        dialog.after(100, lambda: process_next_file(remaining))
        
                show_dialog_and_continue()
    
            # Start processing the first file
            process_next_file(selected_items)

        def on_selection_change(event):
            """Restore original versions for deselected items."""
            if hasattr(dialog, 'original_mappings'):
                current_selection = set(file_listbox.curselection())
        
                # Check each item that has been temporarily replaced
                items_to_restore = []
                for idx, original_version in dialog.original_mappings.items():
                    if idx not in current_selection:
                        items_to_restore.append((idx, original_version))
        
                # Restore items that were deselected
                for idx, original_version in items_to_restore:
                    base_name = extract_base_filename(original_version['filename'])
                    count = len(file_groups[base_name])
                    created_str = original_version['created_at'].strftime('%Y-%m-%d %H:%M')
            
                    if count > 1:
                        display_text = f"{base_name} ({count} versions, latest: {created_str})"
                    else:
                        display_text = f"{base_name} ({created_str})"
            
                    # Restore display and mapping
                    file_listbox.delete(idx)
                    file_listbox.insert(idx, display_text)
                    listbox_to_file_mapping[idx] = original_version
            
                    debug_print(f"DEBUG: Restored {base_name} to latest version (deselected)")
        
                # Remove restored items from tracking
                for idx, _ in items_to_restore:
                    del dialog.original_mappings[idx]
    
            update_selection_info()

        def select_all():
            file_listbox.select_set(0, tk.END)
            update_selection_info()

        def select_none():
            file_listbox.selection_clear(0, tk.END)
            update_selection_info()

        # Define load functions based on mode
        if comparison_mode:
            def on_load():
                selected_items = file_listbox.curselection()
                if len(selected_items) < 2:
                    messagebox.showwarning("Warning", "Please select at least 2 files for comparison.")
                    return

                file_ids = []
                for idx in selected_items:
                    if idx in listbox_to_file_mapping:
                        file_record = listbox_to_file_mapping[idx]
                        file_ids.append(file_record['id'])

                dialog.destroy()
                if len(file_ids) >= 2:
                    original_all_filtered_sheets = self.gui.all_filtered_sheets.copy()
                    self.load_multiple_from_database(file_ids)
                    if len(self.gui.all_filtered_sheets) >= 2:
                        from sample_comparison import SampleComparisonWindow
                        comparison_window = SampleComparisonWindow(self.gui, self.gui.all_filtered_sheets)
                        comparison_window.show()
                    else:
                        messagebox.showwarning("Warning", "Failed to load enough files for comparison.")
                        self.gui.all_filtered_sheets = original_all_filtered_sheets
        else:
            def on_load():
                selected_items = file_listbox.curselection()
                if not selected_items:
                    messagebox.showwarning("Warning", "Please select at least one file to load.")
                    return

                file_ids = []
                for idx in selected_items:
                    if idx in listbox_to_file_mapping:
                        file_record = listbox_to_file_mapping[idx]
                        file_ids.append(file_record['id'])

                dialog.destroy()
                if len(file_ids) == 1:
                    self.load_from_database(file_ids[0])
                elif len(file_ids) > 1:
                    self.load_multiple_from_database(file_ids)

        # Bind events
        file_listbox.bind("<Double-Button-1>", on_double_click)
        file_listbox.bind("<<ListboxSelect>>", on_selection_change)

        # Create buttons
        button_frame = Frame(bottom_frame)
        button_frame.pack(fill="x", pady=10)

        # Add select all/none buttons
        select_frame = Frame(button_frame)
        select_frame.pack(fill="x", pady=(0, 10))

        Button(select_frame, text="Select All", command=select_all, font=FONT).pack(side="left", padx=5)
        Button(select_frame, text="Select None", command=select_none, font=FONT).pack(side="left", padx=5)

        # Add info label
        info_label = Label(select_frame, text="Double-click files to view version history", 
                          font=("Arial", 9), fg="gray")
        info_label.pack(side="right", padx=5)

        # Create main action buttons
        Button(button_frame, text="Load Selected", command=on_load, 
               bg="#4CAF50", fg="black", font=FONT).pack(side="left", padx=5)
    
        # Add delete button in the middle
        Button(button_frame, text="Delete Selected", command=on_delete, 
               bg="#f44336", fg="white", font=FONT).pack(side="left", padx=20)
        
        Button(button_frame, text="Close", command=dialog.destroy, 
               bg="#666666", fg="black", font=FONT).pack(side="right", padx=5)

        # Initialize selection info
        update_selection_info()

    def ensure_main_gui_initialized(self):
        """Ensure the main GUI components are properly initialized before displaying data."""
        debug_print("DEBUG: Ensuring main GUI is properly initialized")
    
        try:
            # Force the main window to update and initialize
            self.gui.root.update_idletasks()
        
            # Ensure basic frames exist
            if not hasattr(self.gui, 'top_frame') or not self.gui.top_frame:
                debug_print("DEBUG: Creating missing top_frame")
                self.gui.create_static_frames()
        
            # Ensure file dropdown exists
            if not hasattr(self.gui, 'file_dropdown') or not self.gui.file_dropdown:
                debug_print("DEBUG: Creating missing file dropdown")
                self.add_or_update_file_dropdown()
        
            # Ensure sheet dropdown exists
            if not hasattr(self.gui, 'drop_down_menu') or not self.gui.drop_down_menu:
                debug_print("DEBUG: Creating missing sheet dropdown")
                self.gui.populate_or_update_sheet_dropdown()
        
            # Ensure display frame exists
            if not hasattr(self.gui, 'display_frame') or not self.gui.display_frame:
                debug_print("DEBUG: Creating missing display_frame")
                self.gui.create_static_frames()
        
            # Force another update
            self.gui.root.update_idletasks()
        
            debug_print("DEBUG: Main GUI initialization check complete")
        
        except Exception as e:
            debug_print(f"ERROR: Failed to ensure main GUI initialization: {e}")
            import traceback
            traceback.print_exc()

    def set_active_file(self, file_name: str) -> None:
        """Set the active file based on the given file name."""
        for file_data in self.gui.all_filtered_sheets:
            if file_data["file_name"] == file_name:
                self.gui.current_file = file_name
                self.gui.file_path = file_data.get("file_path", None)
                self.gui.filtered_sheets = file_data["filtered_sheets"]
                if self.gui.file_path is None:
                    raise ValueError(f"No file path associated with the file '{file_name}'.")
                break
        else:
            raise ValueError(f"File '{file_name}' not found.")

    def update_ui_for_current_file(self) -> None:
        """Update UI components to reflect the currently active file."""
        if not self.gui.current_file:
            return
        
        #debug_print(f"DEBUG: Updating UI for current file: {self.gui.current_file}")
    
        # Ensure main GUI is properly initialized first
        self.ensure_main_gui_initialized()
    
        # Update file dropdown
        if hasattr(self.gui, 'file_dropdown_var'):
            self.gui.file_dropdown_var.set(self.gui.current_file)
            #debug_print(f"DEBUG: Set file dropdown to: {self.gui.current_file}")
    
        # Update sheet dropdown
        self.gui.populate_or_update_sheet_dropdown()
        #debug_print("DEBUG: Updated sheet dropdown")
    
        # Update displayed sheet
        current_sheet = self.gui.selected_sheet.get()
        if current_sheet not in self.gui.filtered_sheets:
            first_sheet = list(self.gui.filtered_sheets.keys())[0] if self.gui.filtered_sheets else None
            if first_sheet:
                self.gui.selected_sheet.set(first_sheet)
                #debug_print(f"DEBUG: Set selected sheet to first sheet: {first_sheet}")
                # Add a small delay to ensure UI is ready
                self.gui.root.after(100, lambda: self.gui.update_displayed_sheet(first_sheet))
            else:
                debug_print("ERROR: No sheets available to display")
        else:
            debug_print(f"DEBUG: Using current sheet: {current_sheet}")
            # Add a small delay to ensure UI is ready
            self.gui.root.after(100, lambda: self.gui.update_displayed_sheet(current_sheet))
    
        #debug_print("DEBUG: UI update for current file complete")

    def add_data(self) -> None:
        """Handle adding a new data file directly and update UI accordingly."""
        file_paths = filedialog.askopenfilenames(
            title="Select Excel File(s)", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_paths:
            show_success_message("Info", "No files were selected", self.gui.root)
            return

        for file_path in file_paths:
            self.load_excel_file(file_path)
            self.gui.all_filtered_sheets.append({
                "file_name": os.path.basename(file_path),
                "file_path": file_path,
                "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
            })
        self.update_file_dropdown()
        last_file = self.gui.all_filtered_sheets[-1]
        self.set_active_file(last_file["file_name"])
        self.update_ui_for_current_file()
        show_success_message("Success", f"Data from {len(file_paths)} file(s) added successfully.", self.gui.root)

    def ask_open_file(self) -> str:
        """Prompt the user to select an Excel file."""
        return filedialog.askopenfilename(
            title="Select Standardized Testing File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

    def reload_excel_file(self) -> None:
        """Reload the Excel file into the program, preserving the state of the UI."""
        try:
            self.load_excel_file(self.gui.file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reload the Excel file: {e}")

    def open_raw_data_in_excel(self, sheet_name=None) -> None:
        """Opens an Excel file for a specific sheet, allowing edits, and then updates the application when Excel closes."""
        try:
            # Validate the current state
            if not hasattr(self.gui, 'filtered_sheets') or not self.gui.filtered_sheets:
                messagebox.showerror("Error", "No data is currently loaded.")
                return
        
            # If no sheet specified, use the currently selected sheet
            if not sheet_name and hasattr(self.gui, 'selected_sheet'):
                sheet_name = self.gui.selected_sheet.get()
        
            if not sheet_name or sheet_name not in self.gui.filtered_sheets:
                messagebox.showerror("Error", "Please select a valid sheet first.")
                return
        
            # Get the sheet data
            sheet_data = self.gui.filtered_sheets[sheet_name]['data']
        
            # Create a temporary Excel file
            import tempfile
            import uuid
            from openpyxl import Workbook
        
            # Generate a unique identifier
            unique_id = str(uuid.uuid4()).split('-')[0]
        
            # Ensure the sheet name only contains valid characters
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c == ' ')
            safe_sheet_name = safe_sheet_name.replace(' ', '_')[:15]
        
            # Create a temporary file
            temp_dir = os.path.abspath(tempfile.gettempdir())
            temp_file = os.path.join(temp_dir, f"dataviewer_{safe_sheet_name}_{unique_id}.xlsx")
        
            # Create a new workbook with just this sheet
            wb = Workbook()
            ws = wb.active
            ws.title = safe_sheet_name[:31]
        
            # Write the headers
            for col_idx, column_name in enumerate(sheet_data.columns, 1):
                ws.cell(row=1, column=col_idx, value=str(column_name))
        
            # Write the data
            for row_idx, row in enumerate(sheet_data.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
            # Save the workbook
            wb.save(temp_file)
        
            if not os.path.exists(temp_file):
                raise FileNotFoundError(f"Failed to create temporary file at {temp_file}")
        
            # Record the file modification time before opening
            original_mod_time = os.path.getmtime(temp_file)
        
            # Create status notification
            status_text = f"Opening {sheet_name} in Excel. Changes will be imported when Excel closes."
            status_label = ttk.Label(self.gui.root, text=status_text, relief="sunken", anchor="w")
            status_label.pack(side="bottom", fill="x")
            self.gui.root.update_idletasks()
        
            # File monitor state
            class FileMonitorState:
                def __init__(self, initial_mod_time):
                    self.last_mod_time = initial_mod_time
                    self.has_changed = False
        
            monitor_state = FileMonitorState(original_mod_time)
        
            # Force a new Excel instance
            import subprocess
            cmd = f'start /wait "" "excel.exe" /x "{os.path.abspath(temp_file)}"'
        
            try:
                subprocess.Popen(cmd, shell=True)
            except Exception as e:
                debug_print(f"Error launching Excel with command: {e}")
                os.startfile(os.path.abspath(temp_file))
            
            time.sleep(2.0)
        
            # Monitor file in background thread
            def monitor_file_lock():
                file_open = True
                while file_open:
                    file_locked = False
                    try:
                        with open(temp_file, 'r+b') as test_lock:
                            pass
                    except PermissionError:
                        file_locked = True
                    except FileNotFoundError:
                        file_locked = False
                        file_open = False
                    except Exception:
                        file_locked = False
                
                    file_open = file_locked
                
                    if os.path.exists(temp_file):
                        try:
                            current_mod_time = os.path.getmtime(temp_file)
                            if current_mod_time > monitor_state.last_mod_time:
                                monitor_state.has_changed = True
                                monitor_state.last_mod_time = current_mod_time
                        except Exception:
                            pass
                            
                    if not file_open:
                        self.gui.root.after(500, lambda: self._process_excel_changes(temp_file, sheet_name, monitor_state.has_changed, status_label))
                        break
                
                    time.sleep(1.0)
        
            # Start monitoring thread
            monitor_thread = threading.Thread(target=monitor_file_lock, daemon=True)
            self.gui.threads.append(monitor_thread)
            monitor_thread.start()
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {e}")
            traceback.print_exc()

    def _process_excel_changes(self, temp_file, sheet_name, file_changed, status_label=None):
        """Process changes made in Excel after it closes."""
        try:
            if status_label and status_label.winfo_exists():
                status_label.destroy()
        
            if file_changed and os.path.exists(temp_file):
                try:
                    max_retries = 3
                    retry_count = 0
                    read_success = False
                    modified_data = None
                
                    while not read_success and retry_count < max_retries:
                        try:
                            modified_data = pd.read_excel(temp_file)
                            read_success = True
                        except Exception:
                            retry_count += 1
                            time.sleep(1.0)
                
                    if not read_success:
                        raise Exception(f"Failed to read Excel file after {max_retries} attempts")
                
                    # Update the filtered sheets with new data
                    self.gui.filtered_sheets[sheet_name]['data'] = modified_data
                
                    # If using VAP3 file, update it
                    if hasattr(self.gui, 'file_path') and self.gui.file_path.endswith('.vap3'):
                        from vap_file_manager import VapFileManager
                        vap_manager = VapFileManager()
                        vap_manager.save_to_vap3(
                            self.gui.file_path,
                            self.gui.filtered_sheets,
                            self.gui.sheet_images,
                            self.gui.plot_options,
                            getattr(self.gui, 'image_crop_states', {})
                        )
                
                    # Refresh the GUI display
                    self.gui.update_displayed_sheet(sheet_name)
                
                    # Show success message
                    success_label = ttk.Label(
                        self.gui.root, 
                        text="Changes from Excel have been imported successfully.",
                        relief="sunken", 
                        anchor="w"
                    )
                    success_label.pack(side="bottom", fill="x")
                    self.gui.root.after(3000, lambda: success_label.destroy() if success_label.winfo_exists() else None)
                
                except Exception as e:
                    traceback.print_exc()
                    messagebox.showerror("Error", f"Failed to read modified Excel file: {e}")
            elif file_changed and not os.path.exists(temp_file):
                show_success_message("Information", "Excel file was modified but appears to have been moved or renamed. Changes could not be imported.", self.gui.root)
            else:
                info_label = ttk.Label(
                    self.gui.root, 
                    text="Excel file was closed without changes.",
                    relief="sunken", 
                    anchor="w"
                )
                info_label.pack(side="bottom", fill="x")
                self.gui.root.after(3000, lambda: info_label.destroy() if info_label.winfo_exists() else None)
            
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to process Excel changes: {e}")
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass

    def create_new_template(self, parent_window=None):
        """Create a new file from the template."""
        return self.create_new_file_with_tests(parent_window)

    def create_new_file_with_tests(self, parent_window=None):
        """Create a new file with only the selected tests."""
        # Get save path before showing test selection dialog
        file_path = get_save_path(".xlsx")
        if not file_path:
            return None

        # Get template path
        template_path = get_resource_path("resources/Standardized Test Template - LATEST VERSION - 2025 Jan.xlsx")
    
        # Load template to get available tests
        wb = openpyxl.load_workbook(template_path)
        available_tests = [sheet for sheet in wb.sheetnames if sheet != "Sheet1"]
    
        # Show test selection dialog
        test_dialog = TestSelectionDialog(self.gui.root, available_tests)
        ok_clicked, selected_tests = test_dialog.show()
    
        if not ok_clicked or not selected_tests:
            return None
    
        # Create a new file with selected tests
        try:
            shutil.copy(template_path, file_path)
            new_wb = openpyxl.load_workbook(file_path)
        
            # Remove unselected tests
            for sheet_name in list(new_wb.sheetnames):
                if sheet_name not in selected_tests and sheet_name != "Sheet1":
                    del new_wb[sheet_name]
        
            new_wb.save(file_path)
        
            # Close parent window if provided
            if parent_window and parent_window.winfo_exists():
                parent_window.destroy()
        
            # Show test start menu
            self.show_test_start_menu(file_path)
        
            return file_path
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while creating the file: {str(e)}")
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
            return None

    def show_test_start_menu(self, file_path, original_filename = None):
        """
        Show the test start menu for a file.
        If file is already loaded, try to extract existing header data and skip the dialog.
        """
        debug_print(f"DEBUG: Showing test start menu for file: {file_path}")
            
        # Store the original filename for later use
        if original_filename:
            debug_print(f"DEBUG: Storing original filename for data collection: {original_filename}")
            # You can store this in a class attribute or pass it through the flow
            self.current_original_filename = original_filename
        else:
            self.current_original_filename = None

        # For .vap3 files or files that don't exist, use loaded sheet data instead
        available_tests = None
        if file_path.endswith('.vap3') or not os.path.exists(file_path):
            # Use the currently loaded sheets instead of trying to read the file
            if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                available_tests = list(self.gui.filtered_sheets.keys())
                debug_print(f"DEBUG: Using loaded sheet names for .vap3 file: {available_tests}")
            else:
                debug_print("ERROR: No loaded sheets available for .vap3 file")
                messagebox.showerror("Error", "No sheet data is currently loaded. Please load a file first.")
                return
    
        # Show the test start menu with available tests
        start_menu = TestStartMenu(self.gui.root, file_path, available_tests)
        result, selected_test = start_menu.show()

        if result == "start_test" and selected_test:
            debug_print(f"DEBUG: Starting test: {selected_test}")
        
            # For existing files, always try to extract header data first
            debug_print("DEBUG: Attempting to extract existing header data from file/loaded data")
        
            # Try to extract existing header data
            existing_header_data = self.extract_existing_header_data(file_path, selected_test)
        
            if existing_header_data and self.validate_header_data(existing_header_data):
                debug_print("DEBUG: Found complete header data - skipping header dialog and going directly to data collection")
                # Go directly to data collection with existing header data
                self.start_data_collection_with_header_data(file_path, selected_test, existing_header_data)
            else:
                debug_print("DEBUG: Header data incomplete or missing - showing header dialog")
                # Only show header data dialog if we couldn't find valid header data
                self.show_header_data_dialog(file_path, selected_test)
            
        elif result == "view_raw_file":
            debug_print("DEBUG: Viewing raw file requested")
            # Load the file for viewing if not already loaded
            if not hasattr(self.gui, 'file_path') or self.gui.file_path != file_path:
                self.load_file(file_path)
        else:
            debug_print("DEBUG: Test start menu was cancelled or closed")

    def extract_existing_header_data(self, file_path, selected_test):
        """Extract existing header data from Excel files or loaded .vap3 data."""
        debug_print(f"DEBUG: Extracting header data from {file_path} for test {selected_test}")
    
        try:
            # Check if this is a .vap3 file or temporary file that doesn't exist
            if file_path.endswith('.vap3') or not os.path.exists(file_path):
                debug_print("DEBUG: Detected .vap3 file or non-existent file, extracting from loaded data")
                return self.extract_header_data_from_loaded_sheets(selected_test)
            else:
                debug_print("DEBUG: Detected Excel file, extracting from file using openpyxl")
                return self.extract_header_data_from_excel_file(file_path, selected_test)
            
        except Exception as e:
            debug_print(f"ERROR: Failed to extract header data: {e}")
            import traceback
            traceback.print_exc()
            return None

    def extract_header_data_from_loaded_sheets(self, selected_test):
        """Extract header data from already-loaded sheet data (for .vap3 files)."""
        debug_print(f"DEBUG: Extracting header data from loaded sheets for test: {selected_test}")

        try:
            # Check if the sheet is loaded
            if not hasattr(self.gui, 'filtered_sheets') or selected_test not in self.gui.filtered_sheets:
                print(f"ERROR: Sheet {selected_test} not found in loaded data")
                return None

            sheet_info = self.gui.filtered_sheets[selected_test]
            debug_print(f"DEBUG: sheet_info keys: {list(sheet_info.keys())}")

            # FOR .VAP3 FILES: Use the stored header_data JSON directly
            if 'header_data' in sheet_info:
                debug_print(f"DEBUG: header_data key exists, value type: {type(sheet_info['header_data'])}")
                debug_print(f"DEBUG: header_data value: {sheet_info['header_data']}")

                if sheet_info['header_data'] is not None:
                    debug_print("DEBUG: Found stored header_data in loaded sheets - using directly")
                    header_data = sheet_info['header_data']

                    # Ensure backwards compatibility for device_type
                    if 'common' in header_data:
                        if 'device_type' not in header_data['common']:
                            header_data['common']['device_type'] = 'T58G'
                            debug_print("DEBUG: Added default device_type T58G for backwards compatibility")

                    # Validate structure
                    debug_print(f"DEBUG: About to validate header_data: {header_data}")
                    if self.validate_header_data(header_data):
                        debug_print(f"DEBUG: Successfully extracted header data from .vap3 for {selected_test}")
                        return header_data
                    else:
                        debug_print("DEBUG: Stored header data failed validation, falling back to extraction")
                else:
                    debug_print("DEBUG: header_data key exists but value is None")
            else:
                debug_print("DEBUG: No header_data key found in sheet_info")

            # FALLBACK: Try to extract from sheet data (for backwards compatibility or corrupted header data)
            print("DEBUG: No valid stored header_data found, attempting extraction from sheet data")
            sheet_data = sheet_info['data']

            if sheet_data.empty:
                debug_print("DEBUG: Sheet data is empty")
                return None

            debug_print(f"DEBUG: Sheet data shape: {sheet_data.shape}")
            debug_print(f"DEBUG: Column headers: {list(sheet_data.columns[:20])}")  # Show first 20 column headers

            # Determine sample count from data structure
            sample_count = self.determine_sample_count_from_data(sheet_data, selected_test)
            debug_print(f"DEBUG: Determined sample count: {sample_count}")

            # Extract header data using simplified structure - only tester and device_type are common
            header_data = {
                'common': {
                    'tester': '',
                    'device_type': 'T58G'
                },
                'samples': [],
                'test': selected_test,
                'num_samples': sample_count
            }

            try:              
                # Extract tester from a common location (row 1, column 3)
                if len(sheet_data) > 2 and len(sheet_data.columns) > 3:
                    tester_value = sheet_data.iloc[1, 3] if not pd.isna(sheet_data.iloc[1, 3]) else ""
                    header_data['common']['tester'] = str(tester_value).strip()
                    debug_print(f"DEBUG: Extracted tester: '{header_data['common']['tester']}'")

                # Extract sample-specific data for each sample - each sample gets its own values
                for i in range(sample_count):
                    # Determine column offset based on test type
                    if selected_test in ["User Test Simulation", "User Simulation Test"]:
                        col_offset = 1 + (i * 8)  # User simulation: 8 columns per sample
                    else:
                        col_offset = 5 + (i * 12)  # Standard: 12 columns per sample, starting after 5 header columns

                    sample_data = {
                        'id': f'Sample {i+1}',
                        'resistance': '',
                        'media': '',
                        'viscosity': '',
                        'voltage': '',
                        'puffing_regime': '60mL/3s/30s',
                        'oil_mass': ''
                    }

                    debug_print(f"DEBUG: Processing sample {i+1} with column offset {col_offset}")

                    try:
                        # Sample ID - look in the header row (row 0) at the sample's column offset
                        if col_offset < len(sheet_data.columns):
                            sample_id_value = sheet_data.columns[col_offset] if col_offset < len(sheet_data.columns) else ""
                            if pd.notna(sample_id_value) and str(sample_id_value).strip() and not str(sample_id_value).startswith('Unnamed'):
                                sample_data['id'] = str(sample_id_value).strip()
                            else:
                                # Try looking in the actual data rows for sample ID
                                for row_idx in range(min(3, len(sheet_data))):
                                    cell_value = sheet_data.iloc[row_idx, col_offset] if col_offset < len(sheet_data.columns) else ""
                                    if pd.notna(cell_value) and str(cell_value).strip():
                                        sample_data['id'] = str(cell_value).strip()
                                        break

                        # Extract sample-specific data from this sample's column block
                        sample_start_col = col_offset - 5  # Go back to start of this sample's block
                    
                        # Resistance - typically at sample_start_col + 3 (row 0)
                        if sample_start_col + 3 < len(sheet_data.columns) and len(sheet_data) > 0:
                            resistance = sheet_data.iloc[0, sample_start_col + 3]
                            if pd.notna(resistance) and str(resistance).strip():
                                sample_data['resistance'] = str(resistance).strip()
                                debug_print(f"DEBUG: Sample {i+1} resistance: '{sample_data['resistance']}'")

                        # Media - typically at sample_start_col + 1 (row 0)
                        if sample_start_col + 1 < len(sheet_data.columns) and len(sheet_data) > 0:
                            media = sheet_data.iloc[0, sample_start_col + 1]
                            if pd.notna(media) and str(media).strip():
                                sample_data['media'] = str(media).strip()
                                debug_print(f"DEBUG: Sample {i+1} media: '{sample_data['media']}'")

                        # Viscosity - typically at sample_start_col + 1 (row 1)
                        if sample_start_col + 1 < len(sheet_data.columns) and len(sheet_data) > 1:
                            viscosity = sheet_data.iloc[1, sample_start_col + 1]
                            if pd.notna(viscosity) and str(viscosity).strip():
                                sample_data['viscosity'] = str(viscosity).strip()
                                debug_print(f"DEBUG: Sample {i+1} viscosity: '{sample_data['viscosity']}'")

                        # Voltage - typically at sample_start_col + 5 (row 1)
                        if sample_start_col + 5 < len(sheet_data.columns) and len(sheet_data) > 1:
                            voltage = sheet_data.iloc[1, sample_start_col + 5]
                            if pd.notna(voltage) and str(voltage).strip():
                                sample_data['voltage'] = str(voltage).strip()
                                debug_print(f"DEBUG: Sample {i+1} voltage: '{sample_data['voltage']}'")

                        # Puffing regime - typically at sample_start_col + 7 (row 0)
                        if sample_start_col + 7 < len(sheet_data.columns) and len(sheet_data) > 0:
                            puffing = sheet_data.iloc[0, sample_start_col + 7]
                            if pd.notna(puffing) and str(puffing).strip():
                                sample_data['puffing_regime'] = str(puffing).strip()
                                debug_print(f"DEBUG: Sample {i+1} puffing regime: '{sample_data['puffing_regime']}'")

                        # Oil mass - typically at sample_start_col + 7 (row 1)
                        if sample_start_col + 7 < len(sheet_data.columns) and len(sheet_data) > 1:
                            oil_mass = sheet_data.iloc[1, sample_start_col + 7]
                            if pd.notna(oil_mass) and str(oil_mass).strip():
                                sample_data['oil_mass'] = str(oil_mass).strip()
                                debug_print(f"DEBUG: Sample {i+1} oil mass: '{sample_data['oil_mass']}'")
        
                        debug_print(f"DEBUG: Sample {i+1} final data: {sample_data}")
        
                    except Exception as e:
                        debug_print(f"DEBUG: Error extracting data for sample {i+1}: {e}")
        
                    # Add the sample data to the header_data
                    header_data['samples'].append(sample_data)

            except Exception as e:
                debug_print(f"DEBUG: Error extracting header data: {e}")

            debug_print(f"DEBUG: Final extracted header data: {sample_count} samples")
            debug_print(f"DEBUG: Final samples: {header_data['samples']}")
            return header_data

        except Exception as e:
            debug_print(f"ERROR: Failed to extract header data from loaded sheets: {e}")
            import traceback
            traceback.print_exc()
            return None

    def extract_header_data_from_excel_file(self, file_path, selected_test):
        """Extract header data from Excel file with enhanced old format support."""
        try:
            debug_print(f"DEBUG: Extracting header data from Excel file for test: {selected_test}")
            wb = load_workbook(file_path, read_only=True)
        
            if selected_test not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet '{selected_test}' not found in workbook")
                return None
            
            ws = wb[selected_test]
        
            # First, detect if this sheet uses old or new format
            format_type = self.detect_sheet_format(ws)
            debug_print(f"DEBUG: Detected sheet format: {format_type}")
        
            if format_type == "old":
                return self.extract_old_format_header_data(ws, selected_test)
            else:
                return self.extract_new_format_header_data(ws, selected_test)
            
        except Exception as e:
            debug_print(f"ERROR: Exception extracting header data from Excel file: {e}")
            traceback.print_exc()
            return None

    def detect_sheet_format(self, ws):
        """Detect if a sheet uses old or new template format."""
        try:
            # Look for old format indicators in first few rows
            old_format_indicators = 0
            new_format_indicators = 0
        
            for row in range(1, 6):  # Check first 5 rows
                for col in range(1, 11):  # Check first 10 columns
                    try:
                        cell_val = str(ws.cell(row=row, column=col).value or "").lower().strip()
                    
                        # Old format indicators
                        if re.search(r"project\s*:", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Project:' at row {row}, col {col}")
                        if re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Ri (Ohms)' at row {row}, col {col}")
                        if re.search(r"rf\s*\(\s*ohms?\s*\)", cell_val):
                            old_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Rf (Ohms)' at row {row}, col {col}")
                    
                        # New format indicators
                        if re.search(r"sample\s*(id|name)\s*:", cell_val):
                            new_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Sample ID/Name:' at row {row}, col {col}")
                        if re.search(r"resistance\s*\(\s*ohms?\s*\)\s*:", cell_val) and "ri" not in cell_val and "rf" not in cell_val:
                            new_format_indicators += 1
                            debug_print(f"DEBUG: Found 'Resistance (Ohms):' at row {row}, col {col}")
                        
                    except Exception:
                        continue
        
            debug_print(f"DEBUG: Format detection - Old indicators: {old_format_indicators}, New indicators: {new_format_indicators}")
        
            if old_format_indicators > new_format_indicators:
                return "old"
            elif new_format_indicators > old_format_indicators:
                return "new"
            else:
                return "new"  # Default to new format
            
        except Exception as e:
            debug_print(f"DEBUG: Error detecting sheet format: {e}")
            return "new"

    def extract_old_format_header_data(self, ws, selected_test):
        """Extract header data from old format sheet - each sample gets individual data."""
        debug_print("DEBUG: Extracting header data using old format logic")

        # Only extract tester and device_type as common data
        common_data = {
            'tester': str(ws.cell(row=1, column=5).value or ""),
            'device_type': 'T58G'
        }

        debug_print(f"DEBUG: Extracted old format common data: {common_data}")

        # Extract sample data by scanning for Project and Sample pairs
        samples = []
        sample_count = 0

        # Keep checking until we don't find any more samples
        i = 0
        while True:
            # Calculate base column for this sample (12 columns per sample)
            base_col = 6 + (i * 12)
        
            # Check if we've gone beyond the worksheet columns
            if base_col > ws.max_column:
                debug_print(f"DEBUG: Reached end of worksheet at column {base_col}")
                break
    
            # Look for Project and Sample in the header area
            project_value = None
            sample_value = None
        
            # Sample-specific data extraction from this sample's column block
            sample_data = {
                'id': f'Sample {i+1}',  # Default, will be overwritten if found
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }
    
            # Search in the first few rows for Project and Sample within this sample's block
            sample_found = False
            for row in range(1, 4):
                for col_offset in range(min(12, ws.max_column - base_col + 1)):  # Don't go beyond max column
                    col = base_col + col_offset
                    if col > ws.max_column:
                        break
                    
                    try:
                        cell_val = str(ws.cell(row=row, column=col).value or "").lower().strip()
                        next_cell_val = str(ws.cell(row=row, column=col+1).value or "").strip() if col+1 <= ws.max_column else ""
                
                        if re.search(r"project\s*:", cell_val) and next_cell_val:
                            project_value = next_cell_val
                            sample_found = True
                            debug_print(f"DEBUG: Found project '{project_value}' at row {row}, col {col}")
                        elif re.search(r"sample\s*:", cell_val) and next_cell_val:
                            sample_value = next_cell_val
                            sample_found = True
                            debug_print(f"DEBUG: Found sample '{sample_value}' at row {row}, col {col}")
                        elif re.search(r"ri\s*\(\s*ohms?\s*\)", cell_val) and next_cell_val:
                            sample_data['resistance'] = next_cell_val
                            debug_print(f"DEBUG: Found resistance '{next_cell_val}' at row {row}, col {col}")
                    except Exception:
                        continue
        
            # If we found a sample, extract all the individual data for this sample
            if sample_found:
                # Create sample ID from project and/or sample values
                if project_value and sample_value:
                    sample_data['id'] = f"{project_value} {sample_value}".strip()
                elif project_value:
                    sample_data['id'] = project_value.strip()
                elif sample_value:
                    sample_data['id'] = sample_value.strip()
            
                # Extract individual sample data from this sample's column block
                sample_base = i * 12  # 0, 12, 24, 36, etc.
            
                # Media - typically at column 2 + sample_base
                if 2 + sample_base <= ws.max_column:
                    media_val = str(ws.cell(row=2, column=2 + sample_base).value or "").strip()
                    if media_val:
                        sample_data['media'] = media_val
                        debug_print(f"DEBUG: Sample {i+1} media: '{media_val}'")
            
                # Viscosity - typically at column 2 + sample_base
                if 2 + sample_base <= ws.max_column:
                    viscosity_val = str(ws.cell(row=3, column=2 + sample_base).value or "").strip()
                    if viscosity_val:
                        sample_data['viscosity'] = viscosity_val
                        debug_print(f"DEBUG: Sample {i+1} viscosity: '{viscosity_val}'")
            
                # Voltage - look for voltage in this sample's area
                if 5 + sample_base <= ws.max_column:
                    voltage_val = str(ws.cell(row=3, column=6 + sample_base).value or "").strip()
                    if voltage_val:
                        sample_data['voltage'] = voltage_val
                        debug_print(f"DEBUG: Sample {i+1} voltage: '{voltage_val}'")
            
                # Puffing regime - typically at column 8 + sample_base
                if 8 + sample_base <= ws.max_column:
                    puffing_val = str(ws.cell(row=2, column=8 + sample_base).value or "").strip()
                    if puffing_val:
                        sample_data['puffing_regime'] = puffing_val
                        debug_print(f"DEBUG: Sample {i+1} puffing regime: '{puffing_val}'")
            
                # Oil mass - look for oil mass in this sample's area  
                if 8 + sample_base <= ws.max_column:
                    oil_mass_val = str(ws.cell(row=3, column=8 + sample_base).value or "").strip()
                    if oil_mass_val:
                        sample_data['oil_mass'] = oil_mass_val
                        debug_print(f"DEBUG: Sample {i+1} oil mass: '{oil_mass_val}'")
        
                samples.append(sample_data)
                sample_count += 1
                debug_print(f"DEBUG: Created old format sample {sample_count} with individual data: {sample_data}")
            
                # Move to next sample
                i += 1
            else:
                # No more samples found
                debug_print(f"DEBUG: No more old format samples found after checking {i+1} positions")
                break

        debug_print(f"DEBUG: Total samples found: {sample_count}")

        if sample_count == 0:
            debug_print("DEBUG: No old format samples found, using default single sample")
            samples = [{
                'id': 'Sample 1', 
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }]
            sample_count = 1

        header_data = {
            'common': common_data,
            'samples': samples,
            'test': selected_test,
            'num_samples': sample_count
        }

        debug_print(f"DEBUG: Final old format header data: {sample_count} samples")
        debug_print(f"DEBUG: Old format samples with individual data: {samples}")
        debug_print(f"DEBUG: Old format common: {common_data}")

        return header_data

    def extract_new_format_header_data(self, ws, selected_test):
        """Extract header data from new format sheet - each sample gets individual data."""
        debug_print("DEBUG: Extracting header data using new format logic")

        # Only extract tester and device_type as common data
        common_data = {
            'tester': str(ws.cell(row=2, column=4).value or ""),
            'device_type': 'T58G'
        }

        debug_print(f"DEBUG: Extracted new format common data: {common_data}")

        # Extract sample data by scanning for sample blocks
        samples = []
        sample_count = 0

        # Keep checking until we don't find any more samples
        i = 0
        while True:
            # Sample ID position: row 1, columns 6, 18, 30, 42, 54, 66, etc.
            sample_id_col = 6 + (i * 12)
        
            # Check if we've gone beyond the worksheet columns
            if sample_id_col > ws.max_column:
                debug_print(f"DEBUG: Reached end of worksheet at column {sample_id_col}")
                break
        
            # Check if there's a sample ID at this position
            sample_id_cell = ws.cell(row=1, column=sample_id_col)
    
            if sample_id_cell.value:
                sample_id = str(sample_id_cell.value).strip()
            
                # Calculate the base column for this sample's block
                base_col = i * 12
        
                # Extract ALL sample-specific data from this sample's column block
                sample_data = {
                    'id': sample_id,
                    'resistance': str(ws.cell(row=2, column=4 + base_col).value or "").strip(),
                    'media': str(ws.cell(row=2, column=2 + base_col).value or "").strip(),
                    'viscosity': str(ws.cell(row=3, column=2 + base_col).value or "").strip(),
                    'voltage': str(ws.cell(row=3, column=6 + base_col).value or "").strip(),
                    'puffing_regime': str(ws.cell(row=2, column=8 + base_col).value or "").strip(),
                    'oil_mass': str(ws.cell(row=3, column=8 + base_col).value or "").strip()
                }
        
                # Clean up the puffing regime value if it has label prefix
                if 'puffing regime:' in sample_data['puffing_regime'].lower():
                    sample_data['puffing_regime'] = sample_data['puffing_regime'].replace('Puffing Regime:', '').strip()
        
                if not sample_data['puffing_regime']:
                    sample_data['puffing_regime'] = '60mL/3s/30s'  # Default

                samples.append(sample_data)
                sample_count += 1
                debug_print(f"DEBUG: Found new format sample {sample_count} with individual data: {sample_data}")
            
                # Move to next sample
                i += 1
            else:
                # If no sample ID, we've reached the end of samples
                debug_print(f"DEBUG: No more new format samples found after checking {i+1} positions")
                break

        debug_print(f"DEBUG: Total samples found: {sample_count}")

        if sample_count == 0:
            debug_print("DEBUG: No new format samples found, using default single sample")
            samples = [{
                'id': 'Sample 1', 
                'resistance': '',
                'media': '',
                'viscosity': '',
                'voltage': '',
                'puffing_regime': '60mL/3s/30s',
                'oil_mass': ''
            }]
            sample_count = 1

        header_data = {
            'common': common_data,
            'samples': samples,
            'test': selected_test,
            'num_samples': sample_count
        }

        debug_print(f"DEBUG: Final new format header data: {sample_count} samples")
        debug_print(f"DEBUG: New format samples with individual data: {samples}")
        debug_print(f"DEBUG: New format common: {common_data}")

        return header_data

    def extract_header_data_from_excel_file_old(self, file_path, selected_test):
        """Extract header data from Excel file using openpyxl (existing method)."""
        debug_print(f"DEBUG: Extracting header data from Excel file: {file_path} for test {selected_test}")
    
        try:
            wb = openpyxl.load_workbook(file_path)
        
            if selected_test not in wb.sheetnames:
                debug_print(f"DEBUG: Sheet {selected_test} not found in file. Available sheets: {wb.sheetnames}")
                return None
            
            ws = wb[selected_test]
            debug_print(f"DEBUG: Successfully opened sheet '{selected_test}'")
        
            # Extract tester name from corrected position: row 3, column 4
            tester_cell = ws.cell(row=3, column=4)
            tester = ""
            if tester_cell.value:
                tester_value = str(tester_cell.value)
                # Remove any label prefix if present
                if "tester:" in tester_value.lower():
                    tester = tester_value.split(":", 1)[1].strip() if ":" in tester_value else ""
                else:
                    tester = tester_value.strip()
        
            debug_print(f"DEBUG: Extracted tester: '{tester}' from cell D3: '{tester_cell.value}'")
        
            # Extract common data from corrected positions
            common_data = {
                'tester': tester,
                'media': str(ws.cell(row=2, column=2).value or ""),          # Row 2, Col B
                'viscosity': str(ws.cell(row=3, column=2).value or ""),      # Row 3, Col B
                'voltage': str(ws.cell(row=3, column=6).value or ""),        # Row 3, Col F (corrected)
                'oil_mass': str(ws.cell(row=3, column=8).value or ""),       # Row 3, Col H (corrected)
                'puffing_regime': str(ws.cell(row=2, column=8).value or "Standard")  # Row 2, Col H
            }
        
            debug_print(f"DEBUG: Extracted common data with corrected positions: {common_data}")
        
            # Extract sample data by scanning the first row for sample IDs
            samples = []
            sample_count = 0
        
            # Scan the first row looking for sample blocks (starting from column 6, then every 12 columns)
            for i in range(6):  # Check up to 6 samples maximum
                # Sample ID position: row 1, columns 6, 18, 30, 42, 54, 66
                sample_id_col = 6 + (i * 12)
                resistance_col = 4 + (i * 12)  # Row 2, columns 4, 16, 28, 40, 52, 64
            
                # Check if there's a sample ID at this position
                sample_id_cell = ws.cell(row=1, column=sample_id_col)
                resistance_cell = ws.cell(row=2, column=resistance_col)
            
                if sample_id_cell.value:
                    sample_id = str(sample_id_cell.value).strip()
                    resistance = str(resistance_cell.value or "").strip()
                
                    samples.append({
                        'id': str(sample_id).strip(),
                        'resistance': str(resistance).strip() if resistance else ""
                    })
                    sample_count += 1
                    debug_print(f"DEBUG: Found valid sample {sample_count}: ID='{sample_id}', Resistance='{resistance}'")
                else:
                    # If no sample ID, we've reached the end of samples
                    debug_print(f"DEBUG: No more samples found after checking {i+1} positions")
                    break
        
            if sample_count == 0:
                debug_print("DEBUG: No samples found, using default single sample")
                samples = [{'id': 'Sample 1', 'resistance': ''}]
                sample_count = 1
        
            header_data = {
                'common': common_data,
                'samples': samples,
                'test': selected_test,
                'num_samples': sample_count
            }
        
            debug_print(f"DEBUG: Final extracted header data from Excel file: {sample_count} samples")
            debug_print(f"DEBUG: Samples: {samples}")
            debug_print(f"DEBUG: Common: {common_data}")
        
            wb.close()
            return header_data
        
        except Exception as e:
            debug_print(f"ERROR: Exception extracting header data from Excel file: {e}")
            traceback.print_exc()
            return None

    def migrate_header_data_for_backwards_compatibility(self, header_data):
        """Ensure header data has device type field for backwards compatibility."""
        if header_data and 'common' in header_data:
            if 'device_type' not in header_data['common']:
                header_data['common']['device_type'] = None
                debug_print("DEBUG: Added device_type: None for backwards compatibility")
        return header_data

    def validate_header_data(self, header_data):
        """Validate that header data has sufficient content for data collection."""
        debug_print("DEBUG: Validating extracted header data")
    
        if not header_data:
            debug_print("DEBUG: Header data is None")
            return False
        
        # Check for samples - this is the only requirement to proceed
        samples = header_data.get('samples', [])
        if not samples:
            debug_print("DEBUG: No samples found")
            return False
        
        # Check that at least one sample has an ID
        has_valid_sample = False
        for sample in samples:
            if sample.get('id', '').strip():
                has_valid_sample = True
                break
            
        if not has_valid_sample:
            debug_print("DEBUG: No samples with valid IDs found")
            return False
    
        debug_print(f"DEBUG: Found {len(samples)} samples with valid IDs - proceeding to data collection")
        debug_print("DEBUG: Header data validation passed")
        return True

    def start_data_collection_with_header_data(self, file_path, selected_test, header_data):
        """Start data collection directly with existing header data."""
        debug_print("DEBUG: Starting data collection with existing header data")
    
        # Ensure header data has all required fields for each sample
        if 'samples' in header_data:
            for i, sample in enumerate(header_data['samples']):
                # Ensure each sample has all required fields
                required_fields = ['id', 'resistance', 'media', 'viscosity', 'voltage', 'puffing_regime', 'oil_mass']
                for field in required_fields:
                    if field not in sample:
                        if field == 'id':
                            sample[field] = f"Sample {i+1}"
                        elif field == 'puffing_regime':
                            sample[field] = '60mL/3s/30s'
                        else:
                            sample[field] = ''
                        
        # Ensure common data has device_type
        if 'common' not in header_data:
            header_data['common'] = {}
        if 'device_type' not in header_data['common']:
            header_data['common']['device_type'] = 'T58G'  # Changed default to T58G
        
        debug_print(f"DEBUG: Validated header data structure with all fields: {header_data}")

        try:
            # Show the data collection window directly
            from data_collection_window import DataCollectionWindow

            # Pass the original filename to the data collection window
            original_filename = getattr(self, 'current_original_filename', None)
            data_collection = DataCollectionWindow(self.gui, file_path, selected_test, header_data, original_filename=original_filename)
            data_result = data_collection.show()

            if data_result in ["load_file", "cancel"]:
                debug_print("DEBUG: Data collection completed - data should already be updated in main GUI")
                # The data collection window should have already updated the main GUI
                # Just refresh the currently displayed sheet if needed
                if hasattr(self.gui, 'selected_sheet') and self.gui.selected_sheet.get() == selected_test:
                    self.gui.root.after(100, lambda: self.gui.update_displayed_sheet(selected_test))

        except Exception as e:
            debug_print(f"DEBUG: Error starting data collection: {e}")
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to start data collection: {e}")

    def ensure_file_is_loaded_in_ui(self, file_path):
        """Ensure the file is properly loaded in the UI without redundant processing."""
        debug_print(f"DEBUG: Ensuring file {file_path} is loaded in UI")
    
        # Special handling for .vap3 files - they're already loaded in memory
        if file_path.endswith('.vap3'):
            debug_print("DEBUG: .vap3 file detected - file should already be loaded in memory")
        
            # For .vap3 files, just verify the UI state is correct
            if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                debug_print("DEBUG: .vap3 file data is already loaded, ensuring UI is updated")
            
                # Check if file is in all_filtered_sheets
                file_name = None
                for file_data in self.gui.all_filtered_sheets:
                    if file_data.get("file_path") == file_path:
                        file_name = file_data["file_name"]
                        break
            
                if file_name:
                    debug_print(f"DEBUG: Found .vap3 file in loaded files: {file_name}")
                    try:
                        self.set_active_file(file_name)
                        self.update_ui_for_current_file()
                        return True
                    except Exception as e:
                        debug_print(f"ERROR: Failed to set active .vap3 file: {e}")
                        return False
                else:
                    debug_print("DEBUG: .vap3 file not found in loaded files list")
                    # The file data is loaded but not in the list - this is OK for database files
                    # Just update the UI with current data
                    if hasattr(self.gui, 'update_displayed_sheet') and hasattr(self.gui, 'selected_sheet'):
                        current_sheet = self.gui.selected_sheet.get()
                        if current_sheet in self.gui.filtered_sheets:
                            self.gui.update_displayed_sheet(current_sheet)
                            debug_print(f"DEBUG: Updated display for current sheet: {current_sheet}")
                            return True
                    debug_print("DEBUG: Could not update .vap3 file UI")
                    return False
            else:
                debug_print("ERROR: .vap3 file should be loaded but no data found")
                return False
    
        # Regular Excel file handling
        # Validate file path
        if not file_path or not os.path.exists(file_path):
            debug_print(f"ERROR: Invalid file path: {file_path}")
            return False
    
        # Check if file is already in the UI state
        file_name = os.path.basename(file_path)
    
        # Look for existing entry in all_filtered_sheets
        existing_entry = None
        for file_data in self.gui.all_filtered_sheets:
            if file_data["file_path"] == file_path:
                existing_entry = file_data
                break
    
        if existing_entry:
            debug_print("DEBUG: File already in UI state, just updating active file")
            try:
                self.set_active_file(existing_entry["file_name"])
                self.update_ui_for_current_file()
                return True
            except Exception as e:
                debug_print(f"ERROR: Failed to set active file: {e}")
                return False
        else:
            debug_print("DEBUG: File not in UI state, adding it")
            try:
                # Load file without database storage (skip_database_storage=True)
                self.load_excel_file(file_path, skip_database_storage=True)
            
                # Add to all_filtered_sheets
                self.gui.all_filtered_sheets.append({
                    "file_name": file_name,
                    "file_path": file_path,
                    "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets)
                })
            
                self.update_file_dropdown()
                self.set_active_file(file_name)
                self.update_ui_for_current_file()
                debug_print(f"DEBUG: Successfully added file {file_name} to UI")
                return True
            
            except Exception as e:
                debug_print(f"ERROR: Failed to load file into UI: {e}")
                traceback.print_exc()
                return False

    def show_header_data_dialog(self, file_path, selected_test):
        """Show the header data dialog for a selected test."""
        debug_print(f"DEBUG: Showing header data dialog for {selected_test}")
    
        # Try to determine sample count from existing data first
        initial_sample_count = 1
        try:
            if hasattr(self.gui, 'filtered_sheets') and selected_test in self.gui.filtered_sheets:
                sheet_data = self.gui.filtered_sheets[selected_test]['data']
                initial_sample_count = self.determine_sample_count_from_data(sheet_data, selected_test)
                debug_print(f"DEBUG: Determined {initial_sample_count} samples from existing data")
        except Exception as e:
            debug_print(f"DEBUG: Could not determine sample count from data: {e}")
    
        # Create initial header data structure with detected sample count
        initial_header_data = {
            'common': {'tester': ''},
            'samples': [
                {
                    'id': f'Sample {i+1}',
                    'resistance': '',
                    'media': '',
                    'viscosity': '',
                    'voltage': '',
                    'puffing_regime': '60mL/3s/30s',
                    'oil_mass': ''
                } for i in range(initial_sample_count)
            ],
            'test': selected_test,
            'num_samples': initial_sample_count
        }
    
        # Show the header data dialog
        header_dialog = HeaderDataDialog(self.gui.root, file_path, selected_test, 
                                       edit_mode=False, current_data=initial_header_data)
        result, header_data = header_dialog.show()

        if result:
            debug_print("DEBUG: Header data dialog completed successfully")
            # Apply header data to the file
            self.apply_header_data_to_file(file_path, header_data)
    
            debug_print("DEBUG: Proceeding to data collection window")
            # Show the data collection window
            from data_collection_window import DataCollectionWindow
            data_collection = DataCollectionWindow(self.gui, file_path, selected_test, header_data)
            data_result = data_collection.show()
    
            if data_result == "load_file":
                # Load the file for viewing
                self.load_file(file_path)
            elif data_result == "cancel":
                # User cancelled data collection, just load the file
                self.load_file(file_path)
        else:
            debug_print("DEBUG: Header data dialog was cancelled")

    def apply_header_data_to_file(self, file_path, header_data):
        """
        Apply the header data to the Excel file.
        Enhanced to correctly apply headers to all sample blocks with proper column mapping.
        """
        try:
            debug_print(f"DEBUG: Applying header data to {file_path} for {header_data['num_samples']} samples")
    
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Get the sheet for the selected test
            if header_data["test"] in wb.sheetnames:
                ws = wb[header_data["test"]]
        
                debug_print(f"DEBUG: Successfully opened sheet '{header_data['test']}'")

                # Set the test name at row 1, column 1 (this should be done once)
                ws.cell(row=1, column=1, value=header_data["test"])
                debug_print(f"DEBUG: Set test name '{header_data['test']}' at row 1, column 1")

                # Get common data once
                common_data = header_data["common"]
                samples_data = header_data.get("samples", [])

                # Apply sample-specific data for each sample block
                num_samples = header_data["num_samples"]
                for i in range(num_samples):
                    # Calculate column offset (12 columns per sample)
                    # Sample blocks start at column 1, so offsets are 0, 12, 24, 36, etc.
                    col_offset = i * 12
                    sample_data = samples_data[i] if i < len(samples_data) else {}
            
                    debug_print(f"DEBUG: Processing sample {i+1} with column offset {col_offset}")
                    debug_print(f"DEBUG: Sample {i+1}: ID='{sample_data.get('id', '')}', Resistance='{sample_data.get('resistance', '')}'")

                    # Row 1, Column F (6) + offset: Sample ID
                    sample_id = sample_data.get('id', f'Sample {i+1}')
                    ws.cell(row=1, column=6 + col_offset, value=sample_id)
                    debug_print(f"DEBUG: Set sample ID '{sample_id}' at row 1, column {6 + col_offset}")

                    # Row 2, Column D (4) + offset: Resistance  
                    resistance = sample_data.get("resistance", "")
                    if resistance:
                        try:
                            resistance_value = float(resistance)
                            ws.cell(row=2, column=4 + col_offset, value=resistance_value)
                        except ValueError:
                            ws.cell(row=2, column=4 + col_offset, value=resistance)
                        debug_print(f"DEBUG: Set resistance '{resistance}' at row 2, column {4 + col_offset}")

                    # Row 3, Column D (4) + offset: Tester name (from common data)
                    tester_name = common_data.get("tester", "")
                    if tester_name:
                        tester_col = 4 + col_offset
                        ws.cell(row=3, column=tester_col, value=tester_name)
                        debug_print(f"DEBUG: Set tester '{tester_name}' at row 3, column {tester_col} for sample {i+1}")

                    # Sample-specific data from the sample record
                    # Row 2, Column B (2) + offset: Media
                    media = sample_data.get("media", "")
                    if media:
                        media_col = 2 + col_offset
                        ws.cell(row=2, column=media_col, value=media)
                        debug_print(f"DEBUG: Set media '{media}' at row 2, column {media_col} for sample {i+1}")
            
                    # Row 3, Column B (2) + offset: Viscosity 
                    viscosity = sample_data.get("viscosity", "")
                    if viscosity:
                        viscosity_col = 2 + col_offset
                        try:
                            viscosity_value = float(viscosity)
                            ws.cell(row=3, column=viscosity_col, value=viscosity_value)
                        except ValueError:
                            ws.cell(row=3, column=viscosity_col, value=viscosity)
                        debug_print(f"DEBUG: Set viscosity '{viscosity}' at row 3, column {viscosity_col} for sample {i+1}")
            
                    # Row 3, Column F (6) + offset: Voltage  
                    voltage = sample_data.get("voltage", "")
                    if voltage:
                        try:
                            voltage_value = float(voltage)
                            ws.cell(row=3, column=6 + col_offset, value=voltage_value)
                        except ValueError:
                            ws.cell(row=3, column=6 + col_offset, value=voltage)

                    # Row 2, Column F (6) + offset: Calculated Power
                    calculated_power = sample_data.get("calculated_power", "")
                    if calculated_power:
                        try:
                            power_value = float(calculated_power)
                            ws.cell(row=2, column=6 + col_offset, value=power_value)
                            debug_print(f"DEBUG: Set calculated power '{power_value}' at row 2, column {6 + col_offset} for sample {i+1}")
                        except ValueError:
                            ws.cell(row=2, column=6 + col_offset, value=calculated_power)
                            debug_print(f"DEBUG: Set calculated power (string) '{calculated_power}' at row 2, column {6 + col_offset} for sample {i+1}")
                    else:
                        debug_print(f"DEBUG: No calculated power available for sample {i+1}")

                    # Row 3, Column H (8) + offset: Oil Mass
                    oil_mass = sample_data.get("oil_mass", "")
                    if oil_mass:
                        oil_mass_col = 8 + col_offset
                        try:
                            oil_mass_value = float(oil_mass)
                            ws.cell(row=3, column=oil_mass_col, value=oil_mass_value)
                        except ValueError:
                            ws.cell(row=3, column=oil_mass_col, value=oil_mass)
                        debug_print(f"DEBUG: Set oil mass '{oil_mass}' at row 3, column {oil_mass_col} for sample {i+1}")

                    # Row 2, Column H + offset: Puffing Regime
                    puffing_regime = sample_data.get("puffing_regime", "60mL/3s/30s")
                    if puffing_regime:
                        puffing_regime_col = 8 + col_offset
                        ws.cell(row=2, column=puffing_regime_col, value=puffing_regime)
                        debug_print(f"DEBUG: Set puffing regime '{puffing_regime}' at row 2, column {puffing_regime_col} for sample {i+1}")

                # Calculate the last column used
                last_sample_column = ((num_samples - 1) * 12) + 12
                max_column = ws.max_column
                debug_print(f"DEBUG: Last sample column: {last_sample_column}, Max column: {max_column}")
        
                # Save the workbook
                wb.save(file_path)
                debug_print(f"DEBUG: Successfully saved workbook to {file_path}")
    
                debug_print(f"SUCCESS: Applied header data for {num_samples} samples to {file_path}")
        
            else:
                error_msg = f"Sheet '{header_data['test']}' not found in the file."
                debug_print(f"ERROR: {error_msg}")
                raise Exception(error_msg)
        
        except Exception as e:
            debug_print(f"ERROR: Error applying header data: {e}")
            debug_print("DEBUG: Full traceback:")
            import traceback
            traceback.print_exc()
            raise

    def determine_sample_count_from_data(self, sheet_data, test_name):
        """Determine the number of samples based on existing data structure."""
        debug_print(f"DEBUG: Determining sample count from data for test: {test_name}")

        try:
            total_columns = len(sheet_data.columns)
            debug_print(f"DEBUG: Total columns in data: {total_columns}")
    
            # Determine columns per sample based on test type
            if test_name in ["User Test Simulation", "User Simulation Test"]:
                columns_per_sample = 8  # User simulation format
            else:
                columns_per_sample = 12  # Standard format
    
            debug_print(f"DEBUG: Using {columns_per_sample} columns per sample for test type")
    
            # Look for actual sample data by checking each potential sample position
            actual_samples = 0
        
            # Start checking from the first sample position and keep going until no more data found
            i = 0
            while True:
                sample_id_col = 5 + (i * columns_per_sample)  # Sample ID column position
            
                if sample_id_col >= total_columns:
                    debug_print(f"DEBUG: Reached end of data at column {sample_id_col}")
                    break
                
                # Check if there's meaningful data in this sample's area
                sample_has_data = False
                for row_idx in range(min(5, len(sheet_data))):
                    for col_offset in range(min(columns_per_sample, total_columns - sample_id_col)):
                        cell_value = sheet_data.iloc[row_idx, sample_id_col + col_offset]
                        if pd.notna(cell_value) and str(cell_value).strip():
                            sample_has_data = True
                            break
                    if sample_has_data:
                        break
        
                if sample_has_data:
                    actual_samples = i + 1
                    debug_print(f"DEBUG: Found data for sample {actual_samples}")
                    i += 1
                else:
                    debug_print(f"DEBUG: No data found for sample {i + 1}, stopping count")
                    break

            final_sample_count = max(1, actual_samples)
            debug_print(f"DEBUG: Final determined sample count: {final_sample_count}")
    
            return final_sample_count
    
        except Exception as e:
            debug_print(f"ERROR: Error determining sample count: {e}")
            return 1  # Default to 1 sample if we can't determine

    def start_file_loading_wrapper(self, startup_menu: tk.Toplevel) -> None:
        """Handle the 'Load' button click in the startup menu."""
        startup_menu.destroy()
        self.load_initial_file()

    def start_file_loading_database_wrapper(self, startup_menu: tk.Toplevel) -> None:
        """Handle the 'Load from Database' button click in the startup menu."""
        startup_menu.destroy()
        self.show_database_browser()

    def update_file_dropdown(self) -> None:
        """Update the file dropdown with loaded file names."""
        file_names = [file_data["file_name"] for file_data in self.gui.all_filtered_sheets]
        self.gui.file_dropdown["values"] = file_names
        if file_names:
            self.gui.file_dropdown_var.set(file_names[-1])
            self.gui.file_dropdown.update_idletasks()

    def add_or_update_file_dropdown(self) -> None:
        """Add a file selection dropdown or update its values if it already exists."""
        if not hasattr(self.gui, 'file_dropdown') or not self.gui.file_dropdown:
            dropdown_frame = ttk.Frame(self.gui.top_frame, width=1400, height=40)
            dropdown_frame.pack(side="left", pady=2, padx=5)
            file_label = ttk.Label(dropdown_frame, text="Select File:", font=FONT, background=APP_BACKGROUND_COLOR)
            file_label.pack(side="left", padx=(0, 0))
            self.gui.file_dropdown_var = tk.StringVar()
            self.gui.file_dropdown = ttk.Combobox(
                dropdown_frame,
                textvariable=self.gui.file_dropdown_var,
                state="readonly",
                font=FONT,
                width=20
            )
            self.gui.file_dropdown.pack(side="left", fill="x", expand=True, padx=(5, 5))
            self.gui.file_dropdown.bind("<<ComboboxSelected>>", self.gui.on_file_selection)
        self.update_file_dropdown()

    # ==================== .vap3 File FUNCTIONS ====================

    def save_as_vap3(self, filepath=None) -> None:
        """Save the current session to a .vap3 file."""
        from vap_file_manager import VapFileManager
    
        if not filepath:
            filepath = filedialog.asksaveasfilename(
                title="Save As VAP3",
                defaultextension=".vap3",
                filetypes=[("VAP3 files", "*.vap3")]
            )
        
        if not filepath:
            return
    
        try:
            self.gui.progress_dialog.show_progress_bar("Saving VAP3 file...")
            self.root.update_idletasks()
        
            vap_manager = VapFileManager()
        
            # Collect plot settings
            plot_settings = {
                'selected_plot_type': self.gui.selected_plot_type.get() if hasattr(self.gui, 'selected_plot_type') else None
            }
        
            # Get image crop states
            image_crop_states = getattr(self.gui, 'image_crop_states', {})
            if hasattr(self.gui, 'image_loader') and self.gui.image_loader:
                image_crop_states.update(self.gui.image_loader.image_crop_states)
        
            success = vap_manager.save_to_vap3(
                filepath,
                self.gui.filtered_sheets,
                self.gui.sheet_images,
                self.gui.plot_options,
                image_crop_states,
                plot_settings
            )
        
            self.gui.progress_dialog.update_progress_bar(100)
            self.root.update_idletasks()
        
            if success:
                show_success_message("Success", f"Data saved successfully to {filepath}", self.gui.root)
            else:
                messagebox.showerror("Error", "Failed to save data")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the file: {e}")
        finally:
            self.gui.progress_dialog.hide_progress_bar()

    def load_vap3_file(self, filepath=None, display_name=None, append_to_existing=False) -> bool:
        """Load a .vap3 file and update the application state."""
        from vap_file_manager import VapFileManager

        if not filepath:
            filepath = filedialog.askopenfilename(
                title="Open VAP3 File",
                filetypes=[("VAP3 files", "*.vap3")]
            )
    
        if not filepath:
            return False

        try:
            self.gui.progress_dialog.show_progress_bar("Loading VAP3 file...")
            self.root.update_idletasks()
    
            vap_manager = VapFileManager()
            result = vap_manager.load_from_vap3(filepath)
    
            # Use display_name if provided, otherwise use the actual filename
            if display_name:
                current_file_name = display_name
                debug_print(f"DEBUG: Using provided display name: {display_name}")
            else:
                current_file_name = os.path.basename(filepath)
                debug_print(f"DEBUG: Using actual filepath basename: {os.path.basename(filepath)}")
        
            # Check if this file is already loaded (avoid duplicates)
            existing_file = None
            for file_data in self.gui.all_filtered_sheets:
                if file_data["file_name"] == current_file_name or file_data["file_path"] == filepath:
                    existing_file = file_data
                    debug_print(f"DEBUG: File already loaded: {current_file_name}")
                    break
            
            vap_data = vap_manager.load_from_vap3(filepath)
            self.gui.load_sample_images_from_vap3(vap_data)

            if existing_file:
                # File already loaded, just make it active
                self.set_active_file(existing_file["file_name"])
                self.update_ui_for_current_file()
                debug_print(f"DEBUG: Switched to existing file: {existing_file['file_name']}")
            else:
                # New file, add it to the collection
                debug_print(f"DEBUG: Adding new file: {current_file_name}")
            
                # Update current session data (this will be the active file)
                self.gui.filtered_sheets = result['filtered_sheets']
                self.gui.sheets = {name: sheet_info['data'] for name, sheet_info in result['filtered_sheets'].items()}
            
                # Handle sheet images - need to be careful about file-specific images
                if 'sheet_images' in result and result['sheet_images']:
                    if not hasattr(self.gui, 'sheet_images'):
                        self.gui.sheet_images = {}
                    self.gui.sheet_images[current_file_name] = result['sheet_images'].get(current_file_name, {})
                    debug_print(f"DEBUG: Loaded sheet images for {current_file_name}")
            
                # Handle image crop states
                if 'image_crop_states' in result and result['image_crop_states']:
                    if not hasattr(self.gui, 'image_crop_states'):
                        self.gui.image_crop_states = {}
                    self.gui.image_crop_states.update(result['image_crop_states'])
        
                # Handle plot options and settings
                if 'plot_options' in result and result['plot_options']:
                    self.gui.plot_options = result['plot_options']
        
                if 'plot_settings' in result and result['plot_settings']:
                    if 'selected_plot_type' in result['plot_settings'] and result['plot_settings']['selected_plot_type']:
                        self.gui.selected_plot_type.set(result['plot_settings']['selected_plot_type'])
            
                # Set current file info
                self.gui.current_file = current_file_name
                self.gui.file_path = filepath
            
                # Add to all_filtered_sheets (append instead of replace if append_to_existing is True)
                new_file_data = {
                    "file_name": current_file_name,
                    "file_path": filepath,
                    "filtered_sheets": copy.deepcopy(result['filtered_sheets'])
                }
            
                if append_to_existing:
                    # Append to existing files
                    self.gui.all_filtered_sheets.append(new_file_data)
                    debug_print(f"DEBUG: Appended file to existing collection. Total files: {len(self.gui.all_filtered_sheets)}")
                else:
                    # Replace existing files (original behavior for single file loading)
                    self.gui.all_filtered_sheets = [new_file_data]
                    debug_print(f"DEBUG: Replaced file collection with single file")
            
                # Update UI
                self.update_file_dropdown()
                self.update_ui_for_current_file()
    
            self.gui.progress_dialog.update_progress_bar(100)
            self.root.update_idletasks()
    
            # Only show success message if this wasn't called from load_from_database
            if not display_name:
                show_success_message("Success", f"VAP3 file loaded successfully: {current_file_name}", self.gui.root)
            else:
                debug_print(f"DEBUG: VAP3 file loaded with display name: {current_file_name}")
            
            return True
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load VAP3 file: {str(e)}")
            return False
        finally:
            self.gui.progress_dialog.hide_progress_bar()

    def load_file(self, file_path):
        """Load a file without showing dialogs - used for data collection flow."""
        debug_print(f"DEBUG: Loading file for data collection: {file_path}")
        try:
            # Use the optimized ensure method instead of full reload
            self.ensure_file_is_loaded_in_ui(file_path)
            debug_print(f"DEBUG: File loaded successfully: {file_path}")
            return True
            
        except Exception as e:
            debug_print(f"DEBUG: Error loading file: {e}")
            messagebox.showerror("Error", f"Failed to load file: {e}")
            return False

### Batch Loading Feature ###

    def batch_load_folder(self):
        """
        Batch load Excel files from a selected folder and all subfolders.
        Only loads files whose names contain test names from the processing dictionary.
        Shows accurate progress during the operation.
        """
        debug_print("DEBUG: Starting batch folder loading process")
    
        # Get folder selection from user
        folder_path = filedialog.askdirectory(
            title="Select Folder for Batch Loading (will scan all subfolders)"
        )
    
        if not folder_path:
            debug_print("DEBUG: No folder selected, canceling batch load")
            return
    
        try:
            # Get test names from processing dictionary for filename matching
            test_names = self._get_test_names_for_matching()
            debug_print(f"DEBUG: Got {len(test_names)} test names for matching: {test_names[:5]}...")
        
            # Scan folder recursively for Excel files
            debug_print(f"DEBUG: Scanning folder: {folder_path}")
            excel_files = self._scan_folder_for_excel_files(folder_path)
            debug_print(f"DEBUG: Found {len(excel_files)} Excel files total")
        
            # Filter files based on test name matching
            matching_files = self._filter_files_by_test_names(excel_files, test_names)
            debug_print(f"DEBUG: Found {len(matching_files)} files matching test names")
        
            if not matching_files:
                messagebox.showinfo("No Matching Files", 
                                   f"No Excel files found in '{folder_path}' with test names in their filename.\n\n"
                                   f"Looking for files containing any of these test names: {', '.join(test_names[:10])}{'...' if len(test_names) > 10 else ''}")
                return
        
            # Confirm with user before proceeding
            proceed = messagebox.askyesno("Batch Load Confirmation", 
                                        f"Found {len(matching_files)} Excel files to load.\n\n"
                                        f"This may take several minutes. Proceed with batch loading?")
            if not proceed:
                debug_print("DEBUG: User cancelled batch loading")
                return
        
            # Perform batch loading with progress tracking
            self._perform_batch_loading(matching_files)
        
        except Exception as e:
            debug_print(f"ERROR: Batch folder loading failed: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Batch Loading Error", f"Failed to complete batch loading: {e}")

    def _get_test_names_for_matching(self):
        """Get test names from processing dictionary for filename matching."""
        debug_print("DEBUG: Extracting test names from processing functions")
    
        try:
            # Import processing module to get test names
            import processing
        
            # Get the processing functions dictionary
            processing_functions = processing.get_processing_functions()
            debug_print(f"DEBUG: Found {len(processing_functions)} processing functions")
        
            # Extract test names, excluding generic ones
            test_names = []
            exclude_terms = ['default', 'sheet1', 'legacy']
        
            for sheet_name in processing_functions.keys():
                if sheet_name.lower() not in exclude_terms:
                    test_names.append(sheet_name.lower())
                    # Also add variations for better matching
                    test_names.append(sheet_name.replace(" ", "").lower())
                    test_names.append(sheet_name.replace(" ", "_").lower())
                    test_names.append(sheet_name.replace(" ", "-").lower())
        
            # Remove duplicates and sort
            test_names = list(set(test_names))
            test_names.sort()
        
            debug_print(f"DEBUG: Generated {len(test_names)} test name variations for matching")
            return test_names
        
        except Exception as e:
            debug_print(f"ERROR: Failed to get test names: {e}")
            # Fallback list of common test names
            return ['test plan', 'lifetime test', 'device life test', 'quick screening', 
                    'aerosol temperature', 'user test', 'horizontal puffing', 'extended test']

    def _scan_folder_for_excel_files(self, folder_path):
        """Recursively scan folder for Excel files."""
        debug_print(f"DEBUG: Starting recursive scan of folder: {folder_path}")
    
        excel_files = []
        excel_extensions = ('.xlsx', '.xls')
    
        try:
            for root, dirs, files in os.walk(folder_path):
                debug_print(f"DEBUG: Scanning directory: {root}")
            
                for file in files:
                    if file.lower().endswith(excel_extensions):
                        full_path = os.path.join(root, file)
                    
                        # Skip temporary Excel files
                        if file.startswith('~$'):
                            debug_print(f"DEBUG: Skipping temporary file: {file}")
                            continue
                    
                        # Check if file is accessible
                        try:
                            if os.access(full_path, os.R_OK):
                                excel_files.append(full_path)
                                debug_print(f"DEBUG: Added Excel file: {file}")
                            else:
                                debug_print(f"DEBUG: Skipping inaccessible file: {file}")
                        except Exception as e:
                            debug_print(f"DEBUG: Error checking file access for {file}: {e}")
                            continue
        
            debug_print(f"DEBUG: Completed folder scan, found {len(excel_files)} Excel files")
            return excel_files
        
        except Exception as e:
            debug_print(f"ERROR: Failed to scan folder: {e}")
            raise

    def _filter_files_by_test_names(self, excel_files, test_names):
        """Filter Excel files by checking if filename contains any test names."""
        debug_print(f"DEBUG: Filtering {len(excel_files)} files against {len(test_names)} test names")
    
        matching_files = []
    
        for file_path in excel_files:
            filename = os.path.basename(file_path).lower()
            debug_print(f"DEBUG: Checking file: {filename}")
        
            # Check if any test name is contained in the filename
            for test_name in test_names:
                if test_name in filename:
                    matching_files.append(file_path)
                    debug_print(f"DEBUG: MATCH - '{filename}' contains '{test_name}'")
                    break
    
        debug_print(f"DEBUG: Filtered to {len(matching_files)} matching files")
        return matching_files

    def _perform_batch_loading(self, file_paths):
        """Perform the actual batch loading with accurate progress tracking."""
        debug_print(f"DEBUG: Starting batch loading of {len(file_paths)} files")
    
        # Show progress dialog
        self.gui.progress_dialog.show_progress_bar("Batch loading files...")
        self.gui.root.update_idletasks()
    
        loaded_files = []
        failed_files = []
        skipped_files = []
        total_files = len(file_paths)
    
        try:
            for index, file_path in enumerate(file_paths, 1):
                try:
                    # Update progress with current file info
                    filename = os.path.basename(file_path)
                    progress = int((index / total_files) * 100)
                    progress_text = f"Loading {index}/{total_files}: {filename[:40]}{'...' if len(filename) > 40 else ''}"
                
                    debug_print(f"DEBUG: {progress_text}")
                
                    # SIMPLIFIED: Only update progress bar, skip label updates
                    try:
                        self.gui.progress_dialog.update_progress_bar(progress)
                    except Exception as e:
                        debug_print(f"DEBUG: Progress update failed: {e}")
                
                    self.gui.root.update_idletasks()
                
                    # Load the file using existing infrastructure
                    debug_print(f"DEBUG: Loading file: {file_path}")
                
                    # Load the file - this stores in database by default
                    self.load_excel_file(file_path, force_reload=True)
                
                    # Store the loaded file data
                    if hasattr(self.gui, 'filtered_sheets') and self.gui.filtered_sheets:
                        file_data = {
                            "file_name": filename,
                            "file_path": file_path,
                            "display_filename": filename,
                            "filtered_sheets": copy.deepcopy(self.gui.filtered_sheets),
                            "source": "batch_folder_load"
                        }
                    
                        # Add to all_filtered_sheets if not already present
                        if not any(f["file_path"] == file_path for f in self.gui.all_filtered_sheets):
                            self.gui.all_filtered_sheets.append(file_data)
                            loaded_files.append(filename)
                            debug_print(f"DEBUG: Successfully loaded and stored: {filename}")
                        else:
                            debug_print(f"DEBUG: File already loaded, skipping: {filename}")
                    else:
                        debug_print(f"ERROR: No filtered_sheets after loading {filename}")
                        failed_files.append(filename)
                
                except ValueError as e:
                    error_msg = str(e).lower()
                    if any(phrase in error_msg for phrase in [
                        "no valid legacy sample data found",
                        "no samples with meaningful data found after filtering",
                        "empty",
                        "no data found"
                    ]):
                        debug_print(f"DEBUG: Gracefully skipping empty file: {filename} - {e}")
                        skipped_files.append(filename)
                        continue
                    else:
                        debug_print(f"ERROR: Failed to load file {file_path}: {e}")
                        failed_files.append(os.path.basename(file_path))
                        continue
                    
                except Exception as e:
                    error_msg = str(e).lower()
                    if any(phrase in error_msg for phrase in [
                        "permission denied",
                        "file not found", 
                        "corrupted",
                        "cannot read",
                        "invalid file format"
                    ]):
                        debug_print(f"DEBUG: Gracefully skipping problematic file: {filename} - {e}")
                        skipped_files.append(filename)
                        continue
                    else:
                        debug_print(f"ERROR: Failed to load file {file_path}: {e}")
                        import traceback
                        traceback.print_exc()
                        failed_files.append(os.path.basename(file_path))
                        continue
        
            # Update UI after successful batch loading
            if loaded_files:
                debug_print("DEBUG: Updating UI after batch loading")
                self.update_file_dropdown()
            
                # Set the last loaded file as active
                if self.gui.all_filtered_sheets:
                    last_file = self.gui.all_filtered_sheets[-1]
                    self.set_active_file(last_file["file_name"])
                    self.update_ui_for_current_file()
        
            # Show completion summary
            self._show_batch_loading_summary(loaded_files, failed_files, skipped_files, total_files)
        
        except Exception as e:
            debug_print(f"ERROR: Batch loading process failed: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            # Always hide progress dialog
            try:
                self.gui.progress_dialog.hide_progress_bar()
            except:
                pass
            self.gui.root.update_idletasks()

    def _show_batch_loading_summary(self, loaded_files, failed_files, skipped_files, total_files):
        """Show summary of batch loading results including skipped files."""
        debug_print(f"DEBUG: Showing batch loading summary - {len(loaded_files)} loaded, {len(failed_files)} failed, {len(skipped_files)} skipped")
    
        success_count = len(loaded_files)
        failure_count = len(failed_files)
        skipped_count = len(skipped_files)
    
        if success_count == 0 and skipped_count == 0:
            messagebox.showerror("Batch Loading Failed", 
                               f"Failed to load any of the {total_files} files.\n\n"
                               f"Check the debug output for specific error details.")
        elif failure_count == 0 and skipped_count == 0:
            # Complete success
            message = f"Batch Loading Complete!\n\n"
            message += f"Successfully loaded {success_count} files:\n\n"
        
            # Show first few filenames
            if len(loaded_files) <= 8:
                message += "\n".join([f"• {name}" for name in loaded_files])
            else:
                message += "\n".join([f"• {name}" for name in loaded_files[:5]])
                message += f"\n• ... and {len(loaded_files) - 5} more files"
        
            message += f"\n\nTotal files now loaded: {len(self.gui.all_filtered_sheets)}"
            message += f"\n\n📁 All files have been stored in the database for future access."
            show_success_message("Batch Loading Success", message, self.gui.root)
        else:
            # Mixed results
            message = f"Batch Loading Completed\n\n"
            message += f"Results Summary:\n"
            message += f"✅ Successfully loaded: {success_count} files\n"
            if skipped_count > 0:
                message += f"⏭️ Skipped (empty): {skipped_count} files\n"
            if failure_count > 0:
                message += f"❌ Failed to load: {failure_count} files\n"
            message += f"\nTotal processed: {total_files} files\n"
        
            if success_count > 0:
                message += f"\n📁 Successfully loaded files have been stored in the database.\n"
                message += "\nLoaded files:\n"
                if len(loaded_files) <= 5:
                    message += "\n".join([f"• {name}" for name in loaded_files])
                else:
                    message += "\n".join([f"• {name}" for name in loaded_files[:3]])
                    message += f"\n• ... and {len(loaded_files) - 3} more"
        
            if skipped_count > 0:
                message += "\n\nSkipped files (empty/no data):\n"
                if len(skipped_files) <= 5:
                    message += "\n".join([f"• {name}" for name in skipped_files])
                else:
                    message += "\n".join([f"• {name}" for name in skipped_files[:3]])
                    message += f"\n• ... and {len(skipped_files) - 3} more"
        
            if failure_count > 0:
                message += "\n\nFailed files:\n"
                if len(failed_files) <= 5:
                    message += "\n".join([f"• {name}" for name in failed_files])
                else:
                    message += "\n".join([f"• {name}" for name in failed_files[:3]])
                    message += f"\n• ... and {len(failed_files) - 3} more"
        
            if failure_count > 0:
                messagebox.showwarning("Batch Loading Complete", message)
            else:
                show_success_message("Batch Loading Complete", message, self.gui.root)